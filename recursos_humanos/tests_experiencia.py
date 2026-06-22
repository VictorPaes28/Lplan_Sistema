from datetime import timedelta
from io import BytesIO

from django.contrib.auth.models import Group, User
from django.test import TestCase
from django.urls import reverse
from django.utils import timezone

from accounts.groups import GRUPOS
from recursos_humanos.models import Colaborador, ContratoAdmissao, PrazoContrato
from recursos_humanos.services.alerts import gerar_alertas
from recursos_humanos.services.prazo_contrato import (
    MARCO_EXPERIENCIA_1,
    MARCO_EXPERIENCIA_2,
    aplicar_data_admissao_oficial,
    calcular_situacao_experiencia,
    formatar_progresso_prazo_teste_clt,
    prazo_teste_clt_deve_exibir,
    sincronizar_prazo_experiencia,
)


class ExperienciaPrazoTests(TestCase):
    def setUp(self):
        self.hoje = timezone.localdate()
        self.admissao = self.hoje - timedelta(days=41)
        self.colab = Colaborador.objects.create(
            nome='Exp Teste',
            cpf='999.888.777-66',
            cargo='Auxiliar',
            status=Colaborador.Status.ATIVO,
            tipo_contrato='CLT',
            data_admissao=self.admissao,
        )
        ContratoAdmissao.objects.create(
            colaborador=self.colab,
            status=ContratoAdmissao.Status.CONCLUIDO,
            data_admissao_oficial=self.admissao,
        )
        self.prazo = PrazoContrato.objects.create(
            colaborador=self.colab,
            tipo=PrazoContrato.Tipo.EXPERIENCIA,
            data_inicio=self.admissao,
            data_fim=self.admissao + timedelta(days=MARCO_EXPERIENCIA_1),
        )

    def test_situacao_primeiro_periodo_atencao(self):
        sit = calcular_situacao_experiencia(self.colab, self.prazo)
        self.assertIsNotNone(sit)
        self.assertEqual(sit.periodo, 1)
        self.assertEqual(sit.prioridade, 'atencao')
        self.assertEqual(sit.proximo_marco, MARCO_EXPERIENCIA_1)

    def test_situacao_segundo_periodo_urgente(self):
        admissao = self.hoje - timedelta(days=86)
        self.colab.data_admissao = admissao
        self.colab.save(update_fields=['data_admissao'])
        self.prazo.data_inicio = admissao
        self.prazo.data_fim = admissao + timedelta(days=MARCO_EXPERIENCIA_2)
        self.prazo.renovacao_numero = 1
        self.prazo.save()
        self.colab.contrato_admissao.data_admissao_oficial = admissao
        self.colab.contrato_admissao.save(update_fields=['data_admissao_oficial'])

        sit = calcular_situacao_experiencia(self.colab, self.prazo)
        self.assertEqual(sit.periodo, 2)
        self.assertEqual(sit.prioridade, 'urgente')

    def test_sincronizar_recalcula_marco_45(self):
        nova_data = self.hoje - timedelta(days=10)
        sincronizar_prazo_experiencia(self.colab, nova_data)
        self.prazo.refresh_from_db()
        self.assertEqual(self.prazo.data_inicio, nova_data)
        self.assertEqual(
            self.prazo.data_fim,
            nova_data + timedelta(days=MARCO_EXPERIENCIA_1),
        )

    def test_calcular_situacao_clt_sem_prazo_registrado(self):
        colab = Colaborador.objects.create(
            nome='CLT Sem Prazo',
            cpf='444.333.222-11',
            cargo='Auxiliar',
            status=Colaborador.Status.ATIVO,
            tipo_contrato='CLT',
            data_admissao=self.admissao,
        )
        ContratoAdmissao.objects.create(
            colaborador=colab,
            status=ContratoAdmissao.Status.CONCLUIDO,
            data_admissao_oficial=self.admissao,
        )
        sit = calcular_situacao_experiencia(colab)
        self.assertIsNotNone(sit)
        self.assertEqual(sit.proximo_marco, MARCO_EXPERIENCIA_1)
        self.assertEqual(sit.periodo, 1)

    def test_formatar_progresso_d45_e_d90(self):
        sit_p1 = calcular_situacao_experiencia(self.colab, self.prazo)
        self.assertEqual(formatar_progresso_prazo_teste_clt(sit_p1), '42/45')

        admissao = self.hoje - timedelta(days=46)
        self.colab.data_admissao = admissao
        self.colab.save(update_fields=['data_admissao'])
        self.prazo.data_inicio = admissao
        self.prazo.save(update_fields=['data_inicio'])
        self.colab.contrato_admissao.data_admissao_oficial = admissao
        self.colab.contrato_admissao.save(update_fields=['data_admissao_oficial'])

        sit_p2 = calcular_situacao_experiencia(self.colab, self.prazo)
        self.assertEqual(formatar_progresso_prazo_teste_clt(sit_p2), '47/90')

    def test_formatar_progresso_prazo_ativo(self):
        from recursos_humanos.services.prazo_contrato import formatar_progresso_prazo_ativo

        prazo = PrazoContrato.objects.create(
            colaborador=self.colab,
            tipo=PrazoContrato.Tipo.ESTAGIO,
            data_inicio=self.hoje - timedelta(days=74),
            data_fim=self.hoje + timedelta(days=125),
        )
        self.assertEqual(formatar_progresso_prazo_ativo(prazo), '75/200')

    def test_resumo_contrato_lista_clt(self):
        from recursos_humanos.services.lista_colaboradores import obter_resumo_contrato_lista

        resumo = obter_resumo_contrato_lista(self.colab)
        self.assertIsNotNone(resumo)
        self.assertEqual(resumo.progresso, '42/45')
        self.assertEqual(resumo.tipo_curto, 'Período')

    def test_resumo_contrato_lista_estagio(self):
        from recursos_humanos.services.lista_colaboradores import (
            enriquecer_lista_colaborador,
            obter_resumo_contrato_lista,
        )

        colab = Colaborador.objects.create(
            nome='Estagiária',
            cpf='222.333.444-55',
            cargo='Estagiária',
            status=Colaborador.Status.ATIVO,
            tipo_contrato='Estágio',
        )
        PrazoContrato.objects.create(
            colaborador=colab,
            tipo=PrazoContrato.Tipo.ESTAGIO,
            data_inicio=self.hoje - timedelta(days=9),
            data_fim=self.hoje + timedelta(days=190),
        )
        resumo = obter_resumo_contrato_lista(colab)
        self.assertIsNotNone(resumo)
        self.assertEqual(resumo.progresso, '10/200')
        self.assertEqual(resumo.tipo_curto, 'Estágio')
        enriquecer_lista_colaborador(colab)
        self.assertEqual(colab.resumo_contrato.progresso, '10/200')

    def test_exibicao_pos_d90_decisao_pendente_permanece(self):
        """Após D90, permanece visível enquanto decisão estiver pendente."""
        admissao_100 = self.hoje - timedelta(days=100)
        self.colab.data_admissao = admissao_100
        self.colab.save(update_fields=['data_admissao'])
        self.prazo.data_inicio = admissao_100
        self.prazo.save(update_fields=['data_inicio'])
        self.colab.contrato_admissao.data_admissao_oficial = admissao_100
        self.colab.contrato_admissao.save(update_fields=['data_admissao_oficial'])

        sit = calcular_situacao_experiencia(self.colab, self.prazo)
        self.assertIsNotNone(sit)
        self.assertTrue(prazo_teste_clt_deve_exibir(sit))
        self.assertEqual(sit.prioridade, 'critico')

    def test_nao_efetiva_automaticamente_apos_d97(self):
        from recursos_humanos.services.prazo_contrato import encerrar_prazos_teste_clt_expirados

        admissao_120 = self.hoje - timedelta(days=120)
        self.colab.data_admissao = admissao_120
        self.colab.save(update_fields=['data_admissao'])
        self.prazo.data_inicio = admissao_120
        self.prazo.save(update_fields=['data_inicio'])
        self.colab.contrato_admissao.data_admissao_oficial = admissao_120
        self.colab.contrato_admissao.save(update_fields=['data_admissao_oficial'])

        encerrar_prazos_teste_clt_expirados()
        self.prazo.refresh_from_db()
        self.assertEqual(self.prazo.status, PrazoContrato.Status.ATIVO)

    def test_acoes_periodo_2_sem_prorrogar(self):
        self.prazo.renovacao_numero = 1
        self.prazo.save(update_fields=['renovacao_numero'])
        acoes = self.prazo.acoes_disponiveis()
        self.assertNotIn('prorrogar', acoes)
        self.assertIn('efetivar', acoes)
        self.assertIn('desligar', acoes)

    def test_preview_d30_gera_alerta(self):
        admissao_30 = self.hoje - timedelta(days=30)
        self.colab.data_admissao = admissao_30
        self.colab.save(update_fields=['data_admissao'])
        self.prazo.data_inicio = admissao_30
        self.prazo.save(update_fields=['data_inicio'])
        self.colab.contrato_admissao.data_admissao_oficial = admissao_30
        self.colab.contrato_admissao.save(update_fields=['data_admissao_oficial'])

        sit = calcular_situacao_experiencia(self.colab, self.prazo)
        self.assertTrue(prazo_teste_clt_deve_exibir(sit))

    def test_alerta_experiencia_gerado(self):
        alertas = gerar_alertas()
        ids = [a.id for a in alertas]
        self.assertIn(f'exp-{self.prazo.pk}', ids)

    def test_aplicar_data_admissao_oficial_cria_prazo_teste_clt(self):
        colab = Colaborador.objects.create(
            nome='Novo Exp',
            cpf='111.222.333-44',
            cargo='Pedreiro',
            status=Colaborador.Status.EM_ADMISSAO,
            etapa_admissao=4,
            tipo_contrato='CLT',
        )
        data = self.hoje - timedelta(days=5)
        ok, _msg = aplicar_data_admissao_oficial(colab, data, None)
        self.assertTrue(ok)
        colab.refresh_from_db()
        self.assertEqual(colab.data_admissao, data)
        contrato = colab.contrato_admissao
        self.assertEqual(contrato.data_admissao_oficial, data)
        prazo = colab.prazos_contrato.filter(
            status=PrazoContrato.Status.ATIVO,
            tipo=PrazoContrato.Tipo.EXPERIENCIA,
        ).first()
        self.assertIsNotNone(prazo)
        self.assertEqual(prazo.data_inicio, data)
        self.assertEqual(prazo.data_fim, data + timedelta(days=MARCO_EXPERIENCIA_1))

    def test_garantir_prazos_teste_clt_ativos_backfill(self):
        from recursos_humanos.services.prazo_contrato import garantir_prazos_teste_clt_ativos

        colab = Colaborador.objects.create(
            nome='Backfill CLT',
            cpf='555.444.333-22',
            cargo='Assistente',
            status=Colaborador.Status.ATIVO,
            tipo_contrato='CLT',
            data_admissao=self.admissao,
        )
        ContratoAdmissao.objects.create(
            colaborador=colab,
            status=ContratoAdmissao.Status.CONCLUIDO,
            data_admissao_oficial=self.admissao,
        )
        self.assertFalse(
            colab.prazos_contrato.filter(tipo=PrazoContrato.Tipo.EXPERIENCIA).exists()
        )
        criados = garantir_prazos_teste_clt_ativos()
        self.assertGreaterEqual(criados, 1)
        self.assertTrue(
            colab.prazos_contrato.filter(
                status=PrazoContrato.Status.ATIVO,
                tipo=PrazoContrato.Tipo.EXPERIENCIA,
            ).exists()
        )


class ExportPrazosContratoTests(TestCase):
    def setUp(self):
        grupo, _ = Group.objects.get_or_create(name=GRUPOS.RECURSOS_HUMANOS)
        self.user = User.objects.create_user('rh_export_prazos', password='test')
        self.user.groups.add(grupo)
        self.hoje = timezone.localdate()
        self.admissao = self.hoje - timedelta(days=41)
        self.colab_exp = Colaborador.objects.create(
            nome='Export Exp',
            cpf='777.666.555-44',
            cargo='Analista',
            status=Colaborador.Status.ATIVO,
            data_admissao=self.admissao,
        )
        ContratoAdmissao.objects.create(
            colaborador=self.colab_exp,
            status=ContratoAdmissao.Status.CONCLUIDO,
            data_admissao_oficial=self.admissao,
        )
        PrazoContrato.objects.create(
            colaborador=self.colab_exp,
            tipo=PrazoContrato.Tipo.EXPERIENCIA,
            data_inicio=self.admissao,
            data_fim=self.admissao + timedelta(days=MARCO_EXPERIENCIA_1),
        )
        self.colab_det = Colaborador.objects.create(
            nome='Export Det',
            cpf='666.555.444-33',
            cargo='Estagiário',
            status=Colaborador.Status.ATIVO,
        )
        PrazoContrato.objects.create(
            colaborador=self.colab_det,
            tipo=PrazoContrato.Tipo.ESTAGIO,
            data_inicio=self.hoje,
            data_fim=self.hoje + timedelta(days=120),
        )
        self.url = reverse('recursos_humanos:alertas_export_prazos')

    def test_export_xlsx_inclui_clt_sem_prazo_previo(self):
        from openpyxl import load_workbook

        colab = Colaborador.objects.create(
            nome='Export Sem Prazo',
            cpf='888.777.666-55',
            cargo='Engenheiro',
            status=Colaborador.Status.ATIVO,
            tipo_contrato='CLT',
            data_admissao=self.hoje - timedelta(days=41),
        )
        ContratoAdmissao.objects.create(
            colaborador=colab,
            status=ContratoAdmissao.Status.CONCLUIDO,
            data_admissao_oficial=colab.data_admissao,
        )

        self.client.force_login(self.user)
        resp = self.client.get(self.url, {'formato': 'xlsx'})
        self.assertEqual(resp.status_code, 200)
        wb = load_workbook(filename=BytesIO(resp.content))
        ws = wb.active
        nomes = [ws.cell(row=r, column=1).value for r in range(9, 27)]
        self.assertIn('Export Sem Prazo', nomes)
        self.assertTrue(
            colab.prazos_contrato.filter(
                status=PrazoContrato.Status.ATIVO,
                tipo=PrazoContrato.Tipo.EXPERIENCIA,
            ).exists()
        )

    def test_export_xlsx_somente_experiencia_clt(self):
        from openpyxl import load_workbook

        self.client.force_login(self.user)
        resp = self.client.get(self.url, {'formato': 'xlsx'})
        self.assertEqual(resp.status_code, 200)
        self.assertIn('spreadsheetml', resp['Content-Type'])
        wb = load_workbook(filename=BytesIO(resp.content))
        ws = wb.active
        self.assertEqual(ws['A8'].value, 'Colaborador')
        nomes = [ws.cell(row=r, column=1).value for r in range(9, 14)]
        self.assertIn('Export Exp', nomes)
        self.assertNotIn('Export Det', nomes)

    def test_export_xlsx_oculta_clt_muito_apos_d90(self):
        from openpyxl import load_workbook

        colab = Colaborador.objects.create(
            nome='Export Antigo',
            cpf='333.222.111-99',
            cargo='Operador',
            status=Colaborador.Status.ATIVO,
            tipo_contrato='CLT',
            data_admissao=self.hoje - timedelta(days=120),
        )
        ContratoAdmissao.objects.create(
            colaborador=colab,
            status=ContratoAdmissao.Status.CONCLUIDO,
            data_admissao_oficial=colab.data_admissao,
        )
        PrazoContrato.objects.create(
            colaborador=colab,
            tipo=PrazoContrato.Tipo.EXPERIENCIA,
            data_inicio=colab.data_admissao,
            data_fim=colab.data_admissao + timedelta(days=MARCO_EXPERIENCIA_2),
        )

        self.client.force_login(self.user)
        resp = self.client.get(self.url, {'formato': 'xlsx'})
        wb = load_workbook(filename=BytesIO(resp.content))
        ws = wb.active
        nomes = [ws.cell(row=r, column=1).value for r in range(9, 27)]
        self.assertIn('Export Antigo', nomes)

    def test_export_pdf_retorna_pdf_valido(self):
        self.client.force_login(self.user)
        resp = self.client.get(self.url, {'formato': 'pdf'})
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp['Content-Type'], 'application/pdf')
        self.assertTrue(resp.content.startswith(b'%PDF'))


class DataAdmissaoOficialTests(TestCase):
    def test_sem_data_etapa4_nao_usa_campo_colaborador(self):
        from recursos_humanos.services.prazo_contrato import obter_data_admissao_oficial

        hoje = timezone.localdate()
        colab = Colaborador.objects.create(
            nome='CLT Legado',
            cpf='999.888.777-66',
            cargo='Servente',
            status=Colaborador.Status.ATIVO,
            tipo_contrato='CLT',
            data_admissao=hoje - timedelta(days=30),
        )
        ContratoAdmissao.objects.create(
            colaborador=colab,
            status=ContratoAdmissao.Status.CONCLUIDO,
        )
        self.assertIsNone(obter_data_admissao_oficial(colab))
        self.assertIsNone(calcular_situacao_experiencia(colab))
