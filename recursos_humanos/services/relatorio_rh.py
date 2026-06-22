"""Cabeçalho, rodapé e identidade visual dos relatórios DP/RH (PDF e Excel)."""

from __future__ import annotations

import os
from xml.sax.saxutils import escape as xml_escape

from django.utils import timezone
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm, mm
from reportlab.platypus import Paragraph, Spacer, Table, TableStyle

EMPRESA_NOME = 'LPLAN ENGENHARIA LTDA.'
MODULO_NOME = 'DP / Recursos Humanos'
COR_PRIMARIA = colors.HexColor('#1A3A5C')
COR_PRIMARIA_HEX = '1A3A5C'
COR_SECUNDARIA = colors.HexColor('#64748B')
COR_SECUNDARIA_HEX = '64748B'
COR_FUNDO_META = colors.HexColor('#F1F5F9')
COR_BORDA = colors.HexColor('#CBD5E1')

URGENTE_CORES_PDF = {
    'Crítico': colors.HexColor('#FEE2E2'),
    'Urgente': colors.HexColor('#FFEDD5'),
    'Atenção': colors.HexColor('#FEF9C3'),
}


def logo_path() -> str | None:
    try:
        from core.utils.pdf_generator import _get_logo_absolute_path

        path = _get_logo_absolute_path()
        if path and os.path.exists(path):
            return path
    except Exception:
        pass
    return None


def _pdf_esc(text) -> str:
    return xml_escape(str(text if text is not None else ''), {'"': '&quot;', "'": '&#39;'})


def _estilos_pdf():
    styles = getSampleStyleSheet()
    return {
        'empresa': ParagraphStyle(
            'RhEmpresa',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=9,
            textColor=COR_PRIMARIA,
            alignment=TA_CENTER,
            spaceAfter=1,
            leading=11,
        ),
        'modulo': ParagraphStyle(
            'RhModulo',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=8,
            textColor=COR_SECUNDARIA,
            alignment=TA_CENTER,
            spaceAfter=4,
            leading=10,
        ),
        'titulo': ParagraphStyle(
            'RhTitulo',
            parent=styles['Heading1'],
            fontName='Helvetica-Bold',
            fontSize=15,
            textColor=COR_PRIMARIA,
            alignment=TA_CENTER,
            spaceAfter=3,
            leading=18,
        ),
        'subtitulo': ParagraphStyle(
            'RhSub',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=8.5,
            textColor=COR_SECUNDARIA,
            alignment=TA_CENTER,
            spaceAfter=0,
            leading=11,
        ),
        'meta': ParagraphStyle(
            'RhMeta',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=8,
            textColor=COR_SECUNDARIA,
            alignment=TA_LEFT,
            leading=10,
        ),
        'th': ParagraphStyle(
            'RhTH',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=8,
            textColor=colors.white,
            alignment=TA_LEFT,
            leading=10,
        ),
        'td': ParagraphStyle(
            'RhTD',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=8,
            textColor=colors.HexColor('#1E293B'),
            alignment=TA_LEFT,
            leading=11,
        ),
        'td_center': ParagraphStyle(
            'RhTDCenter',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=8,
            textColor=colors.HexColor('#1E293B'),
            alignment=TA_CENTER,
            leading=11,
        ),
        'td_right': ParagraphStyle(
            'RhTDRight',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=8,
            textColor=colors.HexColor('#1E293B'),
            alignment=TA_RIGHT,
            leading=11,
        ),
        'empty': ParagraphStyle(
            'RhEmpty',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=9,
            textColor=COR_SECUNDARIA,
            alignment=TA_CENTER,
            spaceBefore=12,
        ),
    }


def montar_cabecalho_pdf(
    story,
    *,
    titulo: str,
    subtitulo: str,
    meta: str,
    largura: float,
) -> None:
    """Cabeçalho vertical: logo centralizado, título e faixa de metadados."""
    from reportlab.platypus import Image as RLImage

    est = _estilos_pdf()

    lp = logo_path()
    if lp:
        try:
            story.append(RLImage(lp, width=5.2 * cm, height=1.25 * cm, hAlign='CENTER'))
            story.append(Spacer(1, 0.3 * cm))
        except Exception:
            pass

    story.append(Paragraph(_pdf_esc(EMPRESA_NOME), est['empresa']))
    story.append(Paragraph(_pdf_esc(MODULO_NOME), est['modulo']))
    story.append(Paragraph(f'<b>{_pdf_esc(titulo.upper())}</b>', est['titulo']))
    story.append(Paragraph(_pdf_esc(subtitulo), est['subtitulo']))
    story.append(Spacer(1, 0.35 * cm))

    faixa_meta = Table(
        [[Paragraph(_pdf_esc(meta), est['meta'])]],
        colWidths=[largura],
        hAlign='LEFT',
    )
    faixa_meta.setStyle(
        TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), COR_FUNDO_META),
            ('BOX', (0, 0), (-1, -1), 0.5, COR_BORDA),
            ('LINEBELOW', (0, 0), (-1, -1), 1.2, COR_PRIMARIA),
            ('LEFTPADDING', (0, 0), (-1, -1), 10),
            ('RIGHTPADDING', (0, 0), (-1, -1), 10),
            ('TOPPADDING', (0, 0), (-1, -1), 7),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 7),
        ])
    )
    story.append(faixa_meta)
    story.append(Spacer(1, 0.45 * cm))


def montar_tabela_pdf(
    story,
    *,
    cabecalhos: list[str],
    linhas: list[list],
    larguras_colunas: list[float],
    col_urgencia: int | None = None,
) -> None:
    """Tabela full-width com Paragraph nas células e destaque na coluna de urgência."""
    est = _estilos_pdf()
    th_row = [Paragraph(f'<b>{_pdf_esc(h)}</b>', est['th']) for h in cabecalhos]
    dados = [th_row]

    for idx_linha, linha in enumerate(linhas):
        row = []
        for col_idx, valor in enumerate(linha):
            texto = _pdf_esc(valor)
            if col_idx == len(linha) - 2 and col_idx != col_urgencia:
                style = est['td']
            elif col_idx == col_urgencia:
                style = est['td_center']
            elif col_idx in (3, 4):
                style = est['td_center']
            elif col_idx == 5:
                style = est['td_right']
            else:
                style = est['td']
            row.append(Paragraph(texto, style))
        dados.append(row)

    tabela = Table(dados, colWidths=larguras_colunas, repeatRows=1, hAlign='LEFT')
    estilos = [
        ('BACKGROUND', (0, 0), (-1, 0), COR_PRIMARIA),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.4, COR_BORDA),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
    ]
    if col_urgencia is not None:
        for row_idx, linha in enumerate(linhas, start=1):
            urgencia = str(linha[col_urgencia]) if col_urgencia < len(linha) else ''
            bg = URGENTE_CORES_PDF.get(urgencia)
            if bg:
                estilos.append(('BACKGROUND', (col_urgencia, row_idx), (col_urgencia, row_idx), bg))
    tabela.setStyle(TableStyle(estilos))
    story.append(tabela)


def larguras_colunas_padrao(largura_total: float, pesos: list[float]) -> list[float]:
    total = sum(pesos)
    return [largura_total * (p / total) for p in pesos]


def rodape_pdf(canvas, doc) -> None:
    """Rodapé institucional em todas as páginas."""
    canvas.saveState()
    try:
        gerado = timezone.localtime().strftime('%d/%m/%Y %H:%M')
        w, _h = doc.pagesize
        canvas.setFont('Helvetica', 7)
        canvas.setFillColorRGB(0.45, 0.45, 0.45)
        canvas.line(20 * mm, 14 * mm, w - 20 * mm, 14 * mm)
        try:
            pagina = canvas.getPageNumber()
        except Exception:
            pagina = 1
        canvas.drawCentredString(
            w / 2,
            9 * mm,
            f'{EMPRESA_NOME}  ·  {MODULO_NOME}  ·  {gerado}  ·  Pág. {pagina}',
        )
    finally:
        canvas.restoreState()


def aplicar_cabecalho_excel(
    ws,
    *,
    titulo: str,
    subtitulo: str,
    num_colunas: int,
    total_registros: int,
) -> tuple[int, object, object]:
    """Cabeçalho vertical alinhado ao PDF."""
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    gerado = timezone.localtime().strftime('%d/%m/%Y %H:%M')
    last_col = get_column_letter(num_colunas)
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    meta_fill = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')
    header_fill = PatternFill(
        start_color=COR_PRIMARIA_HEX,
        end_color=COR_PRIMARIA_HEX,
        fill_type='solid',
    )

    ws.merge_cells(f'A1:{last_col}1')
    ws.merge_cells(f'A2:{last_col}2')
    ws.merge_cells(f'A3:{last_col}3')
    ws.merge_cells(f'A4:{last_col}4')
    ws.merge_cells(f'A5:{last_col}5')

    ws.row_dimensions[1].height = 52
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 22
    ws.row_dimensions[4].height = 28
    ws.row_dimensions[5].height = 18
    ws.row_dimensions[6].height = 8

    lp = logo_path()
    if lp:
        try:
            img = XLImage(lp)
            img.width = 180
            img.height = 42
            ws.add_image(img, 'A1')
        except Exception:
            pass

    c1 = ws['A2']
    c1.value = EMPRESA_NOME
    c1.font = Font(bold=True, size=10, color=COR_PRIMARIA_HEX)
    c1.alignment = align_center

    c2 = ws['A3']
    c2.value = MODULO_NOME
    c2.font = Font(size=9, color=COR_SECUNDARIA_HEX)
    c2.alignment = align_center

    c3 = ws['A4']
    c3.value = titulo.upper()
    c3.font = Font(bold=True, size=14, color=COR_PRIMARIA_HEX)
    c3.alignment = align_center

    c4 = ws['A5']
    c4.value = subtitulo
    c4.font = Font(size=9, color=COR_SECUNDARIA_HEX)
    c4.alignment = align_center

    header_row = 8
    ws.merge_cells(f'A6:{last_col}6')
    c_meta = ws['A6']
    c_meta.value = f'Gerado em {gerado}  ·  {total_registros} registro(s)'
    c_meta.font = Font(size=8, color=COR_SECUNDARIA_HEX)
    c_meta.fill = meta_fill
    c_meta.alignment = align_left
    for col in range(1, num_colunas + 1):
        ws.cell(row=6, column=col).fill = meta_fill

    return header_row, header_fill, align_center
