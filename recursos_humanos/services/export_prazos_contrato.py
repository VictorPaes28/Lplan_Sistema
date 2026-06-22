"""Exportação do período de experiência CLT (Excel e PDF)."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date
from io import BytesIO

from django.utils import timezone
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer

from recursos_humanos.models import Colaborador, PrazoContrato
from recursos_humanos.services.prazo_contrato import (
    NOME_EXPERIENCIA_CLT,
    calcular_situacao_experiencia,
    formatar_progresso_prazo_teste_clt,
    prazo_teste_clt_deve_exibir,
)
from recursos_humanos.services.relatorio_rh import (
    aplicar_cabecalho_excel,
    larguras_colunas_padrao,
    montar_cabecalho_pdf,
    montar_tabela_pdf,
    rodape_pdf,
)

TITULO_RELATORIO = NOME_EXPERIENCIA_CLT
SUBTITULO_RELATORIO = 'Período de experiência CLT — marcos D45 e D90 (decisão obrigatória após D90)'
SUBTITULO_CONSOLIDADO = 'Prazos contratuais ativos — experiência, estágio, PJ e determinado'

CABECALHOS = [
    'Colaborador',
    'Cargo',
    'Tipo',
    'Admissão',
    'Próximo marco',
    'Dias',
    'Progresso / decisão',
    'Urgência',
]

PESOS_COLUNAS = [2.2, 1.5, 1.3, 1.0, 1.6, 0.7, 1.5, 0.9]

PRIORIDADE_LABELS = {
    'normal': 'Normal',
    'atencao': 'Atenção',
    'urgente': 'Urgente',
    'critico': 'Crítico',
}

ORDEM_URGENCIA = {'Crítico': 0, 'Urgente': 1, 'Atenção': 2, 'Normal': 3}


@dataclass(frozen=True)
class LinhaExportPrazoContrato:
    nome: str
    cargo: str
    tipo_prazo: str
    data_inicio: date
    data_fim_label: str
    dias_restantes: int | None
    situacao: str
    urgencia: str

    def as_row(self) -> list:
        dias = self.dias_restantes
        if dias is None:
            dias_txt = '—'
        elif dias < 0:
            dias_txt = f'{abs(dias)} dia(s) de atraso'
        else:
            dias_txt = dias
        return [
            self.nome,
            self.cargo,
            self.tipo_prazo,
            self.data_inicio.strftime('%d/%m/%Y'),
            self.data_fim_label,
            dias_txt,
            self.situacao,
            self.urgencia,
        ]


def _linha_experiencia(prazo: PrazoContrato, colaborador: Colaborador) -> LinhaExportPrazoContrato | None:
    situacao = calcular_situacao_experiencia(colaborador, prazo)
    if not situacao or not prazo_teste_clt_deve_exibir(situacao):
        return None
    decisao = {
        'pendente': 'decisão pendente',
        'efetivado': 'efetivado',
        'desligado': 'desligado',
        'prorrogado': 'prorrogado',
    }.get(situacao.decisao_status, situacao.decisao_status)
    return LinhaExportPrazoContrato(
        nome=colaborador.nome,
        cargo=colaborador.cargo or '',
        tipo_prazo=NOME_EXPERIENCIA_CLT,
        data_inicio=situacao.data_admissao,
        data_fim_label=f'D{situacao.proximo_marco} · {situacao.proximo_marco_data.strftime("%d/%m/%Y")}',
        dias_restantes=situacao.dias_restantes_marco,
        situacao=f'{formatar_progresso_prazo_teste_clt(situacao)} · {decisao}',
        urgencia=PRIORIDADE_LABELS.get(situacao.prioridade, situacao.prioridade.capitalize()),
    )


def coletar_linhas_export_prazos_contrato() -> list[LinhaExportPrazoContrato]:
    from recursos_humanos.services.prazo_contrato import (
        garantir_prazos_teste_clt_ativos,
        sincronizar_datas_prazos_experiencia,
    )

    sincronizar_datas_prazos_experiencia()
    garantir_prazos_teste_clt_ativos()

    prazos = PrazoContrato.objects.filter(
        status=PrazoContrato.Status.ATIVO,
        tipo=PrazoContrato.Tipo.EXPERIENCIA,
        colaborador__status=Colaborador.Status.ATIVO,
        colaborador__tipo_contrato__iexact='CLT',
    ).select_related('colaborador').order_by('colaborador__nome')

    linhas: list[LinhaExportPrazoContrato] = []
    for prazo in prazos:
        linha = _linha_experiencia(prazo, prazo.colaborador)
        if linha:
            linhas.append(linha)

    linhas.sort(
        key=lambda linha: (
            ORDEM_URGENCIA.get(linha.urgencia, 9),
            linha.dias_restantes if linha.dias_restantes is not None else 9999,
            linha.nome.lower(),
        )
    )
    return linhas


def _linha_prazo_generico(prazo: PrazoContrato) -> LinhaExportPrazoContrato:
    colaborador = prazo.colaborador
    dias = prazo.dias_restantes()
    if dias is None:
        dias_txt = None
        fim_label = 'Indeterminado'
    else:
        dias_txt = dias
        fim_label = prazo.data_fim.strftime('%d/%m/%Y') if prazo.data_fim else '—'
    from recursos_humanos.services.prazo_contrato import formatar_progresso_prazo_ativo
    progresso = formatar_progresso_prazo_ativo(prazo) or '—'
    if dias is not None and dias < 0:
        urgencia = 'Crítico'
    elif dias is not None and dias <= 7:
        urgencia = 'Urgente'
    elif dias is not None and dias <= 30:
        urgencia = 'Atenção'
    else:
        urgencia = 'Normal'
    return LinhaExportPrazoContrato(
        nome=colaborador.nome,
        cargo=colaborador.cargo or '',
        tipo_prazo=prazo.get_tipo_display(),
        data_inicio=prazo.data_inicio,
        data_fim_label=fim_label,
        dias_restantes=dias_txt,
        situacao=progresso,
        urgencia=urgencia,
    )


def coletar_linhas_export_contratos_consolidado() -> list[LinhaExportPrazoContrato]:
    linhas = coletar_linhas_export_prazos_contrato()
    prazos = PrazoContrato.objects.filter(
        status=PrazoContrato.Status.ATIVO,
        colaborador__status=Colaborador.Status.ATIVO,
    ).exclude(
        tipo=PrazoContrato.Tipo.EXPERIENCIA,
    ).select_related('colaborador').order_by('colaborador__nome')
    for prazo in prazos:
        linhas.append(_linha_prazo_generico(prazo))
    linhas.sort(
        key=lambda linha: (
            ORDEM_URGENCIA.get(linha.urgencia, 9),
            linha.dias_restantes if linha.dias_restantes is not None else 9999,
            linha.nome.lower(),
        )
    )
    return linhas


def gerar_excel_prazos_contrato(linhas: list[LinhaExportPrazoContrato], *, escopo: str = 'experiencia') -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = 'Prazos contratuais' if escopo == 'todos' else 'Período de experiência'
    titulo = 'Prazos contratuais' if escopo == 'todos' else TITULO_RELATORIO
    subtitulo = SUBTITULO_CONSOLIDADO if escopo == 'todos' else SUBTITULO_RELATORIO

    header_row, header_fill, cell_align = aplicar_cabecalho_excel(
        ws,
        titulo=titulo,
        subtitulo=subtitulo,
        num_colunas=len(CABECALHOS),
        total_registros=len(linhas),
    )
    header_font = Font(bold=True, color='FFFFFF', size=10)

    for col_idx, titulo_col in enumerate(CABECALHOS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=titulo_col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = cell_align

    data_start = header_row + 1
    urgencia_fills = {
        'Crítico': 'FEE2E2',
        'Urgente': 'FFEDD5',
        'Atenção': 'FEF9C3',
    }
    for row_idx, linha in enumerate(linhas, start=data_start):
        for col_idx, valor in enumerate(linha.as_row(), start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.alignment = cell_align
            if col_idx == len(CABECALHOS) and linha.urgencia in urgencia_fills:
                from openpyxl.styles import PatternFill

                cell.fill = PatternFill(
                    start_color=urgencia_fills[linha.urgencia],
                    end_color=urgencia_fills[linha.urgencia],
                    fill_type='solid',
                )

    widths = [30, 22, 18, 12, 20, 8, 26, 12]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.freeze_panes = ws.cell(row=data_start, column=1)
    if linhas:
        ws.auto_filter.ref = (
            f'A{header_row}:{get_column_letter(len(CABECALHOS))}'
            f'{data_start + len(linhas) - 1}'
        )

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def gerar_pdf_prazos_contrato(
    linhas: list[LinhaExportPrazoContrato],
    *,
    escopo: str = 'experiencia',
) -> bytes:
    from recursos_humanos.services.relatorio_rh import _estilos_pdf

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        topMargin=1.0 * cm,
        bottomMargin=1.5 * cm,
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
    )

    gerado = timezone.localtime().strftime('%d/%m/%Y %H:%M')
    meta = f'Gerado em {gerado}  ·  {len(linhas)} registro(s)'
    titulo = 'Prazos contratuais' if escopo == 'todos' else TITULO_RELATORIO
    subtitulo = SUBTITULO_CONSOLIDADO if escopo == 'todos' else SUBTITULO_RELATORIO
    story = []
    montar_cabecalho_pdf(
        story,
        titulo=titulo,
        subtitulo=subtitulo,
        meta=meta,
        largura=doc.width,
    )

    if not linhas:
        est = _estilos_pdf()
        vazio = (
            'Nenhum prazo contratual ativo no momento.'
            if escopo == 'todos'
            else 'Nenhum colaborador CLT em período de experiência no momento.'
        )
        story.append(Paragraph(vazio, est['empty']))
    else:
        larguras = larguras_colunas_padrao(doc.width, PESOS_COLUNAS)
        montar_tabela_pdf(
            story,
            cabecalhos=CABECALHOS,
            linhas=[linha.as_row() for linha in linhas],
            larguras_colunas=larguras,
            col_urgencia=len(CABECALHOS) - 1,
        )

    story.append(Spacer(1, 0.4 * cm))
    doc.build(story, onFirstPage=rodape_pdf, onLaterPages=rodape_pdf)
    return buffer.getvalue()
