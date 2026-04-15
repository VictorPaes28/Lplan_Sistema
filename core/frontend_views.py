"""
Views para frontend do LPlan - Templates Django com HTMX/Alpine.js
"""
from functools import wraps
import re
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import Group, User
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse, Http404
from django.views.decorators.http import require_http_methods
from django.db.models import Q, Count, Avg, Sum, OuterRef, Subquery
from django.db import IntegrityError
from django.utils import timezone
from django.utils.http import url_has_allowed_host_and_scheme
from datetime import datetime, timedelta
from .models import (
    Project,
    ProjectMember,
    ProjectOwner,
    ProjectDiaryApprover,
    SupportTicket,
    SupportTicketMessage,
    SupportTicketAttachment,
    Notification,
    ConstructionDiary,
    DiaryNoReportDay,
    DiaryCorrectionRequestLog,
    DiaryApprovalHistory,
    DiaryComment,
    DiaryImage,
    DailyWorkLog,
    Labor,
    Equipment,
    DiaryStatus,
    Activity,
    ActivityStatus,
)
from django.core.exceptions import PermissionDenied, ValidationError
from django.core.paginator import Paginator
import logging
from accounts.groups import GRUPOS
from accounts.models import UserSignupRequest
from accounts.signup_services import (
    create_signup_request,
    is_allowed_signup_email,
    notify_signup_request_created,
    get_allowed_signup_domains,
)
from .user_messages import flash_message

logger = logging.getLogger(__name__)
# PDFGenerator será importado apenas quando necessário (lazy import)
PDFGenerator = None
REPORTLAB_AVAILABLE = False

SUPPORT_CATEGORY_CHOICES = (
    "Diário de Obra (criação/edição)",
    "RDO - Aprovação de gestor",
    "Relatórios e PDF/Excel",
    "Fotos, vídeos e anexos",
    "GestControll - Pedidos",
    "GestControll - Aprovação/Reprovação",
    "Mapa de Controle / Engenharia",
    "Painel do sistema (obras/usuários)",
    "Acesso e permissões",
    "Notificações e e-mails",
    "Lentidão/Performance",
    "Dúvida de uso do sistema",
    "Outro (descrever no texto)",
)

# Mapeamento obra → contratante para autopreencher o formulário do diário.
# Chave: substring normalizada (lower) do nome ou código da obra.
OBRA_CONTRATANTE_MAP = {
    'entreaguas': 'Incorporadora Adamo',
    'okena': 'JP Empreendimentos',
    'marghot': 'Antonina Hotéis',
    'sunrise': 'Rpontes',
}


@require_http_methods(['GET', 'POST'])
def signup_request_view(request):
    """Tela pública para solicitar cadastro com aprovação manual."""
    groups = []
    for group_name in GRUPOS.TODOS:
        group = Group.objects.filter(name=group_name).first()
        if group:
            groups.append(group)

    projects = Project.objects.filter(is_active=True).order_by('name')
    allowed_domains = get_allowed_signup_domains()

    selected_groups = []
    selected_projects = []
    if request.method == 'POST':
        first_name = (request.POST.get('first_name') or '').strip()
        last_name = (request.POST.get('last_name') or '').strip()
        full_name = f'{first_name} {last_name}'.strip()
        email = (request.POST.get('email') or '').strip().lower()
        username_suggestion = (request.POST.get('username_suggestion') or '').strip()
        notes = (request.POST.get('notes') or '').strip()
        selected_projects = request.POST.getlist('projects')

        if not first_name or not last_name or not email:
            messages.error(request, 'Nome, sobrenome e e-mail são obrigatórios.')
        elif not is_allowed_signup_email(email):
            domains = ', '.join(allowed_domains) if allowed_domains else 'domínios permitidos'
            messages.error(request, f'Este e-mail não é permitido para cadastro. Use um domínio autorizado ({domains}).')
        elif UserSignupRequest.objects.filter(email__iexact=email, status=UserSignupRequest.STATUS_PENDENTE).exists():
            messages.info(request, 'Já existe uma solicitação pendente para este e-mail. Aguarde a análise.')
        else:
            create_signup_request(
                full_name=full_name,
                email=email,
                username_suggestion=username_suggestion,
                notes=notes,
                requested_groups=[],
                requested_project_ids=selected_projects,
                origem=UserSignupRequest.ORIGEM_AUTO,
                requested_by=request.user if request.user.is_authenticated else None,
            )
            signup_req = UserSignupRequest.objects.filter(email=email).order_by('-created_at').first()
            if signup_req:
                notify_signup_request_created(signup_req)
            messages.success(request, 'Solicitação enviada com sucesso! Assim que aprovada, você receberá os dados de acesso por e-mail.')
            return redirect('signup-request')

    return render(
        request,
        'core/signup_request_form.html',
        {
            'groups': groups,
            'projects': projects,
            'allowed_domains': allowed_domains,
            'selected_groups': selected_groups,
            'selected_projects': [str(p) for p in selected_projects],
        },
    )


def get_contractante_for_project(project):
    """Retorna o nome do contratante: para obras mapeadas usa sempre o mapeamento; senão usa project.client_name."""
    if not project:
        return ''
    name = (getattr(project, 'name', None) or '').strip().lower()
    code = (getattr(project, 'code', None) or '').strip().lower()
    for key, contratante in OBRA_CONTRATANTE_MAP.items():
        if key in name or key in code:
            return contratante
    if getattr(project, 'client_name', None) and (project.client_name or '').strip():
        return (project.client_name or '').strip()
    return ''


def _is_truthy_flag(value):
    """Interpreta flags vindas de forms HTML de forma resiliente."""
    normalized = str(value or '').strip().lower()
    return normalized in {'1', 'true', 'on', 'yes'}


def _decode_js_escaped_text(value):
    """
    Decodifica sequências JS literais (ex.: \\u0027) para exibição humana.
    Mantém o texto original quando não há escapes.
    """
    text = str(value or '').strip()
    if not text:
        return ''
    text = re.sub(r'\\u([0-9a-fA-F]{4})', lambda m: chr(int(m.group(1), 16)), text)
    return text.replace("\\'", "'")


def _safe_positive_int(value, default=1, minimum=1):
    """Converte para inteiro positivo sem quebrar fluxo."""
    try:
        number = int(value)
    except (TypeError, ValueError):
        return max(minimum, int(default))
    return max(minimum, number)


def _normalize_equipment_name(value):
    """Normaliza nome de equipamento recebido do front/API."""
    return _decode_js_escaped_text(value).strip()


def _normalized_name_key(value):
    """Chave de comparação para nome de equipamento (case/espacos)."""
    return ' '.join(_normalize_equipment_name(value).lower().split())


def _generate_unique_equipment_code(base_name):
    """
    Gera código único para equipamento custom, evitando colisão em Equipment.code.
    """
    normalized = _normalize_equipment_name(base_name)
    token = re.sub(r'[^A-Za-z0-9]+', '-', normalized.upper()).strip('-')
    token = token[:26] if token else 'CUSTOM'
    base_code = f"EQ-CUSTOM-{token}"
    candidate = base_code
    suffix = 1
    while Equipment.objects.filter(code=candidate).exists():
        suffix += 1
        candidate = f"{base_code}-{suffix}"
    return candidate


def _resolve_equipment_from_payload_item(equipment_item):
    """
    Resolve/Cria Equipment a partir do item do payload de equipamentos.
    Prioriza equipment_id válido; fallback por nome normalizado.
    """
    equipment = None
    payload_name = _normalize_equipment_name(equipment_item.get('name', ''))
    payload_name_key = _normalized_name_key(payload_name)

    # Payload novo do formulário: id de StandardEquipment (não é Equipment.pk).
    standard_equipment_id = equipment_item.get('standard_equipment_id')
    if standard_equipment_id:
        try:
            from .models import StandardEquipment
            std = StandardEquipment.objects.filter(pk=int(standard_equipment_id)).only('name').first()
            if std and getattr(std, 'name', None):
                payload_name = _normalize_equipment_name(std.name)
                payload_name_key = _normalized_name_key(payload_name)
        except (ValueError, TypeError):
            pass

    equipment_id = equipment_item.get('equipment_id')
    if equipment_id:
        try:
            by_id = Equipment.objects.filter(pk=int(equipment_id)).first()
            if by_id:
                by_id_name_key = _normalized_name_key(getattr(by_id, 'name', ''))
                # Evita mapear equipamento errado quando o front envia id de outra tabela.
                if not payload_name_key or by_id_name_key == payload_name_key:
                    equipment = by_id
        except (ValueError, TypeError):
            equipment = None
    if equipment is not None:
        return equipment, payload_name
    if not payload_name:
        return None, ''

    by_name = Equipment.objects.filter(name__iexact=payload_name).order_by('id').first()
    if by_name:
        return by_name, payload_name

    code = _generate_unique_equipment_code(payload_name)
    equipment = Equipment.objects.create(
        code=code,
        name=payload_name,
        equipment_type=payload_name,
        is_active=True,
    )
    return equipment, payload_name


def _build_diary_equipment_list(diary):
    """
    Lista de equipamentos agregada por equipamento (mesma regra do PDF),
    com quantidade total por diário.
    """
    equipment_list = []
    try:
        from core.utils.diary_equipment import aggregate_equipment_for_diary
        rows, _total = aggregate_equipment_for_diary(diary)
        for row in rows:
            eq = row.get('equipment')
            if not eq:
                continue
            equipment_list.append({
                'name': _decode_js_escaped_text(getattr(eq, 'name', '') or ''),
                'code': _decode_js_escaped_text(getattr(eq, 'code', '') or ''),
                'quantity': int(row.get('quantity') or 0),
            })
    except Exception:
        # Fallback defensivo para não quebrar a tela caso a agregação falhe.
        seen = {}
        for work_log in diary.work_logs.prefetch_related('resources_equipment').all():
            for equipment in work_log.resources_equipment.all():
                eid = getattr(equipment, 'pk', None)
                if eid is None:
                    continue
                if eid not in seen:
                    seen[eid] = {
                        'name': _decode_js_escaped_text(getattr(equipment, 'name', '') or ''),
                        'code': _decode_js_escaped_text(getattr(equipment, 'code', '') or ''),
                        'quantity': 0,
                    }
                seen[eid]['quantity'] += 1
        equipment_list = list(seen.values())
    return equipment_list


def login_view(request):
    """View de login."""
    if request.user.is_authenticated:
        # Sempre redireciona para seleção de sistema (não redireciona automaticamente)
        return redirect('select-system')

    next_url = (request.POST.get('next') or request.GET.get('next') or '').strip()

    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)
            # Limpa obra selecionada anterior (se houver) para forçar nova seleção
            if 'selected_project_id' in request.session:
                del request.session['selected_project_id']
            if 'selected_project_name' in request.session:
                del request.session['selected_project_name']
            if 'selected_project_code' in request.session:
                del request.session['selected_project_code']
            if next_url and url_has_allowed_host_and_scheme(
                url=next_url,
                allowed_hosts={request.get_host()},
                require_https=request.is_secure(),
            ):
                return redirect(next_url)
            return redirect('select-system')
        else:
            return render(request, 'core/login.html', {'error': 'Credenciais inválidas', 'next': next_url})

    return render(request, 'core/login.html', {'next': next_url})


@login_required
def logout_view(request):
    """View de logout."""
    # Limpa a obra selecionada (todos os campos para não vazar dados entre sessões)
    for key in ('selected_project_id', 'selected_project_name', 'selected_project_code'):
        request.session.pop(key, None)
    logout(request)
    return redirect('login')


@login_required
def select_system_view(request):
    """View para seleção de sistema após login."""
    user = request.user
    user_groups = set(user.groups.values_list('name', flat=True))
    
    # Determinar acesso por sistema baseado nos grupos
    has_diario = user.is_superuser or user.is_staff or GRUPOS.GERENTES in user_groups
    has_gestao = user.is_superuser or user.is_staff or bool(
        user_groups & {GRUPOS.ADMINISTRADOR, GRUPOS.RESPONSAVEL_EMPRESA, GRUPOS.APROVADOR, GRUPOS.SOLICITANTE}
    )
    has_mapa = user.is_superuser or user.is_staff or GRUPOS.ENGENHARIA in user_groups
    # BI da Obra: mesma base de obras (projeto vinculado); visível para quem usa Diário ou Mapa
    has_bi_obra = user.is_superuser or user.is_staff or has_diario or has_mapa
    from accounts.painel_sistema_access import user_is_painel_sistema_admin

    has_central = user_is_painel_sistema_admin(user)
    # Dono da obra: se só tem acesso ao portal cliente, redireciona direto
    if not (has_diario or has_gestao or has_mapa or has_central) and _is_work_owner(user):
        return redirect('client-diary-list')
    support_projects = list(_get_support_projects_for_user(user))
    context = {
        'has_diario': has_diario,
        'has_gestao': has_gestao,
        'has_mapa': has_mapa,
        'has_bi_obra': has_bi_obra,
        'has_admin': user.is_superuser or user.is_staff,
        'has_central': has_central,
        'can_manage_support_tickets': user.is_superuser or user.is_staff,
        'support_projects': support_projects,
        'support_auto_project': support_projects[0] if len(support_projects) == 1 else None,
        'support_categories': SUPPORT_CATEGORY_CHOICES,
    }
    return render(request, 'core/select_system.html', context)


def _can_manage_support_tickets(user):
    return bool(user and user.is_authenticated and (user.is_superuser or user.is_staff))


def _support_sla_windows(severity):
    if severity == SupportTicket.Severity.BLOCKER:
        return timedelta(minutes=30), timedelta(hours=4)
    if severity == SupportTicket.Severity.IMPORTANT:
        return timedelta(hours=2), timedelta(hours=24)
    if severity == SupportTicket.Severity.LOW:
        return timedelta(hours=24), timedelta(hours=120)
    return timedelta(hours=8), timedelta(hours=72)


def _notify_support_staff(title, message, exclude_user_id=None):
    staff_users = User.objects.filter(
        is_active=True,
    ).filter(
        Q(is_superuser=True) | Q(is_staff=True)
    )
    if exclude_user_id:
        staff_users = staff_users.exclude(id=exclude_user_id)
    notifications = [
        Notification(
            user=u,
            notification_type='system',
            title=title[:255],
            message=message,
        )
        for u in staff_users
    ]
    if notifications:
        Notification.objects.bulk_create(notifications)


def _notify_user(user, title, message):
    if not user:
        return
    Notification.objects.create(
        user=user,
        notification_type='system',
        title=title[:255],
        message=message,
    )


def _support_can_transition(current_status, target_status):
    transitions = {
        SupportTicket.Status.OPEN: {
            SupportTicket.Status.TRIAGE,
            SupportTicket.Status.IN_PROGRESS,
            SupportTicket.Status.WAITING_USER,
            SupportTicket.Status.WAITING_DEPLOY,
            SupportTicket.Status.RESOLVED,
            SupportTicket.Status.CLOSED,
        },
        SupportTicket.Status.TRIAGE: {
            SupportTicket.Status.IN_PROGRESS,
            SupportTicket.Status.WAITING_USER,
            SupportTicket.Status.WAITING_DEPLOY,
            SupportTicket.Status.RESOLVED,
            SupportTicket.Status.CLOSED,
        },
        SupportTicket.Status.IN_PROGRESS: {
            SupportTicket.Status.WAITING_USER,
            SupportTicket.Status.WAITING_DEPLOY,
            SupportTicket.Status.RESOLVED,
            SupportTicket.Status.CLOSED,
        },
        SupportTicket.Status.WAITING_USER: {
            SupportTicket.Status.IN_PROGRESS,
            SupportTicket.Status.RESOLVED,
            SupportTicket.Status.CLOSED,
        },
        SupportTicket.Status.WAITING_DEPLOY: {
            SupportTicket.Status.IN_PROGRESS,
            SupportTicket.Status.RESOLVED,
            SupportTicket.Status.CLOSED,
        },
        SupportTicket.Status.RESOLVED: {
            SupportTicket.Status.CLOSED,
            SupportTicket.Status.REOPENED,
        },
        SupportTicket.Status.CLOSED: {
            SupportTicket.Status.REOPENED,
        },
        SupportTicket.Status.REOPENED: {
            SupportTicket.Status.TRIAGE,
            SupportTicket.Status.IN_PROGRESS,
            SupportTicket.Status.WAITING_USER,
            SupportTicket.Status.WAITING_DEPLOY,
            SupportTicket.Status.RESOLVED,
            SupportTicket.Status.CLOSED,
        },
    }
    return target_status == current_status or target_status in transitions.get(current_status, set())


def _support_attention_tag(ticket, viewer_is_manager=False):
    """
    Tag operacional para leitura rápida do histórico/filas.
    Não substitui o status oficial do chamado.
    """
    if ticket.status == SupportTicket.Status.CLOSED:
        return ('Fechado', 'closed')
    if ticket.status == SupportTicket.Status.RESOLVED:
        return ('Resolvido', 'resolved')
    if ticket.status == SupportTicket.Status.REOPENED:
        return ('Reaberto', 'reopened')
    if ticket.status == SupportTicket.Status.WAITING_USER:
        return ('Aguardando solicitante', 'waiting-requester')
    if ticket.status == SupportTicket.Status.WAITING_DEPLOY:
        return ('Aguardando deploy', 'waiting-requester')

    public_count = getattr(ticket, 'public_message_count', None) or 0
    last_author_id = getattr(ticket, 'last_public_author_id', None)
    last_is_staff = bool(getattr(ticket, 'last_public_author_is_staff', False))
    last_is_superuser = bool(getattr(ticket, 'last_public_author_is_superuser', False))
    last_from_team = bool(last_is_staff or last_is_superuser)

    if ticket.status in (SupportTicket.Status.TRIAGE, SupportTicket.Status.IN_PROGRESS):
        if public_count <= 1:
            return ('Em atendimento', 'in-progress')
        if viewer_is_manager:
            return ('Aguardando solicitante', 'responded') if last_from_team else ('Aguardando equipe', 'waiting-team')
        return ('Respondido', 'responded') if last_from_team else ('Aguardando equipe', 'waiting-team')

    if ticket.status == SupportTicket.Status.OPEN:
        if public_count <= 1:
            return ('Novo', 'new')
        if viewer_is_manager:
            return ('Aguardando solicitante', 'responded') if last_from_team else ('Aguardando equipe', 'waiting-team')
        if last_from_team:
            return ('Respondido', 'responded')
        if last_author_id:
            return ('Aguardando equipe', 'waiting-team')
        return ('Novo', 'new')
    if viewer_is_manager:
        return ('Aguardando solicitante', 'responded') if last_from_team else ('Aguardando equipe', 'waiting-team')
    if last_from_team:
        return ('Respondido', 'responded')
    if last_author_id:
        return ('Aguardando equipe', 'waiting-team')
    return ('Em atendimento', 'in-progress')


@login_required
@require_http_methods(["POST"])
def support_ticket_create_view(request):
    """Cria chamado de suporte interno pelo modal da tela inicial."""
    category = (request.POST.get('category') or '').strip()[:80]
    severity = (request.POST.get('severity') or '').strip()
    title = (request.POST.get('title') or '').strip()[:120]
    description = (request.POST.get('description') or '').strip()
    screen_path = (request.POST.get('screen_path') or '').strip()[:255]
    project_id_raw = (request.POST.get('project_id') or '').strip()
    browser_info = (request.META.get('HTTP_USER_AGENT') or '')[:255]
    accessible_projects_qs = _get_support_projects_for_user(request.user)
    accessible_count = accessible_projects_qs.count()
    related_project = None
    if project_id_raw:
        try:
            project_id = int(project_id_raw)
        except ValueError:
            return JsonResponse({'success': False, 'error': 'Obra inválida.'}, status=400)
        related_project = accessible_projects_qs.filter(pk=project_id).first()
        if related_project is None:
            return JsonResponse({'success': False, 'error': 'Você não tem acesso a esta obra.'}, status=403)
    elif accessible_count == 1:
        related_project = accessible_projects_qs.first()
    elif accessible_count > 1:
        return JsonResponse({'success': False, 'error': 'Selecione a obra relacionada ao chamado.'}, status=400)
    else:
        related_project = None

    if not category or not severity or not title or not description:
        return JsonResponse(
            {'success': False, 'error': 'Preencha categoria, impacto, título e descrição.'},
            status=400,
        )
    if category not in SUPPORT_CATEGORY_CHOICES:
        return JsonResponse({'success': False, 'error': 'Categoria inválida.'}, status=400)

    valid_severities = {choice[0] for choice in SupportTicket.Severity.choices}
    if severity not in valid_severities:
        return JsonResponse({'success': False, 'error': 'Severidade inválida.'}, status=400)

    now = timezone.now()
    first_window, resolution_window = _support_sla_windows(severity)
    ticket = SupportTicket.objects.create(
        created_by=request.user,
        category=category,
        severity=severity,
        title=title,
        description=description,
        related_project=related_project,
        screen_path=screen_path,
        browser_info=browser_info,
        first_response_due_at=now + first_window,
        resolution_due_at=now + resolution_window,
    )
    SupportTicketMessage.objects.create(
        ticket=ticket,
        author=request.user,
        message=description,
        is_internal_note=False,
    )

    for f in request.FILES.getlist('attachments'):
        SupportTicketAttachment.objects.create(
            ticket=ticket,
            uploaded_by=request.user,
            file=f,
            original_name=(getattr(f, 'name', '') or '')[:255],
        )
    _notify_support_staff(
        title=f"Novo chamado #{ticket.pk}",
        message=f"{request.user.get_full_name() or request.user.username} abriu: {ticket.title}",
        exclude_user_id=request.user.id,
    )

    return JsonResponse(
        {
            'success': True,
            'ticket_id': ticket.pk,
            'detail_url': f"/support/tickets/{ticket.pk}/",
            'message': 'Chamado criado com sucesso.',
        }
    )


@login_required
@require_http_methods(["GET"])
def support_ticket_list_view(request):
    """Lista chamados do usuário logado."""
    last_public_author_subq = SupportTicketMessage.objects.filter(
        ticket=OuterRef('pk'),
        is_internal_note=False,
    ).order_by('-created_at').values('author_id')[:1]
    last_public_author_is_staff_subq = SupportTicketMessage.objects.filter(
        ticket=OuterRef('pk'),
        is_internal_note=False,
    ).order_by('-created_at').values('author__is_staff')[:1]
    last_public_author_is_superuser_subq = SupportTicketMessage.objects.filter(
        ticket=OuterRef('pk'),
        is_internal_note=False,
    ).order_by('-created_at').values('author__is_superuser')[:1]
    public_count_subq = SupportTicketMessage.objects.filter(
        ticket=OuterRef('pk'),
        is_internal_note=False,
    ).values('ticket').annotate(c=Count('id')).values('c')[:1]
    tickets = list(
        SupportTicket.objects.filter(created_by=request.user)
        .select_related('assigned_to', 'related_project')
        .annotate(
            last_public_author_id=Subquery(last_public_author_subq),
            last_public_author_is_staff=Subquery(last_public_author_is_staff_subq),
            last_public_author_is_superuser=Subquery(last_public_author_is_superuser_subq),
            public_message_count=Subquery(public_count_subq),
        )
        .order_by('-created_at')
    )
    for t in tickets:
        t.attention_label, t.attention_tone = _support_attention_tag(t, viewer_is_manager=False)
    return render(
        request,
        'core/support_ticket_list.html',
        {
            'tickets': tickets,
            'can_manage_support_tickets': _can_manage_support_tickets(request.user),
            'now': timezone.now(),
        },
    )


@login_required
@require_http_methods(["GET"])
def support_ticket_detail_view(request, pk):
    """Detalhe de chamado com conversa e anexos."""
    ticket = get_object_or_404(
        SupportTicket.objects.select_related('created_by', 'assigned_to', 'related_project'),
        pk=pk,
    )
    can_manage = _can_manage_support_tickets(request.user)
    if not can_manage and ticket.created_by_id != request.user.id:
        raise Http404()

    messages_qs = ticket.messages.select_related('author').order_by('created_at')
    if not can_manage:
        messages_qs = messages_qs.filter(is_internal_note=False)
    public_messages_qs = ticket.messages.filter(is_internal_note=False).order_by('-created_at')
    ticket.public_message_count = public_messages_qs.count()
    last_public_message = public_messages_qs.select_related('author').first()
    ticket.last_public_author_id = getattr(last_public_message, 'author_id', None)
    ticket.last_public_author_is_staff = bool(
        getattr(getattr(last_public_message, 'author', None), 'is_staff', False)
    )
    ticket.last_public_author_is_superuser = bool(
        getattr(getattr(last_public_message, 'author', None), 'is_superuser', False)
    )
    attachments = ticket.attachments.select_related('uploaded_by').order_by('-uploaded_at')
    admins = []
    if can_manage:
        admins = User.objects.filter(
            is_active=True,
        ).filter(
            Q(is_superuser=True) | Q(is_staff=True)
        ).order_by('first_name', 'username')

    return render(
        request,
        'core/support_ticket_detail.html',
        {
            'ticket': ticket,
            'ticket_messages': messages_qs,
            'attachments': attachments,
            'can_manage_support_tickets': can_manage,
            'admin_users': admins,
            'status_choices': SupportTicket.Status.choices,
            'can_reopen': ticket.can_be_reopened_by_user(request.user),
            'attention_tag': _support_attention_tag(
                ticket,
                viewer_is_manager=can_manage,
            ),
        },
    )


@login_required
@require_http_methods(["POST"])
def support_ticket_reply_view(request, pk):
    """Responde chamado e permite gestão por admin/superuser."""
    ticket = get_object_or_404(SupportTicket, pk=pk)
    can_manage = _can_manage_support_tickets(request.user)
    if not can_manage and ticket.created_by_id != request.user.id:
        raise Http404()
    if (not can_manage) and ticket.status in (SupportTicket.Status.RESOLVED, SupportTicket.Status.CLOSED):
        messages.error(request, 'Chamado encerrado. Use a opção de reabrir chamado quando disponível.')
        return redirect('support-ticket-detail', pk=ticket.pk)

    message_text = (request.POST.get('message') or '').strip()
    is_internal_note = bool(request.POST.get('is_internal_note')) and can_manage

    if message_text:
        SupportTicketMessage.objects.create(
            ticket=ticket,
            author=request.user,
            message=message_text,
            is_internal_note=is_internal_note,
        )
        if can_manage and not is_internal_note and not ticket.first_response_at:
            ticket.first_response_at = timezone.now()
    fields_to_update = set()
    if can_manage and not is_internal_note and message_text and ticket.first_response_at:
        fields_to_update.add('first_response_at')

    for f in request.FILES.getlist('attachments'):
        SupportTicketAttachment.objects.create(
            ticket=ticket,
            uploaded_by=request.user,
            file=f,
            original_name=(getattr(f, 'name', '') or '')[:255],
        )

    if can_manage:
        old_status = ticket.status
        old_assigned_to_id = ticket.assigned_to_id
        status = (request.POST.get('status') or '').strip()
        if not status and message_text and not is_internal_note and old_status in (
            SupportTicket.Status.RESOLVED,
            SupportTicket.Status.CLOSED,
        ):
            status = SupportTicket.Status.REOPENED
        if not status and message_text and not is_internal_note:
            # Resposta da equipe sem escolha manual de status: deixa claro que agora aguardamos o solicitante.
            if old_status in (
                SupportTicket.Status.OPEN,
                SupportTicket.Status.TRIAGE,
                SupportTicket.Status.IN_PROGRESS,
                SupportTicket.Status.REOPENED,
            ):
                status = SupportTicket.Status.WAITING_USER
        valid_statuses = {choice[0] for choice in SupportTicket.Status.choices}
        if status in valid_statuses:
            if not _support_can_transition(old_status, status):
                messages.error(request, 'Transição de status inválida para este chamado.')
                return redirect('support-ticket-detail', pk=ticket.pk)
            ticket.status = status
            if status in (SupportTicket.Status.RESOLVED, SupportTicket.Status.CLOSED) and not ticket.resolved_at:
                ticket.resolved_at = timezone.now()
            elif status not in (SupportTicket.Status.RESOLVED, SupportTicket.Status.CLOSED):
                ticket.resolved_at = None
            fields_to_update.update({'status', 'resolved_at'})
        if 'assigned_to' in request.POST:
            assigned_to = (request.POST.get('assigned_to') or '').strip()
            if assigned_to:
                try:
                    ticket.assigned_to_id = int(assigned_to)
                except ValueError:
                    pass
            else:
                ticket.assigned_to = None
            fields_to_update.add('assigned_to')
        if fields_to_update:
            ticket.save(update_fields=list(fields_to_update | {'updated_at'}))

        if ticket.assigned_to_id and ticket.assigned_to_id != old_assigned_to_id:
            _notify_user(
                ticket.assigned_to,
                f"Chamado #{ticket.pk} atribuído a você",
                f"Ticket: {ticket.title}",
            )
        if message_text and not is_internal_note:
            _notify_user(
                ticket.created_by,
                f"Nova resposta no chamado #{ticket.pk}",
                f"A equipe respondeu: {ticket.title}",
            )
        if old_status != ticket.status:
            _notify_user(
                ticket.created_by,
                f"Status do chamado #{ticket.pk} atualizado",
                f"Novo status: {ticket.get_status_display()}",
            )
    else:
        if message_text and not is_internal_note:
            if ticket.status in (
                SupportTicket.Status.OPEN,
                SupportTicket.Status.TRIAGE,
                SupportTicket.Status.WAITING_USER,
                SupportTicket.Status.REOPENED,
            ):
                ticket.status = SupportTicket.Status.IN_PROGRESS
                ticket.save(update_fields=['status', 'updated_at'])
            if ticket.assigned_to_id:
                _notify_user(
                    ticket.assigned_to,
                    f"Solicitante respondeu chamado #{ticket.pk}",
                    f"{request.user.get_full_name() or request.user.username} enviou nova mensagem.",
                )
            else:
                _notify_support_staff(
                    title=f"Atualização no chamado #{ticket.pk}",
                    message=f"Solicitante enviou nova mensagem em: {ticket.title}",
                    exclude_user_id=request.user.id,
                )

    messages.success(request, 'Chamado atualizado com sucesso.')
    return redirect('support-ticket-detail', pk=ticket.pk)


@login_required
@require_http_methods(["POST"])
def support_ticket_reopen_view(request, pk):
    """Permite reabertura pelo solicitante em até 7 dias após resolução/fechamento."""
    ticket = get_object_or_404(SupportTicket, pk=pk)
    if ticket.created_by_id != request.user.id:
        raise Http404()
    if not ticket.can_be_reopened_by_user(request.user):
        messages.error(request, 'Prazo de reabertura expirado ou chamado não elegível.')
        return redirect('support-ticket-detail', pk=ticket.pk)

    reason = (request.POST.get('reason') or '').strip()
    ticket.status = SupportTicket.Status.REOPENED
    ticket.resolved_at = None
    ticket.save(update_fields=['status', 'resolved_at', 'updated_at'])
    if reason:
        SupportTicketMessage.objects.create(
            ticket=ticket,
            author=request.user,
            message=f"Reabertura solicitada: {reason}",
            is_internal_note=False,
        )
    _notify_support_staff(
        title=f"Chamado #{ticket.pk} reaberto",
        message=f"{request.user.get_full_name() or request.user.username} reabriu o chamado: {ticket.title}",
        exclude_user_id=request.user.id,
    )
    messages.success(request, 'Chamado reaberto com sucesso.')
    return redirect('support-ticket-detail', pk=ticket.pk)


@login_required
@require_http_methods(["GET"])
def support_ticket_admin_list_view(request):
    """Fila administrativa de chamados para superuser/admin."""
    if not _can_manage_support_tickets(request.user):
        raise PermissionDenied("Acesso restrito ao painel de suporte.")

    last_public_author_subq = SupportTicketMessage.objects.filter(
        ticket=OuterRef('pk'),
        is_internal_note=False,
    ).order_by('-created_at').values('author_id')[:1]
    last_public_author_is_staff_subq = SupportTicketMessage.objects.filter(
        ticket=OuterRef('pk'),
        is_internal_note=False,
    ).order_by('-created_at').values('author__is_staff')[:1]
    last_public_author_is_superuser_subq = SupportTicketMessage.objects.filter(
        ticket=OuterRef('pk'),
        is_internal_note=False,
    ).order_by('-created_at').values('author__is_superuser')[:1]
    public_count_subq = SupportTicketMessage.objects.filter(
        ticket=OuterRef('pk'),
        is_internal_note=False,
    ).values('ticket').annotate(c=Count('id')).values('c')[:1]
    tickets = SupportTicket.objects.select_related('created_by', 'assigned_to', 'related_project').annotate(
        last_public_author_id=Subquery(last_public_author_subq),
        last_public_author_is_staff=Subquery(last_public_author_is_staff_subq),
        last_public_author_is_superuser=Subquery(last_public_author_is_superuser_subq),
        public_message_count=Subquery(public_count_subq),
    ).all()
    q = (request.GET.get('q') or '').strip()
    status = (request.GET.get('status') or '').strip()
    severity = (request.GET.get('severity') or '').strip()
    if q:
        tickets = tickets.filter(
            Q(title__icontains=q)
            | Q(description__icontains=q)
            | Q(created_by__username__icontains=q)
            | Q(created_by__first_name__icontains=q)
        )
    if status:
        tickets = tickets.filter(status=status)
    if severity:
        tickets = tickets.filter(severity=severity)

    tickets = list(tickets.order_by('-created_at'))
    for t in tickets:
        t.attention_label, t.attention_tone = _support_attention_tag(t, viewer_is_manager=True)
    return render(
        request,
        'core/support_ticket_admin_list.html',
        {
            'tickets': tickets,
            'status_choices': SupportTicket.Status.choices,
            'severity_choices': SupportTicket.Severity.choices,
            'current_q': q,
            'current_status': status,
            'current_severity': severity,
            'now': timezone.now(),
        },
    )


@login_required
def central_hub_view(request):
    """Redireciona para o Painel do sistema (hub unificado). Staff/superuser sempre usam o Painel."""
    if not (request.user.is_staff or request.user.is_superuser):
        from django.core.exceptions import PermissionDenied
        raise PermissionDenied('Acesso restrito.')
    from django.shortcuts import redirect
    return redirect('accounts:admin_central')


@login_required
def teams_chat_embed_view(request):
    """
    Tela para chat do Teams embutido no LPLAN.
    Suporta:
    - acs_iframe: embute uma app web ACS hospedada pela empresa.
    - embedded_sdk: usa SDK JS (quando configurado no tenant).
    """
    from django.conf import settings

    mode = (getattr(settings, 'TEAMS_CHAT_EMBED_MODE', 'acs_iframe') or 'acs_iframe').strip()
    context = {
        'teams_chat_enabled': bool(getattr(settings, 'TEAMS_CHAT_EMBED_ENABLED', False)),
        'teams_chat_mode': mode,
        'teams_chat_app_url': (getattr(settings, 'TEAMS_CHAT_APP_URL', '') or '').strip(),
        'teams_embedded_sdk_url': (getattr(settings, 'TEAMS_EMBEDDED_SDK_URL', '') or '').strip(),
        'teams_chat_entity_prefix': (getattr(settings, 'TEAMS_CHAT_ENTITY_PREFIX', 'LPLAN') or 'LPLAN').strip(),
    }
    return render(request, 'core/teams_chat_embed.html', context)


def _is_work_owner(user):
    """True se o usuário é dono de alguma obra (acesso restrito só às suas obras)."""
    from core.models import ProjectOwner
    return ProjectOwner.objects.filter(user=user).exists()


def _get_projects_for_user(request):
    """Obras que o usuário pode acessar no Diário: staff/superuser vê todas; donos só as que possuem; demais só as vinculadas."""
    from core.models import ProjectOwner
    if request.user.is_staff or request.user.is_superuser:
        return Project.objects.filter(is_active=True).order_by('-created_at')
    # Donos da obra: só veem as obras das quais são donos
    owner_project_ids = list(ProjectOwner.objects.filter(user=request.user).values_list('project_id', flat=True))
    if owner_project_ids:
        return Project.objects.filter(pk__in=owner_project_ids, is_active=True).order_by('-created_at')
    project_ids = list(ProjectMember.objects.filter(user=request.user).values_list('project_id', flat=True))
    # Se não tem vínculo no Diário mas tem obras no GestControll, sincronizar (lista única)
    if not project_ids:
        try:
            from gestao_aprovacao.models import Obra, WorkOrderPermission
            for project_id in Obra.objects.filter(
                permissoes__usuario=request.user,
                permissoes__ativo=True,
            ).exclude(project__isnull=True).values_list('project_id', flat=True).distinct():
                ProjectMember.objects.get_or_create(user=request.user, project_id=project_id)
            project_ids = list(ProjectMember.objects.filter(user=request.user).values_list('project_id', flat=True))
        except Exception:
            pass
    approver_project_ids = list(
        ProjectDiaryApprover.objects.filter(
            user=request.user,
            is_active=True,
        ).values_list('project_id', flat=True)
    )
    combined_ids = sorted(set(project_ids + approver_project_ids))
    return Project.objects.filter(pk__in=combined_ids, is_active=True).order_by('-created_at')


def _get_support_projects_for_user(user):
    """
    Obras permitidas no chamado de suporte.
    Regra: somente obras em que o usuário está vinculado.
    """
    if not user or not getattr(user, 'is_authenticated', False):
        return Project.objects.none()

    owner_project_ids = list(
        ProjectOwner.objects.filter(user=user).values_list('project_id', flat=True)
    )
    member_project_ids = list(
        ProjectMember.objects.filter(user=user).values_list('project_id', flat=True)
    )
    approver_project_ids = list(
        ProjectDiaryApprover.objects.filter(user=user, is_active=True).values_list('project_id', flat=True)
    )
    linked_ids = sorted(set(owner_project_ids + member_project_ids + approver_project_ids))
    return Project.objects.filter(pk__in=linked_ids, is_active=True).order_by('-created_at')


def _user_can_access_project(user, project):
    """Verifica se o usuário pode acessar a obra (dono, vinculado ou staff/superuser)."""
    if user.is_staff or user.is_superuser:
        return True
    from core.models import ProjectOwner
    if ProjectOwner.objects.filter(user=user, project=project).exists():
        return True
    if ProjectDiaryApprover.objects.filter(user=user, project=project, is_active=True).exists():
        return True
    return ProjectMember.objects.filter(user=user, project=project).exists()


def _is_project_rdo_approver(user, project):
    """True se o usuário é aprovador ativo de RDO para a obra."""
    if not user or not getattr(user, 'is_authenticated', False):
        return False
    # Superusuário pode atuar como contingência operacional.
    # Staff comum só aprova quando estiver explicitamente cadastrado na obra.
    if user.is_superuser:
        return True
    return ProjectDiaryApprover.objects.filter(
        user=user,
        project=project,
        is_active=True,
    ).exists()


@login_required
def select_project_view(request):
    """View para seleção de obra após login. Só aparecem obras às quais o usuário está vinculado (staff vê todas)."""
    projects = _get_projects_for_user(request)

    if request.method == 'POST':
        project_id = request.POST.get('project_id')
        if project_id:
            try:
                project = Project.objects.get(pk=project_id, is_active=True)
                if not _user_can_access_project(request.user, project):
                    messages.error(request, 'Você não está vinculado a esta obra.')
                    return render(request, 'core/select_project.html', {
                        'projects': projects,
                        'selected_project_id': request.session.get('selected_project_id'),
                    })
                request.session['selected_project_id'] = project.id
                request.session['selected_project_name'] = project.name
                request.session['selected_project_code'] = project.code
                return redirect('dashboard')
            except (Project.DoesNotExist, ValueError, TypeError):
                return render(request, 'core/select_project.html', {
                    'error': 'Obra não encontrada ou inativa.',
                    'projects': projects,
                    'selected_project_id': request.session.get('selected_project_id'),
                })

    selected_project_id = request.session.get('selected_project_id')
    return render(request, 'core/select_project.html', {
        'projects': projects,
        'selected_project_id': selected_project_id,
    })


def get_selected_project(request):
    """Helper function para obter a obra selecionada da sessão."""
    project_id = request.session.get('selected_project_id')
    if project_id:
        try:
            return Project.objects.get(pk=project_id, is_active=True)
        except Project.DoesNotExist:
            # Limpa sessão se obra não existe mais
            for key in ('selected_project_id', 'selected_project_name', 'selected_project_code'):
                request.session.pop(key, None)
    return None


def project_required(view_func):
    """Decorator para garantir que uma obra foi selecionada e que o usuário ainda tem acesso a ela."""
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('login')
        if 'selected_project_id' not in request.session:
            return redirect('select-project')
        project = get_selected_project(request)
        if not project:
            return redirect('select-project')
        if not _user_can_access_project(request.user, project):
            for key in ('selected_project_id', 'selected_project_name', 'selected_project_code'):
                request.session.pop(key, None)
            messages.warning(request, 'Você não está mais vinculado a essa obra. Selecione outra.')
            return redirect('select-project')
        return view_func(request, *args, **kwargs)
    return wrapper


@login_required
@project_required
def dashboard_view(request):
    """View do dashboard com KPIs e calendário."""
    project = get_selected_project(request)

    # KPIs filtrados pela obra selecionada
    total_diaries = ConstructionDiary.objects.filter(project=project).count()
    # Relatórios pendentes (inclui rascunho, preenchendo, aguardando aprovação e reprovados)
    pending_reports = ConstructionDiary.objects.filter(
        project=project,
        status__in=[
            DiaryStatus.PREENCHENDO,
            DiaryStatus.SALVAMENTO_PARCIAL,
            DiaryStatus.AGUARDANDO_APROVACAO_GESTOR,
            DiaryStatus.REPROVADO_GESTOR,
        ],
    ).count()
    approved_reports = ConstructionDiary.objects.filter(
        project=project,
        status=DiaryStatus.APROVADO
    ).count()
    
    # Clima médio (últimos 7 dias) - filtrado pela obra
    last_week = timezone.now().date() - timedelta(days=7)
    recent_diaries = ConstructionDiary.objects.filter(
        project=project,
        date__gte=last_week
    )
    
    # Calcula clima médio real
    weather_counts = {}
    for diary in recent_diaries:
        if diary.weather_conditions:
            weather_lower = diary.weather_conditions.lower()
            if 'sol' in weather_lower:
                weather_counts['sol'] = weather_counts.get('sol', 0) + 1
            elif 'nuvem' in weather_lower or 'nublado' in weather_lower:
                weather_counts['nublado'] = weather_counts.get('nublado', 0) + 1
            elif 'chuva' in weather_lower:
                weather_counts['chuva'] = weather_counts.get('chuva', 0) + 1
    
    if weather_counts:
        avg_weather = max(weather_counts, key=weather_counts.get).title()
    else:
        avg_weather = "N/A"
    
    # Efetivo total (conta todos os funcionários ativos)
    total_workers = Labor.objects.filter(is_active=True).count()
    
    # Estatísticas adicionais
    total_work_hours = ConstructionDiary.objects.filter(
        project=project,
        work_hours__isnull=False
    ).aggregate(total=Sum('work_hours'))['total'] or 0
    
    # Média de horas trabalhadas
    avg_work_hours = ConstructionDiary.objects.filter(
        project=project,
        work_hours__isnull=False
    ).aggregate(avg=Avg('work_hours'))['avg']
    
    # KPIs adicionais baseados no diariodeobra.app
    from core.models import DiaryImage, DiaryVideo, DiaryAttachment, Activity, DailyWorkLog, DiaryOccurrence
    
    # Total de fotos
    total_photos = DiaryImage.objects.filter(
        diary__project=project
    ).count()
    
    # Total de atividades
    total_activities = Activity.objects.filter(project=project).count()
    
    # Total de ocorrências (usando modelo DiaryOccurrence)
    # Usa select_related para otimizar a query
    total_occurrences = DiaryOccurrence.objects.filter(
        diary__project=project
    ).select_related('diary', 'diary__project').count()

    # Total de comentários (notas gerais dos diários)
    total_comments = ConstructionDiary.objects.filter(
        project=project
    ).exclude(general_notes='').count()
    
    # Total de vídeos
    # Usa select_related para otimizar a query
    total_videos = DiaryVideo.objects.filter(
        diary__project=project
    ).select_related('diary', 'diary__project').count()

    # Total de anexos
    total_attachments = DiaryAttachment.objects.filter(
        diary__project=project
    ).count()
    
    # Relatórios recentes (últimos 7 para cards; até 25 para modo tabela)
    recent_reports = ConstructionDiary.objects.filter(
        project=project
    ).select_related('project').prefetch_related(
        'images', 'videos', 'occurrences'
    ).order_by('-date', '-created_at')[:7]
    recent_reports_table = ConstructionDiary.objects.filter(
        project=project
    ).select_related('project').prefetch_related(
        'images', 'videos', 'occurrences'
    ).order_by('-date', '-created_at')[:25]
    
    # Fotos recentes (últimas 9)
    recent_photos = DiaryImage.objects.filter(
        diary__project=project
    ).select_related('diary').order_by('-uploaded_at')[:9]
    
    # Vídeos recentes (últimos 6)
    recent_videos = DiaryVideo.objects.filter(
        diary__project=project
    ).select_related('diary').order_by('-uploaded_at')[:6]

    # Cálculos para Informações da Obra
    project_days_elapsed = 0
    project_days_total = 0
    project_days_remaining = 0
    project_progress_percent = 0
    
    if project.start_date:
        from datetime import date
        today = date.today()
        project_days_elapsed = (today - project.start_date).days
        
        if project.end_date:
            project_days_total = (project.end_date - project.start_date).days
            project_days_remaining = (project.end_date - today).days
            
            if project_days_total > 0:
                project_progress_percent = min(100, max(0, (project_days_elapsed / project_days_total) * 100))
    
    from django.conf import settings
    context = {
        'project': project,
        'total_diaries': total_diaries,
        'pending_reports': pending_reports,
        'approved_reports': approved_reports,
        'avg_weather': avg_weather,
        'total_workers': total_workers,
        'total_work_hours': total_work_hours,
        'avg_work_hours': round(avg_work_hours, 1) if avg_work_hours else None,
        'total_photos': total_photos,
        'total_activities': total_activities,
        'total_occurrences': total_occurrences,
        'total_comments': total_comments,
        'total_videos': total_videos,
        'total_attachments': total_attachments,
        'recent_reports': recent_reports,
        'recent_reports_table': recent_reports_table,
        'recent_photos': recent_photos,
        'recent_videos': recent_videos,
        # Informações da Obra
        'project_days_elapsed': project_days_elapsed,
        'project_days_total': project_days_total,
        'project_days_remaining': project_days_remaining,
        'project_progress_percent': round(project_progress_percent, 1),
        # Para sidebar
        'total_reports_count': total_diaries,
        'total_photos_count': total_photos,
        'total_videos_count': total_videos,
        'total_activities_count': total_activities,
        'total_occurrences_count': total_occurrences,
        'total_comments_count': total_comments,
        'total_attachments_count': total_attachments,
        'DEBUG': settings.DEBUG,
        'no_report_reasons': DiaryNoReportDay.Reason.choices,
    }
    
    return render(request, 'core/dashboard.html', context)


@login_required
@project_required
def calendar_events_view(request):
    """API endpoint para eventos do FullCalendar."""
    from datetime import date as date_type
    
    project = get_selected_project(request)
    start = request.GET.get('start')
    end = request.GET.get('end')
    
    try:
        start_date = datetime.fromisoformat(start.replace('Z', '+00:00'))
        end_date = datetime.fromisoformat(end.replace('Z', '+00:00'))
    except (ValueError, AttributeError, TypeError):
        # start/end ausentes ou inválidos: usa período padrão (datetime para .date() abaixo)
        start_date = timezone.now()
        end_date = start_date + timedelta(days=30)
    
    # Ajusta o período para considerar o período da obra
    view_start = start_date.date()
    view_end = end_date.date()

    # Início: não mostrar antes do início do projeto
    if project.start_date:
        view_start = max(view_start, project.start_date)
    # Término previsto: não limitar view_end à data prevista, para que obras atrasadas
    # possam exibir e registrar diários além do prazo original
    
    # Busca todos os diários no período
    diaries = ConstructionDiary.objects.filter(
        project=project,
        date__gte=view_start,
        date__lte=view_end
    ).select_related('project', 'created_by')
    
    # Cria um conjunto de datas que já têm relatórios
    dates_with_diaries = set(diary.date for diary in diaries)
    
    events = []
    
    # Adiciona eventos para dias com relatórios
    for diary in diaries:
        # Determina cor baseada no status
        if diary.status == DiaryStatus.APROVADO:
            color = '#10b981'  # Verde - Preenchido/Finalizado
            title_status = 'Preenchido'
        elif diary.status == DiaryStatus.AGUARDANDO_APROVACAO_GESTOR:
            color = '#2563eb'  # Azul - aguardando aprovação
            title_status = 'Aguardando aprovação'
        elif diary.status == DiaryStatus.REPROVADO_GESTOR:
            color = '#dc2626'  # Vermelho - reprovado
            title_status = 'Reprovado'
        elif diary.status == DiaryStatus.SALVAMENTO_PARCIAL:
            color = '#f59e0b'  # Âmbar - Salvamento parcial (rascunho)
            title_status = 'Salvamento Parcial'
        elif diary.status == DiaryStatus.PREENCHENDO:
            # Preenchendo = relatório existe mas ainda está sendo preenchido
            color = '#10b981'  # Verde - Preenchido
            title_status = 'Preenchido'
        else:
            # Status desconhecido ou inválido - trata como preenchido mas com status indefinido
            color = '#6b7280'  # Cinza
            title_status = 'Indefinido'
        
        # Título completo e curto (para preview compacto no calendário)
        if diary.report_number:
            title = f"RDO #{diary.report_number} - {title_status}"
            short_title = f"RDO #{diary.report_number}"
        else:
            creator = getattr(diary, 'created_by', None)
            if creator:
                creator_name = creator.get_full_name() or creator.username
                title = f"{creator_name[:15]}... - {title_status}"
            else:
                title = f"RDO - {title_status}"
            short_title = "RDO"
        
        events.append({
            'id': diary.id,
            'title': title,
            'start': diary.date.isoformat(),
            'allDay': True,
            'color': color,
            'display': 'block',
            'extendedProps': {
                'status': title_status,
                'diary_id': diary.id,
                'has_diary': True,
                'short_title': short_title,
            },
        })
    
    # Adiciona eventos para dias sem relatórios (dias faltantes) ou já justificados (sem RDO)
    # Considera todo o intervalo exibido (view_start até view_end), incluindo após o término previsto
    if view_start and view_end:
        today = timezone.now().date()
        justified_by_date = {
            row.date: row
            for row in DiaryNoReportDay.objects.filter(
                project=project,
                date__gte=view_start,
                date__lte=view_end,
            ).only('date', 'reason', 'note')
        }
        current_date = view_start

        while current_date <= view_end:
            # Se não tem relatório neste dia e está dentro do período da obra
            # Só exibe evento para hoje ou dias passados (Falta/Atraso). Dias futuros ficam em branco.
            if current_date not in dates_with_diaries and current_date <= today:
                if current_date in justified_by_date:
                    nrd = justified_by_date[current_date]
                    reason_label = nrd.get_reason_display()
                    note = (nrd.note or '').strip()
                    if note:
                        title = f'{reason_label} — {note}'
                        # Na célula: prioriza o nome livre (ex.: feriado); senão o motivo
                        short_raw = note
                    else:
                        title = reason_label
                        short_raw = reason_label
                    if len(short_raw) > 22:
                        short_title = short_raw[:19] + '…'
                    else:
                        short_title = short_raw
                    events.append({
                        'id': f'justified_{current_date.isoformat()}',
                        'title': title,
                        'start': current_date.isoformat(),
                        'allDay': True,
                        'color': '#94a3b8',
                        'display': 'block',
                        'extendedProps': {
                            'status': 'Justificado',
                            'diary_id': None,
                            'has_diary': False,
                            'missing': False,
                            'no_report_justified': True,
                            'no_report_reason': reason_label,
                            'no_report_note': note,
                            'short_title': short_title,
                        },
                    })
                else:
                    color = '#dc2626'  # Vermelho escuro para borda
                    title_status = 'Atraso'
                    title = f"Falta relatório - {title_status}"
                    short_title = 'Falta'
                    events.append({
                        'id': f'missing_{current_date.isoformat()}',
                        'title': title,
                        'start': current_date.isoformat(),
                        'allDay': True,
                        'color': color,
                        'display': 'block',
                        'extendedProps': {
                            'status': title_status,
                            'diary_id': None,
                            'has_diary': False,
                            'missing': True,
                            'short_title': short_title,
                        },
                    })

            current_date += timedelta(days=1)
    
    # Evita resposta em cache (proxy/navegador): o calendário ficava com “Falta”
    # mesmo após justificar o dia no servidor.
    response = JsonResponse(events, safe=False)
    response['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    response['Pragma'] = 'no-cache'
    response['Vary'] = 'Cookie'
    return response


_REPORT_LIST_SORT_KEYS = frozenset({'date', 'report_number', 'status'})
_REPORT_LIST_SORT_DEFAULT_DIR = {
    'date': 'desc',
    'report_number': 'desc',
    'status': 'asc',
}


def _report_list_sort_query_string(request, column: str) -> str:
    """Monta query string para alternar ordenação ao clicar numa coluna (preserva filtros, remove page)."""
    if column not in _REPORT_LIST_SORT_KEYS:
        column = 'date'
    qd = request.GET.copy()
    qd.pop('page', None)
    cur_sort = qd.get('sort') or 'date'
    if cur_sort not in _REPORT_LIST_SORT_KEYS:
        cur_sort = 'date'
    cur_dir = qd.get('dir')
    if cur_dir not in ('asc', 'desc'):
        cur_dir = _REPORT_LIST_SORT_DEFAULT_DIR[cur_sort]
    if cur_sort == column:
        qd['sort'] = column
        qd['dir'] = 'asc' if cur_dir == 'desc' else 'desc'
    else:
        qd['sort'] = column
        qd['dir'] = _REPORT_LIST_SORT_DEFAULT_DIR[column]
    return qd.urlencode()


def _report_list_parse_sort(request):
    sort = request.GET.get('sort') or 'date'
    if sort not in _REPORT_LIST_SORT_KEYS:
        sort = 'date'
    sort_dir = request.GET.get('dir')
    if sort_dir not in ('asc', 'desc'):
        sort_dir = _REPORT_LIST_SORT_DEFAULT_DIR[sort]
    return sort, sort_dir


def _report_list_merge_rows(diary_list, no_report_list, sort, sort_dir):
    """Mistura RDOs com dias justificados (sem RDO); ordenação alinhada à coluna ativa."""
    diary_rows = [('diary', d) for d in diary_list]
    no_report_rows = [('no_report', j) for j in no_report_list]
    reverse = sort_dir == 'desc'

    if sort == 'date':
        rows = diary_rows + no_report_rows

        def k(item):
            kind, o = item
            sub_kind = 0 if kind == 'diary' else 1
            return (o.date, sub_kind, o.created_at, o.pk)

        rows.sort(key=k, reverse=reverse)
        return rows

    if sort == 'report_number':

        def dk(x):
            o = x[1]
            n = o.report_number
            return (1 if n is None else 0, n or 0, o.date, o.pk)

        diary_rows.sort(key=dk, reverse=reverse)
        no_report_rows.sort(key=lambda x: (x[1].date, x[1].pk), reverse=reverse)
        return diary_rows + no_report_rows

    diary_rows.sort(
        key=lambda x: (x[1].status, x[1].date, x[1].pk),
        reverse=reverse,
    )
    no_report_rows.sort(
        key=lambda x: (x[1].reason, x[1].date, x[1].pk),
        reverse=reverse,
    )
    return diary_rows + no_report_rows


@login_required
@project_required
def report_list_view(request):
    """View de listagem de relatórios com filtros HTMX."""
    project = get_selected_project(request)
    report_list_per_page = 30
    is_htmx = bool(request.headers.get('HX-Request'))

    diaries = (
        ConstructionDiary.objects.filter(project=project)
        .select_related('project')
        .annotate(image_count=Count('images'))
    )

    if is_htmx:
        can_review_diaries = False
        pending_approval_diaries = []
    else:
        can_review_diaries = _is_project_rdo_approver(request.user, project) if project else False
        pending_approval_diaries = ConstructionDiary.objects.filter(
            project=project,
            status=DiaryStatus.AGUARDANDO_APROVACAO_GESTOR,
        ).select_related('created_by').order_by('date', 'report_number')
    
    # Filtros
    search = request.GET.get('search')
    if search:
        diaries = diaries.filter(
            Q(project__code__icontains=search) |
            Q(project__name__icontains=search) |
            Q(general_notes__icontains=search)
        )
    
    date_start = request.GET.get('date_start')
    if date_start:
        try:
            diaries = diaries.filter(date__gte=date_start)
        except ValueError:
            pass
    
    date_end = request.GET.get('date_end')
    if date_end:
        try:
            diaries = diaries.filter(date__lte=date_end)
        except ValueError:
            pass
    
    status = request.GET.get('status')
    if status:
        diaries = diaries.filter(status=status)

    sort, sort_dir = _report_list_parse_sort(request)

    no_report_qs = DiaryNoReportDay.objects.filter(
        project=project,
        date__gte=timezone.now().date() - timedelta(days=366),
    ).select_related('created_by')
    if date_start:
        try:
            no_report_qs = no_report_qs.filter(date__gte=date_start)
        except ValueError:
            pass
    if date_end:
        try:
            no_report_qs = no_report_qs.filter(date__lte=date_end)
        except ValueError:
            pass
    if search:
        no_report_qs = no_report_qs.filter(Q(note__icontains=search))
    if status:
        no_report_qs = DiaryNoReportDay.objects.none()

    diary_list = list(diaries.order_by())
    no_report_list = list(no_report_qs)
    merged_rows = _report_list_merge_rows(diary_list, no_report_list, sort, sort_dir)

    paginator = Paginator(merged_rows, report_list_per_page)
    page_obj = paginator.get_page(request.GET.get('page') or 1)

    qd = request.GET.copy()
    qd.pop('page', None)
    report_list_pagination_query = qd.urlencode()
    
    # Último relatório para o modal (qualquer status)
    if is_htmx:
        last_diary = None
        all_projects = []
        no_report_reasons = []
    else:
        last_diary = ConstructionDiary.objects.filter(
            project=project
        ).order_by('-date', '-created_at').first()
        all_projects = _get_projects_for_user(request)
        no_report_reasons = DiaryNoReportDay.Reason.choices

    context = {
        'report_list_total': paginator.count,
        'report_list_page_obj': page_obj,
        'report_list_paginator': paginator,
        'report_list_per_page': report_list_per_page,
        'report_list_pagination_query': report_list_pagination_query,
        'report_list_sort': sort,
        'report_list_sort_dir': sort_dir,
        'report_list_sort_q': {
            'date': _report_list_sort_query_string(request, 'date'),
            'report_number': _report_list_sort_query_string(request, 'report_number'),
            'status': _report_list_sort_query_string(request, 'status'),
        },
        'last_diary': last_diary,
        'user': request.user,  # Adiciona user ao contexto para can_be_edited_by
        'project': project,  # Adiciona projeto ao contexto para o modal
        'all_projects': all_projects,  # Projetos acessíveis para o select do modal
        'can_review_diaries': can_review_diaries,
        'pending_approval_diaries': pending_approval_diaries,
        'no_report_reasons': no_report_reasons,
    }
    
    # Se for requisição HTMX, retorna apenas o conteúdo
    if request.headers.get('HX-Request'):
        return render(request, 'core/report_list_partial.html', context)
    
    return render(request, 'core/report_list.html', context)


@login_required
@project_required
@require_http_methods(['POST'])
def diary_no_report_day_create_view(request):
    """Registo rápido: dia sem RDO (feriado, fim de semana, etc.)."""
    from datetime import datetime as dt_mod

    def _redirect_after_no_report_day():
        if request.POST.get('return_to', '').strip() == 'dashboard':
            return redirect('dashboard')
        return redirect('report-list')

    project = get_selected_project(request)
    if not project or not _user_can_access_project(request.user, project):
        raise PermissionDenied()
    date_str = request.POST.get('date', '').strip()
    reason = request.POST.get('reason', '').strip()
    note = (request.POST.get('note') or '').strip()[:300]
    valid_reasons = {c.value for c in DiaryNoReportDay.Reason}
    if reason not in valid_reasons:
        messages.error(request, 'Selecione um motivo válido.')
        return _redirect_after_no_report_day()
    try:
        day = dt_mod.strptime(date_str, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        messages.error(request, 'Data inválida.')
        return _redirect_after_no_report_day()
    today = timezone.now().date()
    if day > today:
        messages.error(request, 'Não é possível justificar datas futuras.')
        return _redirect_after_no_report_day()
    if ConstructionDiary.objects.filter(project=project, date=day).exists():
        messages.error(
            request,
            f'Já existe relatório em {day.strftime("%d/%m/%Y")}. Abra ou edite o RDO; não é possível justificar o mesmo dia.',
        )
        return _redirect_after_no_report_day()
    existing = DiaryNoReportDay.objects.filter(project=project, date=day).first()
    if existing:
        if not request.user.is_superuser and existing.created_by_id != request.user.id:
            messages.error(
                request,
                'Só quem registou esta justificativa (ou um administrador) pode alterá-la. Peça a remoção ou a correção a essa pessoa.',
            )
            return _redirect_after_no_report_day()
        existing.reason = reason
        existing.note = note
        existing.save(update_fields=['reason', 'note', 'updated_at'])
        messages.success(
            request,
            f'Justificativa atualizada: {day.strftime("%d/%m/%Y")} — {existing.get_reason_display()}.',
        )
    else:
        nrd = DiaryNoReportDay.objects.create(
            project=project,
            date=day,
            reason=reason,
            note=note,
            created_by=request.user,
        )
        messages.success(
            request,
            f'Registado: {day.strftime("%d/%m/%Y")} — {nrd.get_reason_display()}.',
        )
    return _redirect_after_no_report_day()


@login_required
@project_required
@require_http_methods(['POST'])
def diary_no_report_day_delete_view(request, pk):
    project = get_selected_project(request)
    if not project or not _user_can_access_project(request.user, project):
        raise PermissionDenied()
    obj = get_object_or_404(DiaryNoReportDay, pk=pk, project=project)
    if not request.user.is_superuser and obj.created_by_id != request.user.id:
        raise PermissionDenied()
    obj.delete()
    messages.success(request, 'Justificativa removida.')
    return redirect('report-list')


@login_required
def diary_detail_view(request, pk):
    """View de detalhe do diário."""
    from collections import defaultdict
    from core.models import DiaryView, DiaryEditLog
    
    # SEMPRE retorna HTML para esta view (é uma view frontend)
    # Ignora qualquer header Accept que possa fazer o DRF interceptar
    # Se vier de /api/diaries/, redireciona para /diaries/
    if request.path.startswith('/api/diaries/'):
        from django.http import HttpResponseRedirect
        return HttpResponseRedirect(f'/diaries/{pk}/')
    
    # Força Content-Type HTML explícito para evitar que DRF intercepte
    # Remove qualquer header Accept que possa fazer o DRF pensar que é JSON
    # SEMPRE força HTML para requisições do frontend
    request.META['HTTP_ACCEPT'] = 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8'
    if 'Accept' in request.META:
        request.META['Accept'] = 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8'
    
    diary = get_object_or_404(
        ConstructionDiary.objects.select_related('project', 'created_by', 'reviewed_by')
        .prefetch_related(
            'images', 'videos',
            'work_logs__activity', 'work_logs__resources_labor', 'work_logs__resources_equipment',
            'occurrences', 'occurrences__tags',
            'owner_comments__author',
            'approval_history__decided_by',
        ),
        pk=pk
    )
    project = get_selected_project(request)
    # Se há projeto na sessão e é diferente do diário, faz auto-switch seguro
    # quando o usuário tem permissão na obra do diário.
    if project is not None and diary.project_id != project.id:
        if not _user_can_access_project(request.user, diary.project):
            raise Http404('Relatório não encontrado.')
        project = diary.project
        request.session['selected_project_id'] = project.id
        request.session['selected_project_name'] = project.name
        request.session['selected_project_code'] = getattr(project, 'code', '')
    # Se não há projeto na sessão (ex.: link direto), só usa o projeto do diário se o usuário tiver acesso
    if project is None:
        if not _user_can_access_project(request.user, diary.project):
            raise Http404('Relatório não encontrado.')
        project = diary.project
        request.session['selected_project_id'] = project.id
        request.session['selected_project_name'] = project.name
        request.session['selected_project_code'] = getattr(project, 'code', '')
    
    # Registra visualização
    DiaryView.objects.create(
        diary=diary,
        viewed_by=request.user,
        ip_address=request.META.get('REMOTE_ADDR')
    )
    
    # Conta visualizações e edições
    view_count = DiaryView.objects.filter(diary=diary).count()
    edit_count = DiaryEditLog.objects.filter(diary=diary).count()
    
    # Processa condições climáticas - melhorado para usar campos específicos
    weather_morning = None
    weather_afternoon = None
    weather_night = None
    
    # Processa clima da manhã usando campos específicos
    if diary.weather_morning_condition:
        condition = diary.weather_morning_condition
        workable = diary.weather_morning_workable
        
        if condition == 'B':  # Bom
            if 'sol' in (diary.weather_conditions or '').lower():
                weather_morning = {'icon': 'sun', 'label': 'Claro', 'condition': 'Trabalhável' if workable == 'T' else 'Não Trabalhável', 'color': 'yellow'}
            else:
                weather_morning = {'icon': 'cloud-sun', 'label': 'Bom', 'condition': 'Trabalhável' if workable == 'T' else 'Não Trabalhável', 'color': 'yellow'}
        elif condition == 'R':  # Ruim
            weather_morning = {'icon': 'cloud-rain', 'label': 'Ruim', 'condition': 'Não Trabalhável', 'color': 'blue'}
        else:
            # Fallback para processamento antigo
            if diary.weather_conditions:
                parts = diary.weather_conditions.split('|')
                for part in parts:
                    part = part.strip()
                    if 'manhã' in part.lower() or 'manha' in part.lower():
                        if 'sol' in part.lower():
                            weather_morning = {'icon': 'sun', 'label': 'Claro', 'condition': 'Praticável', 'color': 'yellow'}
                        elif 'nuvem' in part.lower() or 'nublado' in part.lower():
                            weather_morning = {'icon': 'cloud', 'label': 'Nublado', 'condition': 'Praticável', 'color': 'gray'}
                        elif 'chuva' in part.lower():
                            weather_morning = {'icon': 'cloud-rain', 'label': 'Chuvoso', 'condition': 'Impraticável', 'color': 'blue'}
    
    # Processa clima da tarde usando campos específicos
    if diary.weather_afternoon_condition:
        condition = diary.weather_afternoon_condition
        workable = diary.weather_afternoon_workable
        
        if condition == 'B':  # Bom
            if 'sol' in (diary.weather_conditions or '').lower():
                weather_afternoon = {'icon': 'sun', 'label': 'Claro', 'condition': 'Trabalhável' if workable == 'T' else 'Não Trabalhável', 'color': 'yellow'}
            else:
                weather_afternoon = {'icon': 'cloud-sun', 'label': 'Bom', 'condition': 'Trabalhável' if workable == 'T' else 'Não Trabalhável', 'color': 'yellow'}
        elif condition == 'R':  # Ruim
            weather_afternoon = {'icon': 'cloud-rain', 'label': 'Ruim', 'condition': 'Não Trabalhável', 'color': 'blue'}
        else:
            # Fallback para processamento antigo
            if diary.weather_conditions:
                parts = diary.weather_conditions.split('|')
                for part in parts:
                    part = part.strip()
                    if 'tarde' in part.lower():
                        if 'sol' in part.lower():
                            weather_afternoon = {'icon': 'sun', 'label': 'Claro', 'condition': 'Praticável', 'color': 'yellow'}
                        elif 'nuvem' in part.lower() or 'nublado' in part.lower():
                            weather_afternoon = {'icon': 'cloud', 'label': 'Nublado', 'condition': 'Praticável', 'color': 'gray'}
                        elif 'chuva' in part.lower():
                            weather_afternoon = {'icon': 'cloud-rain', 'label': 'Chuvoso', 'condition': 'Impraticável', 'color': 'blue'}
    
    # Processa clima da noite se habilitado
    if diary.weather_night_enabled and diary.weather_night_type:
        night_type = diary.weather_night_type
        night_workable = diary.weather_night_workable
        
        if night_type == 'C':  # Claro
            weather_night = {'icon': 'moon', 'label': 'Claro', 'condition': 'Praticável' if night_workable == 'P' else 'Impraticável', 'color': 'yellow'}
        elif night_type == 'N':  # Nublado
            weather_night = {'icon': 'cloud-moon', 'label': 'Nublado', 'condition': 'Praticável' if night_workable == 'P' else 'Impraticável', 'color': 'gray'}
        elif night_type == 'CH':  # Chuvoso
            weather_night = {'icon': 'cloud-moon-rain', 'label': 'Chuvoso', 'condition': 'Impraticável', 'color': 'blue'}
    
    # Agrupa mão de obra por tipo
    labor_by_type = {
        'Direto': {},
        'Indireto': {},
        'Terceiros': {},
    }
    
    # Coleta mão de obra de todos os work_logs
    for work_log in diary.work_logs.all():
        for labor in work_log.resources_labor.all():
            labor_type = labor.labor_type
            if labor_type == 'D':
                if labor.name not in labor_by_type['Direto']:
                    labor_by_type['Direto'][labor.name] = 0
                labor_by_type['Direto'][labor.name] += 1
            elif labor_type == 'I':
                if labor.name not in labor_by_type['Indireto']:
                    labor_by_type['Indireto'][labor.name] = 0
                labor_by_type['Indireto'][labor.name] += 1
            elif labor_type == 'T':
                if labor.name not in labor_by_type['Terceiros']:
                    labor_by_type['Terceiros'][labor.name] = 0
                labor_by_type['Terceiros'][labor.name] += 1
    
    # Calcula métricas do projeto
    project = diary.project
    project_days_total = (project.end_date - project.start_date).days if project.end_date and project.start_date else 0
    project_days_elapsed = (diary.date - project.start_date).days if project.end_date and project.start_date else 0
    project_days_remaining = (project.end_date - diary.date).days if project.end_date and project.start_date else 0
    
    # Nome do dia da semana
    from datetime import datetime
    weekday_names = ['Segunda-Feira', 'Terça-Feira', 'Quarta-Feira', 'Quinta-Feira', 'Sexta-Feira', 'Sábado', 'Domingo']
    weekday_name = weekday_names[diary.date.weekday()] if diary.date else ''
    
    # Equipamentos agregados por diário (mesma regra do PDF)
    equipment_list = _build_diary_equipment_list(diary)

    # Mão de obra por categorias (DiaryLaborEntry) - preferência sobre work_log.resources_labor
    labor_entries_by_category = None
    try:
        from .models import DiaryLaborEntry
        entries = DiaryLaborEntry.objects.filter(diary=diary).select_related('cargo', 'cargo__category').order_by('cargo__category__order', 'company', 'cargo__name')
        if entries.exists():
            labor_entries_by_category = {'indireta': [], 'direta': [], 'terceirizada': {}}
            for e in entries:
                slug = e.cargo.category.slug
                item = {'cargo_name': e.cargo.name, 'quantity': e.quantity}
                if slug == 'terceirizada':
                    company = e.company or '(Sem empresa)'
                    if company not in labor_entries_by_category['terceirizada']:
                        labor_entries_by_category['terceirizada'][company] = []
                    labor_entries_by_category['terceirizada'][company].append(item)
                elif slug in labor_entries_by_category:
                    labor_entries_by_category[slug].append(item)
            labor_entries_by_category['terceirizada'] = [{'company': k, 'items': v} for k, v in labor_entries_by_category['terceirizada'].items()]
    except Exception:
        pass
    
    # Assinaturas
    signatures = diary.signatures.all().select_related('signer')
    signature_inspection = signatures.filter(signature_type='inspection').first()
    signature_production = signatures.filter(signature_type='production').first()
    
    diary_provisional_edit_unlocked = bool(getattr(diary, 'provisional_edit_granted_at', None))
    diary_edit_request_pending = bool(
        getattr(diary, 'edit_requested_at', None) and not diary_provisional_edit_unlocked
    )
    show_request_edit_diary = (
        diary.is_approved()
        and _user_can_access_project(request.user, diary.project)
        and not diary_provisional_edit_unlocked
        and not diary_edit_request_pending
    )
    can_decide_rdo_approval = (
        diary.status == DiaryStatus.AGUARDANDO_APROVACAO_GESTOR
        and _is_project_rdo_approver(request.user, diary.project)
    )
    approval_history = list(
        diary.approval_history.select_related('decided_by').order_by('-created_at')[:20]
    )
    rdo_approvers = list(
        diary.project.rdo_approvers.filter(is_active=True).select_related('user').order_by('order', 'user__first_name', 'user__username')
    )

    nav_prev_pk, nav_next_pk, nav_position, nav_total = _diary_adjacent_ids_for_project(project, diary.pk)

    context = {
        'diary': diary,
        'user': request.user,
        'show_request_edit_diary': show_request_edit_diary,
        'diary_edit_request_pending': diary_edit_request_pending,
        'diary_provisional_edit_unlocked': diary_provisional_edit_unlocked,
        'weather_morning': weather_morning,
        'weather_afternoon': weather_afternoon,
        'weather_night': weather_night,
        'labor_by_type': labor_by_type,
        'project_days_total': project_days_total,
        'project_days_elapsed': project_days_elapsed,
        'project_days_remaining': project_days_remaining,
        'total_indirect_labor': sum(labor_by_type['Indireto'].values()),
        'total_direct_labor': sum(labor_by_type['Direto'].values()),
        'total_third_party_labor': sum(labor_by_type['Terceiros'].values()),
        'equipment_list': equipment_list,
        'labor_entries_by_category': labor_entries_by_category,
        'weekday_name': weekday_name,
        'view_count': view_count,
        'edit_count': edit_count,
        'edit_logs': DiaryEditLog.objects.filter(diary=diary).select_related('edited_by')[:10],
        'signature_inspection': signature_inspection,
        'signature_production': signature_production,
        'owner_comments': list(diary.owner_comments.select_related('author').order_by('created_at')),
        'can_add_owner_comment': diary.is_approved() and (request.user.is_staff or request.user.groups.filter(name=GRUPOS.GERENTES).exists()),
        'can_decide_rdo_approval': can_decide_rdo_approval,
        'approval_history': approval_history,
        'rdo_approvers': rdo_approvers,
        'diary_nav_prev_pk': nav_prev_pk,
        'diary_nav_next_pk': nav_next_pk,
        'diary_nav_position': nav_position,
        'diary_nav_total': nav_total,
    }
    
    # Força retorno HTML explícito (evita conflito com API REST)
    # Remove qualquer header que possa fazer o DRF interceptar
    response = render(request, 'core/diary_detail.html', context)
    response['Content-Type'] = 'text/html; charset=utf-8'
    # Remove headers que podem fazer o DRF pensar que é JSON
    if 'Content-Type' in response:
        response['Content-Type'] = 'text/html; charset=utf-8'
    # Garante que não seja tratado como API
    response['X-Content-Type-Options'] = 'nosniff'
    return response


# ==================== PORTAL DO DONO DA OBRA (CLIENTE) ====================

def _client_can_access_diary(user, diary):
    """True se o usuário é dono da obra do diário (pode acessar a página cliente)."""
    return ProjectOwner.objects.filter(project=diary.project, user=user).exists()


def _client_can_comment(diary):
    """True se ainda está na janela de 24h úteis para enviar comentários."""
    deadline = _client_comment_deadline(diary)
    return bool(deadline and timezone.now() <= deadline)


def _client_comment_deadline(diary):
    """
    Calcula o prazo de comentários em 24 horas úteis.
    Sábado e domingo não contam para a contagem.
    """
    if not diary.sent_to_owner_at:
        return None

    hours_remaining = 24.0
    current = diary.sent_to_owner_at

    # Regra operacional: envio muito tarde na sexta-feira começa a contar na segunda.
    if current.weekday() == 4 and current.hour >= 18:
        current = (current + timedelta(days=3)).replace(hour=0, minute=0, second=0, microsecond=0)

    while hours_remaining > 0:
        # Pula finais de semana inteiros sem consumir horas.
        while current.weekday() >= 5:  # 5=sábado, 6=domingo
            current = (current + timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)

        next_midnight = (current + timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
        available_today = (next_midnight - current).total_seconds() / 3600.0

        step = min(hours_remaining, available_today)
        current = current + timedelta(hours=step)
        hours_remaining -= step

    return current


def _client_comment_rate_limited(request, diary_id, window_seconds=12):
    """
    Limite simples anti-spam por sessão/diário.
    Evita múltiplos envios em sequência muito rápida.
    """
    now_ts = int(timezone.now().timestamp())
    key = f'client_comment_last_{diary_id}'
    last_ts = request.session.get(key)
    if last_ts and now_ts - int(last_ts) < window_seconds:
        return True
    request.session[key] = now_ts
    return False


@login_required
def client_diary_list_view(request):
    """Lista de diários disponíveis para o dono da obra (só obras que ele possui)."""
    if not _is_work_owner(request.user):
        raise Http404("Acesso restrito.")
    projects = _get_projects_for_user(request)
    from datetime import timedelta
    now = timezone.now()
    diaries_by_project = []
    for project in projects:
        diaries = (
            ConstructionDiary.objects.filter(
                project=project,
                status=DiaryStatus.APROVADO,
                sent_to_owner_at__isnull=False,
            )
            .order_by('-date')[:10]
        )
        enriched = []
        for d in diaries:
            deadline = _client_comment_deadline(d)
            enriched.append({
                'diary': d,
                'can_comment': bool(deadline and now <= deadline),
                'deadline': deadline,
                'comment_count': d.owner_comments.count(),
            })
        if enriched:
            diaries_by_project.append({'project': project, 'diaries': enriched})
    context = {
        'diaries_by_project': diaries_by_project,
        'user': request.user,
    }
    return render(request, 'core/client_diary_list.html', context)


@login_required
def client_diary_detail_view(request, pk):
    """Visualização do diário pelo dono da obra: leitura + comentários (24h)."""
    diary = get_object_or_404(
        ConstructionDiary.objects.select_related('project', 'created_by', 'reviewed_by').prefetch_related(
            'images', 'videos', 'work_logs__activity', 'work_logs__resources_labor', 'work_logs__resources_equipment',
            'owner_comments__author',
        ),
        pk=pk,
    )
    if not _client_can_access_diary(request.user, diary):
        # Usuários internos (staff/membro da obra) podem cair aqui por link do cliente.
        # Nesses casos, leva para a tela interna em vez de retornar 404.
        if _user_can_access_project(request.user, diary.project):
            # Evita colisão de nome de rota com a API (diary-detail do DRF),
            # que pode causar loop de redirects no portal do cliente.
            return redirect(f'/diaries/{pk}/')
        raise Http404("Você não tem acesso a este diário.")
    if diary.status != DiaryStatus.APROVADO:
        raise Http404("Diário não disponível para visualização.")
    comments = list(diary.owner_comments.select_related('author').order_by('created_at'))
    can_comment = _client_can_comment(diary)
    comment_deadline = _client_comment_deadline(diary)
    # Contexto mínimo para exibir o diário (reutiliza lógica de clima/labor se necessário)
    labor_by_type = {'Direto': {}, 'Indireto': {}, 'Terceiros': {}}
    for work_log in diary.work_logs.all():
        for labor in work_log.resources_labor.all():
            labor_type = labor.labor_type
            if labor_type == 'D':
                labor_by_type['Direto'][labor.name] = labor_by_type['Direto'].get(labor.name, 0) + 1
            elif labor_type == 'I':
                labor_by_type['Indireto'][labor.name] = labor_by_type['Indireto'].get(labor.name, 0) + 1
            elif labor_type == 'T':
                labor_by_type['Terceiros'][labor.name] = labor_by_type['Terceiros'].get(labor.name, 0) + 1
    equipment_list = _build_diary_equipment_list(diary)
    project = diary.project
    project_days_total = (project.end_date - project.start_date).days if project.end_date and project.start_date else 0
    project_days_elapsed = (diary.date - project.start_date).days if project.end_date and project.start_date else 0
    project_days_remaining = (project.end_date - diary.date).days if project.end_date and project.start_date else 0
    from datetime import datetime as dt
    weekday_names = ['Segunda-Feira', 'Terça-Feira', 'Quarta-Feira', 'Quinta-Feira', 'Sexta-Feira', 'Sábado', 'Domingo']
    weekday_name = weekday_names[diary.date.weekday()] if diary.date else ''
    context = {
        'diary': diary,
        'project': project,
        'comments': comments,
        'can_comment': can_comment,
        'comment_deadline': comment_deadline,
        'labor_by_type': labor_by_type,
        'equipment_list': equipment_list,
        'project_days_total': project_days_total,
        'project_days_elapsed': project_days_elapsed,
        'project_days_remaining': project_days_remaining,
        'weekday_name': weekday_name,
        'total_indirect_labor': sum(labor_by_type['Indireto'].values()),
        'total_direct_labor': sum(labor_by_type['Direto'].values()),
        'total_third_party_labor': sum(labor_by_type['Terceiros'].values()),
    }
    return render(request, 'core/client_diary_detail.html', context)


@login_required
@project_required
@require_http_methods(["POST"])
def diary_add_owner_comment_view(request, pk):
    """POST: adiciona comentário da LPLAN no diário (staff/gerentes). Diário deve estar aprovado."""
    diary = get_object_or_404(ConstructionDiary.objects.select_related('project'), pk=pk)
    if diary.project_id != get_selected_project(request).id:
        raise Http404()
    if not diary.is_approved():
        messages.error(request, "Só é possível comentar em diários aprovados.")
        return redirect('diary-detail', pk=pk)
    if not (request.user.is_staff or request.user.groups.filter(name=GRUPOS.GERENTES).exists()):
        raise PermissionDenied("Sem permissão para comentar.")
    text = (request.POST.get('text') or '').strip()
    if not text:
        messages.error(request, "Escreva um comentário.")
        return redirect('diary-detail', pk=pk)
    DiaryComment.objects.create(diary=diary, author=request.user, text=text)
    messages.success(request, "Comentário enviado. O dono da obra poderá ver na página de visualização.")
    return redirect('diary-detail', pk=pk)


@login_required
@project_required
@require_http_methods(["POST"])
def diary_delete_view(request, pk):
    """Exclui um relatório (diário) de obra. Apenas superuser pode excluir."""
    diary = get_object_or_404(ConstructionDiary.objects.select_related('project'), pk=pk)
    if diary.project_id != get_selected_project(request).id:
        raise Http404()
    if not request.user.is_superuser:
        raise PermissionDenied("Apenas administradores podem excluir relatórios.")
    date_str = diary.date.strftime('%d/%m/%Y')
    diary.delete()
    messages.success(request, f"Relatório de {date_str} foi excluído.")
    return redirect('report-list')


@login_required
@project_required
@require_http_methods(['POST'])
def diary_request_edit_view(request, pk):
    """
    Pedido de correção em relatório aprovado: regista o pedido para o staff liberar edição provisória.
    """
    diary = get_object_or_404(ConstructionDiary.objects.select_related('project'), pk=pk)
    if diary.project_id != get_selected_project(request).id:
        raise Http404()
    if not _user_can_access_project(request.user, diary.project):
        raise Http404()
    if not diary.is_approved():
        messages.error(request, 'Só é possível pedir correção em relatórios aprovados.')
        return redirect('diary-detail', pk=pk)
    if diary.provisional_edit_granted_at:
        messages.info(request, 'Este relatório já tem edição liberada.')
        return redirect('diary-detail', pk=pk)
    if diary.edit_requested_at and not diary.provisional_edit_granted_at:
        messages.info(request, 'Já existe um pedido de correção pendente.')
        return redirect('diary-detail', pk=pk)
    note = (request.POST.get('note') or '').strip()[:2000]
    now = timezone.now()
    ConstructionDiary.objects.filter(pk=diary.pk).update(
        edit_requested_at=now,
        edit_requested_by=request.user,
        edit_request_note=note,
    )
    DiaryCorrectionRequestLog.objects.create(
        diary_id=diary.pk,
        requested_at=now,
        requested_by=request.user,
        note=note,
    )
    messages.success(
        request,
        'Pedido de correção enviado. Quando um administrador liberar, o botão Editar ficará disponível.',
    )
    return redirect('diary-detail', pk=pk)


@login_required
@require_http_methods(["POST"])
def diary_review_decision_view(request, pk):
    """
    Aprovação/reprovação do RDO por aprovador da obra.
    Só se aplica quando o diário está em AGUARDANDO_APROVACAO_GESTOR.
    """
    diary = get_object_or_404(ConstructionDiary.objects.select_related('project'), pk=pk)
    if not _user_can_access_project(request.user, diary.project):
        raise Http404()
    if not _is_project_rdo_approver(request.user, diary.project):
        raise PermissionDenied("Você não tem permissão para decidir a aprovação deste RDO.")
    if diary.status != DiaryStatus.AGUARDANDO_APROVACAO_GESTOR:
        messages.info(request, 'Este RDO não está pendente de aprovação.')
        return redirect('diary-detail', pk=pk)

    decision = (request.POST.get('decision') or '').strip().lower()
    comment = (request.POST.get('comment') or '').strip()[:2000]
    now = timezone.now()

    if decision == 'approve':
        diary.status = DiaryStatus.APROVADO
        diary.reviewed_by = request.user
        diary.approved_at = now
        if not diary.sent_to_owner_at:
            diary.sent_to_owner_at = now
        diary.save(update_fields=['status', 'reviewed_by', 'approved_at', 'sent_to_owner_at', 'updated_at'])

        DiaryApprovalHistory.objects.create(
            diary=diary,
            decided_by=request.user,
            decision=DiaryApprovalHistory.DECISAO_APROVAR,
            comment=comment,
        )
        try:
            from .diary_email import send_diary_to_owners, send_diary_pdf_to_recipients
            send_diary_to_owners(diary)
            send_diary_pdf_to_recipients(diary)
        except Exception as exc:
            logger.exception("Erro ao enviar RDO aprovado aos destinatários: %s", exc)
        messages.success(request, 'RDO aprovado e enviado ao cliente.')
        return redirect('diary-detail', pk=pk)

    if decision == 'reject':
        if not comment:
            messages.error(request, 'Informe o motivo da reprovação.')
            return redirect('diary-detail', pk=pk)
        diary.status = DiaryStatus.REPROVADO_GESTOR
        diary.reviewed_by = request.user
        diary.approved_at = None
        diary.sent_to_owner_at = None
        diary.save(update_fields=['status', 'reviewed_by', 'approved_at', 'sent_to_owner_at', 'updated_at'])

        DiaryApprovalHistory.objects.create(
            diary=diary,
            decided_by=request.user,
            decision=DiaryApprovalHistory.DECISAO_REPROVAR,
            comment=comment,
        )
        messages.success(request, 'RDO reprovado. O responsável poderá ajustar e reenviar para aprovação.')
        return redirect('diary-detail', pk=pk)

    messages.error(request, 'Decisão inválida.')
    return redirect('diary-detail', pk=pk)


@login_required
@require_http_methods(["POST"])
def client_diary_add_comment_view(request, pk):
    """POST: adiciona comentário ao diário (apenas dono da obra, dentro da janela de 24h)."""
    diary = get_object_or_404(ConstructionDiary.objects.select_related('project'), pk=pk)
    if not _client_can_access_diary(request.user, diary):
        raise PermissionDenied("Você não tem acesso a este diário.")
    if diary.status != DiaryStatus.APROVADO:
        raise PermissionDenied("Diário não disponível para comentário.")
    if not _client_can_comment(diary):
        messages.error(request, "O prazo de 24 horas úteis para enviar comentários foi encerrado.")
        return redirect('client-diary-detail', pk=pk)
    if _client_comment_rate_limited(request, diary.pk):
        messages.warning(request, "Aguarde alguns segundos antes de enviar outro comentário.")
        return redirect('client-diary-detail', pk=pk)
    text = (request.POST.get('text') or '').strip()
    if not text:
        messages.error(request, "Escreva um comentário.")
        return redirect('client-diary-detail', pk=pk)
    DiaryComment.objects.create(diary=diary, author=request.user, text=text)
    messages.success(request, "Comentário enviado.")
    return redirect('client-diary-detail', pk=pk)


# ==================== FILTROS DE BUSCA ====================

@login_required
@project_required
def filter_photos_view(request):
    """View para filtro de fotos."""
    project = get_selected_project(request)
    
    # Busca todas as fotos do projeto
    photos = DiaryImage.objects.filter(
        diary__project=project,
        is_approved_for_report=True
    ).select_related('diary').order_by('-diary__date', '-uploaded_at')
    
    # Filtros
    search = request.GET.get('search', '')
    date_start = request.GET.get('date_start')
    date_end = request.GET.get('date_end')
    diary_id = request.GET.get('diary_id')
    
    if search:
        photos = photos.filter(
            Q(caption__icontains=search) |
            Q(diary__project__name__icontains=search) |
            Q(diary__project__code__icontains=search)
        )
    
    if date_start:
        try:
            photos = photos.filter(diary__date__gte=date_start)
        except ValueError:
            pass
    
    if date_end:
        try:
            photos = photos.filter(diary__date__lte=date_end)
        except ValueError:
            pass
    
    if diary_id:
        photos = photos.filter(diary_id=diary_id)
    
    # Remove itens com arquivo ausente/inacessível para evitar 404/500 no carregamento da galeria.
    visible_photos = []
    for photo in photos:
        try:
            if photo.image and photo.image.name and photo.image.storage.exists(photo.image.name):
                visible_photos.append(photo)
        except Exception:
            # Se o storage falhar para um item específico, omite o card sem quebrar a página.
            continue

    # Paginação
    from django.core.paginator import Paginator
    paginator = Paginator(visible_photos, 24)  # 24 fotos por página (grid 4x6)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    # Estatísticas
    total_photos = len(visible_photos)
    by_date_counter = {}
    for photo in visible_photos:
        day = getattr(photo.diary, 'date', None)
        by_date_counter[day] = by_date_counter.get(day, 0) + 1
    photos_by_date = [{'diary__date': d, 'count': c} for d, c in sorted(by_date_counter.items(), key=lambda x: (x[0] is None, x[0]), reverse=True)[:10]]
    
    context = {
        'photos': page_obj,
        'total_photos': total_photos,
        'photos_by_date': photos_by_date,
        'search': search,
        'date_start': date_start or '',
        'date_end': date_end or '',
        'diary_id': diary_id,
    }
    
    return render(request, 'core/filters/photos.html', context)


@login_required
@project_required
def filter_videos_view(request):
    """View para filtro de vídeos."""
    project = get_selected_project(request)
    from core.models import DiaryVideo
    
    # Busca todos os vídeos do projeto
    videos = DiaryVideo.objects.filter(
        diary__project=project
    ).select_related('diary').order_by('-diary__date', '-uploaded_at')
    
    # Filtros
    search = request.GET.get('search', '')
    date_start = request.GET.get('date_start')
    date_end = request.GET.get('date_end')
    diary_id = request.GET.get('diary_id')
    
    if search:
        videos = videos.filter(
            Q(caption__icontains=search) |
            Q(diary__project__name__icontains=search) |
            Q(diary__project__code__icontains=search)
        )
    
    if date_start:
        try:
            videos = videos.filter(diary__date__gte=date_start)
        except ValueError:
            pass
    
    if date_end:
        try:
            videos = videos.filter(diary__date__lte=date_end)
        except ValueError:
            pass
    
    if diary_id:
        videos = videos.filter(diary_id=diary_id)
    
    # Paginação
    from django.core.paginator import Paginator
    paginator = Paginator(videos, 24)  # 24 vídeos por página
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    # Estatísticas
    total_videos = videos.count()
    videos_by_date = videos.values('diary__date').annotate(count=Count('id')).order_by('-diary__date')[:10]
    
    context = {
        'videos': page_obj,
        'total_videos': total_videos,
        'videos_by_date': videos_by_date,
        'search': search,
        'date_start': date_start or '',
        'date_end': date_end or '',
        'diary_id': diary_id,
    }
    
    return render(request, 'core/filters/videos.html', context)


@login_required
@project_required
def filter_activities_view(request):
    """View para filtro de atividades."""
    project = get_selected_project(request)
    
    # Busca todas as atividades do projeto
    activities = Activity.objects.filter(project=project).order_by('path')
    
    # Filtros
    search = request.GET.get('search', '')
    status_filter = request.GET.get('status')
    progress_min = request.GET.get('progress_min')
    progress_max = request.GET.get('progress_max')
    
    if search:
        activities = activities.filter(
            Q(name__icontains=search) |
            Q(code__icontains=search) |
            Q(description__icontains=search)
        )
    
    if status_filter:
        activities = activities.filter(status=status_filter)
    
    # Filtro por progresso (se implementado)
    # TODO: Implementar filtro por progresso quando o campo estiver disponível
    
    # Estatísticas
    total_activities = activities.count()
    activities_by_status = activities.values('status').annotate(count=Count('id'))
    
    context = {
        'activities': activities,
        'total_activities': total_activities,
        'activities_by_status': activities_by_status,
        'search': search,
        'status_filter': status_filter,
        'progress_min': progress_min,
        'progress_max': progress_max,
    }
    
    return render(request, 'core/filters/activities.html', context)


@login_required
@project_required
def filter_occurrences_view(request):
    """View para filtro de ocorrências (modelo DiaryOccurrence + campo incidents)."""
    from .models import DiaryOccurrence
    from django.core.paginator import Paginator

    project = get_selected_project(request)

    # Ocorrências do formulário (DiaryOccurrence) – mesma fonte que o dashboard
    occurrences = DiaryOccurrence.objects.filter(
        diary__project=project
    ).select_related('diary', 'diary__project').prefetch_related('tags').order_by('-diary__date', '-created_at')

    # Filtros
    search = request.GET.get('search', '')
    date_start = request.GET.get('date_start')
    date_end = request.GET.get('date_end')

    if search:
        occurrences = occurrences.filter(description__icontains=search)
    if date_start:
        try:
            occurrences = occurrences.filter(diary__date__gte=date_start)
        except ValueError:
            pass
    if date_end:
        try:
            occurrences = occurrences.filter(diary__date__lte=date_end)
        except ValueError:
            pass

    # Estatísticas (total e por data, com os mesmos filtros)
    total_occurrences = occurrences.count()
    occurrences_by_date = list(
        occurrences.values('diary__date')
        .annotate(count=Count('id'))
        .order_by('-diary__date')[:10]
    )

    # Paginação
    paginator = Paginator(occurrences, 20)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    context = {
        'occurrences': page_obj,
        'total_occurrences': total_occurrences,
        'occurrences_by_date': occurrences_by_date,
        'search': search,
        'date_start': date_start,
        'date_end': date_end,
    }

    return render(request, 'core/filters/occurrences.html', context)


@login_required
@project_required
def filter_comments_view(request):
    """View para filtro de comentários."""
    project = get_selected_project(request)
    
    # Busca comentários (general_notes) de todos os diários do projeto
    diaries = ConstructionDiary.objects.filter(
        project=project
    ).exclude(general_notes='').exclude(general_notes__isnull=True)
    
    # Filtros
    search = request.GET.get('search', '')
    date_start = request.GET.get('date_start')
    date_end = request.GET.get('date_end')
    
    if search:
        diaries = diaries.filter(general_notes__icontains=search)
    
    if date_start:
        try:
            diaries = diaries.filter(date__gte=date_start)
        except ValueError:
            pass
    
    if date_end:
        try:
            diaries = diaries.filter(date__lte=date_end)
        except ValueError:
            pass
    
    # Paginação
    from django.core.paginator import Paginator
    paginator = Paginator(diaries, 20)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    # Estatísticas
    total_comments = diaries.count()
    
    context = {
        'diaries': page_obj,
        'total_comments': total_comments,
        'search': search,
        'date_start': date_start,
        'date_end': date_end,
    }
    
    return render(request, 'core/filters/comments.html', context)


@login_required
@project_required
def filter_attachments_view(request):
    """View para filtro de anexos."""
    from .models import DiaryAttachment
    
    project = get_selected_project(request)
    
    # Busca anexos
    attachments = DiaryAttachment.objects.filter(
        diary__project=project
    ).select_related('diary').order_by('-diary__date', '-uploaded_at')
    
    # Filtros
    search = request.GET.get('search', '')
    date_start = request.GET.get('date_start')
    date_end = request.GET.get('date_end')
    
    if search:
        attachments = attachments.filter(
            Q(name__icontains=search) |
            Q(description__icontains=search) |
            Q(diary__project__name__icontains=search)
        )
    
    if date_start:
        try:
            attachments = attachments.filter(diary__date__gte=date_start)
        except ValueError:
            pass
    
    if date_end:
        try:
            attachments = attachments.filter(diary__date__lte=date_end)
        except ValueError:
            pass
    
    # Paginação
    from django.core.paginator import Paginator
    paginator = Paginator(attachments, 20)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    context = {
        'attachments': page_obj,
        'total_attachments': attachments.count(),
        'search': search,
        'date_start': date_start,
        'date_end': date_end,
    }
    
    return render(request, 'core/filters/attachments.html', context)


@login_required
@project_required
def weather_conditions_view(request):
    """View para análise de condições climáticas."""
    project = get_selected_project(request)
    
    # Busca todos os diários com condições climáticas
    diaries = ConstructionDiary.objects.filter(
        project=project
    ).exclude(weather_conditions='').exclude(weather_conditions__isnull=True).order_by('-date')
    
    # Filtros
    date_start = request.GET.get('date_start')
    date_end = request.GET.get('date_end')
    
    if date_start:
        try:
            diaries = diaries.filter(date__gte=date_start)
        except ValueError:
            pass
    
    if date_end:
        try:
            diaries = diaries.filter(date__lte=date_end)
        except ValueError:
            pass
    
    # Análise de condições climáticas
    weather_stats = {
        'sunny_days': 0,
        'cloudy_days': 0,
        'rainy_days': 0,
        'total_days': diaries.count(),
    }
    
    for diary in diaries:
        weather = diary.weather_conditions.lower()
        if 'sol' in weather or 'claro' in weather:
            weather_stats['sunny_days'] += 1
        elif 'nuvem' in weather or 'nublado' in weather:
            weather_stats['cloudy_days'] += 1
        elif 'chuva' in weather or 'chuvoso' in weather:
            weather_stats['rainy_days'] += 1
    
    # Estatísticas de chuva
    rain_diaries = diaries.exclude(rain_occurrence='').exclude(rain_occurrence__isnull=True)
    rain_stats = {
        'weak': rain_diaries.filter(rain_occurrence='F').count(),
        'medium': rain_diaries.filter(rain_occurrence='M').count(),
        'strong': rain_diaries.filter(rain_occurrence='S').count(),
        'total': rain_diaries.count(),
    }
    
    context = {
        'diaries': diaries[:50],  # Limita a 50 para performance
        'weather_stats': weather_stats,
        'rain_stats': rain_stats,
        'date_start': date_start,
        'date_end': date_end,
    }
    
    return render(request, 'core/filters/weather_conditions.html', context)


@login_required
@project_required
def labor_histogram_view(request):
    """View para histograma de mão de obra."""
    project = get_selected_project(request)
    
    # Busca todos os work_logs do projeto
    work_logs = DailyWorkLog.objects.filter(
        diary__project=project
    ).select_related('diary', 'activity').prefetch_related('resources_labor')
    
    # Filtros
    date_start = request.GET.get('date_start')
    date_end = request.GET.get('date_end')
    activity_id = request.GET.get('activity_id')
    
    if date_start:
        try:
            work_logs = work_logs.filter(diary__date__gte=date_start)
        except ValueError:
            pass
    
    if date_end:
        try:
            work_logs = work_logs.filter(diary__date__lte=date_end)
        except ValueError:
            pass
    
    if activity_id:
        work_logs = work_logs.filter(activity_id=activity_id)
    
    # Agrupa mão de obra por tipo e nome
    labor_stats = {
        'Direto': {},
        'Indireto': {},
        'Terceiros': {},
    }
    
    for work_log in work_logs:
        for labor in work_log.resources_labor.all():
            labor_type = 'Direto' if labor.labor_type == 'D' else ('Indireto' if labor.labor_type == 'I' else 'Terceiros')
            if labor.name not in labor_stats[labor_type]:
                labor_stats[labor_type][labor.name] = 0
            labor_stats[labor_type][labor.name] += 1
    
    # Estatísticas por data
    labor_by_date = {}
    for work_log in work_logs:
        date_key = work_log.diary.date.isoformat()
        if date_key not in labor_by_date:
            labor_by_date[date_key] = {'Direto': 0, 'Indireto': 0, 'Terceiros': 0}
        
        for labor in work_log.resources_labor.all():
            labor_type = 'Direto' if labor.labor_type == 'D' else ('Indireto' if labor.labor_type == 'I' else 'Terceiros')
            labor_by_date[date_key][labor_type] += 1
    
    context = {
        'labor_stats': labor_stats,
        'labor_by_date': labor_by_date,
        'date_start': date_start,
        'date_end': date_end,
        'activity_id': activity_id,
        'activities': Activity.objects.filter(project=project).order_by('name'),
    }
    
    return render(request, 'core/filters/labor_histogram.html', context)


@login_required
@project_required
def equipment_histogram_view(request):
    """View para histograma de equipamentos."""
    project = get_selected_project(request)
    
    # Busca todos os work_logs do projeto
    work_logs = DailyWorkLog.objects.filter(
        diary__project=project
    ).select_related('diary', 'activity').prefetch_related('resources_equipment')
    
    # Filtros
    date_start = request.GET.get('date_start')
    date_end = request.GET.get('date_end')
    activity_id = request.GET.get('activity_id')
    
    if date_start:
        try:
            work_logs = work_logs.filter(diary__date__gte=date_start)
        except ValueError:
            pass
    
    if date_end:
        try:
            work_logs = work_logs.filter(diary__date__lte=date_end)
        except ValueError:
            pass
    
    if activity_id:
        work_logs = work_logs.filter(activity_id=activity_id)
    
    # Agrupa por quantidade do through (DailyWorkLogEquipment), evitando contagem por presença.
    from .models import DailyWorkLogEquipment
    equipment_stats = {}
    equipment_by_date = {}
    through_rows = DailyWorkLogEquipment.objects.filter(work_log__in=work_logs).select_related('work_log__diary', 'equipment')
    for row in through_rows:
        eq_name = _normalize_equipment_name(getattr(row.equipment, 'name', ''))
        if not eq_name:
            continue
        qty = _safe_positive_int(getattr(row, 'quantity', 1), default=1, minimum=1)
        equipment_stats[eq_name] = equipment_stats.get(eq_name, 0) + qty
        date_key = row.work_log.diary.date.isoformat()
        if date_key not in equipment_by_date:
            equipment_by_date[date_key] = {}
        equipment_by_date[date_key][eq_name] = equipment_by_date[date_key].get(eq_name, 0) + qty
    
    context = {
        'equipment_stats': equipment_stats,
        'equipment_by_date': equipment_by_date,
        'date_start': date_start,
        'date_end': date_end,
        'activity_id': activity_id,
        'activities': Activity.objects.filter(project=project).order_by('name'),
    }
    
    return render(request, 'core/filters/equipment_histogram.html', context)


def _diary_form_context_from_post(request, project, form, image_formset, worklog_formset, occurrence_formset, diary=None):
    """
    Monta existing_diary_equipment e existing_diary_labor a partir do POST
    para repopular o formulário quando há erro de validação (evita perda de tags de equipamentos e mão de obra).
    """
    import json
    existing_diary_labor = []
    existing_diary_equipment = []
    try:
        labor_json = (request.POST.get('diary_labor_data') or '').strip()
        if labor_json and labor_json != '[]':
            data = json.loads(labor_json)
            if isinstance(data, list):
                for item in data:
                    if isinstance(item, dict) and (
                        'cargo_id' in item or 'cargoId' in item or 'cargo_name' in item or 'cargoName' in item
                    ):
                        existing_diary_labor.append({
                            'cargo_id': item.get('cargo_id') or item.get('cargoId'),
                            'cargo_name': (item.get('cargo_name') or item.get('cargoName') or '').strip(),
                            'quantity': int(item.get('quantity') or 1),
                            'company': (item.get('company') or '').strip(),
                        })
    except (json.JSONDecodeError, TypeError, ValueError):
        pass
    try:
        eq_json = (request.POST.get('equipment_data') or '').strip()
        if eq_json and eq_json != '[]':
            data = json.loads(eq_json)
            if isinstance(data, list):
                for item in data:
                    if isinstance(item, dict) and (item.get('name') or '').strip():
                        name = (item.get('name') or '').strip()
                        existing_diary_equipment.append({
                            'name': name,
                            'quantity': int(item.get('quantity') or 1),
                            'equipment_id': item.get('equipment_id'),
                            'standard_equipment_id': item.get('standard_equipment_id'),
                        })
    except (json.JSONDecodeError, TypeError, ValueError):
        pass
    from .models import OccurrenceTag, LaborCategory, EquipmentCategory
    try:
        occurrence_tags = OccurrenceTag.objects.filter(is_active=True)
    except Exception:
        occurrence_tags = []
    try:
        labor_categories = LaborCategory.objects.prefetch_related('cargos').order_by('order')
        labor_terceirizada_cargos = []
        for cat in labor_categories:
            if cat.slug == 'terceirizada':
                labor_terceirizada_cargos = [{'id': c.id, 'name': c.name} for c in cat.cargos.all()]
                break
    except Exception:
        labor_categories = []
        labor_terceirizada_cargos = []
    try:
        equipment_categories = EquipmentCategory.objects.prefetch_related('items').order_by('order')
    except Exception:
        equipment_categories = []
    next_report_number = None
    if project and (not diary or not diary.pk):
        last_d = ConstructionDiary.objects.filter(project=project).only('report_number').order_by('-report_number').first()
        if last_d and last_d.report_number:
            next_report_number = last_d.report_number + 1
        else:
            next_report_number = 1
    last_diary_for_copy = None
    if project:
        qs = ConstructionDiary.objects.filter(project=project).order_by('-date', '-report_number')
        if diary and getattr(diary, 'pk', None):
            qs = qs.exclude(pk=diary.pk)
        d = qs.only('pk', 'date', 'report_number').first()
        if d:
            try:
                date_str = d.date.strftime('%d/%m/%Y') if d.date else ''
            except Exception:
                date_str = str(d.date) if d.date else ''
            last_diary_for_copy = {'id': d.pk, 'date': date_str, 'report_number': d.report_number or '-'}
    signature_inspection_value = (request.POST.get('signature_inspection') or '').strip()
    signature_production_value = (request.POST.get('signature_production') or '').strip()
    if diary and getattr(diary, 'pk', None):
        if not signature_inspection_value:
            sig = diary.signatures.filter(signature_type='inspection').only('signature_data').first()
            signature_inspection_value = sig.signature_data if sig else ''
        if not signature_production_value:
            sig_prod = diary.signatures.filter(signature_type='production').only('signature_data').first()
            signature_production_value = sig_prod.signature_data if sig_prod else ''
    return {
        'diary': diary if diary and getattr(diary, 'pk', None) else None,
        'form': form,
        'image_formset': image_formset,
        'worklog_formset': worklog_formset,
        'occurrence_formset': occurrence_formset,
        'occurrence_tags': occurrence_tags,
        'labor_categories': labor_categories,
        'labor_terceirizada_cargos': labor_terceirizada_cargos,
        'existing_diary_labor': existing_diary_labor,
        'existing_diary_equipment': existing_diary_equipment,
        'equipment_categories': equipment_categories,
        'project': project,
        'next_report_number': next_report_number,
        'initial_contractante': request.POST.get('project_client_name') or (get_contractante_for_project(project) if project else ''),
        'last_diary_for_copy': last_diary_for_copy,
        'copy_from_id': '',
        'copy_options': '',
        'copy_source_diary': None,
        'signature_inspection_value': signature_inspection_value,
        'signature_production_value': signature_production_value,
    }


@login_required
@project_required
def diary_form_view(request, pk=None):
    """View de formulário de diário de obra."""
    import logging
    from django.forms import inlineformset_factory
    logger = logging.getLogger(__name__)
    from .forms import (
        ConstructionDiaryForm,
        DiaryImageFormSet,
        DailyWorkLogFormSet,
        DiaryOccurrenceFormSet,
        DailyWorkLogForm,
        DiaryOccurrenceForm,
    )
    from .services import ProgressService
    
    project = get_selected_project(request)
    
    if pk:
        # Otimiza queries com select_related e prefetch_related
        diary = get_object_or_404(
            ConstructionDiary.objects.select_related('project', 'created_by', 'reviewed_by')
            .prefetch_related('images', 'videos', 'attachments', 'work_logs__activity', 'occurrences__tags'),
            pk=pk, 
            project=project
        )
        if not diary.can_be_edited_by(request.user):
            messages.warning(request, 'Este diário não pode ser editado no momento.')
            return redirect('diary-detail', pk=pk)
    else:
        diary = None
    
    copy_source_diary = None
    copy_opts_list = []
    copy_from_id = None
    copy_options_raw = ''
    
    if request.method == 'POST':
        # Liberação provisória: após guardar com sucesso, limpa estes campos (captura antes do POST)
        had_provisional_unlock = bool(
            diary and diary.pk and getattr(diary, 'provisional_edit_granted_at', None)
        )
        # Verifica permissão de edição antes de processar (se for edição)
        if diary and not diary.can_be_edited_by(request.user):
            flash_message(request, "error", "core.diary.edit.no_permission")
            return redirect('diary-detail', pk=pk)
        
        # Valida que projeto existe
        if not project:
            flash_message(request, "error", "core.diary.project_not_selected")
            return redirect('select-project')
        
        form = ConstructionDiaryForm(request.POST, instance=diary, user=request.user, project=project)
        if request.FILES:
            for key, file in request.FILES.items():
                logger.info(f"  Arquivo: {key} -> {file.name} ({file.size} bytes)")
        
        # IMPORTANTE: Preserva request.FILES antes de qualquer processamento
        # request.FILES pode ser consumido apenas uma vez, então precisamos preservá-lo
        files_dict = {}
        if request.FILES:
            # Cria uma cópia dos arquivos para poder reutilizar
            for key, file_obj in request.FILES.items():
                # Cria uma cópia do arquivo em memória
                from django.core.files.uploadedfile import InMemoryUploadedFile
                from io import BytesIO
                
                # Lê o conteúdo do arquivo
                file_obj.seek(0)  # Garante que está no início
                file_content = file_obj.read()
                file_obj.seek(0)  # Volta ao início para uso posterior
                
                # Cria uma nova instância do arquivo
                files_dict[key] = InMemoryUploadedFile(
                    BytesIO(file_content),
                    None,
                    file_obj.name,
                    file_obj.content_type,
                    file_obj.size,
                    file_obj.charset
                )
        
        # Cria formsets iniciais (antes de salvar o diário)
        image_formset = DiaryImageFormSet(
            request.POST,
            request.FILES,
            instance=diary if diary else None
        )
        
        logger.info(f"Formset criado. Total de forms: {image_formset.total_form_count()}")
        logger.info(f"INITIAL_FORMS: {image_formset.initial_form_count()}")
        logger.info(f"TOTAL_FORMS do POST: {request.POST.get('diaryimage_set-TOTAL_FORMS', 'N/A')}")
        logger.info(f"Arquivos em request.FILES: {list(request.FILES.keys())}")
        
        worklog_formset = DailyWorkLogFormSet(
            request.POST,
            instance=diary if diary else None,
            form_kwargs={'diary': diary if diary else None},
            prefix='work_logs'
        )
        # POST para ocorrências: aceita tanto 'ocorrencias-*' quanto 'occurrences-*' (normaliza para ocorrencias)
        from copy import deepcopy
        _post_occ = deepcopy(request.POST)
        if hasattr(_post_occ, '_mutable'):
            _post_occ._mutable = True
        if 'occurrences-TOTAL_FORMS' in request.POST and 'ocorrencias-TOTAL_FORMS' not in _post_occ:
            for key, value in request.POST.items():
                if key.startswith('occurrences-'):
                    _post_occ[key.replace('occurrences-', 'ocorrencias-', 1)] = value
        occurrence_formset = DiaryOccurrenceFormSet(
            _post_occ,
            instance=diary if diary else None,
            prefix='ocorrencias'
        )
        if not worklog_formset.is_valid():
            pass  # erros em worklog_formset.errors
        if not occurrence_formset.is_valid():
            pass  # erros em occurrence_formset.errors

        # Valida o form primeiro
        import logging
        logger = logging.getLogger(__name__)
        
        if form.is_valid():
            logger.info(f"Form principal válido. Salvando diário...")
            # Se o form for válido, salva o diário primeiro (mesmo que seja None)
            # Isso é necessário para que o formset tenha uma instância válida
            diary = form.save(commit=False)

            # Valida que diary foi criado corretamente (não deve ser None)
            if diary is None:
                logger.error("Form.save(commit=False) retornou None! Isso não deveria acontecer.")
                flash_message(request, "error", "core.diary.form_process_error")
                # Retorna para o formulário com erros
                form = ConstructionDiaryForm(request.POST, instance=diary, user=request.user, project=project)
                if diary and diary.pk:
                    image_formset = DiaryImageFormSet(instance=diary)
                    worklog_formset = DailyWorkLogFormSet(instance=diary, form_kwargs={'diary': diary}, prefix='work_logs')
                    occurrence_formset = DiaryOccurrenceFormSet(instance=diary, prefix='ocorrencias')
                else:
                    image_formset = DiaryImageFormSet()
                    worklog_formset = DailyWorkLogFormSet(form_kwargs={'diary': None}, prefix='work_logs')
                    occurrence_formset = DiaryOccurrenceFormSet(prefix='ocorrencias')
                context = _diary_form_context_from_post(request, project, form, image_formset, worklog_formset, occurrence_formset, diary)
                return render(request, 'core/daily_log_form.html', context)
            
            is_new = not diary.pk if diary else True
            logger.info(f"Diário {'novo' if is_new else 'existente'} (pk={diary.pk if diary else None})")
            
            # IMPORTANTE: Prepara o diário mas NÃO salva ainda
            # O diário será salvo dentro da transação atomic() após validar os formsets
            if diary:
                is_partial_save = (
                    _is_truthy_flag(request.POST.get('partial_save')) or
                    _is_truthy_flag(request.POST.get('as_partial_checkbox'))
                )
                if not diary.pk:
                    diary.created_by = request.user
                    diary.project = project  # Associa à obra selecionada
                elif getattr(diary, 'created_by_id', None) is None:
                    # Diário já existe mas sem criador (legado): define para permitir edição
                    diary.created_by = request.user
                if is_partial_save:
                    diary.status = DiaryStatus.SALVAMENTO_PARCIAL
                    logger.info("Salvamento parcial: status definido como SALVAMENTO_PARCIAL")
                else:
                    has_project_approver = diary.project.rdo_approvers.filter(is_active=True).exists()
                    if has_project_approver:
                        # Novo fluxo: aguarda aprovação antes de enviar ao cliente.
                        diary.status = DiaryStatus.AGUARDANDO_APROVACAO_GESTOR
                        diary.approved_at = None
                        diary.reviewed_by = None
                        diary.sent_to_owner_at = None
                        logger.info("Salvar diário: status definido como AGUARDANDO_APROVACAO_GESTOR")
                    else:
                        # Compatibilidade: sem aprovadores cadastrados, mantém fluxo direto.
                        diary.status = DiaryStatus.APROVADO
                        diary.approved_at = timezone.now()
                        if not diary.sent_to_owner_at:
                            diary.sent_to_owner_at = timezone.now()
                        diary.reviewed_by = request.user
                        logger.info("Salvar diário: status definido como APROVADO (sem aprovadores configurados)")
            
            # Recria os formsets ANTES de salvar o diário
            # Isso permite validar os formsets antes de criar o diário no banco
            # Usa a cópia dos arquivos preservada
            from django.utils.datastructures import MultiValueDict
            from copy import deepcopy
            
            # Normaliza os dados do POST para usar o prefixo correto do formset
            # O formset espera 'work_logs' mas os testes podem enviar 'dailyworklog_set'
            normalized_post = deepcopy(request.POST)
            if hasattr(normalized_post, '_mutable'):
                normalized_post._mutable = True
            
            # Converte dailyworklog_set para work_logs se necessário
            worklog_prefix_old = 'dailyworklog_set'
            worklog_prefix_new = 'work_logs'
            
            # Verifica se há dados com o prefixo antigo (dailyworklog_set); só normaliza se
            # o front ainda enviar esse prefixo (ex.: cache). O formset usa prefix='work_logs',
            # então o esperado é work_logs-* no POST; não sobrescrever work_logs com valores antigos.
            if f'{worklog_prefix_old}-TOTAL_FORMS' in request.POST and f'{worklog_prefix_new}-TOTAL_FORMS' not in request.POST:
                total_forms = request.POST.get(f'{worklog_prefix_old}-TOTAL_FORMS', '0')
                initial_forms = request.POST.get(f'{worklog_prefix_old}-INITIAL_FORMS', '0')
                normalized_post[f'{worklog_prefix_new}-TOTAL_FORMS'] = total_forms
                normalized_post[f'{worklog_prefix_new}-INITIAL_FORMS'] = initial_forms
                normalized_post[f'{worklog_prefix_new}-MIN_NUM_FORMS'] = request.POST.get(f'{worklog_prefix_old}-MIN_NUM_FORMS', '0')
                normalized_post[f'{worklog_prefix_new}-MAX_NUM_FORMS'] = request.POST.get(f'{worklog_prefix_old}-MAX_NUM_FORMS', '1000')
                for key, value in request.POST.items():
                    if key.startswith(f'{worklog_prefix_old}-'):
                        new_key = key.replace(f'{worklog_prefix_old}-', f'{worklog_prefix_new}-', 1)
                        normalized_post[new_key] = value
            
            # Normaliza occurrences -> ocorrencias (formset usa prefix 'ocorrencias')
            if 'occurrences-TOTAL_FORMS' in request.POST and 'ocorrencias-TOTAL_FORMS' not in normalized_post:
                for key, value in request.POST.items():
                    if key.startswith('occurrences-'):
                        new_key = key.replace('occurrences-', 'ocorrencias-', 1)
                        normalized_post[new_key] = value
            
            # Garante TOTAL_FORMS mínimo se o POST tiver dados de atividades/ocorrências (evita perder dados)
            wl_total = int(normalized_post.get('work_logs-TOTAL_FORMS', '0'))
            occ_total = int(normalized_post.get('ocorrencias-TOTAL_FORMS', '0'))
            for key in request.POST:
                if key.startswith('work_logs-'):
                    parts = key.split('-')
                    if len(parts) >= 2 and parts[1].isdigit():
                        idx = int(parts[1])
                        if idx + 1 > wl_total:
                            wl_total = idx + 1
                if key.startswith('ocorrencias-'):
                    parts = key.split('-')
                    if len(parts) >= 2 and parts[1].isdigit():
                        idx = int(parts[1])
                        if idx + 1 > occ_total:
                            occ_total = idx + 1
            if wl_total > int(normalized_post.get('work_logs-TOTAL_FORMS', '0')):
                normalized_post['work_logs-TOTAL_FORMS'] = str(wl_total)
            if occ_total > int(normalized_post.get('ocorrencias-TOTAL_FORMS', '0')):
                normalized_post['ocorrencias-TOTAL_FORMS'] = str(occ_total)
            
            # Cria um novo MultiValueDict com os arquivos preservados
            preserved_files = MultiValueDict()
            for key, file_obj in files_dict.items():
                preserved_files.appendlist(key, file_obj)
            
            image_formset = DiaryImageFormSet(
                normalized_post,
                preserved_files,
                instance=diary
            )
            worklog_formset = DailyWorkLogFormSet(
                normalized_post,
                instance=diary,
                form_kwargs={'diary': diary},
                prefix='work_logs'
            )
            occurrence_formset = DiaryOccurrenceFormSet(
                normalized_post,
                instance=diary,
                prefix='ocorrencias'
            )
            
            # IMPORTANTE: Formsets inline precisam de instância salva (com PK)
            # Vamos salvar o diário dentro da transação primeiro, depois validar os formsets
            # Se os formsets críticos falharem, lançamos exceção para fazer rollback
            from django.db import transaction
            from core.models import DiarySignature
            
            try:
                with transaction.atomic():
                    # Inicializa variáveis de validação (serão atualizadas dentro do bloco condicional)
                    image_valid = False
                    worklog_valid = False
                    occurrence_valid = False
                    
                    # 0. SALVA O DIÁRIO PRIMEIRO (dentro da transação para permitir rollback se necessário)
                    # IMPORTANTE: Persiste sempre (novo ou edição) para gravar dados do form e status (ex.: Salvamento Parcial)
                    if diary:
                        try:
                            diary.save()
                            logger.info(f"Diário salvo (id={diary.pk}) para validação de formsets e processamento de dados.")
                        except Exception as e:
                            logger.error(f"ERRO ao salvar diário: {e}", exc_info=True)
                            messages.error(request, f'Erro ao salvar diário: {str(e)}')
                            raise  # Re-raise para fazer rollback da transação
                    
                    # Recria os formsets com a instância salva (agora tem PK)
                    if diary and diary.pk:
                        image_formset = DiaryImageFormSet(
                            normalized_post,
                            preserved_files,
                            instance=diary
                        )
                        worklog_formset = DailyWorkLogFormSet(
                            normalized_post,
                            instance=diary,
                            form_kwargs={'diary': diary},
                            prefix='work_logs'
                        )
                        occurrence_formset = DiaryOccurrenceFormSet(
                            normalized_post,
                            instance=diary,
                            prefix='ocorrencias'
                        )

                        # Re-valida os formsets agora que o diário tem PK
                        total_image_forms = int(normalized_post.get('diaryimage_set-TOTAL_FORMS', '0'))
                        total_worklog_forms = int(normalized_post.get('work_logs-TOTAL_FORMS', '0'))
                        total_occurrence_forms = int(normalized_post.get('ocorrencias-TOTAL_FORMS', '0'))
                        
                        if total_image_forms == 0:
                            image_valid_final = True
                        else:
                            has_image_data = False
                            for i in range(total_image_forms):
                                image_key = f'diaryimage_set-{i}-image'
                                caption_key = f'diaryimage_set-{i}-caption'
                                id_key = f'diaryimage_set-{i}-id'
                                if (image_key in preserved_files and preserved_files.get(image_key)) or \
                                   (caption_key in normalized_post and normalized_post.get(caption_key, '').strip()) or \
                                   (id_key in normalized_post and normalized_post.get(id_key, '').strip()):
                                    has_image_data = True
                                    break
                            image_valid_final = True if not has_image_data else image_formset.is_valid()
                        
                        if total_worklog_forms == 0:
                            worklog_valid_final = True
                        else:
                            has_worklog_data = False
                            for i in range(total_worklog_forms):
                                # Verifica múltiplos campos para detectar se há dados
                                activity_key = f'work_logs-{i}-activity_description'
                                location_key = f'work_logs-{i}-location'
                                notes_key = f'work_logs-{i}-notes'
                                percentage_key = f'work_logs-{i}-percentage_executed_today'
                                progress_key = f'work_logs-{i}-accumulated_progress_snapshot'
                                id_key = f'work_logs-{i}-id'
                                delete_key = f'work_logs-{i}-DELETE'
                                
                                # Verifica se há dados (ignora se estiver marcado para deletar)
                                is_deleted = normalized_post.get(delete_key, '').strip() == 'on'
                                
                                # IMPORTANTE: activity_description é obrigatório, então só considera como tendo dados
                                # se o campo obrigatório estiver preenchido OU se for uma edição (tem ID)
                                activity_description = normalized_post.get(activity_key, '').strip() if activity_key in normalized_post else ''
                                has_id = id_key in normalized_post and normalized_post.get(id_key, '').strip()
                                
                                # Considera como tendo dados apenas se:
                                # 1. Tem ID (é uma edição) OU
                                # 2. Tem activity_description preenchido (campo obrigatório)
                                if not is_deleted and (has_id or activity_description):
                                    has_worklog_data = True
                                    break
                            
                            # Se não há dados válidos, considera válido (formset vazio)
                            # Se há dados, valida o formset
                            if not has_worklog_data:
                                worklog_valid_final = True
                            else:
                                worklog_valid_final = worklog_formset.is_valid()
                                # Log detalhado se falhar
                                if not worklog_valid_final:
                                    logger.debug(f"Formset de worklogs inválido. Erros: {worklog_formset.errors}")
                                    for i, form in enumerate(worklog_formset.forms):
                                        if form.errors:
                                            logger.debug(f"  Form {i} erros: {form.errors}")
                                    if worklog_formset.non_form_errors():
                                        logger.debug(f"  Erros não-form: {worklog_formset.non_form_errors()}")
                        
                        if total_occurrence_forms == 0:
                            occurrence_valid_final = True
                        else:
                            has_occurrence_data = False
                            for i in range(total_occurrence_forms):
                                delete_key = f'ocorrencias-{i}-DELETE'
                                is_deleted = normalized_post.get(delete_key, '').strip() == 'on'
                                if is_deleted:
                                    continue
                                description_key = f'ocorrencias-{i}-description'
                                id_key = f'ocorrencias-{i}-id'
                                if (description_key in normalized_post and normalized_post.get(description_key, '').strip()) or \
                                   (id_key in normalized_post and normalized_post.get(id_key, '').strip()):
                                    has_occurrence_data = True
                                    break
                            occurrence_valid_final = True if not has_occurrence_data else occurrence_formset.is_valid()

                        logger.info(f"Re-validação dos formsets (com PK): imagens={image_valid_final}, worklogs={worklog_valid_final}, ocorrências={occurrence_valid_final}")
                        if not worklog_valid_final and total_worklog_forms > 0:
                            pass  # erros em worklog_formset
                        if not occurrence_valid_final and total_occurrence_forms > 0:
                            pass  # erros em occurrence_formset
                        
                        # Se o formset de worklogs é crítico e falhou com dados, coleta erros antes de fazer rollback
                        if not worklog_valid_final and total_worklog_forms > 0:
                            logger.error("Formset de worklogs falhou e há dados. Coletando erros antes do rollback.")
                            # Coleta erros específicos do formset para mostrar ao usuário
                            worklog_errors = []
                            for i, form_obj in enumerate(worklog_formset.forms):
                                if form_obj.errors:
                                    for field, errors in form_obj.errors.items():
                                        for error in errors:
                                            worklog_errors.append(f'Atividade {i+1} - {field}: {error}')
                            if worklog_formset.non_form_errors():
                                for error in worklog_formset.non_form_errors():
                                    worklog_errors.append(f'Erro geral: {error}')
                            
                            error_message = "Erro ao processar atividades. "
                            if worklog_errors:
                                error_message += "Detalhes: " + "; ".join(worklog_errors[:5])  # Limita a 5 erros
                                if len(worklog_errors) > 5:
                                    error_message += f" (e mais {len(worklog_errors) - 5} erro(s))"
                            else:
                                error_message += "Por favor, verifique os dados das atividades e tente novamente."
                            
                            logger.error(f"Erros do formset de worklogs: {worklog_errors}")
                            raise ValueError(error_message)
                        
                        # Atualiza as variáveis de validação para usar os valores finais
                        image_valid = image_valid_final
                        worklog_valid = worklog_valid_final
                        occurrence_valid = occurrence_valid_final
                    else:
                        # Se o diário não existe ou não tem PK, considera formsets vazios (válidos)
                        image_valid = True
                        worklog_valid = True
                        occurrence_valid = True
                        logger.debug("Diário não existe ou não tem PK, considerando formsets vazios como válidos")
                    
                    # Se for edição de diário existente, registra no log
                    if diary and diary.pk and not is_new:
                        # Registra edição
                        try:
                            from core.models import DiaryEditLog
                            DiaryEditLog.objects.create(
                                diary=diary,
                                edited_by=request.user,
                                notes=f"Diário editado via formulário web"
                            )
                        except Exception as e:
                            logger.debug(f"Erro ao criar DiaryEditLog: {e}")
                    
                    # 1. PROCESSAMENTO DE FOTOS
                    from core.models import DiaryImage
                    saved_images = []
                    manually_saved_images = []
                    
                    # Se formset é válido, usa ele para salvar (evita duplicação)
                    if image_valid:
                        saved_images = image_formset.save()
                        logger.info(f"Imagens salvas pelo formset: {len(saved_images)} imagens")
                        
                        # Coleta IDs das imagens salvas pelo formset para evitar processamento manual duplicado
                        formset_saved_ids = {img.id for img in saved_images if img.id}
                        formset_processed_indices = set()
                        
                        # Identifica quais índices do formset foram processados
                        for form in image_formset.forms:
                            if form.instance.pk:
                                # Tenta identificar o índice do form
                                prefix = form.prefix
                                if prefix and prefix.startswith('diaryimage_set-'):
                                    try:
                                        idx = int(prefix.replace('diaryimage_set-', '').split('-')[0])
                                        formset_processed_indices.add(idx)
                                    except (ValueError, IndexError):
                                        pass
                    else:
                        saved_images = []
                        formset_saved_ids = set()
                        formset_processed_indices = set()
                        logger.debug("Formset de imagens inválido, processando manualmente...")
                    
                    # Processa fotos manualmente APENAS se:
                    # 1. Formset falhou (image_valid = False), OU
                    # 2. Há fotos que não foram processadas pelo formset (fotos adicionadas dinamicamente)
                    total_forms = int(normalized_post.get('diaryimage_set-TOTAL_FORMS', '0'))
                    logger.info(f"TOTAL_FORMS do formset de imagens: {total_forms}")
                    
                    # Processa manualmente apenas se formset falhou ou para fotos não processadas
                    if not image_valid or total_forms > len(formset_processed_indices):
                        processed_image_indices = set()
                        for i in range(total_forms):
                            # Se formset processou este índice e foi válido, pula
                            if image_valid and i in formset_processed_indices:
                                continue
                            
                            if i in processed_image_indices:
                                continue
                            
                            image_key = f'diaryimage_set-{i}-image'
                            id_key = f'diaryimage_set-{i}-id'
                            delete_key = f'diaryimage_set-{i}-DELETE'
                            caption_key = f'diaryimage_set-{i}-caption'
                            approved_key = f'diaryimage_set-{i}-is_approved_for_report'
                            
                            if delete_key in request.POST:
                                delete_val = request.POST.get(delete_key, '').strip().lower()
                                if delete_val in ('on', 'true', '1', 'yes'):
                                    if id_key in request.POST:
                                        image_id = request.POST.get(id_key, '').strip()
                                        if image_id and image_id not in formset_saved_ids:
                                            try:
                                                existing_image = DiaryImage.objects.get(id=image_id, diary=diary)
                                                existing_image.delete()
                                                logger.info(f"Imagem deletada manualmente: ID={image_id}")
                                            except DiaryImage.DoesNotExist:
                                                pass
                                continue
                            
                            # Processa apenas se há arquivo E não foi processado pelo formset
                            if image_key in preserved_files:
                                image_file = preserved_files[image_key]
                                
                                # Valida arquivo antes de processar
                                try:
                                    from .utils.file_validators import validate_image_file
                                    image_file = validate_image_file(image_file)
                                except ValidationError as e:
                                    logger.error(f"Erro de validação na imagem {image_key}: {e}")
                                    messages.error(request, f'Erro ao processar imagem: {e}')
                                    continue
                                
                                image_id = request.POST.get(id_key, '').strip()
                                
                                # Se tem ID e já foi salvo pelo formset, pula (evita duplicação)
                                if image_id and image_id in formset_saved_ids:
                                    logger.info(f"Imagem {image_id} já processada pelo formset, pulando processamento manual")
                                    continue
                                
                                # Se tem ID mas não foi salvo pelo formset, atualiza
                                if image_id:
                                    try:
                                        existing = DiaryImage.objects.get(id=image_id, diary=diary)
                                        existing.image = image_file
                                        if caption_key in request.POST:
                                            existing.caption = request.POST[caption_key].strip()
                                        if approved_key in request.POST:
                                            existing.is_approved_for_report = (request.POST[approved_key] == 'on')
                                        existing.save()
                                        manually_saved_images.append(existing)
                                        logger.info(f"Imagem atualizada manualmente: ID={existing.id}")
                                    except DiaryImage.DoesNotExist:
                                        # Cria nova se não encontrou (legenda obrigatória)
                                        caption = request.POST.get(caption_key, '').strip()
                                        if not caption:
                                            continue
                                        is_approved = approved_key in request.POST and request.POST[approved_key] == 'on'
                                        new_image = DiaryImage.objects.create(
                                            diary=diary,
                                            image=image_file,
                                            caption=caption,
                                            is_approved_for_report=is_approved
                                        )
                                        manually_saved_images.append(new_image)
                                        logger.info(f"Imagem nova criada manualmente: ID={new_image.id}")
                                else:
                                    # Nova imagem sem ID - verifica se não foi salva pelo formset
                                    already_saved = False
                                    if saved_images:
                                        # Compara pelo tamanho e nome do arquivo
                                        for saved_img in saved_images:
                                            if (hasattr(saved_img, 'image') and saved_img.image and 
                                                saved_img.image.size == image_file.size):
                                                already_saved = True
                                                break
                                    
                                    if not already_saved:
                                        caption = request.POST.get(caption_key, '').strip()
                                        if not caption:
                                            continue
                                        is_approved = approved_key in request.POST and request.POST[approved_key] == 'on'
                                        new_image = DiaryImage.objects.create(
                                            diary=diary,
                                            image=image_file,
                                            caption=caption,
                                            is_approved_for_report=is_approved
                                        )
                                        manually_saved_images.append(new_image)
                                        logger.info(f"Imagem nova salva manualmente: ID={new_image.id}")
                                
                                processed_image_indices.add(i)
                    
                    all_saved_images = list(saved_images) + manually_saved_images
                    logger.info(f"Total de imagens salvas: {len(all_saved_images)} (formset: {len(saved_images)}, manual: {len(manually_saved_images)})")
                    
                    # 2. PROCESSAMENTO DE VÍDEOS (sempre processa)
                    from core.models import DiaryVideo
                    saved_videos = []
                    
                    if diary and diary.pk:
                        for video in diary.videos.all():
                            video_id = str(video.id)
                            delete_key = f'video_delete_{video_id}'
                            video_file_key = f'video_{video_id}'
                            caption_key = f'video_caption_{video_id}'
                            
                            if delete_key in request.POST:
                                if video.diary_id != diary.pk:
                                    continue
                                logger.info(f"Deletando vídeo ID={video_id}")
                                video.delete()
                                continue
                            
                            if caption_key in request.POST:
                                video.caption = request.POST[caption_key].strip()
                                video.save()
                                if video not in saved_videos:
                                    saved_videos.append(video)
                            
                            if video_file_key in preserved_files:
                                logger.info(f"Atualizando vídeo ID={video_id}")
                                video_file = preserved_files[video_file_key]
                                # Valida arquivo antes de processar
                                try:
                                    from .utils.file_validators import validate_video_file
                                    validate_video_file(video_file)
                                except ValidationError as e:
                                    logger.error(f"Erro de validação no vídeo {video_id}: {e}")
                                    messages.error(request, f'Erro ao processar vídeo: {e}')
                                    continue
                                video.video = video_file
                                if caption_key in request.POST:
                                    video.caption = request.POST[caption_key].strip()
                                video.save()
                                if video not in saved_videos:
                                    saved_videos.append(video)
                    
                    processed_video_indices = set()
                    for key in preserved_files.keys():
                        if key.startswith('video_new_'):
                            try:
                                index = int(key.replace('video_new_', ''))
                                if index in processed_video_indices:
                                    continue
                                processed_video_indices.add(index)
                                
                                video_file = preserved_files[key]
                                if not video_file or (getattr(video_file, 'size', 0) or 0) == 0:
                                    continue
                                # Valida arquivo antes de processar
                                try:
                                    from .utils.file_validators import validate_video_file
                                    validate_video_file(video_file)
                                except ValidationError as e:
                                    logger.error(f"Erro de validação no vídeo novo {key}: {e}")
                                    messages.error(request, f'Erro ao processar vídeo: {e}')
                                    continue
                                
                                caption_key = f'video_caption_new_{index}'
                                caption = request.POST.get(caption_key, '').strip()
                                if not caption:
                                    continue
                                video = DiaryVideo.objects.create(
                                    diary=diary,
                                    video=video_file,
                                    caption=caption,
                                    is_approved_for_report=True
                                )
                                saved_videos.append(video)
                                logger.info(f"Vídeo novo salvo: ID={video.id}")
                            except (ValueError, IndexError) as e:
                                logger.debug(f"Erro ao processar vídeo {key}: {e}")
                    
                    logger.info(f"Vídeos salvos: {len(saved_videos)} vídeos")
                    
                    # 3. PROCESSAMENTO DE ANEXOS (sempre processa)
                    from core.models import DiaryAttachment
                    saved_attachments = []
                    
                    if diary and diary.pk:
                        kept_attachment_ids = set()
                        for key in request.POST.keys():
                            if key.startswith('attachment_name_') and not key.startswith('attachment_name_new_'):
                                try:
                                    attachment_id = int(key.replace('attachment_name_', ''))
                                    kept_attachment_ids.add(attachment_id)
                                except ValueError:
                                    pass
                        
                        for key in preserved_files.keys():
                            if key.startswith('attachment_') and not key.startswith('attachment_new_'):
                                try:
                                    attachment_id = int(key.replace('attachment_', ''))
                                    kept_attachment_ids.add(attachment_id)
                                except ValueError:
                                    pass
                        
                        for attachment in list(diary.attachments.all()):
                            if attachment.id not in kept_attachment_ids:
                                logger.info(f"Deletando anexo ID={attachment.id}")
                                attachment.delete()
                                continue
                            
                            attachment_id = str(attachment.id)
                            file_key = f'attachment_{attachment_id}'
                            name_key = f'attachment_name_{attachment_id}'
                            
                            if name_key in request.POST:
                                attachment.name = request.POST[name_key].strip()
                                attachment.save()
                                if attachment not in saved_attachments:
                                    saved_attachments.append(attachment)
                            
                            if file_key in preserved_files:
                                logger.info(f"Atualizando anexo ID={attachment_id}")
                                attachment_file = preserved_files[file_key]
                                # Valida arquivo antes de processar
                                try:
                                    from .utils.file_validators import validate_attachment_file
                                    validate_attachment_file(attachment_file)
                                except ValidationError as e:
                                    logger.error(f"Erro de validação no anexo {attachment_id}: {e}")
                                    messages.error(request, f'Erro ao processar anexo: {e}')
                                    continue
                                attachment.file = attachment_file
                                if name_key in request.POST:
                                    attachment.name = request.POST[name_key].strip()
                                attachment.save()
                                if attachment not in saved_attachments:
                                    saved_attachments.append(attachment)
                    
                    processed_attachment_indices = set()
                    for key in preserved_files.keys():
                        if key.startswith('attachment_new_'):
                            try:
                                index = int(key.replace('attachment_new_', ''))
                                if index in processed_attachment_indices:
                                    continue
                                processed_attachment_indices.add(index)
                                
                                attachment_file = preserved_files[key]
                                # Valida arquivo antes de processar
                                try:
                                    from .utils.file_validators import validate_attachment_file
                                    validate_attachment_file(attachment_file)
                                except ValidationError as e:
                                    logger.error(f"Erro de validação no anexo novo {key}: {e}")
                                    messages.error(request, f'Erro ao processar anexo: {e}')
                                    continue
                                
                                name_key = f'attachment_name_new_{index}'
                                name = request.POST.get(name_key, '').strip() or attachment_file.name
                                
                                attachment = DiaryAttachment.objects.create(
                                    diary=diary,
                                    file=attachment_file,
                                    name=name
                                )
                                saved_attachments.append(attachment)
                                logger.info(f"Anexo novo salvo: ID={attachment.id}")
                            except (ValueError, IndexError) as e:
                                logger.debug(f"Erro ao processar anexo {key}: {e}")
                    
                    logger.info(f"Anexos salvos: {len(saved_attachments)} anexos")
                    
                    # 4. PROCESSAMENTO DE MÃO DE OBRA E EQUIPAMENTOS (sempre processa)
                    import json
                    from core.models import Labor, Equipment, DiaryLaborEntry, LaborCargo, LaborCategory
                    
                    # Novo sistema: mão de obra por categorias/cargos (diary_labor_data)
                    diary_labor_json = request.POST.get('diary_labor_data', '')
                    if diary_labor_json:
                        try:
                            diary_labor_data = json.loads(diary_labor_json) if diary_labor_json else []
                            DiaryLaborEntry.objects.filter(diary=diary).delete()
                            terceirizada_category = LaborCategory.objects.filter(slug='terceirizada').first()
                            for item in diary_labor_data:
                                cargo_id = item.get('cargo_id')
                                cargo_name = (item.get('cargo_name') or item.get('cargoName') or '').strip()
                                quantity = max(1, int(item.get('quantity') or 1))
                                company = (item.get('company') or '').strip()

                                selected_cargo_id = None
                                if cargo_id and LaborCargo.objects.filter(pk=cargo_id).exists():
                                    selected_cargo_id = int(cargo_id)
                                elif cargo_name and terceirizada_category:
                                    existing_cargo = LaborCargo.objects.filter(
                                        category=terceirizada_category,
                                        name__iexact=cargo_name
                                    ).only('id').first()
                                    if existing_cargo:
                                        selected_cargo_id = existing_cargo.id
                                    else:
                                        selected_cargo_id = LaborCargo.objects.create(
                                            category=terceirizada_category,
                                            name=cargo_name,
                                            order=0
                                        ).id

                                if not selected_cargo_id:
                                    continue
                                DiaryLaborEntry.objects.create(
                                    diary=diary,
                                    cargo_id=selected_cargo_id,
                                    quantity=quantity,
                                    company=company
                                )
                            logger.info(f"DiaryLaborEntry: {len(diary_labor_data)} itens salvos")
                            request._labor_objects = []
                        except (json.JSONDecodeError, ValueError, TypeError) as e:
                            logger.debug(f"Erro ao processar diary_labor_data: {e}")
                            request._labor_objects = []
                    else:
                        # Legado: labor_data (nome/role/quantidade)
                        labor_data_json = request.POST.get('labor_data', '[]')
                        try:
                            labor_data = json.loads(labor_data_json) if labor_data_json else []
                            logger.info(f"Dados de mão de obra recebidos: {len(labor_data)} itens")
                            
                            labor_objects = []
                            for labor_item in labor_data:
                                labor_name = labor_item.get('name', '').strip()
                                labor_role = labor_item.get('role', '')
                                labor_quantity = int(labor_item.get('quantity') or 1)
                                
                                if not labor_name or not labor_role:
                                    continue
                                
                                labor, created = Labor.objects.get_or_create(
                                    name=labor_name,
                                    role=labor_role,
                                    defaults={
                                        'is_active': True
                                    }
                                )
                                
                                for _ in range(labor_quantity):
                                    labor_objects.append(labor)
                                
                                logger.info(f"Labor processado: {labor.name} ({labor.get_role_display()}) x{labor_quantity}")
                            
                            request._labor_objects = labor_objects
                        except (json.JSONDecodeError, ValueError) as e:
                            logger.debug(f"Erro ao processar dados de mão de obra: {e}")
                            request._labor_objects = []
                    
                    # Equipamentos: só processa se campo foi enviado no POST.
                    # Isso evita limpar dados por acidente em submits parciais/legados.
                    if 'equipment_data' in request.POST:
                        equipment_data_json = request.POST.get('equipment_data', '[]')
                        try:
                            equipment_data = json.loads(equipment_data_json) if equipment_data_json else []
                            if not isinstance(equipment_data, list):
                                equipment_data = []
                            logger.info(f"Dados de equipamentos recebidos: {len(equipment_data)} itens")

                            equipment_qty_by_id = {}  # equipment_id -> {'equipment': Equipment, 'quantity': int}
                            for equipment_item in equipment_data:
                                if not isinstance(equipment_item, dict):
                                    continue
                                equipment, parsed_name = _resolve_equipment_from_payload_item(equipment_item)
                                if equipment is None:
                                    continue
                                equipment_quantity = _safe_positive_int(
                                    equipment_item.get('quantity'),
                                    default=1,
                                    minimum=1,
                                )

                                if equipment.pk not in equipment_qty_by_id:
                                    equipment_qty_by_id[equipment.pk] = {
                                        'equipment': equipment,
                                        'quantity': 0,
                                    }
                                equipment_qty_by_id[equipment.pk]['quantity'] += equipment_quantity
                                logger.info("Equipment processado: %s x%s", equipment.name, equipment_quantity)

                            equipment_items = [
                                (v['equipment'], v['quantity'])
                                for v in equipment_qty_by_id.values()
                                if v['quantity'] > 0
                            ]
                            request._equipment_items = equipment_items
                        except (json.JSONDecodeError, TypeError) as e:
                            logger.debug(f"Erro ao processar dados de equipamentos: {e}")
                            request._equipment_items = []
                    else:
                        request._equipment_items = None
                    
                    # 5. PROCESSAMENTO DE WORKLOGS (prioridade: formset; JSON apenas fallback não destrutivo)
                    saved_worklogs = []
                    work_logs_json_str = (request.POST.get('work_logs_json') or '').strip()
                    occurrences_json_str = (request.POST.get('occurrences_json') or '').strip()
                    has_worklogs_json = 'work_logs_json' in request.POST and work_logs_json_str not in ('', '[]')
                    has_occurrences_json = 'occurrences_json' in request.POST and occurrences_json_str not in ('', '[]')
                    total_worklog_forms_post = int(normalized_post.get('work_logs-TOTAL_FORMS', '0'))
                    total_occurrence_forms_post = int(normalized_post.get('ocorrencias-TOTAL_FORMS', '0'))

                    if worklog_valid:
                        saved_worklogs = worklog_formset.save()
                        logger.info(f"Worklogs salvos pelo formset: {len(saved_worklogs)} worklogs")
                        # Fallback de compatibilidade:
                        # se o front enviou apenas JSON (sem linhas no formset), processa JSON sem apagar dados existentes.
                        if not saved_worklogs and has_worklogs_json and total_worklog_forms_post == 0:
                            from core.diary_json_services import create_worklogs_from_json
                            create_worklogs_from_json(
                                diary,
                                project,
                                work_logs_json_str,
                                replace_existing=False,
                            )
                            saved_worklogs = list(diary.work_logs.all())
                            logger.info(
                                "Worklogs processados por fallback JSON (não destrutivo); total no diário: %s",
                                len(saved_worklogs),
                            )
                    else:
                        logger.warning("Formset de worklogs inválido, pulando processamento")
                    
                    # Se não há worklogs mas há mão de obra ou equipamentos, cria um worklog padrão
                    # IMPORTANTE: Isso deve acontecer mesmo se o formset falhou
                    if (not saved_worklogs and 
                        ((hasattr(request, '_labor_objects') and request._labor_objects) or 
                         (hasattr(request, '_equipment_items') and request._equipment_items))):
                        from core.models import DailyWorkLog, Activity
                        
                        # Valida que project existe
                        if not project:
                            logger.error("Tentativa de criar worklog padrão sem projeto")
                            raise ValueError("Projeto é obrigatório para criar worklog padrão")
                        
                        # Cria uma Activity genérica para o worklog padrão
                        # Treebeard requer usar add_root() para criar nós raiz
                        with transaction.atomic():
                            # Tenta buscar Activity existente primeiro
                            try:
                                default_activity = Activity.objects.get(
                                    project=project,
                                    name='Registro Geral de Mão de Obra e Equipamentos'
                                )
                                created = False
                            except Activity.DoesNotExist:
                                # Se não existe, cria como raiz usando add_root()
                                from decimal import Decimal
                                default_activity = Activity.add_root(
                                    project=project,
                                    name='Registro Geral de Mão de Obra e Equipamentos',
                                    code='GEN-MAO-OBRA-EQUIP',
                                    description='Atividade genérica para registro de mão de obra e equipamentos sem atividade específica',
                                    weight=Decimal('0.00'),
                                    status=ActivityStatus.NOT_STARTED
                                )
                                created = True
                                logger.info(f"Activity padrão criada como raiz: {default_activity.code}")
                            
                            # Verifica se já existe worklog com esta activity e diary (unique_together)
                            # Usa get_or_create para evitar race condition
                            try:
                                default_worklog, created_worklog = DailyWorkLog.objects.get_or_create(
                                    diary=diary,
                                    activity=default_activity,
                                    defaults={
                                        'location': 'Geral',
                                        'notes': 'Registro de mão de obra e equipamentos do dia',
                                        'percentage_executed_today': 0,
                                        'accumulated_progress_snapshot': 0
                                    }
                                )
                                if created_worklog:
                                    logger.info(f"Worklog padrão criado (ID={default_worklog.id}) com Activity {default_activity.id}")
                                else:
                                    logger.info(f"Worklog padrão já existe (ID={default_worklog.id}), reutilizando")
                            except IntegrityError as e:
                                # Se ainda assim houver erro de integridade, tenta buscar novamente
                                logger.debug(f"Erro de integridade ao criar worklog padrão: {e}. Tentando buscar existente...")
                                try:
                                    default_worklog = DailyWorkLog.objects.get(
                                        diary=diary,
                                        activity=default_activity
                                    )
                                    logger.info(f"Worklog padrão encontrado após erro (ID={default_worklog.id})")
                                except DailyWorkLog.DoesNotExist:
                                    logger.error(f"Não foi possível criar ou encontrar worklog padrão: {e}")
                                    raise
                        
                        saved_worklogs = [default_worklog]
                    
                    # Associa mão de obra e equipamentos aos worklogs (sempre, mesmo se formset falhou)
                    if hasattr(request, '_labor_objects') and request._labor_objects and saved_worklogs:
                        for worklog in saved_worklogs:
                            if worklog.pk:
                                # Remove duplicados antes de adicionar
                                existing_labor_ids = set(worklog.resources_labor.values_list('id', flat=True))
                                new_labor_objects = [lab for lab in request._labor_objects if lab.id not in existing_labor_ids]
                                if new_labor_objects:
                                    worklog.resources_labor.add(*new_labor_objects)
                                    logger.info(f"Mão de obra associada ao worklog {worklog.id}: {len(new_labor_objects)} novos itens (total: {len(request._labor_objects)})")
                                else:
                                    logger.info(f"Mão de obra já associada ao worklog {worklog.id}, nenhum item novo")
                    
                    if hasattr(request, '_equipment_items') and request._equipment_items is not None and saved_worklogs:
                        from .models import DailyWorkLogEquipment
                        # A UI atual coleta equipamentos no nível do diário (não por atividade).
                        # Para evitar multiplicação indevida no PDF quando existem vários worklogs,
                        # mantém os equipamentos em um único worklog "alvo" e limpa os demais.
                        target_worklog = None
                        for wl in saved_worklogs:
                            if wl.pk and getattr(getattr(wl, 'activity', None), 'code', '') == 'GEN-MAO-OBRA-EQUIP':
                                target_worklog = wl
                                break
                        if target_worklog is None:
                            target_worklog = next((wl for wl in saved_worklogs if wl.pk), None)

                        for worklog in saved_worklogs:
                            if worklog.pk:
                                DailyWorkLogEquipment.objects.filter(work_log=worklog).delete()

                        if target_worklog and request._equipment_items:
                            for equipment, qty in request._equipment_items:
                                DailyWorkLogEquipment.objects.create(
                                    work_log=target_worklog,
                                    equipment=equipment,
                                    quantity=max(1, int(qty or 1)),
                                )
                            logger.info(
                                "Equipamentos associados ao worklog %s (alvo): %s itens (com quantidade)",
                                target_worklog.id,
                                len(request._equipment_items),
                            )
                    
                    # Recalcula progresso quando worklogs foram salvos (formset ou fallback JSON)
                    if saved_worklogs:
                        try:
                            from .services import ProgressService
                            for wl in saved_worklogs:
                                if wl.pk and wl.activity_id:
                                    try:
                                        ProgressService.calculate_rollup_progress(wl.activity_id)
                                    except Exception as e:
                                        logger.debug(f"Erro ao recalcular progresso da atividade {wl.activity_id}: {e}", exc_info=True)
                        except Exception as e:
                            logger.debug(f"Erro ao recalcular progresso: {e}", exc_info=True)
                    elif worklog_valid:
                        for form_obj in worklog_formset.forms:
                            if form_obj.instance.pk and form_obj.instance.activity:
                                try:
                                    from .services import ProgressService
                                    ProgressService.calculate_rollup_progress(form_obj.instance.activity_id)
                                except Exception as e:
                                    logger.debug(f"Erro ao recalcular progresso da atividade {form_obj.instance.activity_id}: {e}", exc_info=True)
                    
                    # 6. PROCESSAMENTO DE OCORRÊNCIAS (prioridade: formset; JSON apenas fallback não destrutivo)
                    if occurrence_valid:
                        occurrences = occurrence_formset.save(commit=False)
                        for occurrence in occurrences:
                            if not occurrence.pk:
                                occurrence.created_by = request.user
                            occurrence.save()
                        occurrence_formset.save_m2m()
                        logger.info(f"Ocorrências salvas: {len(occurrences)} ocorrências")
                        # Fallback de compatibilidade:
                        # se o front enviou apenas JSON (sem linhas no formset), processa JSON sem apagar dados existentes.
                        if has_occurrences_json and total_occurrence_forms_post == 0:
                            from core.diary_json_services import create_occurrences_from_json
                            created_occurrences = create_occurrences_from_json(
                                diary,
                                occurrences_json_str,
                                request.user,
                                replace_existing=False,
                            )
                            logger.info(
                                "Ocorrências processadas por fallback JSON (não destrutivo): %s itens (diary_id=%s)",
                                len(created_occurrences),
                                diary.pk,
                            )
                    else:
                        logger.warning("Formset de ocorrências inválido, pulando processamento")
                    
                    # 7. ATUALIZA INFORMAÇÕES DO PROJETO
                    if project:
                        project_updated = False
                        # Segurança: nesta tela os campos estruturais da obra não são editáveis.
                        # Não confiar em valores de POST para nome/cliente/endereço.
                        if 'project_responsible' in request.POST:
                            responsible_value = (request.POST.get('project_responsible') or '').strip()
                            if (project.responsible or '') != responsible_value:
                                project.responsible = responsible_value
                                project_updated = True
                        
                        if project_updated:
                            project.save()
                    
                    # 8. SALVA ASSINATURAS (assinatura obrigatória, exceto em Salvamento Parcial)
                    signature_inspection = request.POST.get('signature_inspection')
                    signature_production = request.POST.get('signature_production')
                    is_partial_save = (
                        _is_truthy_flag(request.POST.get('partial_save')) or
                        _is_truthy_flag(request.POST.get('as_partial_checkbox'))
                    )
                    
                    if not is_partial_save and (not signature_inspection or not signature_inspection.strip()):
                        raise ValueError("A assinatura do responsável pelo preenchimento é obrigatória. Preencha a seção Assinaturas.")
                    
                    if signature_inspection:
                        DiarySignature.objects.update_or_create(
                            diary=diary,
                            signature_type='inspection',
                            defaults={
                                'signer': request.user,
                                'signature_data': signature_inspection
                            }
                        )
                    
                    if signature_production:
                        DiarySignature.objects.update_or_create(
                            diary=diary,
                            signature_type='production',
                            defaults={
                                'signer': request.user,
                                'signature_data': signature_production
                            }
                        )
                    
            except Exception as e:
                # 1. Mensagem específica por tipo de erro
                if isinstance(e, ValueError):
                    messages.error(request, str(e))
                elif isinstance(e, (ValidationError, PermissionDenied)):
                    messages.error(request, str(e) if str(e) else 'Dados inválidos ou sem permissão.')
                elif isinstance(e, IntegrityError):
                    messages.error(request, 'Erro de consistência dos dados. Verifique duplicidades ou dependências.')
                else:
                    messages.error(request, f'Erro ao processar dados: {str(e)}')

                if isinstance(e, ValueError):
                    # Erro de validação funcional (ex.: assinatura obrigatória), não falha interna do servidor.
                    logger.warning("Validação ao salvar diário: %s", e)
                else:
                    logger.error("Erro ao salvar diário: %s", e, exc_info=True)

                # 2. Reconstrução do form uma única vez
                if diary and diary.pk:
                    try:
                        diary.refresh_from_db()
                    except Exception:
                        diary = None

                form = ConstructionDiaryForm(request.POST, instance=diary, user=request.user, project=project)
                files_for_formset = preserved_files if 'preserved_files' in locals() else request.FILES

                from copy import deepcopy
                normalized_post = deepcopy(request.POST)
                if hasattr(normalized_post, '_mutable'):
                    normalized_post._mutable = True

                if 'dailyworklog_set-TOTAL_FORMS' in request.POST and 'work_logs-TOTAL_FORMS' not in request.POST:
                    normalized_post['work_logs-TOTAL_FORMS'] = request.POST.get('dailyworklog_set-TOTAL_FORMS', '0')
                    normalized_post['work_logs-INITIAL_FORMS'] = request.POST.get('dailyworklog_set-INITIAL_FORMS', '0')
                    for key, value in request.POST.items():
                        if key.startswith('dailyworklog_set-'):
                            normalized_post[key.replace('dailyworklog_set-', 'work_logs-', 1)] = value

                if 'occurrences-TOTAL_FORMS' in request.POST and 'ocorrencias-TOTAL_FORMS' not in normalized_post:
                    for key, value in request.POST.items():
                        if key.startswith('occurrences-'):
                            normalized_post[key.replace('occurrences-', 'ocorrencias-', 1)] = value

                image_formset = DiaryImageFormSet(request.POST, files_for_formset, instance=diary)
                worklog_formset = DailyWorkLogFormSet(normalized_post, instance=diary, form_kwargs={'diary': diary}, prefix='work_logs')
                occurrence_formset = DiaryOccurrenceFormSet(normalized_post, instance=diary, prefix='ocorrencias')
                context = _diary_form_context_from_post(request, project, form, image_formset, worklog_formset, occurrence_formset, diary)
                return render(request, 'core/daily_log_form.html', context)

            # Mensagem de sucesso ou aviso
            # Se chegou aqui, a transação foi commitada com sucesso
            if image_valid and worklog_valid and occurrence_valid:
                from django.urls import reverse
                if had_provisional_unlock and diary and diary.pk:
                    DiaryCorrectionRequestLog.objects.filter(
                        diary_id=diary.pk,
                        granted_at__isnull=False,
                        closed_at__isnull=True,
                    ).update(closed_at=timezone.now())
                    ConstructionDiary.objects.filter(pk=diary.pk).update(
                        provisional_edit_granted_at=None,
                        provisional_edit_granted_by_id=None,
                        edit_requested_at=None,
                        edit_requested_by_id=None,
                        edit_request_note='',
                    )
                # Salvar diário (não rascunho) = diário aprovado → enviar e-mail ao dono da obra
                if not is_partial_save and diary and diary.status == DiaryStatus.APROVADO:
                    try:
                        from .diary_email import send_diary_to_owners, send_diary_pdf_to_recipients
                        send_diary_to_owners(diary)
                        send_diary_pdf_to_recipients(diary)
                    except Exception as e:
                        logger.exception("Erro ao enviar diário aos donos da obra: %s", e)
                if is_partial_save:
                    messages.success(request, 'Diário salvo parcialmente. Você pode continuar o preenchimento depois.')
                    return redirect('report-list')
                if diary and diary.status == DiaryStatus.AGUARDANDO_APROVACAO_GESTOR:
                    messages.success(request, 'RDO enviado para aprovação dos gestores. O envio ao cliente ocorrerá após aprovação.')
                    return redirect(reverse('diary-detail', kwargs={'pk': diary.pk}))
                if is_new:
                    messages.success(request, f'Diário criado com sucesso! Relatório #{diary.report_number or "em processamento"}')
                    return redirect('report-list')
                else:
                    messages.success(request, f'Diário atualizado com sucesso!')
                    return redirect(reverse('diary-detail', kwargs={'pk': diary.pk}))
            else:
                # Alguns formsets falharam, mas dados foram salvos na transação: re-renderiza com erros
                error_parts = []
                if not image_valid:
                    error_parts.append('fotos')
                if not worklog_valid:
                    error_parts.append('atividades')
                if not occurrence_valid:
                    error_parts.append('ocorrências')
                messages.warning(request, f'Diário salvo, mas alguns dados não puderam ser processados: {", ".join(error_parts)}. Verifique os erros abaixo.')
                form = ConstructionDiaryForm(request.POST, instance=diary, user=request.user, project=project)
                files_for_retry = request.FILES
                if locals().get('preserved_files'):
                    files_for_retry = preserved_files
                image_formset = DiaryImageFormSet(normalized_post, files_for_retry, instance=diary)
                worklog_formset = DailyWorkLogFormSet(normalized_post, instance=diary, form_kwargs={'diary': diary}, prefix='work_logs')
                occurrence_formset = DiaryOccurrenceFormSet(normalized_post, instance=diary, prefix='ocorrencias')
                context = _diary_form_context_from_post(request, project, form, image_formset, worklog_formset, occurrence_formset, diary)
                return render(request, 'core/daily_log_form.html', context)
        else:
            # Form principal inválido - coleta erros
            import logging
            logger = logging.getLogger(__name__)
            logger.warning(f"Form principal inválido; form.errors={dict(form.errors)}")
            
            errors = []
            error_details = []
            
            if not form.is_valid():
                errors.append('Erros no formulário principal')
                for field, field_errors in form.errors.items():
                    # Remove mensagens duplicadas de unique_together do Django
                    unique_error_msg = "Diário de Obra com este Projeto e Data já existe."
                    filtered_errors = [e for e in field_errors if unique_error_msg not in str(e)]
                    if filtered_errors:
                        error_details.append(f'{field}: {", ".join(filtered_errors)}')
                        logger.warning(f"Erro no campo {field}: {', '.join(filtered_errors)}")
            
            if not image_formset.is_valid():
                errors.append('Erros nas imagens')
                # Adiciona detalhes dos erros de imagem
                for i, form_obj in enumerate(image_formset.forms):
                    if form_obj.errors:
                        error_details.append(f'Foto {i+1}: {", ".join([", ".join(errors) for errors in form_obj.errors.values()])}')
                # Adiciona erros não-form (erros do formset)
                if image_formset.non_form_errors():
                    error_details.append(f'Erros gerais nas fotos: {", ".join(image_formset.non_form_errors())}')
            
            if not worklog_formset.is_valid():
                errors.append('Erros nas atividades')
                for i, form_obj in enumerate(worklog_formset.forms):
                    if form_obj.errors:
                        error_details.append(f'Atividade {i+1}: {", ".join([", ".join(errors) for errors in form_obj.errors.values()])}')
            
            if not occurrence_formset.is_valid():
                errors.append('Erros nas ocorrências')
                for i, form_obj in enumerate(occurrence_formset.forms):
                    if form_obj.errors:
                        error_details.append(f'Ocorrência {i+1}: {", ".join([", ".join(errors) for errors in form_obj.errors.values()])}')
            
            if errors:
                error_message = 'Por favor, corrija os erros abaixo: ' + ', '.join(errors)
                if error_details:
                    error_message += '\n\nDetalhes:\n' + '\n'.join(error_details[:10])  # Limita a 10 detalhes
                messages.error(request, error_message)
            # Re-renderiza o formulário com os dados do POST e os erros (evita perda de dados)
            from django.utils.datastructures import MultiValueDict
            from copy import deepcopy
            normalized_post = deepcopy(request.POST)
            if hasattr(normalized_post, '_mutable'):
                normalized_post._mutable = True
            if 'dailyworklog_set-TOTAL_FORMS' in request.POST and 'work_logs-TOTAL_FORMS' not in request.POST:
                normalized_post['work_logs-TOTAL_FORMS'] = request.POST.get('dailyworklog_set-TOTAL_FORMS', '0')
                normalized_post['work_logs-INITIAL_FORMS'] = request.POST.get('dailyworklog_set-INITIAL_FORMS', '0')
                for key, value in request.POST.items():
                    if key.startswith('dailyworklog_set-'):
                        normalized_post[key.replace('dailyworklog_set-', 'work_logs-', 1)] = value
            if 'occurrences-TOTAL_FORMS' in request.POST and 'ocorrencias-TOTAL_FORMS' not in normalized_post:
                for key, value in request.POST.items():
                    if key.startswith('occurrences-'):
                        normalized_post[key.replace('occurrences-', 'ocorrencias-', 1)] = value
            files_for_formset = request.FILES
            if locals().get('files_dict'):
                files_for_formset = MultiValueDict()
                for k, v in files_dict.items():
                    files_for_formset.appendlist(k, v)
            image_formset = DiaryImageFormSet(request.POST, files_for_formset, instance=diary)
            worklog_formset = DailyWorkLogFormSet(normalized_post, instance=diary, form_kwargs={'diary': diary}, prefix='work_logs')
            occurrence_formset = DiaryOccurrenceFormSet(normalized_post, instance=diary, prefix='ocorrencias')
            context = _diary_form_context_from_post(request, project, form, image_formset, worklog_formset, occurrence_formset, diary)
            return render(request, 'core/daily_log_form.html', context)
    else:
        # GET request - mostra formulário
        # Se vier da lista com projeto/data (modal "Adicionar relatório"), troca a obra na sessão e redireciona
        get_project_id = request.GET.get('project')
        if get_project_id and not pk:
            try:
                proj = Project.objects.get(pk=get_project_id, is_active=True)
                if _user_can_access_project(request.user, proj):
                    request.session['selected_project_id'] = proj.id
                    request.session['selected_project_name'] = proj.name
                    request.session['selected_project_code'] = getattr(proj, 'code', '') or ''
                    get_date = (request.GET.get('date') or '').strip()
                    if get_date:
                        from django.urls import reverse
                        from urllib.parse import urlencode
                        return redirect(reverse('diary-new') + '?' + urlencode({'date': get_date}))
                    return redirect('diary-new')
            except (Project.DoesNotExist, ValueError, TypeError):
                pass  # Mantém projeto atual e ignora parâmetro inválido
        # Sincroniza contratante da obra: se o mapeamento define um valor, grava no projeto (corrige vazio ou valor antigo)
        if project:
            default_cli = get_contractante_for_project(project)
            if default_cli and (getattr(project, 'client_name', None) or '').strip() != default_cli:
                project.client_name = default_cli
                project.save(update_fields=['client_name'])
        # Verifica se há data passada via GET; para novo diário, preenche data (e dia da semana) automaticamente
        get_date_str = request.GET.get('date')
        form = ConstructionDiaryForm(instance=diary, user=request.user, project=project)
        
        initial_date_obj = None
        if not diary:
            from datetime import date as date_type
            if get_date_str:
                try:
                    from datetime import datetime
                    if '/' in get_date_str:
                        initial_date_obj = datetime.strptime(get_date_str, '%d/%m/%Y').date()
                    else:
                        initial_date_obj = datetime.strptime(get_date_str, '%Y-%m-%d').date()
                except (ValueError, TypeError):
                    pass
            if initial_date_obj is None:
                initial_date_obj = date_type.today()
            form.initial['date'] = initial_date_obj
        
        if diary and diary.pk:
            image_formset = DiaryImageFormSet(instance=diary)
            worklog_formset = DailyWorkLogFormSet(instance=diary, form_kwargs={'diary': diary}, prefix='work_logs')
            occurrence_formset = DiaryOccurrenceFormSet(instance=diary, prefix='ocorrencias')
        else:
            image_formset = DiaryImageFormSet()
            worklog_formset = DailyWorkLogFormSet(form_kwargs={'diary': None}, prefix='work_logs')
            occurrence_formset = DiaryOccurrenceFormSet(prefix='ocorrencias')
        
        # Copiar de relatório anterior: copy_from=<pk>&copy=climate,labor,equipment,activities,ocorrencias,all
        copy_from_id = request.GET.get('copy_from')
        copy_options_raw = request.GET.get('copy', '') or ''
        copy_source_diary = None
        copy_opts_list = []
        if project and copy_from_id and copy_options_raw:
            copy_opts = [x.strip().lower() for x in copy_options_raw.split(',') if x.strip()]
            if 'all' in copy_opts:
                copy_opts = ['climate', 'labor', 'equipment', 'activities', 'ocorrencias', 'interrupcoes']
            try:
                from .models import ConstructionDiary as CD
                src = CD.objects.filter(project=project, pk=copy_from_id).select_related('project').prefetch_related(
                    'work_logs__activity', 'work_logs__resources_equipment', 'occurrences__tags'
                ).first()
                if src and (not diary or src.pk != diary.pk):
                    copy_source_diary = src
                    copy_opts_list = copy_opts
                    copy_formset_instance = ConstructionDiary(project=project)
                    # Form initial a partir do relatório fonte
                    try:
                        if any(o in copy_opts for o in ('climate',)):
                            climate_fields = ('weather_conditions', 'weather_morning_condition', 'weather_morning_workable',
                                             'weather_afternoon_condition', 'weather_afternoon_workable',
                                             'weather_night_enabled', 'weather_night_type', 'weather_night_workable',
                                             'pluviometric_index', 'rain_occurrence', 'rain_observations')
                            for f in climate_fields:
                                if f in form.fields and hasattr(src, f):
                                    form.initial[f] = getattr(src, f)
                        if any(o in copy_opts for o in ('interrupcoes',)):
                            try:
                                interrupcoes_fields = (
                                    'accidents', 'stoppages', 'imminent_risks', 'incidents',
                                    'inspections', 'dds', 'general_notes',
                                )
                                for f in interrupcoes_fields:
                                    if f in form.fields and hasattr(src, f):
                                        val = getattr(src, f)
                                        form.initial[f] = val if val is not None else ''
                            except Exception:
                                pass
                        # Atividades e ocorrências: preencher formsets iniciais
                        # (também em edição, para permitir sobrescrever com dados do relatório fonte).
                        if 'activities' in copy_opts and src.work_logs.exists():
                            worklog_initial = []
                            for wl in src.work_logs.prefetch_related('activity').all():
                                worklog_initial.append({
                                    'location': wl.location or '',
                                    'work_stage': getattr(wl, 'work_stage', 'AN') or 'AN',
                                    'percentage_executed_today': wl.percentage_executed_today,
                                    'accumulated_progress_snapshot': wl.accumulated_progress_snapshot,
                                    'notes': wl.notes or '',
                                    'activity_description': wl.activity.name if wl.activity else '',
                                })
                            if worklog_initial:
                                # Inline formsets com extra=0 ignoram initial em GET.
                                # Criamos um formset dinâmico com extra proporcional ao que foi copiado.
                                WorklogCopyFormSet = inlineformset_factory(
                                    ConstructionDiary,
                                    DailyWorkLog,
                                    form=DailyWorkLogForm,
                                    extra=len(worklog_initial),
                                    can_delete=True,
                                )
                                worklog_formset = WorklogCopyFormSet(
                                    instance=copy_formset_instance,
                                    initial=worklog_initial,
                                    form_kwargs={'diary': diary if diary and diary.pk else None},
                                    prefix='work_logs',
                                )
                        if 'ocorrencias' in copy_opts and src.occurrences.exists():
                            occ_initial = []
                            for o in src.occurrences.prefetch_related('tags').all():
                                occ_initial.append({
                                    'description': o.description or '',
                                    'tags': list(o.tags.values_list('pk', flat=True)),
                                })
                            if occ_initial:
                                # Inline formsets com extra=0 ignoram initial em GET.
                                # Criamos um formset dinâmico com extra proporcional ao que foi copiado.
                                OccurrenceCopyFormSet = inlineformset_factory(
                                    ConstructionDiary,
                                    DiaryOccurrence,
                                    form=DiaryOccurrenceForm,
                                    extra=len(occ_initial),
                                    can_delete=True,
                                )
                                occurrence_formset = OccurrenceCopyFormSet(
                                    instance=copy_formset_instance,
                                    initial=occ_initial,
                                    prefix='ocorrencias',
                                )
                    except Exception:
                        pass  # não perder copy_source_diary se preenchimento falhar
            except Exception:
                copy_source_diary = None
    
    # Prepara dados para o template
    from .models import OccurrenceTag, LaborCategory
    try:
        occurrence_tags = OccurrenceTag.objects.filter(is_active=True)
    except Exception:
        # Se as tabelas ainda não existirem (migration não aplicada)
        occurrence_tags = []

    # Categorias e cargos para seleção de mão de obra por blocos
    try:
        labor_categories = LaborCategory.objects.prefetch_related('cargos').order_by('order')
        labor_terceirizada_cargos = []
        for cat in labor_categories:
            if cat.slug == 'terceirizada':
                labor_terceirizada_cargos = [{'id': c.id, 'name': c.name} for c in cat.cargos.all()]
                break
    except Exception:
        labor_categories = []
        labor_terceirizada_cargos = []

    # Categorias e equipamentos padrão para seleção no diário
    try:
        from .models import EquipmentCategory
        equipment_categories = EquipmentCategory.objects.prefetch_related('items').order_by('order')
    except Exception:
        equipment_categories = []

    # Registros de mão de obra já salvos (para edição ou cópia de relatório anterior)
    existing_diary_labor = []
    labor_source = (copy_source_diary if copy_source_diary and 'labor' in copy_opts_list else None) or (diary if diary and diary.pk else None)
    if labor_source:
        try:
            from .models import DiaryLaborEntry
            for e in DiaryLaborEntry.objects.filter(diary=labor_source).select_related('cargo'):
                existing_diary_labor.append({
                    'cargo_id': e.cargo_id,
                    'cargo_name': getattr(e.cargo, 'name', ''),
                    'quantity': e.quantity,
                    'company': e.company or '',
                })
        except Exception as e:
            logger.warning("Erro ao montar existing_diary_labor (cópia): %s", e, exc_info=True)
    
    # Equipamentos já salvos no diário (para edição ou cópia) – agregados por ID para evitar colisão por nome
    existing_diary_equipment = []
    equipment_source = (copy_source_diary if copy_source_diary and 'equipment' in copy_opts_list else None) or (diary if diary and diary.pk else None)
    if equipment_source:
        try:
            from core.utils.diary_equipment import aggregate_equipment_for_diary
            rows_agg, _tot = aggregate_equipment_for_diary(equipment_source)
            for r in rows_agg:
                eq = r['equipment']
                existing_diary_equipment.append({
                    'name': getattr(eq, 'name', '') or '',
                    'quantity': r['quantity'],
                    'equipment_id': r['equipment_id'],
                })
        except Exception as e:
            logger.warning("Erro ao montar existing_diary_equipment (cópia): %s", e, exc_info=True)
    
    # Calcula próximo número do relatório se for novo diário (otimizado)
    next_report_number = None
    if not diary or not diary.pk:
        last_diary_for_number = ConstructionDiary.objects.filter(
            project=project
        ).only('report_number').order_by('-report_number').first()
        
        if last_diary_for_number and last_diary_for_number.report_number:
            next_report_number = last_diary_for_number.report_number + 1
        else:
            next_report_number = 1
    
    # Último relatório do projeto (para "Copiar dados de relatório anterior" — sempre só o último, escala sem dropdown)
    last_diary_for_copy = None
    if project:
        qs = ConstructionDiary.objects.filter(project=project).order_by('-date', '-report_number')
        if diary and diary.pk:
            qs = qs.exclude(pk=diary.pk)
        d = qs.only('pk', 'date', 'report_number').first()
        if d:
            try:
                date_str = d.date.strftime('%d/%m/%Y') if d.date else ''
            except Exception:
                date_str = str(d.date) if d.date else ''
            last_diary_for_copy = {
                'id': d.pk,
                'date': date_str,
                'report_number': d.report_number or '-',
            }

    context = {
        'diary': diary if diary and diary.pk else None,  # Só passa se já estiver salvo
        'form': form,
        'image_formset': image_formset,
        'worklog_formset': worklog_formset,
        'occurrence_formset': occurrence_formset,
        'occurrence_tags': occurrence_tags,
        'labor_categories': labor_categories,
        'labor_terceirizada_cargos': labor_terceirizada_cargos,
        'existing_diary_labor': existing_diary_labor,
        'existing_diary_equipment': existing_diary_equipment,
        'equipment_categories': equipment_categories,
        'project': project,  # Adiciona projeto ao contexto
        'next_report_number': next_report_number,  # Próximo número do relatório
        'initial_contractante': get_contractante_for_project(project),
        'project_responsible_initial': getattr(project, 'responsible', '') if project else '',
        'last_diary_for_copy': last_diary_for_copy,
        'copy_from_id': copy_from_id,
        'copy_options': copy_options_raw if copy_source_diary else '',
        'copy_source_diary': copy_source_diary,
        # Data e dia da semana: para novo diário preenchidos automaticamente (hoje ou GET ?date=)
        'initial_date': initial_date_obj,
        'diary_date_display': (diary.date if diary and diary.pk else initial_date_obj),
    }
    
    return render(request, 'core/daily_log_form.html', context)


def _ensure_diary_pdf_generator_loaded(request):
    """
    Importação lazy do PDFGenerator (ReportLab). Mantém o mesmo comportamento de erro da view de PDF.
    """
    global PDFGenerator, REPORTLAB_AVAILABLE

    if PDFGenerator is None:
        try:
            from .utils.pdf_generator import (
                PDFGenerator as PDFGen,
                REPORTLAB_AVAILABLE as RL_AVAILABLE,
            )
            PDFGenerator = PDFGen
            REPORTLAB_AVAILABLE = RL_AVAILABLE
        except Exception as e:
            import logging
            logging.getLogger(__name__).exception(
                "Falha ao importar gerador de PDF: %s. Instale reportlab e Pillow: pip install reportlab Pillow",
                e,
            )
            REPORTLAB_AVAILABLE = False
            messages.error(request, "Geração de PDF não disponível neste ambiente.")
            return False

    return True


def _diary_pdf_http_response(diary, pdf_type, disposition='attachment'):
    """
    Monta HttpResponse com o PDF gerado do diário.
    disposition: 'attachment' (download, fluxo existente) ou 'inline' (visualização/embed no navegador).
    """
    pdf_buffer = PDFGenerator.generate_diary_pdf(diary.id, pdf_type=pdf_type)

    if not pdf_buffer:
        return None

    pdf_buffer.seek(0)
    pdf_bytes = pdf_buffer.read()
    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    type_suffix = {
        'normal': '',
        'detailed': '_detalhado',
        'no_photos': '_sem_fotos'
    }.get(pdf_type, '')
    try:
        from .utils.pdf_generator import get_rdo_pdf_filename
        filename = get_rdo_pdf_filename(diary.project, diary.date, suffix=type_suffix)
    except Exception:
        filename = f"RDO_{diary.project.code}_{diary.date.strftime('%Y%m%d')}{type_suffix}.pdf"
    safe_name = "".join(c if c.isalnum() or c in '-_.' else '_' for c in filename)
    disp = 'attachment' if disposition == 'attachment' else 'inline'
    response['Content-Disposition'] = f'{disp}; filename="{safe_name}"'
    response['Content-Length'] = len(pdf_bytes)
    return response


def _diary_pdf_sequence_for_project(project):
    """
    Ordem cronológica dos relatórios da obra para navegação entre PDFs.
    Critério: data do diário; desempate: data/hora de criação; depois id.
    (Alinhado ao modelo: no máximo um diário por dia por obra; desempates cobrem migrações/dados legados.)
    """
    return ConstructionDiary.objects.filter(project=project).order_by(
        'date', 'created_at', 'pk',
    )


def _diary_adjacent_ids_for_project(project, diary_pk):
    """
    Retorna (prev_pk, next_pk, position, total) na mesma ordem de _diary_pdf_sequence_for_project.
    Anterior = relatório mais antigo; próximo = mais recente.
    """
    ids = list(_diary_pdf_sequence_for_project(project).values_list('pk', flat=True))
    idx = None
    try:
        idx = ids.index(diary_pk)
    except ValueError:
        # Evita spans "desativados" por mismatch str/int em edge cases (URL vs ORM).
        try:
            idx = ids.index(int(diary_pk))
        except (TypeError, ValueError):
            pass
    if idx is None:
        return None, None, 0, len(ids)
    total = len(ids)
    position = idx + 1
    prev_pk = ids[idx - 1] if idx > 0 else None
    next_pk = ids[idx + 1] if idx < total - 1 else None
    return prev_pk, next_pk, position, total


def _diaries_queryset_for_report_filters(project, get_dict):
    """
    Mesmos filtros da listagem de relatórios (report_list), ordenação cronológica para exportação ZIP.
    get_dict: request.GET (QueryDict ou dict-like).
    """
    diaries = ConstructionDiary.objects.filter(project=project).select_related('project')
    search = get_dict.get('search')
    if search:
        diaries = diaries.filter(
            Q(project__code__icontains=search) |
            Q(project__name__icontains=search) |
            Q(general_notes__icontains=search)
        )
    date_start = get_dict.get('date_start')
    if date_start:
        try:
            diaries = diaries.filter(date__gte=date_start)
        except ValueError:
            pass
    date_end = get_dict.get('date_end')
    if date_end:
        try:
            diaries = diaries.filter(date__lte=date_end)
        except ValueError:
            pass
    status = get_dict.get('status')
    if status:
        diaries = diaries.filter(status=status)
    return diaries.order_by('date', 'created_at', 'pk')


def _client_portal_accessible_diary(user, pk):
    """
    Diário visível no portal do cliente: dono da obra, aprovado e já enviado ao dono.
    """
    diary = get_object_or_404(
        ConstructionDiary.objects.select_related('project'),
        pk=pk,
    )
    if not _client_can_access_diary(user, diary):
        raise Http404('Diário não encontrado.')
    if diary.status != DiaryStatus.APROVADO or diary.sent_to_owner_at is None:
        raise Http404('Diário não disponível.')
    return diary


def _client_diary_pdf_sequence_for_project(project):
    """Sequência de navegação no portal: só diários aprovados já enviados ao dono."""
    return ConstructionDiary.objects.filter(
        project=project,
        status=DiaryStatus.APROVADO,
        sent_to_owner_at__isnull=False,
    ).order_by('date', 'created_at', 'pk')


@login_required
@project_required
def diary_pdf_view(request, pk, pdf_type='normal'):
    """
    View para gerar e retornar PDF do diário (download — Content-Disposition: attachment).
    pdf_type: 'normal', 'detailed', 'no_photos'
    """
    if not _ensure_diary_pdf_generator_loaded(request):
        return redirect('diary-detail', pk=pk)

    project = get_selected_project(request)
    diary = get_object_or_404(ConstructionDiary, pk=pk, project=project)

    if not REPORTLAB_AVAILABLE:
        messages.error(request, "Geração de PDF não disponível neste ambiente.")
        return redirect('diary-detail', pk=pk)

    try:
        response = _diary_pdf_http_response(diary, pdf_type, disposition='attachment')
        if response:
            return response
        messages.error(request, "Erro ao gerar PDF. Tente novamente.")
        return redirect('diary-detail', pk=pk)
    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        logger.error("Erro ao gerar PDF do diário %s: %s\n%s", diary.id, e, tb, exc_info=False)
        messages.error(request, f"Erro ao gerar PDF: {str(e)}")
        return redirect('diary-detail', pk=pk)


@login_required
@project_required
def diary_pdf_inline_view(request, pk, pdf_type='normal'):
    """
    Mesmo PDF que diary_pdf_view, porém com Content-Disposition: inline para exibição no navegador (iframe / nova aba).
    Não substitui o fluxo de download; rota adicional.
    """
    if not _ensure_diary_pdf_generator_loaded(request):
        return redirect('diary-detail', pk=pk)

    project = get_selected_project(request)
    diary = get_object_or_404(ConstructionDiary, pk=pk, project=project)

    if not REPORTLAB_AVAILABLE:
        messages.error(request, "Geração de PDF não disponível neste ambiente.")
        return redirect('diary-detail', pk=pk)

    try:
        response = _diary_pdf_http_response(diary, pdf_type, disposition='inline')
        if response:
            return response
        messages.error(request, "Erro ao gerar PDF. Tente novamente.")
        return redirect('diary-detail', pk=pk)
    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        logger.error("Erro ao gerar PDF inline do diário %s: %s\n%s", diary.id, e, tb, exc_info=False)
        messages.error(request, f"Erro ao gerar PDF: {str(e)}")
        return redirect('diary-detail', pk=pk)


@login_required
@project_required
def diary_pdf_reader_view(request, pk):
    """
    Tela de leitura: PDF embutido na interface com navegação anterior/próximo na sequência do diário da obra.
    O download continua disponível pela rota diary-pdf (attachment).
    """
    project = get_selected_project(request)
    diary = get_object_or_404(ConstructionDiary, pk=pk, project=project)

    qs = _diary_pdf_sequence_for_project(project)
    ids = list(qs.values_list('pk', flat=True))
    idx = ids.index(diary.pk)

    total = len(ids)
    position = idx + 1 if total else 0
    prev_pk = ids[idx - 1] if idx > 0 else None
    next_pk = ids[idx + 1] if idx < len(ids) - 1 else None

    context = {
        'diary': diary,
        'project': project,
        'reader_position': position,
        'reader_total': total,
        'reader_prev_pk': prev_pk,
        'reader_next_pk': next_pk,
    }
    return render(request, 'core/diary_pdf_reader.html', context)


@login_required
def client_diary_pdf_view(request, pk, pdf_type='normal'):
    """
    PDF do RDO para o portal do cliente (download). Mesma geração do diário interno; sem obra na sessão.
    """
    if not _ensure_diary_pdf_generator_loaded(request):
        return redirect('client-diary-detail', pk=pk)
    diary = _client_portal_accessible_diary(request.user, pk)
    if not REPORTLAB_AVAILABLE:
        messages.error(request, 'Geração de PDF não disponível neste ambiente.')
        return redirect('client-diary-detail', pk=pk)
    try:
        response = _diary_pdf_http_response(diary, pdf_type, disposition='attachment')
        if response:
            return response
        messages.error(request, 'Erro ao gerar PDF. Tente novamente.')
        return redirect('client-diary-detail', pk=pk)
    except Exception as e:
        logger.error('Erro ao gerar PDF (cliente) diário %s: %s', diary.id, e, exc_info=False)
        messages.error(request, f'Erro ao gerar PDF: {str(e)}')
        return redirect('client-diary-detail', pk=pk)


@login_required
def client_diary_pdf_inline_view(request, pk, pdf_type='normal'):
    """PDF inline para iframe / nova aba no portal do cliente."""
    if not _ensure_diary_pdf_generator_loaded(request):
        return redirect('client-diary-detail', pk=pk)
    diary = _client_portal_accessible_diary(request.user, pk)
    if not REPORTLAB_AVAILABLE:
        messages.error(request, 'Geração de PDF não disponível neste ambiente.')
        return redirect('client-diary-detail', pk=pk)
    try:
        response = _diary_pdf_http_response(diary, pdf_type, disposition='inline')
        if response:
            return response
        messages.error(request, 'Erro ao gerar PDF. Tente novamente.')
        return redirect('client-diary-detail', pk=pk)
    except Exception as e:
        logger.error('Erro ao gerar PDF inline (cliente) diário %s: %s', diary.id, e, exc_info=False)
        messages.error(request, f'Erro ao gerar PDF: {str(e)}')
        return redirect('client-diary-detail', pk=pk)


@login_required
def client_diary_pdf_reader_view(request, pk):
    """Modo leitura com navegação apenas entre diários visíveis no portal do cliente."""
    diary = _client_portal_accessible_diary(request.user, pk)
    qs = _client_diary_pdf_sequence_for_project(diary.project)
    ids = list(qs.values_list('pk', flat=True))
    idx = ids.index(diary.pk)
    total = len(ids)
    position = idx + 1 if total else 0
    prev_pk = ids[idx - 1] if idx > 0 else None
    next_pk = ids[idx + 1] if idx < len(ids) - 1 else None
    context = {
        'diary': diary,
        'project': diary.project,
        'reader_position': position,
        'reader_total': total,
        'reader_prev_pk': prev_pk,
        'reader_next_pk': next_pk,
    }
    return render(request, 'core/client_diary_pdf_reader.html', context)


BULK_PDF_ZIP_MAX_DIARIES = 250


@login_required
@project_required
def diary_bulk_pdf_zip_view(request):
    """
    Exporta vários PDFs de RDO da obra atual em um arquivo ZIP (mesmos filtros GET da listagem).
    """
    import io
    import zipfile

    if not _ensure_diary_pdf_generator_loaded(request):
        return redirect('report-list')
    project = get_selected_project(request)
    if not REPORTLAB_AVAILABLE:
        messages.error(request, 'Geração de PDF não disponível neste ambiente.')
        return redirect('report-list')

    diaries = _diaries_queryset_for_report_filters(project, request.GET)
    count = diaries.count()
    if count == 0:
        messages.warning(request, 'Nenhum relatório encontrado com os filtros atuais.')
        return redirect('report-list')
    if count > BULK_PDF_ZIP_MAX_DIARIES:
        messages.error(
            request,
            f'Há {count} relatórios no filtro. O limite por arquivo ZIP é {BULK_PDF_ZIP_MAX_DIARIES}. '
            'Reduza o período ou refine a pesquisa.',
        )
        return redirect('report-list')

    buf = io.BytesIO()
    added = 0
    try:
        with zipfile.ZipFile(buf, 'w', compression=zipfile.ZIP_DEFLATED) as zf:
            for diary in diaries:
                try:
                    resp = _diary_pdf_http_response(diary, 'normal', disposition='attachment')
                    if not resp:
                        continue
                    pdf_bytes = resp.content
                    try:
                        from .utils.pdf_generator import get_rdo_pdf_filename
                        fname = get_rdo_pdf_filename(diary.project, diary.date, suffix='')
                    except Exception:
                        fname = f"RDO_{diary.project.code}_{diary.date.strftime('%Y%m%d')}.pdf"
                    safe_inner = ''.join(c if c.isalnum() or c in '-_.' else '_' for c in fname)
                    zf.writestr(safe_inner, pdf_bytes)
                    added += 1
                except Exception as ex:
                    logger.warning('ZIP RDO: falha no diário %s: %s', diary.pk, ex)
        if added == 0:
            messages.error(request, 'Não foi possível gerar nenhum PDF para o ZIP.')
            return redirect('report-list')
        buf.seek(0)
        zip_name = f"RDOs_{project.code}_{timezone.now().strftime('%Y%m%d_%H%M')}.zip"
        zip_safe = ''.join(c if c.isalnum() or c in '-_.' else '_' for c in zip_name)
        response = HttpResponse(buf.read(), content_type='application/zip')
        response['Content-Disposition'] = f'attachment; filename="{zip_safe}"'
        return response
    except Exception as e:
        logger.exception('Erro ao montar ZIP de RDOs: %s', e)
        messages.error(request, f'Erro ao gerar arquivo ZIP: {str(e)}')
        return redirect('report-list')


@login_required
@project_required
def diary_excel_view(request, pk):
    """View para gerar e retornar Excel do diário."""
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    
    project = get_selected_project(request)
    diary = get_object_or_404(ConstructionDiary, pk=pk, project=project)
    
    # Cria workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"Diario {diary.date.strftime('%d-%m-%Y')}"  # Remove / para evitar erro no Excel
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=14)
    
    # Título
    ws['A1'] = f"Relatório Diário de Obra - {diary.project.name}"
    ws['A1'].font = title_font
    ws.merge_cells('A1:D1')
    
    row = 3
    
    # Informações do Relatório
    ws[f'A{row}'] = 'N° do Relatório:'
    ws[f'B{row}'] = diary.report_number or '-'
    row += 1
    ws[f'A{row}'] = 'Data:'
    ws[f'B{row}'] = diary.date.strftime('%d/%m/%Y')
    row += 1
    ws[f'A{row}'] = 'Obra:'
    ws[f'B{row}'] = diary.project.name
    row += 2
    
    # Condições Climáticas
    ws[f'A{row}'] = 'Condições Climáticas'
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws.merge_cells(f'A{row}:D{row}')
    row += 1
    ws[f'A{row}'] = 'Manhã:'
    ws[f'B{row}'] = diary.weather_conditions or '-'
    row += 1
    ws[f'A{row}'] = 'Tarde:'
    ws[f'B{row}'] = diary.weather_conditions or '-'
    row += 2
    
    # Atividades
    ws[f'A{row}'] = 'Atividades'
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws.merge_cells(f'A{row}:D{row}')
    row += 1
    ws[f'A{row}'] = 'Atividade'
    ws[f'B{row}'] = 'Progresso (%)'
    ws[f'C{row}'] = 'Local'
    for cell in [ws[f'A{row}'], ws[f'B{row}'], ws[f'C{row}']]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    row += 1
    
    for work_log in diary.work_logs.all():
        ws[f'A{row}'] = work_log.activity.display_name
        ws[f'B{row}'] = float(work_log.percentage_executed_today)
        ws[f'C{row}'] = work_log.location or '-'
        row += 1
    
    # Ajusta largura das colunas
    ws.column_dimensions['A'].width = 48
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 25
    
    # Resposta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"diario_{diary.project.code}_{diary.date.strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def profile_view(request):
    """View de perfil do usuário."""
    from django.contrib import messages
    from .forms import ProfileEditForm
    
    user = request.user
    
    # Estatísticas do usuário
    from .models import ConstructionDiary, DiaryImage
    
    project = get_selected_project(request)
    user_diaries_count = 0
    user_photos_count = 0
    
    if project:
        user_diaries_count = ConstructionDiary.objects.filter(
            project=project,
            created_by=user
        ).count()
        
        user_photos_count = DiaryImage.objects.filter(
            diary__project=project,
            diary__created_by=user
        ).count()
    
    # Processa o formulário se for POST
    if request.method == 'POST':
        form = ProfileEditForm(request.POST, user=user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Perfil atualizado com sucesso!')
            # Se a senha foi alterada, faz logout e redireciona para login
            if form.cleaned_data.get('new_password'):
                from django.contrib.auth import logout
                logout(request)
                messages.info(request, 'Sua senha foi alterada. Por favor, faça login novamente.')
                return redirect('login')
            return redirect('profile')
        else:
            flash_message(request, "error", "core.form.fix_errors.profile")
    else:
        form = ProfileEditForm(user=user)
    
    context = {
        'user': user,
        'user_diaries_count': user_diaries_count,
        'user_photos_count': user_photos_count,
        'form': form,
    }
    
    return render(request, 'core/profile.html', context)


@login_required
@project_required
def activity_form_view(request, project_id, pk=None, parent_id=None):
    """View para criar/editar atividades na EAP."""
    from django.contrib import messages
    from .forms import ActivityForm
    
    project = get_selected_project(request)
    
    # Verifica se o projeto da URL corresponde ao selecionado
    if project.id != project_id:
        from django.core.exceptions import PermissionDenied
        raise PermissionDenied("Projeto não corresponde ao selecionado.")
    
    if pk:
        activity = get_object_or_404(Activity, pk=pk, project=project)
        form_title = "Editar Atividade"
        parent = activity.get_parent() if not activity.is_root() else None
    else:
        activity = None
        form_title = "Nova Atividade"
        if parent_id:
            parent = get_object_or_404(Activity, pk=parent_id, project=project)
        else:
            parent = None
    
    if request.method == 'POST':
        form = ActivityForm(request.POST, instance=activity)
        if form.is_valid():
            activity = form.save(commit=False)
            activity.project = project
            
            if not activity.pk:  # Nova atividade
                if parent:
                    activity = parent.add_child(instance=activity)
                else:
                    # Cria como raiz
                    activity = Activity.add_root(instance=activity)
            else:
                activity.save()
            
            messages.success(request, f'Atividade "{activity.name}" foi salva com sucesso!')
            return redirect('project-activities-tree', project_id=project.id)
        else:
            flash_message(request, "error", "core.form.fix_errors.activity")
    else:
        form = ActivityForm(instance=activity)
    
    # Lista de atividades para seleção de pai (se edição)
    available_parents = Activity.objects.filter(project=project)
    if activity:
        # Exclui a própria atividade e seus descendentes
        available_parents = available_parents.exclude(
            path__startswith=activity.path
        )
    
    context = {
        'form': form,
        'activity': activity,
        'project': project,
        'parent': parent,
        'form_title': form_title,
        'available_parents': available_parents,
    }
    
    return render(request, 'core/activity_form.html', context)


@login_required
@project_required
def activity_delete_view(request, project_id, pk):
    """View para deletar atividade."""
    from django.contrib import messages
    
    project = get_selected_project(request)
    activity = get_object_or_404(Activity, pk=pk, project=project)
    
    if request.method == 'POST':
        # Verifica se a atividade tem filhos
        if activity.get_children().exists():
            flash_message(request, "error", "core.activity.delete.has_children", {"atividade": activity.name})
            return redirect('project-activities-tree', project_id=project.id)
        
        activity_name = activity.name
        activity.delete()
        messages.success(request, f'Atividade "{activity_name}" foi deletada com sucesso!')
        return redirect('project-activities-tree', project_id=project.id)
    
    context = {
        'activity': activity,
        'project': project,
    }
    
    return render(request, 'core/activity_delete_confirm.html', context)


@login_required
@project_required
def labor_list_view(request):
    """View para listar mão de obra."""
    project = get_selected_project(request)
    
    from .models import Labor
    
    labor_list = Labor.objects.filter(is_active=True).order_by('labor_type', 'name')
    
    # Filtros
    search = request.GET.get('search', '')
    labor_type_filter = request.GET.get('labor_type')
    
    if search:
        labor_list = labor_list.filter(
            Q(name__icontains=search) |
            Q(role_custom__icontains=search) |
            Q(company__icontains=search)
        )
    
    if labor_type_filter:
        labor_list = labor_list.filter(labor_type=labor_type_filter)
    
    # Paginação
    from django.core.paginator import Paginator
    paginator = Paginator(labor_list, 20)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    # Estatísticas
    total_labor = labor_list.count()
    labor_by_type = labor_list.values('labor_type').annotate(count=Count('id'))
    
    context = {
        'labor_list': page_obj,
        'total_labor': total_labor,
        'labor_by_type': labor_by_type,
        'search': search,
        'labor_type_filter': labor_type_filter,
    }
    
    return render(request, 'core/labor_list.html', context)


@login_required
@project_required
def labor_form_view(request, pk=None):
    """View para criar/editar mão de obra."""
    from django.contrib import messages
    from .forms import LaborForm
    
    if pk:
        labor = get_object_or_404(Labor, pk=pk)
        form_title = "Editar Mão de Obra"
    else:
        labor = None
        form_title = "Nova Mão de Obra"
    
    if request.method == 'POST':
        form = LaborForm(request.POST, instance=labor)
        if form.is_valid():
            labor = form.save()
            messages.success(request, f'Mão de obra "{labor.name}" foi salva com sucesso!')
            return redirect('labor-list')
        else:
            flash_message(request, "error", "core.form.fix_errors.labor")
    else:
        form = LaborForm(instance=labor)
    
    context = {
        'form': form,
        'labor': labor,
        'form_title': form_title,
    }
    
    return render(request, 'core/labor_form.html', context)


@login_required
@project_required
def equipment_list_view(request):
    """View para listar equipamentos."""
    project = get_selected_project(request)
    
    from .models import Equipment
    
    equipment_list = Equipment.objects.filter(is_active=True).order_by('name')
    
    # Filtros
    search = request.GET.get('search', '')
    
    if search:
        equipment_list = equipment_list.filter(
            Q(name__icontains=search) |
            Q(equipment_type__icontains=search) |
            Q(description__icontains=search)
        )
    
    # Paginação
    from django.core.paginator import Paginator
    paginator = Paginator(equipment_list, 20)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    # Estatísticas
    total_equipment = equipment_list.count()
    
    context = {
        'equipment_list': page_obj,
        'total_equipment': total_equipment,
        'search': search,
    }
    
    return render(request, 'core/equipment_list.html', context)


@login_required
@project_required
def equipment_form_view(request, pk=None):
    """View para criar/editar equipamentos."""
    from django.contrib import messages
    from .forms import EquipmentForm
    
    if pk:
        equipment = get_object_or_404(Equipment, pk=pk)
        form_title = "Editar Equipamento"
    else:
        equipment = None
        form_title = "Novo Equipamento"
    
    if request.method == 'POST':
        form = EquipmentForm(request.POST, instance=equipment)
        if form.is_valid():
            equipment = form.save()
            messages.success(request, f'Equipamento "{equipment.name}" foi salvo com sucesso!')
            return redirect('equipment-list')
        else:
            flash_message(request, "error", "core.form.fix_errors.equipment")
    else:
        form = EquipmentForm(instance=equipment)
    
    context = {
        'form': form,
        'equipment': equipment,
        'form_title': form_title,
    }
    
    return render(request, 'core/equipment_form.html', context)


@login_required
def notifications_view(request):
    """View para listar notificações do usuário."""
    from .models import Notification
    
    notifications = Notification.objects.filter(
        user=request.user
    ).select_related('related_diary', 'related_diary__project').order_by('-created_at')
    
    # Marca notificações como lidas
    unread_count = notifications.filter(is_read=False).count()
    
    # Filtros
    filter_type = request.GET.get('type')
    filter_read = request.GET.get('read')
    
    if filter_type:
        notifications = notifications.filter(notification_type=filter_type)
    
    if filter_read == 'unread':
        notifications = notifications.filter(is_read=False)
    elif filter_read == 'read':
        notifications = notifications.filter(is_read=True)
    
    # Paginação
    from django.core.paginator import Paginator
    paginator = Paginator(notifications, 20)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    context = {
        'notifications': page_obj,
        'unread_count': unread_count,
        'filter_type': filter_type,
        'filter_read': filter_read,
    }
    
    return render(request, 'core/notifications.html', context)


@login_required
def notification_mark_read_view(request, pk):
    """View para marcar notificação como lida."""
    from django.http import JsonResponse
    from .models import Notification
    
    notification = get_object_or_404(Notification, pk=pk, user=request.user)
    notification.is_read = True
    notification.save()
    
    if request.headers.get('HX-Request'):
        return render(request, 'core/partials/notification_item.html', {'notification': notification})
    
    return JsonResponse({'status': 'success'})


@login_required
def notification_mark_all_read_view(request):
    """View para marcar todas as notificações como lidas."""
    from django.http import JsonResponse
    from django.contrib import messages
    from django.shortcuts import redirect
    from .models import Notification
    
    updated = Notification.objects.filter(
        user=request.user,
        is_read=False
    ).update(is_read=True)
    
    messages.success(request, f'{updated} notificação(ões) marcada(s) como lida(s).')
    
    if request.headers.get('HX-Request'):
        return redirect('notifications')
    
    return redirect('notifications')


@login_required
@project_required
def analytics_view(request):
    """View de análise de dados e estatísticas."""
    project = get_selected_project(request)
    
    from .models import ConstructionDiary, DiaryImage, Activity, DailyWorkLog, Labor, Equipment
    from django.db.models import Count, Avg, Sum, Q
    from datetime import datetime, timedelta
    
    # Estatísticas gerais
    total_diaries = ConstructionDiary.objects.filter(project=project).count()
    total_photos = DiaryImage.objects.filter(diary__project=project).count()
    total_activities = Activity.objects.filter(project=project).count()
    
    # Relatórios por status
    diaries_by_status = ConstructionDiary.objects.filter(
        project=project
    ).values('status').annotate(count=Count('id'))
    
    status_labels = {
        'PR': 'Preenchendo',
        'RV': 'Revisar',
        'AP': 'Aprovado',
    }
    
    status_data = {}
    for item in diaries_by_status:
        status_data[status_labels.get(item['status'], item['status'])] = item['count']
    
    # Relatórios por mês (últimos 6 meses)
    six_months_ago = timezone.now().date() - timedelta(days=180)
    from django.db import connection
    if connection.vendor == 'sqlite':
        diaries_by_month = ConstructionDiary.objects.filter(
            project=project,
            date__gte=six_months_ago
        ).extra(
            select={'month': "strftime('%%Y-%%m', date)"}
        ).values('month').annotate(count=Count('id')).order_by('month')
    else:
        # PostgreSQL ou MySQL
        from django.db.models.functions import TruncMonth
        diaries_by_month = ConstructionDiary.objects.filter(
            project=project,
            date__gte=six_months_ago
        ).annotate(
            month=TruncMonth('date')
        ).values('month').annotate(count=Count('id')).order_by('month')
    
    # Top atividades mais frequentes
    top_activities = DailyWorkLog.objects.filter(
        activity__project=project
    ).values('activity__name').annotate(
        count=Count('id')
    ).order_by('-count')[:10]
    
    # Distribuição de mão de obra
    labor_by_type = Labor.objects.filter(
        work_logs__activity__project=project
    ).values('labor_type').annotate(
        count=Count('id', distinct=True)
    ).order_by('labor_type')
    
    # Total de horas trabalhadas
    total_hours = ConstructionDiary.objects.filter(
        project=project,
        work_hours__isnull=False
    ).aggregate(total=Sum('work_hours'))['total'] or 0
    
    context = {
        'project': project,
        'total_diaries': total_diaries,
        'total_photos': total_photos,
        'total_activities': total_activities,
        'status_data': status_data,
        'diaries_by_month': list(diaries_by_month),
        'top_activities': list(top_activities),
        'labor_by_type': list(labor_by_type),
        'total_hours': total_hours,
    }
    
    return render(request, 'core/analytics.html', context)


@login_required
def project_form_view(request, pk=None):
    """View para criar/editar projetos."""
    from django.contrib import messages
    from accounts.painel_sistema_access import user_can_central_obras_diario_e_mapa
    from .forms import ProjectForm
    from django.core.exceptions import PermissionDenied

    if not user_can_central_obras_diario_e_mapa(request.user):
        raise PermissionDenied('Você não tem permissão para criar ou editar projetos.')
    
    if pk:
        project = get_object_or_404(Project, pk=pk)
        form_title = "Editar Obra"
    else:
        project = None
        form_title = "Nova Obra"
    
    if request.method == 'POST':
        form = ProjectForm(request.POST, instance=project)
        if form.is_valid():
            project = form.save(commit=False)
            # Nova obra: sempre criar como ativa para aparecer no Diário, Painel e GestControll
            if project.pk is None:
                project.is_active = True
            project.save()
            from core.sync_obras import sync_project_to_gestao_and_mapa
            sync_project_to_gestao_and_mapa(project)
            messages.success(request, f'Obra "{project.name}" foi salva e sincronizada com GestControll e Mapa.')
            return redirect('central_project_list')
        else:
            flash_message(request, "error", "core.form.fix_errors.project")
    else:
        form = ProjectForm(instance=project)
    
    context = {
        'form': form,
        'project': project,
        'form_title': form_title,
    }
    
    return render(request, 'core/project_form.html', context)


def _redirect_anonymous_to_login(view_func):
    """Redireciona anônimos para o login (302). Garante comportamento consistente em testes e produção."""
    @wraps(view_func)
    def _wrapped(request, *args, **kwargs):
        if not getattr(request.user, 'is_authenticated', False):
            from django.conf import settings
            from django.contrib.auth.views import redirect_to_login
            return redirect_to_login(request.get_full_path())
        return view_func(request, *args, **kwargs)
    return _wrapped


@_redirect_anonymous_to_login
@login_required
def project_list_view(request):
    """
    View de listagem de projetos (Central).
    Staff/superuser ou grupo Administrador (mesmo critério do antigo «Gerenciar obras»).
    """
    from accounts.painel_sistema_access import user_can_central_obras_diario_e_mapa
    from django.core.exceptions import PermissionDenied
    from django.db.models import Count, IntegerField, OuterRef, Subquery
    from django.db.models.functions import Coalesce
    from mapa_obras.models import LocalObra

    if not getattr(request.user, 'is_authenticated', False):
        from django.contrib.auth.views import redirect_to_login
        return redirect_to_login(request.get_full_path())
    if not user_can_central_obras_diario_e_mapa(request.user):
        raise PermissionDenied('Você não tem permissão para acessar esta página.')

    # Contagem de locais em subconsulta: evita JOIN cartesiano com diaries/activities
    # (vários Count em relações diferentes na mesma query podem distorcer resultados).
    # Usa codigo_sienge da obra mapa = project.code (válido com ou sem FK project preenchida).
    _locais_sq = Subquery(
        LocalObra.objects.filter(obra__codigo_sienge=OuterRef('code'))
        .values('obra')
        .annotate(_n=Count('id'))
        .values('_n')[:1],
        output_field=IntegerField(),
    )

    projects = Project.objects.annotate(
        diaries_count=Count('diaries', distinct=True),
        activities_count=Count('activities', distinct=True),
        n_locais_mapa=Coalesce(_locais_sq, 0),
    ).order_by('-created_at')

    context = {
        'projects': projects,
    }

    return render(request, 'core/project_list.html', context)


@login_required
@require_http_methods(["POST"])
def project_delete_view(request, pk):
    """Exclui uma obra. Remove também a obra no GestControll."""
    from django.contrib import messages
    from accounts.painel_sistema_access import user_can_central_obras_diario_e_mapa
    from django.core.exceptions import PermissionDenied

    if not user_can_central_obras_diario_e_mapa(request.user):
        raise PermissionDenied('Você não tem permissão para excluir obras.')

    project = get_object_or_404(Project, pk=pk)
    name = project.name

    # Remove a obra correspondente no GestControll para sumir da listagem de lá também
    from gestao_aprovacao.models import Obra as ObraGestao
    from django.db.models import ProtectedError

    obras_gestao = ObraGestao.objects.filter(project=project)
    try:
        obras_gestao.delete()
        removido_gestao = True
    except ProtectedError:
        # Há pedidos vinculados; não é possível excluir a obra no GestControll
        removido_gestao = False

    project.delete()

    # Se a obra excluída era a selecionada na sessão, limpa para evitar referência inválida
    if request.session.get('selected_project_id') == pk:
        for key in ('selected_project_id', 'selected_project_name', 'selected_project_code'):
            request.session.pop(key, None)

    if removido_gestao:
        messages.success(request, f'Obra "{name}" foi excluída do Diário e do GestControll.')
    else:
        messages.warning(
            request,
            f'Obra "{name}" foi excluída do Painel. No GestControll ela permanece porque há pedidos vinculados; exclua por lá se desejar.'
        )
    return redirect('central_project_list')


