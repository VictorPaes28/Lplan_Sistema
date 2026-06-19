"""
Auditoria de integridade do módulo RDO (Diário de Obra).

Valida paridade entre banco, payload de salvamento, contexto de views e HTML
para os fluxos principais: criar, editar, copiar, detalhe, listagem, dashboard,
exportação Excel e exclusão.
"""
from __future__ import annotations

import json
from datetime import date, timedelta
from decimal import Decimal
from io import BytesIO

from django.contrib.auth.models import User
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import Client, TestCase
from django.urls import reverse

from core.diary_copy_services import GEN_MAO_OBRA_EQUIP_CODE, serialize_work_logs_for_copy
from core.models import (
    Activity,
    ActivityStatus,
    ConstructionDiary,
    DailyWorkLog,
    DailyWorkLogEquipment,
    DiaryAttachment,
    DiaryLaborEntry,
    DiaryOccurrence,
    DiaryStatus,
    Equipment,
    Labor,
    LaborCargo,
    LaborCategory,
    OccurrenceTag,
    Project,
)
from core.tests.test_diary_copy import DiaryCopyTestMixin
from core.utils.diary_equipment import aggregate_equipment_for_diary
from core.utils.diary_labor import (
    build_labor_entries_by_category,
    merge_labor_entries_m2m_fallback_for_html,
)


def _detail_url(pk: int) -> str:
    """Detalhe HTML interno."""
    return f'/diaries/{pk}/'


def _full_post(
    project,
    diary_date,
    *,
    partial: bool = True,
    payload_ready: bool = True,
    work_logs_json=None,
    occurrences_json=None,
    labor_json=None,
    equipment_json=None,
    **extra,
) -> dict:
    post = {
        'project': str(project.id),
        'date': diary_date.isoformat(),
        'partial_save': '1' if partial else '',
        'work_logs-TOTAL_FORMS': '0',
        'work_logs-INITIAL_FORMS': '0',
        'work_logs-MIN_NUM_FORMS': '0',
        'work_logs-MAX_NUM_FORMS': '1000',
        'ocorrencias-TOTAL_FORMS': '0',
        'ocorrencias-INITIAL_FORMS': '0',
        'ocorrencias-MIN_NUM_FORMS': '0',
        'ocorrencias-MAX_NUM_FORMS': '1000',
        'diaryimage_set-TOTAL_FORMS': '0',
        'diaryimage_set-INITIAL_FORMS': '0',
        'diaryimage_set-MIN_NUM_FORMS': '0',
        'diaryimage_set-MAX_NUM_FORMS': '1000',
    }
    if payload_ready:
        post['diary_payload_ready'] = '1'
        post['work_logs_json'] = json.dumps(work_logs_json if work_logs_json is not None else [])
        post['occurrences_json'] = json.dumps(occurrences_json if occurrences_json is not None else [])
        post['diary_labor_data'] = json.dumps(labor_json if labor_json is not None else [])
        post['equipment_data'] = json.dumps(equipment_json if equipment_json is not None else [])
    post.update(extra)
    return post


def _labor_qty_from_db(diary) -> dict[str, int]:
    """Quantidades por cargo_name a partir de DiaryLaborEntry."""
    out: dict[str, int] = {}
    for entry in diary.labor_entries.select_related('cargo').all():
        name = entry.cargo.name
        out[name] = out.get(name, 0) + int(entry.quantity or 0)
    return out


def _labor_qty_from_detail_context(labor_entries_by_category) -> dict[str, int]:
    out: dict[str, int] = {}
    for slug in ('indireta', 'direta'):
        for row in labor_entries_by_category.get(slug) or []:
            name = (row.get('cargo_name') or '').strip()
            if name:
                out[name] = int(row.get('quantity') or 0)
    for block in labor_entries_by_category.get('terceirizada') or []:
        for row in block.get('items') or []:
            name = (row.get('cargo_name') or '').strip()
            if name:
                out[name] = int(row.get('quantity') or 0)
    return out


class DiaryCreateEditIntegrityTests(DiaryCopyTestMixin, TestCase):
    """Criação e edição: valor salvo = valor exibido no detalhe."""

    def test_create_draft_persists_all_sections_detail_matches_db(self):
        """POST rascunho completo → detalhe exibe os mesmos dados do banco."""
        client = self._session_client()
        target = date.today() - timedelta(days=4)
        act = self._make_activity('Atividade Auditoria')
        tag, _ = OccurrenceTag.objects.get_or_create(
            name='Tag Audit', defaults={'color': '#111', 'is_active': True},
        )
        from core.project_labor_catalog import ensure_project_labor_catalog

        ensure_project_labor_catalog(self.project)
        dir_cat = LaborCategory.objects.filter(slug='direta').first()
        cargo, _ = LaborCargo.objects.get_or_create(
            category=dir_cat, name='Carpinteiro Audit', defaults={'order': 0},
        )
        eq, _ = Equipment.objects.get_or_create(
            name='Guindaste Audit', defaults={'code': 'EQ-AUDIT-01'},
        )

        wl = [{
            'activity_description': act.name,
            'work_stage': 'IN',
            'percentage_executed_today': '12.5',
            'accumulated_progress_snapshot': '30',
            'location': 'Bloco A',
            'notes': 'Nota atividade audit',
        }]
        occ = [{'description': 'Ocorrência audit', 'tags': [tag.pk], 'tag_ids': [tag.pk]}]
        labor = [{
            'cargo_name': cargo.name,
            'quantity': 6,
            'company': '',
            'category_slug': 'direta',
            'cargo_id': cargo.id,
        }]
        equip = [{'equipment_id': eq.id, 'name': eq.name, 'quantity': 4}]

        post = _full_post(
            self.project, target,
            work_logs_json=wl, occurrences_json=occ,
            labor_json=labor, equipment_json=equip,
            general_notes='Obs audit',
            accidents='Acidente audit',
            inspections='Fiscal audit',
            weather_morning_condition='B',
            weather_morning_workable='T',
            weather_afternoon_condition='R',
            weather_afternoon_workable='N',
        )
        resp = client.post(reverse('diary-new'), post)
        self.assertIn(resp.status_code, (302, 200))

        diary = ConstructionDiary.objects.get(project=self.project, date=target)
        self.assertEqual(diary.status, DiaryStatus.SALVAMENTO_PARCIAL)
        self.assertEqual(diary.general_notes, 'Obs audit')
        self.assertEqual(diary.accidents, 'Acidente audit')
        self.assertEqual(diary.weather_morning_condition, 'B')
        self.assertEqual(diary.occurrences.count(), 1)
        self.assertEqual(diary.labor_entries.count(), 1)
        self.assertEqual(diary.labor_entries.first().quantity, 6)

        wl_db = diary.work_logs.exclude(activity__code=GEN_MAO_OBRA_EQUIP_CODE).first()
        self.assertIsNotNone(wl_db)
        self.assertEqual(wl_db.notes, 'Nota atividade audit')
        self.assertEqual(str(wl_db.percentage_executed_today), '12.50')

        rows, total = aggregate_equipment_for_diary(diary)
        self.assertEqual(total, 4)
        self.assertEqual(rows[0]['equipment_id'], eq.id)

        detail = client.get(_detail_url(diary.pk))
        self.assertEqual(detail.status_code, 200)
        html = detail.content.decode()
        self.assertIn('Atividade Auditoria', html)
        self.assertIn('Ocorrência audit', html)
        self.assertIn('Obs audit', html)
        self.assertIn('Acidente audit', html)
        self.assertIn('Guindaste Audit', html)
        self.assertTrue(detail.context['diary_has_extra_info'])
        self.assertEqual(detail.context['total_direct_labor'], 6)
        self.assertEqual(detail.context['equipment_total_quantity'], 4)

        db_labor = _labor_qty_from_db(diary)
        ctx_labor = _labor_qty_from_detail_context(detail.context['labor_entries_by_category'])
        self.assertEqual(db_labor, ctx_labor)

    def test_edit_changes_persist_after_reopen(self):
        """Editar campos → salvar → reabrir detalhe e edição com mesmos valores."""
        client = self._session_client()
        diary = self._make_source_diary(with_attachment=False)
        diary.status = DiaryStatus.SALVAMENTO_PARCIAL
        diary.save(update_fields=['status'])

        wl = serialize_work_logs_for_copy(diary)
        for item in wl:
            if item['activity_description'] == 'teste':
                item['notes'] = 'EDITADO audit'

        post = _full_post(
            self.project, diary.date,
            work_logs_json=wl,
            occurrences_json=[],
            labor_json=[],
            equipment_json=[],
            general_notes='Notas EDITADAS audit',
        )
        url = reverse('diary-edit', kwargs={'pk': diary.pk})
        client.post(url, post)

        diary.refresh_from_db()
        self.assertEqual(diary.general_notes, 'Notas EDITADAS audit')
        wl_obj = diary.work_logs.filter(activity__name='teste').first()
        self.assertEqual(wl_obj.notes, 'EDITADO audit')

        detail = client.get(_detail_url(diary.pk))
        self.assertIn('Notas EDITADAS audit', detail.content.decode())
        self.assertIn('EDITADO audit', detail.content.decode())

    def test_double_save_does_not_duplicate_activities(self):
        """Dois saves com mesmo JSON não duplicam atividades (replace_existing)."""
        client = self._session_client()
        target = date.today() - timedelta(days=6)
        act = self._make_activity('Ativ Única')
        wl = [{
            'activity_description': act.name,
            'work_stage': 'AN',
            'percentage_executed_today': '0',
            'accumulated_progress_snapshot': '0',
            'location': '',
            'notes': '',
        }]
        post = _full_post(self.project, target, work_logs_json=wl)
        client.post(reverse('diary-new'), post)
        diary = ConstructionDiary.objects.get(project=self.project, date=target)
        self.assertEqual(
            diary.work_logs.exclude(activity__code=GEN_MAO_OBRA_EQUIP_CODE).count(), 1,
        )
        client.post(reverse('diary-edit', kwargs={'pk': diary.pk}), post)
        diary.refresh_from_db()
        self.assertEqual(
            diary.work_logs.exclude(activity__code=GEN_MAO_OBRA_EQUIP_CODE).count(), 1,
        )

    def test_equipment_duplicate_payload_aggregates_to_single_row(self):
        """Payload com mesmo equipment_id duas vezes → 1 linha, qty somada no banco."""
        client = self._session_client()
        target = date.today() - timedelta(days=7)
        eq = Equipment.objects.create(code='EQ-AGG-AUDIT', name='Rolo Audit')
        equip = [
            {'equipment_id': eq.id, 'name': eq.name, 'quantity': 2},
            {'equipment_id': eq.id, 'name': eq.name, 'quantity': 3},
        ]
        post = _full_post(self.project, target, equipment_json=equip)
        client.post(reverse('diary-new'), post)
        diary = ConstructionDiary.objects.get(project=self.project, date=target)
        through = DailyWorkLogEquipment.objects.filter(work_log__diary=diary, equipment=eq)
        self.assertEqual(through.count(), 1)
        self.assertEqual(through.first().quantity, 5)

        detail = client.get(_detail_url(diary.pk))
        self.assertEqual(detail.context['equipment_total_quantity'], 5)

    def test_equipment_without_payload_ready_not_saved(self):
        """Sem diary_payload_ready, equipment_data no POST é ignorado (gate de segurança)."""
        client = self._session_client()
        target = date.today() - timedelta(days=8)
        eq = Equipment.objects.create(code='EQ-GATE', name='Gate Equip')
        post = _full_post(
            self.project, target, payload_ready=False,
            equipment_data=json.dumps([{'equipment_id': eq.id, 'name': eq.name, 'quantity': 2}]),
        )
        client.post(reverse('diary-new'), post)
        diary = ConstructionDiary.objects.filter(project=self.project, date=target).first()
        self.assertIsNotNone(diary)
        self.assertFalse(
            DailyWorkLogEquipment.objects.filter(work_log__diary=diary).exists(),
        )

    def test_legacy_m2m_labor_visible_in_detail(self):
        """MO só em resources_labor (legado) aparece no detalhe via fallback."""
        client = self._session_client()
        diary = ConstructionDiary.objects.create(
            project=self.project,
            date=date.today() - timedelta(days=9),
            status=DiaryStatus.APROVADO,
            created_by=self.user,
        )
        act = self._make_activity('Ativ Legado MO')
        wl = DailyWorkLog.objects.create(diary=diary, activity=act, work_stage='AN')
        lab = Labor.objects.create(name='Operador Legado Audit', role='ME', labor_type='D')
        wl.resources_labor.add(lab)

        detail = client.get(_detail_url(diary.pk))
        self.assertEqual(detail.status_code, 200)
        merged = merge_labor_entries_m2m_fallback_for_html(
            build_labor_entries_by_category(diary), diary,
        )
        ctx = detail.context['labor_entries_by_category']
        self.assertIn('Operador Legado Audit', detail.content.decode())
        self.assertGreater(
            sum(len(ctx.get(s) or []) for s in ('direta', 'indireta'))
            + sum(len(b.get('items') or []) for b in (ctx.get('terceirizada') or [])),
            0,
        )
        self.assertEqual(merged, ctx)

    def test_attachment_persists_and_shows_in_detail(self):
        """Anexo salvo no POST aparece na seção de anexos do detalhe."""
        client = self._session_client()
        diary = ConstructionDiary.objects.create(
            project=self.project,
            date=date.today() - timedelta(days=10),
            status=DiaryStatus.SALVAMENTO_PARCIAL,
            created_by=self.user,
        )
        DiaryAttachment.objects.create(
            diary=diary,
            name='relatorio_audit.pdf',
            file=SimpleUploadedFile('relatorio_audit.pdf', b'%PDF audit'),
        )
        detail = client.get(_detail_url(diary.pk))
        self.assertIn('relatorio_audit.pdf', detail.content.decode())
        self.assertIn('Anexos', detail.content.decode())

    def test_two_distinct_activities_persist_on_detail(self):
        """Duas atividades com nomes diferentes devem aparecer no detalhe."""
        client = self._session_client()
        target = date.today() - timedelta(days=16)
        act_a = self._make_activity('Atividade Alpha')
        act_b = self._make_activity('Atividade Beta')
        post = _full_post(
            self.project, target,
            work_logs_json=[
                {
                    'activity_description': act_a.name,
                    'work_stage': 'IN',
                    'percentage_executed_today': '0',
                    'accumulated_progress_snapshot': '0',
                    'location': '',
                    'notes': '',
                },
                {
                    'activity_description': act_b.name,
                    'work_stage': 'AN',
                    'percentage_executed_today': '0',
                    'accumulated_progress_snapshot': '0',
                    'location': '',
                    'notes': '',
                },
            ],
        )
        client.post(reverse('diary-new'), post)
        diary = ConstructionDiary.objects.get(project=self.project, date=target)
        names = list(
            diary.work_logs.exclude(activity__code=GEN_MAO_OBRA_EQUIP_CODE)
            .values_list('activity__name', flat=True)
        )
        self.assertEqual(len(names), 2)
        self.assertIn('Atividade Alpha', names)
        self.assertIn('Atividade Beta', names)

        detail = client.get(_detail_url(diary.pk))
        html = detail.content.decode()
        self.assertIn('Atividade Alpha', html)
        self.assertIn('Atividade Beta', html)
        self.assertEqual(len(detail.context['display_work_logs']), 2)

    def test_same_activity_name_twice_collapses_to_one_worklog(self):
        """Limitação do modelo: mesma atividade no mesmo RDO = 1 work_log (unique activity+diary)."""
        client = self._session_client()
        target = date.today() - timedelta(days=17)
        act = self._make_activity('Atividade Duplicada')
        post = _full_post(
            self.project, target,
            work_logs_json=[
                {'activity_description': act.name, 'work_stage': 'IN', 'notes': 'primeira'},
                {'activity_description': act.name, 'work_stage': 'AN', 'notes': 'segunda'},
            ],
        )
        client.post(reverse('diary-new'), post)
        diary = ConstructionDiary.objects.get(project=self.project, date=target)
        self.assertEqual(
            diary.work_logs.exclude(activity__code=GEN_MAO_OBRA_EQUIP_CODE).count(), 1,
        )
        wl = diary.work_logs.filter(activity__name=act.name).first()
        self.assertEqual(wl.notes, 'segunda')

    def test_approved_diary_blocks_edit_get(self):
        """RDO aprovado redireciona em diary-edit (sem unlock provisório)."""
        client = self._session_client()
        diary = ConstructionDiary.objects.create(
            project=self.project,
            date=date.today() - timedelta(days=11),
            status=DiaryStatus.APROVADO,
            created_by=self.user,
        )
        resp = client.get(reverse('diary-edit', kwargs={'pk': diary.pk}))
        self.assertEqual(resp.status_code, 302)


class DiaryListDashboardExportTests(DiaryCopyTestMixin, TestCase):
    """Listagem, dashboard e exportações."""

    def test_report_list_contains_diary_with_status(self):
        client = self._session_client()
        diary = ConstructionDiary.objects.create(
            project=self.project,
            date=date.today() - timedelta(days=12),
            status=DiaryStatus.SALVAMENTO_PARCIAL,
            created_by=self.user,
            general_notes='Marcador listagem audit',
        )
        resp = client.get(reverse('report-list'))
        self.assertEqual(resp.status_code, 200)
        html = resp.content.decode()
        self.assertIn(diary.date.strftime('%d/%m/%Y'), html)

    def test_dashboard_kpis_match_db_counts(self):
        client = self._session_client()
        ConstructionDiary.objects.filter(project=self.project).delete()
        ConstructionDiary.objects.create(
            project=self.project, date=date.today(),
            status=DiaryStatus.SALVAMENTO_PARCIAL, created_by=self.user,
        )
        ConstructionDiary.objects.create(
            project=self.project, date=date.today() - timedelta(days=1),
            status=DiaryStatus.APROVADO, created_by=self.user,
        )
        resp = client.get(reverse('dashboard'))
        self.assertEqual(resp.status_code, 200)
        ctx = resp.context
        self.assertEqual(ctx['total_diaries'], 2)
        self.assertEqual(ctx['pending_reports'], 1)
        self.assertEqual(ctx['approved_reports'], 1)

    def test_excel_export_activity_matches_db(self):
        client = self._session_client()
        diary = self._make_source_diary(with_occurrences=False, with_labor_entries=False, with_equipment=False)
        resp = client.get(reverse('diary-excel', kwargs={'pk': diary.pk}))
        self.assertEqual(resp.status_code, 200)
        self.assertIn('spreadsheetml', resp['Content-Type'])

        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(resp.content))
        ws = wb.active
        cells = [str(c.value) for row in ws.iter_rows() for c in row if c.value]
        self.assertIn('teste', cells)
        self.assertIn(diary.date.strftime('%d/%m/%Y'), cells)

    def test_superuser_delete_removes_diary(self):
        superuser = User.objects.create_superuser('audit_admin', 'a@b.com', 'pass')
        client = Client()
        client.force_login(superuser)
        session = client.session
        session['selected_project_id'] = self.project.id
        session['selected_project_name'] = self.project.name
        session['selected_project_code'] = self.project.code
        session.save()

        diary = ConstructionDiary.objects.create(
            project=self.project,
            date=date.today() - timedelta(days=13),
            status=DiaryStatus.SALVAMENTO_PARCIAL,
            created_by=self.user,
        )
        pk = diary.pk
        resp = client.post(reverse('diary-delete', kwargs={'pk': pk}))
        self.assertEqual(resp.status_code, 302)
        self.assertFalse(ConstructionDiary.objects.filter(pk=pk).exists())


class DiaryDisplayConditionalTests(DiaryCopyTestMixin, TestCase):
    """Regras condicionais de exibição no detalhe."""

    def test_extra_info_section_shows_dashes_when_empty(self):
        client = self._session_client()
        diary = ConstructionDiary.objects.create(
            project=self.project,
            date=date.today() - timedelta(days=14),
            status=DiaryStatus.APROVADO,
            created_by=self.user,
        )
        detail = client.get(_detail_url(diary.pk))
        self.assertFalse(detail.context['diary_has_extra_info'])
        html = detail.content.decode()
        self.assertIn('Informações adicionais', html)
        self.assertIn('diary-detail-extra__text--empty', html)

    def test_gen_mao_obra_equip_hidden_from_display_work_logs(self):
        client = self._session_client()
        diary = self._make_source_diary(with_occurrences=False, with_labor_entries=False, with_equipment=False)
        detail = client.get(_detail_url(diary.pk))
        names = [wl.activity.name for wl in detail.context['display_work_logs']]
        self.assertNotIn('Registro Geral de Mão de Obra e Equipamentos', names)


class DiaryInfrastructureTests(DiaryCopyTestMixin, TestCase):
    """Infraestrutura de rotas e contexto de erro."""

    def test_reverse_diary_detail_points_to_frontend_html(self):
        from django.urls import reverse
        self.assertEqual(reverse('diary-detail', kwargs={'pk': 42}), '/diaries/42/')

    def test_validation_error_after_copy_loses_copied_activity_data(self):
        """
        Falha de validação no POST após cópia: _diary_form_context_from_post não repõe
        copy_*_seed; atividades copiadas somem da tela (work_logs_json vazio no POST).
        """
        src = self._make_source_diary()
        client = self._session_client()
        get_url = (
            reverse('diary-new')
            + f'?copy_from={src.pk}&copy=climate,activities,ocorrencias'
        )
        get_resp = client.get(get_url)
        self.assertIn('teste', get_resp.content.decode())

        target = date.today() - timedelta(days=15)
        post = _full_post(self.project, target, partial=False)
        post['signature_inspection'] = ''
        post['work_logs_json'] = '[]'
        post['occurrences_json'] = '[]'
        post.pop('partial_save', None)
        resp = client.post(reverse('diary-new'), post)
        self.assertEqual(resp.status_code, 200)
        content = resp.content.decode()
        self.assertNotIn('Ocorrência teste1', content)
        self.assertNotIn('id="copy-work-logs-seed">[{"activity_description": "teste"', content)
