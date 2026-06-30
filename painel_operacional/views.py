import json
import csv
import logging
import re
import unicodedata
import warnings
from datetime import date, datetime
from io import BytesIO
from uuid import uuid4

from django.db import transaction
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, render
from django.urls import reverse
from django.utils import timezone
from django.utils.text import slugify
from django.views.decorators.cache import cache_control
from django.views.decorators.http import require_http_methods
from openpyxl import Workbook, load_workbook

from accounts.decorators import login_required, require_group
from accounts.groups import GRUPOS
from mapa_obras.models import Obra
from mapa_obras.contexto_obra import resolve_obra_context
from suprimentos.views_controle import _normalize_ambiente_layout

from .models import (
    AmbienteCelula,
    AmbienteElemento,
    AmbienteHistorico,
    AmbienteModoEditor,
    AmbienteOperacional,
    AmbienteTipo,
    AmbienteVersao,
    SemanticaIndicador,
    VersaoEstado,
)

logger = logging.getLogger(__name__)

# TEMPORÁRIO: importação de planilha na criação do mapa de controle desabilitada
# devido a erro no fluxo de interpretação/importação. Reativar quando corrigido.
PO_IMPORTACAO_PLANILHA_CRIACAO_DESABILITADA = True


def _importacao_planilha_criacao_bloqueada_response() -> JsonResponse:
    return JsonResponse(
        {
            "success": False,
            "error": (
                "Importação de planilha temporariamente indisponível na criação do mapa de controle. "
                "Crie o ambiente vazio e monte a matriz manualmente."
            ),
            "import_disabled": True,
        },
        status=503,
    )


def _resolve_editor_mode(ambiente: AmbienteOperacional) -> str:
    modo = str(getattr(ambiente, "modo_editor", "") or "").strip()
    if modo in {AmbienteModoEditor.MAPA_DEDICADO, AmbienteModoEditor.QUADRO}:
        return modo
    return AmbienteModoEditor.MAPA_DEDICADO if ambiente.tipo == AmbienteTipo.MAPA_CONTROLE else AmbienteModoEditor.QUADRO


def _resolver_obra(request):
    return resolve_obra_context(request, allow_post=True)


def _parse_json_body(request):
    try:
        return json.loads(request.body.decode("utf-8") or "{}")
    except json.JSONDecodeError:
        return {}


def _mapa_controle_rows_canonico(colunas: int = 20, linhas: int = 0):
    """
    Estrutura oficial canônica do Mapa de Controle.
    Hierarquia BLOCO → PAVIMENTO → APTO (como planilha / import Excel), depois atividades.
    Por padrão inicia vazio (sem unidades automáticas).
    """
    colunas = max(5, int(colunas))
    linhas = max(0, int(linhas))
    header = ["BLOCO", "PAVIMENTO", "APTO"]
    for i in range(colunas):
        header.append(f"Atividade {i + 1}")
    header.append("Total")
    rows = [header]
    for i in range(linhas):
        rows.append(["", "", ""] + [""] * colunas + [""])
    return rows


def _mapa_controle_weights(rows: list[list], totals_row_auto: bool = True):
    col_count = max((len(r) for r in rows if isinstance(r, list)), default=1)
    row_count = len(rows)
    vis_row_count = row_count + (1 if totals_row_auto else 0)

    # Eixo (Bloco/Local) mais estreito; o resto permanece amplo para atividades.
    col_weights = [0.7] + [1.0] * max(0, col_count - 2) + ([0.9] if col_count > 1 else [])
    # Cabeçalho superior mais alto para leitura com títulos verticais.
    row_weights = [2.2] + [1.0] * max(0, vis_row_count - 2) + ([1.1] if vis_row_count > 1 else [])
    return {"colWeights": col_weights, "rowWeights": row_weights}


def _preset_layout(tipo: str, obra: Obra | None = None):
    if tipo == AmbienteTipo.MAPA_CONTROLE:
        rows = _mapa_controle_rows_canonico(20, 0)
        weights = _mapa_controle_weights(rows, totals_row_auto=True)
        header = rows[0] if rows else []
        activity_cols = list(range(3, max(3, len(header) - 1)))
        import_meta = {
            "strategy": "manual_template",
            "axis_cols_interpreted": [0, 1, 2],
            "axis_headers_interpreted": ["BLOCO", "PAVIMENTO", "APTO"],
            "activity_cols_interpreted": activity_cols,
            "row_axis_key": "bloco",
        }
        # Um único bloco matriz; o utilizador adiciona KPI / detalhe pela barra se precisar.
        return {
            "title": "Mapa de Controle",
            "sections": [
                {
                    "id": "matriz",
                    "kind": "matrix_table",
                    "title": "Matriz de Controle",
                    "x": 80,
                    "y": 80,
                    "width": 680,
                    "height": 400,
                    "layer": {},
                    "data": {
                        "mapaControleTemplate": True,
                        "headerBandCount": 1,
                        "heatmap": True,
                        "totalsColumnAuto": True,
                        "totalsRowAuto": True,
                        "verticalHeaders": True,
                        "rows": rows,
                        "colWeights": weights["colWeights"],
                        "rowWeights": weights["rowWeights"],
                        "importMeta": import_meta,
                    },
                },
            ],
        }
    return {"title": "Ambiente Operacional", "sections": []}


def _serializar_ambiente(ambiente: AmbienteOperacional):
    versao_publicada = ambiente.versoes.filter(estado=VersaoEstado.PUBLISHED).order_by("-numero").first()
    versao_draft = ambiente.versoes.filter(estado=VersaoEstado.DRAFT).order_by("-numero").first()
    versao_atual = versao_draft or versao_publicada
    return {
        "id": ambiente.id,
        "nome": ambiente.nome,
        "tipo": ambiente.tipo,
        "modo_editor": _resolve_editor_mode(ambiente),
        "descricao": ambiente.descricao,
        "obra_id": ambiente.obra_id,
        "ativo": ambiente.ativo,
        "updated_at": ambiente.updated_at.isoformat(),
        "versao_atual": versao_atual.numero if versao_atual else None,
        "versao_publicada": versao_publicada.numero if versao_publicada else None,
        "versao_rascunho": versao_draft.numero if versao_draft else None,
    }


def _serializar_versao(versao: AmbienteVersao | None):
    if not versao:
        return None
    return {
        "id": versao.id,
        "numero": versao.numero,
        "estado": versao.estado,
        "layout": versao.layout,
        "metadados": versao.metadados,
        "updated_at": versao.updated_at.isoformat(),
    }


PO_MAX_IMPORT_ROWS = 30000
PO_MAX_IMPORT_COLS = 400
PO_MAX_IMPORT_CELLS = 600_000


def _excel_value_to_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value).strip()


def _normalize_import_rows(rows: list[list[str]]) -> list[list[str]]:
    compact = []
    max_cols = 0
    for row in rows:
        out = [_excel_value_to_text(cell) for cell in row]
        while out and not out[-1]:
            out.pop()
        if not out:
            continue
        compact.append(out)
        max_cols = max(max_cols, len(out))
    if not compact or max_cols <= 0:
        return []
    for row in compact:
        if len(row) < max_cols:
            row.extend([""] * (max_cols - len(row)))
    return compact


def _read_excel_rows(uploaded_file, sheet_name: str = "") -> tuple[list[list[str]], str, dict]:
    suffix = (uploaded_file.name or "").lower().strip()
    data = uploaded_file.read()
    if not data:
        return [], "", {}

    if suffix.endswith(".csv"):
        csv_encoding = "utf-8-sig"
        try:
            text = data.decode("utf-8-sig")
        except UnicodeDecodeError:
            # Fallback comum em exportações legadas do Excel no Windows.
            text = data.decode("latin-1", errors="replace")
            csv_encoding = "latin-1"
        sample = text[:4096]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=";,")
            sep = dialect.delimiter
        except csv.Error:
            sep = ";" if sample.count(";") >= sample.count(",") else ","
        reader = csv.reader(text.splitlines(), delimiter=sep)
        raw_rows = []
        for idx, row in enumerate(reader):
            if idx >= PO_MAX_IMPORT_ROWS + 1:
                break
            raw_rows.append(list(row)[: PO_MAX_IMPORT_COLS + 1])
        rows = _normalize_import_rows(raw_rows)
        return rows, "CSV", {"selected_sheet": "CSV", "sheet_mode": "csv", "encoding": csv_encoding}

    if suffix.endswith(".xls"):
        raise ValueError("Formato .xls antigo não suportado diretamente. Salve como .xlsx e tente novamente.")

    with warnings.catch_warnings():
        # Planilhas reais frequentemente trazem extensões de Excel que o openpyxl ignora;
        # não queremos poluir logs/terminal com esses avisos esperados.
        warnings.filterwarnings(
            "ignore",
            category=UserWarning,
            message=r".*(Unknown extension|Conditional Formatting extension|Slicer List extension|Data Validation extension).*",
        )
        wb = load_workbook(filename=BytesIO(data), read_only=True, data_only=True)
    try:
        sheets = list(wb.sheetnames)
        explicit_sheet = sheet_name if sheet_name and sheet_name in sheets else ""
        diagnostics = {"sheet_mode": "explicit" if explicit_sheet else "auto", "candidate_sheets": sheets[:20]}

        def read_rows_from_ws(ws, max_rows: int | None = None, max_cols: int | None = None) -> list[list[str]]:
            if not hasattr(ws, "iter_rows"):
                return []
            kwargs = {"values_only": True}
            row_cap = max_rows if max_rows and max_rows > 0 else (PO_MAX_IMPORT_ROWS + 1)
            if row_cap > 0:
                kwargs["max_row"] = row_cap
            col_cap = max_cols if max_cols and max_cols > 0 else (PO_MAX_IMPORT_COLS + 1)
            raw = []
            with warnings.catch_warnings():
                warnings.filterwarnings(
                    "ignore",
                    category=UserWarning,
                    message=r".*(Unknown extension|Conditional Formatting extension|Slicer List extension|Data Validation extension).*",
                )
                for row in ws.iter_rows(**kwargs):
                    vals = list(row)
                    if col_cap and col_cap > 0:
                        vals = vals[:col_cap]
                    raw.append(vals)
            return _normalize_import_rows(raw)

        if explicit_sheet:
            ws = wb[explicit_sheet]
            rows = read_rows_from_ws(ws)
            if not rows:
                raise ValueError("A aba selecionada não contém células tabulares para importação.")
            diagnostics["selected_sheet"] = explicit_sheet
            return rows, str(ws.title or "Planilha"), diagnostics

        best_name = None
        best_score = -1.0
        best_sample_rows = []
        best_strategy = ""
        best_confidence = 0.0
        for name in sheets:
            ws = wb[name]
            sample_rows = read_rows_from_ws(ws, max_rows=700, max_cols=140)
            score = _score_rows_for_auto_pick(sample_rows)
            try:
                _tmp_rows, strategy, report = _interpret_import_rows(sample_rows, mode="auto")
                confidence = float(report.get("confidence") or 0.0)
            except Exception:
                strategy = ""
                confidence = 0.0
            score += confidence * 7.0
            if strategy and strategy != "fallback_bruto":
                score += 3.0
            name_norm = _norm_token(name)
            if name_norm in {"DADOS", "EXECUCAO", "SERVICOS", "SERVICO", "STATUS"}:
                score += 1.8
            if score > best_score:
                best_score = score
                best_name = name
                best_sample_rows = sample_rows
                best_strategy = strategy
                best_confidence = confidence

        if not best_name:
            ws = wb.active
            rows = read_rows_from_ws(ws)
            diagnostics["selected_sheet"] = str(ws.title or "Planilha")
            return rows, str(ws.title or "Planilha"), diagnostics

        ws = wb[best_name]
        rows = read_rows_from_ws(ws)
        diagnostics["selected_sheet"] = best_name
        diagnostics["auto_score"] = round(best_score, 3)
        diagnostics["sample_rows"] = len(best_sample_rows)
        if best_strategy:
            diagnostics["auto_strategy"] = best_strategy
            diagnostics["auto_confidence"] = round(best_confidence, 3)
        return rows, str(ws.title or "Planilha"), diagnostics
    finally:
        wb.close()


def _norm_token(value: object) -> str:
    text = _excel_value_to_text(value).upper()
    if not text:
        return ""
    text = unicodedata.normalize("NFKD", text).encode("ASCII", "ignore").decode("ASCII")
    text = re.sub(r"[^A-Z0-9% ]+", " ", text)
    return " ".join(text.split())


def _parse_percent_value(text: str) -> float | None:
    raw = _excel_value_to_text(text)
    if not raw:
        return None
    t = raw.replace(" ", "").replace("\u00a0", "")
    t = t.replace("%", "")
    t = t.replace("R$", "").replace("r$", "")
    if "," in t and "." in t:
        # Suporta formatos BR (1.234,56) e EN (1,234.56) usando o último separador como decimal.
        if t.rfind(".") > t.rfind(","):
            t = t.replace(",", "")
        else:
            t = t.replace(".", "").replace(",", ".")
    elif "," in t:
        t = t.replace(",", ".")
    try:
        v = float(t)
    except ValueError:
        norm = _norm_token(raw)
        if norm in {"OK", "CONCLUIDO", "CONCLUIDA", "SIM", "DONE"}:
            return 100.0
        if norm in {"N", "NAO", "NAO INICIADO", "PENDENTE"}:
            return 0.0
        return None
    if 0 <= v <= 1:
        v *= 100
    return max(0.0, min(100.0, v))


def _binary_ratio_rows(rows: list[list[str]], max_cells: int = 5000) -> float:
    if not rows:
        return 0.0
    seen = 0
    binary = 0
    for row in rows:
        if not isinstance(row, list):
            continue
        for cell in row:
            txt = _excel_value_to_text(cell)
            if not txt:
                continue
            seen += 1
            t = txt.strip()
            if t in {"0", "1"}:
                binary += 1
            if seen >= max_cells:
                return binary / max(1, seen)
    return binary / max(1, seen)


def _score_rows_for_auto_pick(rows: list[list[str]]) -> float:
    if not rows:
        return -1.0
    row_count = len(rows)
    col_count = max((len(r) for r in rows), default=0)
    if row_count < 2 or col_count < 2:
        return -0.5

    header_idx, col_map = _find_header_and_map(rows)
    token_hits = 0
    if header_idx is not None and col_map:
        token_hits = len([k for k in col_map.keys() if not str(k).startswith("_")])
    has_activity = col_map.get("atividade") is not None if isinstance(col_map, dict) else False

    non_empty = 0
    numeric_like = 0
    scan_rows = rows[: min(180, row_count)]
    for row in scan_rows:
        for cell in row[: min(80, len(row))]:
            txt = _excel_value_to_text(cell)
            if not txt:
                continue
            non_empty += 1
            if _parse_percent_value(txt) is not None:
                numeric_like += 1
    density = non_empty / max(1, len(scan_rows) * max(1, min(80, col_count)))
    numeric_ratio = numeric_like / max(1, non_empty)
    binary_ratio = _binary_ratio_rows(scan_rows, max_cells=3200)

    semantic_bonus = 0.0
    if isinstance(col_map, dict):
        if col_map.get("atividade") is not None:
            semantic_bonus += 9.0
        if col_map.get("status") is not None:
            semantic_bonus += 4.0
        if col_map.get("_axis_cols"):
            semantic_bonus += 5.0

    score = (
        token_hits * 5.0
        + min(4.0, row_count / 120.0)
        + min(4.0, col_count / 25.0)
        + density * 6.0
        + numeric_ratio * 2.0
        + semantic_bonus
    )
    # Matrizes binárias densas tendem a ser resultados consolidados sem semântica de cabeçalho.
    if binary_ratio > 0.72 and not has_activity:
        score -= 8.5
    return score


def _find_header_and_map(rows: list[list[str]]) -> tuple[int | None, dict]:
    if not rows:
        return None, {}
    aliases = {
        "setor": {"SETOR", "AREA", "ZONA", "TORRE", "REGIAO"},
        "bloco": {"BLOCO", "BL", "BLC"},
        "pavimento": {"PAVIMENTO", "PAV", "ANDAR", "NIVEL"},
        "unidade": {"APTO", "UNIDADE", "LOCAL", "AMBIENTE", "SALA", "APARTAMENTO", "APART", "UND", "UH", "UNID"},
        "atividade": {"ATIVIDADE", "SERVICO", "SERVICOS", "ITEM", "ETAPA"},
        "grupo_servicos": {"GRUPO DE SERVICO", "GRUPO DE SERVICOS", "GRUPO SERVICO", "GRUPO SERVICOS"},
        "status": {"STATUS", "AVANCO", "PROGRESSO", "PERCENTUAL", "%", "MEDICAO"},
        "custo": {"CUSTO", "VALOR", "PRECO"},
        "observacao": {"OBS", "OBSERVACAO", "COMENTARIO", "NOTA"},
        "data_termino": {"DATA DE TERMINO", "DATA TERMINO", "TERMINO", "DATA FINAL"},
    }

    def match_field(token: str) -> str | None:
        if not token:
            return None
        if "GRUPO" in token and "SERVICO" in token:
            return "grupo_servicos"
        for field, terms in aliases.items():
            for term in terms:
                if token == term or token.startswith(f"{term} ") or f" {term} " in f" {token} ":
                    return field
        return None

    best_row = None
    best_score = -1
    best_map = {}
    scan_limit = min(len(rows), 60)
    for idx in range(scan_limit):
        row = rows[idx]
        current = {}
        for col_idx, raw in enumerate(row[:PO_MAX_IMPORT_COLS]):
            field = match_field(_norm_token(raw))
            if field and field not in current:
                current[field] = col_idx
        score = len(current)
        if "atividade" in current:
            score += 2
        if "status" in current:
            score += 1
        if score > best_score:
            best_score = score
            best_row = idx
            best_map = current

    if best_row is None or best_score < 3 or "atividade" not in best_map:
        return None, {}
    row_src = rows[best_row] if best_row < len(rows) and isinstance(rows[best_row], list) else []
    axis_terms = {
        "SETOR",
        "BLOCO",
        "PAVIMENTO",
        "ANDAR",
        "NIVEL",
        "APTO",
        "APARTAMENTO",
        "APART",
        "UNIDADE",
        "UNID",
        "UND",
        "UH",
        "LOCAL",
        "TORRE",
        "ALA",
        "VILA",
        "QUADRA",
        "LOTE",
        "FASE",
        "MODULO",
        "NUCLEO",
        "TIPOLOGIA",
    }
    ignore_fields = {"atividade", "status", "grupo_servicos", "custo", "observacao", "data_termino"}
    axis_cols = []
    for col_idx, raw in enumerate(row_src):
        token = _norm_token(raw)
        if not token:
            continue
        mapped_field = match_field(token)
        if mapped_field and mapped_field not in ignore_fields:
            axis_cols.append((col_idx, _excel_value_to_text(raw) or mapped_field.title()))
            continue
        if any(term == token or token.startswith(f"{term} ") or f" {term} " in f" {token} " for term in axis_terms):
            axis_cols.append((col_idx, _excel_value_to_text(raw) or "Local"))
    if axis_cols:
        dedup = []
        seen_cols = set()
        for idx, label in axis_cols:
            if idx in seen_cols:
                continue
            seen_cols.add(idx)
            dedup.append((idx, label))
        best_map["_axis_cols"] = dedup
    return best_row, best_map


def _find_minimal_activity_header(rows: list[list[str]]) -> tuple[int | None, dict]:
    if not rows:
        return None, {}
    scan_limit = min(len(rows), 40)
    for idx in range(scan_limit):
        row = rows[idx] if isinstance(rows[idx], list) else []
        for col_idx, raw in enumerate(row[:PO_MAX_IMPORT_COLS]):
            token = _norm_token(raw)
            if token in {"ATIVIDADE", "SERVICO", "SERVICOS", "ITEM", "ETAPA"}:
                return idx, {"atividade": col_idx}
    return None, {}


def _infer_status_column(rows: list[list[str]], header_idx: int, col_map: dict) -> int | None:
    mapped = col_map.get("status")
    if mapped is not None:
        return mapped
    activity_col = col_map.get("atividade")
    ignore = {activity_col}
    for key in ("setor", "bloco", "pavimento", "unidade"):
        if col_map.get(key) is not None:
            ignore.add(col_map.get(key))
    for idx, _ in col_map.get("_axis_cols") or []:
        ignore.add(idx)

    if not rows or header_idx >= len(rows) - 1:
        return None

    col_count = max((len(r) for r in rows), default=0)
    best_idx = None
    best_score = -1
    data_rows = rows[header_idx + 1 : header_idx + 1 + 1200]
    for c in range(col_count):
        if c in ignore:
            continue
        hit = 0
        non_empty = 0
        for row in data_rows:
            if c >= len(row):
                continue
            txt = _excel_value_to_text(row[c])
            if not txt:
                continue
            non_empty += 1
            if _parse_percent_value(txt) is not None:
                hit += 1
        if non_empty < 6:
            continue
        score = hit * 2 + non_empty
        if score > best_score and (hit / max(1, non_empty)) >= 0.35:
            best_score = score
            best_idx = c
    return best_idx


def _build_matrix_from_records(rows: list[list[str]], header_idx: int, col_map: dict) -> list[list[str]]:
    activities = []
    activities_seen = set()
    locals_axis = []
    locals_seen = set()
    cell_data = {}

    def add_activity(name: str) -> str:
        key = _norm_token(name)
        if not key:
            return ""
        if key not in activities_seen:
            activities_seen.add(key)
            activities.append((key, name[:120]))
        return key

    axis_cols = list(col_map.get("_axis_cols") or [])
    if not axis_cols:
        for k in ("setor", "bloco", "pavimento", "unidade"):
            idx = col_map.get(k)
            if idx is not None:
                axis_cols.append((idx, k.title()))
    axis_headers = [str(label).strip() or "Local" for _, label in axis_cols]
    if not axis_headers:
        axis_headers = ["Local"]

    status_col = _infer_status_column(rows, header_idx, col_map)
    for row in rows[header_idx + 1 :]:
        if not isinstance(row, list):
            continue
        atividade = _excel_value_to_text(row[col_map["atividade"]] if col_map["atividade"] < len(row) else "")
        if not atividade:
            continue
        atividade_key = add_activity(atividade)
        if not atividade_key:
            continue

        axis_values = []
        for col_idx, _label in axis_cols:
            if col_idx is None or col_idx >= len(row):
                axis_values.append("")
                continue
            axis_values.append(_excel_value_to_text(row[col_idx]))
        if not axis_values:
            axis_values = [_excel_value_to_text(row[0] if row else "")]
        axis_key = tuple(_norm_token(v) for v in axis_values)
        if not any(axis_key):
            axis_values = ["Sem local"] + [""] * (len(axis_headers) - 1)
            axis_key = tuple(_norm_token(v) for v in axis_values)
        if axis_key not in locals_seen:
            locals_seen.add(axis_key)
            locals_axis.append((axis_key, axis_values))

        status_raw = _excel_value_to_text(row[status_col] if status_col is not None and status_col < len(row) else "")
        pct = _parse_percent_value(status_raw)
        display = f"{round(pct)}%" if pct is not None else status_raw[:30]
        key = (axis_key, atividade_key)
        prev = cell_data.get(key)
        if prev is None:
            cell_data[key] = {"display": display, "pct": pct}
        else:
            prev_pct = prev.get("pct")
            if pct is not None and (prev_pct is None or pct > prev_pct):
                prev["pct"] = pct
                prev["display"] = display
            elif not prev.get("display") and display:
                prev["display"] = display

    if not activities or not locals_axis:
        return []

    header = axis_headers + [label for _, label in activities]
    out = [header]
    for axis_key, axis_values in locals_axis:
        row = list(axis_values)
        for act_key, _ in activities:
            item = cell_data.get((axis_key, act_key))
            val = item["display"] if item else ""
            row.append(val)
        out.append(row)

    return _normalize_import_rows(out)


def _build_matrix_from_activity_columns(rows: list[list[str]], header_idx: int, col_map: dict) -> list[list[str]]:
    if not rows or header_idx >= len(rows):
        return []
    head = rows[header_idx]
    activity_col = col_map.get("atividade")
    if activity_col is None:
        return []

    ignore_cols = {activity_col}
    for key in ("status", "setor", "bloco", "pavimento", "unidade"):
        if col_map.get(key) is not None:
            ignore_cols.add(col_map.get(key))

    data_rows = rows[header_idx + 1 :]
    col_count = max((len(r) for r in rows), default=0)
    axis_cols = []
    for c in range(col_count):
        if c in ignore_cols:
            continue
        label = _excel_value_to_text(head[c] if c < len(head) else "")
        if not label:
            continue
        token = _norm_token(label)
        if token in {"GRUPO", "GRUPO DE SERVICO", "OBS", "OBSERVACAO"}:
            continue
        non_empty = 0
        pct_hits = 0
        for row in data_rows[:1200]:
            if c >= len(row):
                continue
            val = _excel_value_to_text(row[c])
            if not val:
                continue
            non_empty += 1
            if _parse_percent_value(val) is not None:
                pct_hits += 1
        if non_empty >= 3 and pct_hits >= 2:
            axis_cols.append((c, label[:120]))

    if len(axis_cols) < 2:
        return []

    activities = []
    act_seen = set()
    data = {}
    for row in data_rows:
        if activity_col >= len(row):
            continue
        activity = _excel_value_to_text(row[activity_col])
        if not activity:
            continue
        akey = _norm_token(activity)
        if not akey:
            continue
        if akey not in act_seen:
            act_seen.add(akey)
            activities.append((akey, activity[:120]))
        for c, axis_label in axis_cols:
            raw = _excel_value_to_text(row[c] if c < len(row) else "")
            if not raw:
                continue
            pct = _parse_percent_value(raw)
            disp = f"{round(pct)}%" if pct is not None else raw[:30]
            key = (_norm_token(axis_label), akey)
            prev = data.get(key)
            if prev is None:
                data[key] = {"disp": disp, "pct": pct}
            else:
                prev_pct = prev.get("pct")
                if pct is not None and (prev_pct is None or pct > prev_pct):
                    prev["pct"] = pct
                    prev["disp"] = disp

    if not activities:
        return []
    out = [["Unidade / eixo"] + [lbl for _, lbl in activities]]
    for c, axis_label in axis_cols:
        row = [axis_label]
        axis_key = _norm_token(axis_label)
        for akey, _ in activities:
            item = data.get((axis_key, akey))
            row.append(item["disp"] if item else "")
        out.append(row)
    return _normalize_import_rows(out)


def _is_already_matrix_shape(rows: list[list[str]]) -> bool:
    if len(rows) < 2:
        return False
    head = [_norm_token(c) for c in rows[0][:12]]
    tabular_markers = {"SETOR", "BLOCO", "PAVIMENTO", "APTO", "APARTAMENTO", "UNIDADE", "ATIVIDADE", "STATUS", "SERVICO"}
    if any(token in tabular_markers for token in head):
        return False
    semantic_tokens = [
        t for t in head if t and len(t) >= 3 and not t.isdigit() and t not in {"0", "1", "OK"}
    ]
    binary_ratio = _binary_ratio_rows(rows[:120], max_cells=2600)
    if binary_ratio > 0.72 and len(semantic_tokens) < 4:
        return False
    has_axis = any("BLOCO" in c or "LOCAL" in c or "EIXO" in c for c in head)
    wide = len(rows[0]) >= 5
    return wide and (has_axis or len(semantic_tokens) >= 5)


def _interpret_import_rows(rows: list[list[str]], mode: str = "auto") -> tuple[list[list[str]], str, dict]:
    report = {
        "mode": mode or "auto",
        "header_idx": None,
        "mapped_fields": [],
        "confidence": 0.0,
        "reason": "",
        "strategy_scores": {},
    }
    if not rows:
        return [], "vazio", report

    mode_norm = str(mode or "auto").strip().lower()
    if mode_norm == "raw":
        report["confidence"] = 1.0
        report["reason"] = "Modo bruto forçado pelo utilizador."
        return rows, "forcado_bruto", report

    candidates = []

    if _is_already_matrix_shape(rows):
        candidates.append(("matriz_detectada", rows, 0.82, "A planilha já possui formato de matriz."))

    header_idx, col_map = _find_header_and_map(rows)
    if header_idx is not None and col_map:
        report["header_idx"] = header_idx
        report["mapped_fields"] = sorted([k for k in col_map.keys() if not str(k).startswith("_")])
        has_axis_field = bool(col_map.get("_axis_cols")) or any(
            col_map.get(k) is not None for k in ("setor", "bloco", "pavimento", "unidade")
        )
        if has_axis_field:
            matrix = _build_matrix_from_records(rows, header_idx, col_map)
            if matrix:
                reason = "Detectado formato tabular por registros com eixo local + atividade."
                candidates.append(("pivot_registros", matrix, 0.90, reason))

        matrix2 = _build_matrix_from_activity_columns(rows, header_idx, col_map)
        if matrix2:
            reason = "Detectado formato com atividades em linhas e progresso em colunas."
            candidates.append(("pivot_atividade_colunas", matrix2, 0.76, reason))
    else:
        # Fallback inteligente: alguns modelos trazem apenas cabeçalho "ATIVIDADE" + colunas de unidades.
        h2, m2 = _find_minimal_activity_header(rows)
        if h2 is not None and m2:
            report["header_idx"] = h2
            report["mapped_fields"] = ["atividade"]
            matrix3 = _build_matrix_from_activity_columns(rows, h2, m2)
            if matrix3:
                reason = "Detectado cabeçalho mínimo de atividade com colunas percentuais."
                candidates.append(("pivot_atividade_colunas", matrix3, 0.68, reason))

    for name, _data, conf, _reason in candidates:
        report["strategy_scores"][name] = conf

    if candidates:
        chosen_name, chosen_rows, chosen_conf, chosen_reason = max(candidates, key=lambda it: it[2])
        report["confidence"] = chosen_conf
        report["reason"] = chosen_reason
        return chosen_rows, chosen_name, report

    if mode_norm == "pivot":
        report["reason"] = "Modo pivot forçado, mas nenhuma estratégia alcançou confiança mínima."
        return [], "pivot_sem_confianca", report
    report["confidence"] = 0.35
    report["reason"] = "Sem padrão confiável detectado; mantendo importação bruta."
    return rows, "fallback_bruto", report


def _detect_total_col_idx_from_header(header: list[str]) -> int | None:
    if not isinstance(header, list):
        return None
    for idx in range(len(header) - 1, -1, -1):
        token = _norm_token(header[idx])
        if token == "TOTAL" or token == "TOTAL GERAL" or token.startswith("TOTAL"):
            return idx
    return None


def _build_interpretation_metadata(
    *,
    raw_rows: list[list[str]],
    interpreted_rows: list[list[str]],
    strategy: str,
    report: dict,
    read_diag: dict,
) -> dict:
    meta: dict = {
        "strategy": str(strategy or "").strip(),
        "confidence": float((report or {}).get("confidence") or 0.0),
        "sheet": str((read_diag or {}).get("selected_sheet") or ""),
        "header_idx": (report or {}).get("header_idx"),
        "status_col_source": None,
        "service_group_col_source": None,
        "total_col_source": None,
        "axis_cols_source": [],
        "axis_headers_source": [],
        "activity_col_source": None,
        "auxiliary_cols_source": [],
        "activity_group_map": {},
        "activity_cols_interpreted": [],
        "activity_headers_interpreted": [],
        "axis_cols_interpreted": [],
        "axis_headers_interpreted": [],
        "total_col_interpreted": None,
        "ignored_auxiliary_cols_source": [],
    }
    if not raw_rows:
        return meta

    header_idx, col_map = _find_header_and_map(raw_rows)
    if header_idx is not None and isinstance(col_map, dict):
        src_header = raw_rows[header_idx] if header_idx < len(raw_rows) else []
        axis_cols = list(col_map.get("_axis_cols") or [])
        axis_cols_norm = []
        axis_headers_norm = []
        for col_idx, label in axis_cols:
            if not isinstance(col_idx, int):
                continue
            axis_cols_norm.append(col_idx)
            label_txt = str(label or "").strip()
            if not label_txt and col_idx < len(src_header):
                label_txt = _excel_value_to_text(src_header[col_idx]) or f"Eixo {col_idx + 1}"
            axis_headers_norm.append(label_txt or f"Eixo {col_idx + 1}")
        meta["axis_cols_source"] = axis_cols_norm
        meta["axis_headers_source"] = axis_headers_norm
        meta["activity_col_source"] = col_map.get("atividade")
        meta["service_group_col_source"] = col_map.get("grupo_servicos")
        meta["status_col_source"] = col_map.get("status")
        meta["total_col_source"] = _detect_total_col_idx_from_header(src_header)
        aux_cols = []
        for key, reason in (
            ("custo", "custo_auxiliar"),
            ("observacao", "observacao_auxiliar"),
            ("data_termino", "data_auxiliar"),
        ):
            col_idx = col_map.get(key)
            if not isinstance(col_idx, int):
                continue
            head_txt = _excel_value_to_text(src_header[col_idx] if col_idx < len(src_header) else "")
            aux_cols.append({"key": key, "col": col_idx, "header": head_txt, "reason": reason})
        meta["auxiliary_cols_source"] = aux_cols

        act_col = col_map.get("atividade")
        grp_col = col_map.get("grupo_servicos")
        if isinstance(act_col, int) and isinstance(grp_col, int):
            act_group: dict[str, str] = {}
            start = header_idx + 1 if isinstance(header_idx, int) else 1
            for row in raw_rows[start:]:
                if not isinstance(row, list):
                    continue
                atividade = _excel_value_to_text(row[act_col] if act_col < len(row) else "")
                grupo = _excel_value_to_text(row[grp_col] if grp_col < len(row) else "")
                if not atividade:
                    continue
                key = atividade.strip().upper()
                if key and key not in act_group:
                    act_group[key] = grupo
            meta["activity_group_map"] = act_group

        ignored = []
        keep = set(axis_cols_norm)
        if isinstance(col_map.get("atividade"), int):
            keep.add(col_map["atividade"])
        if isinstance(col_map.get("grupo_servicos"), int):
            keep.add(col_map["grupo_servicos"])
        if isinstance(col_map.get("status"), int):
            keep.add(col_map["status"])
        for key in ("custo", "observacao", "data_termino"):
            if isinstance(col_map.get(key), int):
                keep.add(col_map[key])
        if isinstance(meta["total_col_source"], int):
            keep.add(meta["total_col_source"])
        for col_idx, raw in enumerate(src_header):
            txt = _excel_value_to_text(raw)
            if not txt:
                continue
            if col_idx in keep:
                continue
            ignored.append({"col": col_idx, "header": txt, "reason": "coluna_auxiliar_fora_do_eixo_principal"})
        meta["ignored_auxiliary_cols_source"] = ignored

    if interpreted_rows and isinstance(interpreted_rows[0], list):
        ih = interpreted_rows[0]
        total_interpreted = _detect_total_col_idx_from_header(ih)
        meta["total_col_interpreted"] = total_interpreted
        axis_cols_interpreted = []
        axis_headers_interpreted = []
        if strategy == "pivot_registros" and meta["axis_headers_source"]:
            axis_count = min(len(meta["axis_headers_source"]), len(ih))
            axis_cols_interpreted = list(range(axis_count))
            axis_headers_interpreted = [str(ih[i] or "").strip() or f"Eixo {i + 1}" for i in axis_cols_interpreted]
        meta["axis_cols_interpreted"] = axis_cols_interpreted
        meta["axis_headers_interpreted"] = axis_headers_interpreted

        activity_cols = []
        activity_headers = []
        for idx, raw in enumerate(ih):
            if idx in axis_cols_interpreted:
                continue
            if total_interpreted is not None and idx == total_interpreted:
                continue
            activity_cols.append(idx)
            activity_headers.append(str(raw or "").strip() or f"Atividade {len(activity_cols)}")
        meta["activity_cols_interpreted"] = activity_cols
        meta["activity_headers_interpreted"] = activity_headers

    return meta


def _extrair_primeira_matriz_rows(layout: dict) -> list[list[str]] | None:
    if not isinstance(layout, dict):
        return None
    sections = layout.get("sections")
    if not isinstance(sections, list):
        return None
    for section in sections:
        if not isinstance(section, dict):
            continue
        if str(section.get("kind") or "").strip() not in {"matrix_table", "table"}:
            continue
        data = section.get("data")
        if not isinstance(data, dict):
            continue
        rows = data.get("rows")
        if not isinstance(rows, list) or not rows:
            continue
        normalized = []
        col_count = 0
        for row in rows:
            if isinstance(row, list):
                out = [str(cell or "") for cell in row]
            else:
                out = [str(row or "")]
            normalized.append(out)
            col_count = max(col_count, len(out))
        if col_count <= 0:
            continue
        for row in normalized:
            if len(row) < col_count:
                row.extend([""] * (col_count - len(row)))
        return normalized
    return None


PO_MAX_MATRIX_CELLS_SYNC = 50_000


def _count_matrix_cells_in_sync_payload(raw_items):
    total = 0
    if not isinstance(raw_items, list):
        return 0
    for item in raw_items:
        if not isinstance(item, dict):
            continue
        kind = str(item.get("kind") or item.get("tipo") or "").strip()
        if kind not in ("matrix_table", "table"):
            continue
        data = item.get("data")
        if not isinstance(data, dict):
            continue
        rows = data.get("rows")
        if not isinstance(rows, list):
            continue
        for row in rows:
            if isinstance(row, list):
                total += len(row)
    return total


def _serializar_semanticas():
    return [
        {
            "id": item.id,
            "nome_canonico": item.nome_canonico,
            "descricao_operacional": item.descricao_operacional,
            "dominio": item.dominio,
            "origem_dado": item.origem_dado,
        }
        for item in SemanticaIndicador.objects.filter(ativo=True).order_by("dominio", "nome_canonico")
    ]


def _map_kind_to_element_type(kind: str):
    if kind == "matrix_table":
        return "table"
    if kind == "kpi_strip":
        return "kpi"
    if kind == "detail_panel":
        return "area"
    return "block"


def _map_element_type_to_kind(element_type: str):
    if element_type == "table":
        return "matrix_table"
    if element_type == "kpi":
        return "kpi_strip"
    if element_type == "area":
        return "detail_panel"
    return "block"


def _sync_layout_to_elementos(ambiente: AmbienteOperacional, versao: AmbienteVersao):
    layout = versao.layout if isinstance(versao.layout, dict) else {}
    sections = layout.get("sections", [])
    if not isinstance(sections, list):
        return

    keys = []
    for idx, section in enumerate(sections):
        if not isinstance(section, dict):
            continue
        section_id = (section.get("id") or "").strip() or f"sec_{uuid4().hex[:8]}"
        keys.append(section_id)
        elemento, _ = AmbienteElemento.objects.get_or_create(
            ambiente=ambiente,
            chave_externa=section_id,
            defaults={"titulo": section.get("title") or "", "tipo": _map_kind_to_element_type(section.get("kind") or "")},
        )
        elemento.versao = versao
        elemento.titulo = (section.get("title") or "").strip()
        elemento.tipo = _map_kind_to_element_type((section.get("kind") or "").strip())
        elemento.x = int(section.get("x") or 0)
        elemento.y = int(section.get("y") or 0)
        elemento.width = int(section.get("width") or 320)
        elemento.height = int(section.get("height") or 180)
        elemento.z_index = idx
        camada = section.get("layer")
        if not isinstance(camada, dict):
            camada = {}
        elemento.camada = camada
        # Preservar payload completo da matriz (rows, heatmap, cabeçalhos, etc.). Sobrescrever só com
        # kind/semantica do section estraga o JSON em disco e, no próximo GET, o editor repõe o modelo vazio.
        matrix_payload = dict(section.get("data")) if isinstance(section.get("data"), dict) else {}
        matrix_payload["kind"] = section.get("kind")
        matrix_payload["semantica"] = (section.get("semantica") or matrix_payload.get("semantica") or "").strip()
        elemento.dados = matrix_payload
        elemento.ativo = True
        elemento.origem_layout = True
        elemento.save()

        if elemento.tipo == "table":
            matrix_data = section.get("data") if isinstance(section.get("data"), dict) else {}
            rows = matrix_data.get("rows", [])
            if isinstance(rows, list):
                elemento.celulas.all().delete()
                new_cells = []
                for r_idx, row in enumerate(rows):
                    if not isinstance(row, list):
                        continue
                    for c_idx, value in enumerate(row):
                        new_cells.append(
                            AmbienteCelula(
                                elemento=elemento,
                                linha_idx=r_idx,
                                coluna_idx=c_idx,
                                valor=str(value) if value is not None else "",
                                tipo="texto",
                            )
                        )
                if new_cells:
                    AmbienteCelula.objects.bulk_create(new_cells, batch_size=500)

    # Proteção contra payload vazio/inválido: evita inativação em massa por acidente.
    if not keys:
        return
    AmbienteElemento.objects.filter(ambiente=ambiente, origem_layout=True).exclude(chave_externa__in=keys).update(ativo=False)


@login_required
@require_group(GRUPOS.FERRAMENTA_OPERACIONAL)
@cache_control(no_cache=True, no_store=True, must_revalidate=True)
def ferramenta_shell(request):
    ctx = _resolver_obra(request)
    obras, obra = ctx
    ambientes = []
    importar_url = reverse("engenharia:importar_mapa_controle")
    if obra:
        ambientes = [
            _serializar_ambiente(amb)
            for amb in AmbienteOperacional.objects.filter(obra=obra, ativo=True).order_by("-updated_at")[:20]
        ]
        importar_url = f"{importar_url}?obra={obra.id}"

    return render(
        request,
        "painel_operacional/ferramenta_shell.html",
        {
            "ambientes_json": json.dumps(ambientes),
            "tipos_ambiente": AmbienteTipo.choices,
            "importar_mapa_url": importar_url,
            "importacao_planilha_criacao_desabilitada": PO_IMPORTACAO_PLANILHA_CRIACAO_DESABILITADA,
            **ctx.to_template_context(),
        },
    )


@login_required
@require_group(GRUPOS.FERRAMENTA_OPERACIONAL)
def editor_ambiente(request, ambiente_id: int):
    _, obra = _resolver_obra(request)
    ambiente = get_object_or_404(AmbienteOperacional, id=ambiente_id, ativo=True)
    if not obra or ambiente.obra_id != obra.id:
        return render(
            request,
            "painel_operacional/editor_ambiente.html",
            {"erro_acesso": "Ambiente não pertence à obra ativa."},
            status=403,
        )

    draft = ambiente.versoes.filter(estado=VersaoEstado.DRAFT).order_by("-numero").first()
    is_mapa_controle = ambiente.tipo == AmbienteTipo.MAPA_CONTROLE
    obra_id = obra.id
    mapa_url = f"{reverse('engenharia:mapa_controle')}?obra={obra_id}&ambiente_id={ambiente.id}&embed=1"
    importar_url = f"{reverse('engenharia:importar_mapa_controle')}?obra={obra_id}"
    exportar_url = reverse("suprimentos:po_api_exportar_matriz_excel", args=[ambiente.id])
    editor_url = reverse("engenharia:ferramenta_editor_ambiente", args=[ambiente.id])
    mapa_edit_url = reverse("engenharia:ferramenta_editar_mapa_controle", args=[ambiente.id])

    editor_mode = _resolve_editor_mode(ambiente)
    modo = (request.GET.get("modo") or "").strip().lower()
    force_quadro = modo == "quadro"
    embed_mode = (request.GET.get("embed") or "").strip() == "1"

    base_ctx = {
        "ambiente": ambiente,
        "obra_selecionada": obra,
        "versao_json": json.dumps(_serializar_versao(draft) or {}),
        "semanticas_json": json.dumps(_serializar_semanticas()),
        "is_mapa_controle": is_mapa_controle,
        "editor_mode": editor_mode,
        "mapa_atual_url": mapa_url,
        "mapa_edit_url": mapa_edit_url,
        "importar_mapa_url": importar_url,
        "exportar_mapa_url": exportar_url,
        "force_quadro": force_quadro,
        "embed_mode": embed_mode,
        "editor_url": editor_url,
        "editor_quadro_url": f"{editor_url}?modo=quadro",
        "editor_quadro_embed_url": f"{editor_url}?modo=quadro&embed=1",
    }
    if is_mapa_controle and not force_quadro:
        return render(request, "painel_operacional/editor_mapa_controle.html", base_ctx)
    if (not is_mapa_controle) and editor_mode == AmbienteModoEditor.MAPA_DEDICADO and not force_quadro:
        return render(request, "painel_operacional/editor_mapa_controle.html", base_ctx)
    return render(request, "painel_operacional/editor_ambiente.html", base_ctx)


@login_required
@require_group(GRUPOS.FERRAMENTA_OPERACIONAL)
@cache_control(no_cache=True, no_store=True, must_revalidate=True)
def editar_mapa_controle(request, ambiente_id: int):
    _, obra = _resolver_obra(request)
    ambiente = get_object_or_404(AmbienteOperacional, id=ambiente_id, ativo=True)
    if not obra or ambiente.obra_id != obra.id:
        return render(
            request,
            "painel_operacional/editor_ambiente.html",
            {"erro_acesso": "Ambiente não pertence à obra ativa."},
            status=403,
        )
    if ambiente.tipo != AmbienteTipo.MAPA_CONTROLE:
        return render(
            request,
            "painel_operacional/editor_ambiente.html",
            {"erro_acesso": "Este ambiente não é do tipo Mapa de Controle/Serviço."},
            status=400,
        )

    mapa_base = reverse("engenharia:mapa_controle")
    mapa_default_url = f"{mapa_base}?obra={obra.id}&ambiente_id={ambiente.id}&embed=1"
    next_mapa = (request.GET.get("next_mapa") or "").strip()
    if next_mapa.startswith(mapa_base):
        try:
            u = next_mapa
            sep = "&" if "?" in u else "?"
            mapa_url = u if "embed=" in u else f"{u}{sep}embed=1"
            if "ambiente_id=" not in mapa_url:
                mapa_url = f"{mapa_url}&ambiente_id={ambiente.id}" if "?" in mapa_url else f"{mapa_url}?ambiente_id={ambiente.id}"
        except Exception:
            mapa_url = mapa_default_url
    else:
        mapa_url = mapa_default_url
    return render(
        request,
        "painel_operacional/editar_mapa_controle.html",
        {
            "ambiente": ambiente,
            "obra_selecionada": obra,
            "mapa_atual_url": mapa_url,
            "mapa_view_url": reverse("engenharia:ferramenta_editor_ambiente", args=[ambiente.id]),
            "ambiente_id": ambiente.id,
            "api_detalhe_url": reverse("suprimentos:po_api_detalhe_ambiente", args=[ambiente.id]),
            "api_salvar_rascunho_url": reverse("suprimentos:po_api_salvar_rascunho", args=[ambiente.id]),
        },
    )


@login_required
@require_group(GRUPOS.FERRAMENTA_OPERACIONAL)
@require_http_methods(["GET"])
@cache_control(no_cache=True, no_store=True, must_revalidate=True)
def api_listar_ambientes(request):
    obras, obra = _resolver_obra(request)
    if not obra:
        return JsonResponse({"success": True, "items": [], "message": "Nenhuma obra disponível."})

    items = [
        _serializar_ambiente(amb)
        for amb in AmbienteOperacional.objects.filter(obra=obra, ativo=True).order_by("-updated_at")[:100]
    ]
    return JsonResponse(
        {"success": True, "items": items, "obra": {"id": obra.id, "nome": obra.nome, "codigo": obra.codigo_sienge}}
    )


@login_required
@require_group(GRUPOS.FERRAMENTA_OPERACIONAL)
@require_http_methods(["GET"])
def api_detalhe_ambiente(request, ambiente_id: int):
    _, obra = _resolver_obra(request)
    ambiente = get_object_or_404(AmbienteOperacional, id=ambiente_id, ativo=True)
    if not obra or ambiente.obra_id != obra.id:
        return JsonResponse({"success": False, "error": "Ambiente não pertence à obra ativa."}, status=403)

    draft = ambiente.versoes.filter(estado=VersaoEstado.DRAFT).order_by("-numero").first()
    if draft and isinstance(draft.layout, dict):
        draft.layout = _normalize_ambiente_layout(draft.layout)
    elementos = (
        AmbienteElemento.objects.filter(ambiente=ambiente, ativo=True)
        .order_by("z_index", "id")
        .values(
            "id",
            "chave_externa",
            "titulo",
            "tipo",
            "x",
            "y",
            "width",
            "height",
            "z_index",
            "camada",
            "dados",
        )
    )
    return JsonResponse(
        {
            "success": True,
            "ambiente": _serializar_ambiente(ambiente),
            "modo_editor": _resolve_editor_mode(ambiente),
            "versao": _serializar_versao(draft),
            "draft": _serializar_versao(draft),
            "semanticas": _serializar_semanticas(),
            "elementos": list(elementos),
        }
    )


@login_required
@require_group(GRUPOS.FERRAMENTA_OPERACIONAL)
@require_http_methods(["POST"])
def api_criar_ambiente(request):
    payload = _parse_json_body(request)
    _, obra = _resolver_obra(request)
    if not obra:
        return JsonResponse({"success": False, "error": "Selecione uma obra válida."}, status=400)

    nome = (payload.get("nome") or "").strip() or "Novo ambiente"
    tipo = (payload.get("tipo") or AmbienteTipo.MAPA_CONTROLE).strip()
    tipos_validos = {choice[0] for choice in AmbienteTipo.choices}
    if tipo not in tipos_validos:
        tipo = AmbienteTipo.CUSTOM
    modo_editor = (
        AmbienteModoEditor.MAPA_DEDICADO
        if tipo == AmbienteTipo.MAPA_CONTROLE
        else AmbienteModoEditor.QUADRO
    )

    with transaction.atomic():
        ambiente = AmbienteOperacional.objects.create(
            obra=obra,
            nome=nome,
            tipo=tipo,
            modo_editor=modo_editor,
            descricao=(payload.get("descricao") or "").strip(),
            criado_por=request.user,
        )
        versao = AmbienteVersao.objects.create(
            ambiente=ambiente,
            numero=1,
            estado=VersaoEstado.DRAFT,
            layout=_preset_layout(tipo, obra=obra),
            metadados={"preset": tipo, "modo_editor": modo_editor},
        )
        AmbienteHistorico.objects.create(
            ambiente=ambiente,
            versao=versao,
            usuario=request.user,
            acao=AmbienteHistorico.ACAO_CRIAR,
            detalhes={"preset": tipo, "modo_editor": modo_editor},
        )

    return JsonResponse({"success": True, "item": _serializar_ambiente(ambiente)})


@login_required
@require_group(GRUPOS.FERRAMENTA_OPERACIONAL)
@require_http_methods(["POST"])
def api_excluir_ambiente(request, ambiente_id: int):
    _, obra = _resolver_obra(request)
    if not obra:
        return JsonResponse({"success": False, "error": "Selecione uma obra válida."}, status=400)
    ambiente = get_object_or_404(AmbienteOperacional, id=ambiente_id)
    if ambiente.obra_id != obra.id:
        return JsonResponse({"success": False, "error": "Ambiente não pertence à obra ativa."}, status=403)
    if not ambiente.ativo:
        return JsonResponse({"success": True, "message": "Ambiente já estava removido."})
    ambiente.ativo = False
    ambiente.save(update_fields=["ativo", "updated_at"])
    return JsonResponse({"success": True})


@login_required
@require_group(GRUPOS.FERRAMENTA_OPERACIONAL)
@require_http_methods(["POST"])
def api_adicionar_secao(request, ambiente_id: int):
    payload = _parse_json_body(request)
    _, obra = _resolver_obra(request)
    ambiente = get_object_or_404(AmbienteOperacional, id=ambiente_id, ativo=True)
    if not obra or ambiente.obra_id != obra.id:
        return JsonResponse({"success": False, "error": "Ambiente não pertence à obra ativa."}, status=403)

    titulo = (payload.get("title") or "").strip()
    tipo = (payload.get("kind") or "").strip()
    if not titulo or not tipo:
        return JsonResponse({"success": False, "error": "Informe título e tipo da seção."}, status=400)

    with transaction.atomic():
        draft = ambiente.versoes.filter(estado=VersaoEstado.DRAFT).order_by("-numero").first()
        if not draft:
            draft = AmbienteVersao.objects.create(
                ambiente=ambiente,
                numero=AmbienteVersao.proximo_numero(ambiente.id),
                estado=VersaoEstado.DRAFT,
                layout=_preset_layout(ambiente.tipo, obra=ambiente.obra),
                metadados={},
            )

        layout = draft.layout if isinstance(draft.layout, dict) else {}
        sections = layout.get("sections", [])
        if not isinstance(sections, list):
            sections = []
        section = {
            "id": payload.get("id") or f"sec_{uuid4().hex[:8]}",
            "title": titulo,
            "kind": tipo,
            "x": payload.get("x", 80 + ((len(sections) % 4) * 280)),
            "y": payload.get("y", 80 + ((len(sections) // 4) * 220)),
            "width": payload.get("width", 320 if tipo != "matrix_table" else 560),
            "height": payload.get("height", 180 if tipo != "matrix_table" else 320),
            "layer": payload.get("layer") if isinstance(payload.get("layer"), dict) else {},
        }
        semantica = (payload.get("semantica") or "").strip()
        if semantica:
            section["semantica"] = semantica
        if tipo == "matrix_table":
            rows_base = (
                _mapa_controle_rows_canonico(20, 0)
                if ambiente.tipo == AmbienteTipo.MAPA_CONTROLE
                else [
                    ["", "Grupo A", "Grupo A", "Grupo B", "Grupo B", "Total"],
                    ["Eixo (linhas)", "Etapa 1", "Etapa 2", "Etapa 3", "Etapa 4", ""],
                    ["Local 1", "", "", "", "", ""],
                    ["Local 2", "", "", "", "", ""],
                ]
            )
            w_base = (
                _mapa_controle_weights(rows_base, totals_row_auto=True)
                if ambiente.tipo == AmbienteTipo.MAPA_CONTROLE
                else {}
            )
            section["data"] = {
                "mapaControleTemplate": ambiente.tipo == AmbienteTipo.MAPA_CONTROLE,
                "headerBandCount": 1,
                "heatmap": True,
                "totalsColumnAuto": True,
                "totalsRowAuto": True,
                "verticalHeaders": True,
                "rows": rows_base,
                "colWeights": w_base.get("colWeights"),
                "rowWeights": w_base.get("rowWeights"),
            }
        sections.append(section)
        layout["sections"] = sections
        draft.layout = layout
        draft.save(update_fields=["layout", "updated_at"])

        AmbienteHistorico.objects.create(
            ambiente=ambiente,
            versao=draft,
            usuario=request.user,
            acao=AmbienteHistorico.ACAO_SALVAR,
            detalhes={"acao_editor": "adicionar_secao", "secao_id": section["id"], "kind": tipo},
        )

    return JsonResponse({"success": True, "draft": _serializar_versao(draft)})


@login_required
@require_group(GRUPOS.FERRAMENTA_OPERACIONAL)
@require_http_methods(["POST"])
def api_salvar_rascunho(request, ambiente_id: int):
    payload = _parse_json_body(request)
    _, obra = _resolver_obra(request)

    ambiente = get_object_or_404(AmbienteOperacional, id=ambiente_id, ativo=True)
    if not obra or ambiente.obra_id != obra.id:
        return JsonResponse({"success": False, "error": "Ambiente não pertence à obra ativa."}, status=403)

    with transaction.atomic():
        draft = ambiente.versoes.filter(estado=VersaoEstado.DRAFT).order_by("-numero").first()
        if not draft:
            draft = AmbienteVersao.objects.create(
                ambiente=ambiente,
                numero=AmbienteVersao.proximo_numero(ambiente.id),
                estado=VersaoEstado.DRAFT,
                layout={},
                metadados={},
            )

        layout = payload.get("layout")
        metadados = payload.get("metadados")
        if isinstance(layout, dict):
            draft.layout = _normalize_ambiente_layout(layout)
        if isinstance(metadados, dict):
            draft.metadados = metadados
        draft.save(update_fields=["layout", "metadados", "updated_at"])
        _sync_layout_to_elementos(ambiente, draft)

        AmbienteHistorico.objects.create(
            ambiente=ambiente,
            versao=draft,
            usuario=request.user,
            acao=AmbienteHistorico.ACAO_SALVAR,
            detalhes={"keys_layout": sorted(list(draft.layout.keys()))},
        )

    versao_ref = {"numero": draft.numero, "updated_at": draft.updated_at.isoformat()}
    return JsonResponse({"success": True, "versao": versao_ref, "rascunho": versao_ref})


@login_required
@require_group(GRUPOS.FERRAMENTA_OPERACIONAL)
@require_http_methods(["GET"])
def api_listar_elementos(request, ambiente_id: int):
    _, obra = _resolver_obra(request)
    ambiente = get_object_or_404(AmbienteOperacional, id=ambiente_id, ativo=True)
    if not obra or ambiente.obra_id != obra.id:
        return JsonResponse({"success": False, "error": "Ambiente não pertence à obra ativa."}, status=403)
    elementos = list(
        AmbienteElemento.objects.filter(ambiente=ambiente, ativo=True)
        .order_by("z_index", "id")
        .values(
            "id",
            "chave_externa",
            "titulo",
            "tipo",
            "x",
            "y",
            "width",
            "height",
            "z_index",
            "camada",
            "dados",
        )
    )
    return JsonResponse({"success": True, "items": elementos})


@login_required
@require_group(GRUPOS.FERRAMENTA_OPERACIONAL)
@require_http_methods(["POST"])
def api_sync_elementos(request, ambiente_id: int):
    payload = _parse_json_body(request)
    _, obra = _resolver_obra(request)
    ambiente = get_object_or_404(AmbienteOperacional, id=ambiente_id, ativo=True)
    if not obra or ambiente.obra_id != obra.id:
        return JsonResponse({"success": False, "error": "Ambiente não pertence à obra ativa."}, status=403)

    raw_items = payload.get("items")
    if not isinstance(raw_items, list):
        return JsonResponse({"success": False, "error": "Payload inválido: items deve ser lista."}, status=400)
    if not raw_items:
        return JsonResponse(
            {"success": False, "error": "Payload inválido: informe ao menos um item para sincronização."},
            status=400,
        )

    if _count_matrix_cells_in_sync_payload(raw_items) > PO_MAX_MATRIX_CELLS_SYNC:
        return JsonResponse(
            {
                "success": False,
                "error": f"Soma de células das matrizes excede o limite ({PO_MAX_MATRIX_CELLS_SYNC}). Reduza linhas/colunas ou divida blocos.",
            },
            status=400,
        )

    with transaction.atomic():
        draft = ambiente.versoes.filter(estado=VersaoEstado.DRAFT).order_by("-numero").first()
        if not draft:
            draft = AmbienteVersao.objects.create(
                ambiente=ambiente,
                numero=AmbienteVersao.proximo_numero(ambiente.id),
                estado=VersaoEstado.DRAFT,
                layout=_preset_layout(ambiente.tipo, obra=ambiente.obra),
                metadados={},
            )

        existing_by_id = {
            row.id: row for row in AmbienteElemento.objects.filter(ambiente=ambiente)
        }
        existing_by_key = {
            row.chave_externa: row
            for row in AmbienteElemento.objects.filter(ambiente=ambiente).exclude(chave_externa="")
        }

        kept_ids = []
        sections = []

        for idx, item in enumerate(raw_items):
            if not isinstance(item, dict):
                continue
            raw_id = item.get("id")
            raw_key = str(item.get("chave_externa") or item.get("key") or "").strip()
            key = raw_key or f"sec_{uuid4().hex[:8]}"
            titulo = str(item.get("titulo") or item.get("title") or "").strip()
            kind = str(item.get("kind") or item.get("tipo") or "block").strip()
            tipo = _map_kind_to_element_type(kind)

            elemento = None
            if raw_id:
                try:
                    elemento = existing_by_id.get(int(raw_id))
                except (TypeError, ValueError):
                    elemento = None
            if not elemento:
                elemento = existing_by_key.get(key)
            if not elemento:
                elemento = AmbienteElemento(ambiente=ambiente)

            elemento.versao = draft
            elemento.chave_externa = key
            elemento.titulo = titulo
            elemento.tipo = tipo
            elemento.x = int(item.get("x") or 0)
            elemento.y = int(item.get("y") or 0)
            elemento.width = max(80, int(item.get("width") or 320))
            elemento.height = max(60, int(item.get("height") or 180))
            elemento.z_index = idx
            layer = item.get("layer")
            elemento.camada = layer if isinstance(layer, dict) else {}
            data = item.get("data")
            element_data = data if isinstance(data, dict) else {}
            semantica = str(item.get("semantica") or element_data.get("semantica") or "").strip()
            element_data["semantica"] = semantica
            element_data["kind"] = kind
            elemento.dados = element_data
            elemento.ativo = True
            elemento.origem_layout = True
            elemento.save()
            kept_ids.append(elemento.id)

            if elemento.tipo == "table":
                rows = []
                if isinstance(element_data.get("rows"), list):
                    rows = element_data.get("rows")
                elemento.celulas.all().delete()
                cells = []
                for r_idx, row in enumerate(rows):
                    if not isinstance(row, list):
                        continue
                    for c_idx, value in enumerate(row):
                        cells.append(
                            AmbienteCelula(
                                elemento=elemento,
                                linha_idx=r_idx,
                                coluna_idx=c_idx,
                                valor=str(value) if value is not None else "",
                                tipo="texto",
                            )
                        )
                if cells:
                    AmbienteCelula.objects.bulk_create(cells, batch_size=500)

            sections.append(
                {
                    "id": key,
                    "title": titulo,
                    "kind": kind,
                    "x": elemento.x,
                    "y": elemento.y,
                    "width": elemento.width,
                    "height": elemento.height,
                    "layer": elemento.camada,
                    "semantica": semantica,
                    "data": element_data if isinstance(element_data, dict) else {},
                }
            )

        if not kept_ids:
            return JsonResponse(
                {"success": False, "error": "Nenhum item válido foi recebido para sincronização."},
                status=400,
            )
        AmbienteElemento.objects.filter(ambiente=ambiente).exclude(id__in=kept_ids).update(ativo=False)

        layout = draft.layout if isinstance(draft.layout, dict) else {}
        layout["sections"] = sections
        draft.layout = layout
        draft.save(update_fields=["layout", "updated_at"])

        AmbienteHistorico.objects.create(
            ambiente=ambiente,
            versao=draft,
            usuario=request.user,
            acao=AmbienteHistorico.ACAO_SALVAR,
            detalhes={"acao_editor": "sync_elementos", "qtd": len(sections)},
        )

    return JsonResponse(
        {
            "success": True,
            "items": sections,
            "versao": {"numero": draft.numero, "updated_at": draft.updated_at.isoformat()},
            "rascunho": {"numero": draft.numero, "updated_at": draft.updated_at.isoformat()},
        }
    )


@login_required
@require_group(GRUPOS.FERRAMENTA_OPERACIONAL)
@require_http_methods(["POST"])
def api_publicar_ambiente(request, ambiente_id: int):
    _, obra = _resolver_obra(request)
    ambiente = get_object_or_404(AmbienteOperacional, id=ambiente_id, ativo=True)
    if not obra or ambiente.obra_id != obra.id:
        return JsonResponse({"success": False, "error": "Ambiente não pertence à obra ativa."}, status=403)
    # Compatibilidade: endpoint mantido para clientes antigos, mas o fluxo atual é somente "Salvar".
    draft = ambiente.versoes.filter(estado=VersaoEstado.DRAFT).order_by("-numero").first()
    return JsonResponse(
        {
            "success": True,
            "message": "Fluxo de publicação desativado: use salvar para persistir alterações.",
            "versao": _serializar_versao(draft),
            "rascunho": _serializar_versao(draft),
        }
    )


@login_required
@require_group(GRUPOS.FERRAMENTA_OPERACIONAL)
@require_http_methods(["POST"])
def api_importar_matriz_excel(request, ambiente_id: int):
    # Bloqueio alinhado à UI do modal de criação (ferramenta_shell); ver PO_IMPORTACAO_PLANILHA_CRIACAO_DESABILITADA.
    if PO_IMPORTACAO_PLANILHA_CRIACAO_DESABILITADA:
        return _importacao_planilha_criacao_bloqueada_response()

    _, obra = _resolver_obra(request)
    ambiente = AmbienteOperacional.objects.filter(id=ambiente_id, ativo=True).first()
    if not ambiente:
        return JsonResponse({"success": False, "error": "Ambiente não encontrado para importação."}, status=404)
    if not obra or ambiente.obra_id != obra.id:
        return JsonResponse({"success": False, "error": "Ambiente não pertence à obra ativa."}, status=403)

    arquivo = request.FILES.get("arquivo")
    sheet = (request.POST.get("sheet") or "").strip()
    mode = (request.POST.get("mode") or "auto").strip().lower()
    if mode not in {"auto", "pivot", "raw"}:
        mode = "auto"
    if not arquivo:
        return JsonResponse({"success": False, "error": "Selecione um arquivo para importar."}, status=400)

    try:
        rows, sheet_lida, read_diag = _read_excel_rows(arquivo, sheet_name=sheet)
    except ValueError as exc:
        return JsonResponse({"success": False, "error": str(exc)}, status=400)
    except Exception:
        logger.exception("Falha inesperada ao importar matriz Excel", extra={"ambiente_id": ambiente_id, "sheet": sheet})
        return JsonResponse(
            {"success": False, "error": "Não foi possível ler a planilha. Verifique o formato do arquivo."},
            status=400,
        )

    if not rows:
        return JsonResponse({"success": False, "error": "A planilha não possui dados utilizáveis."}, status=400)

    rows_raw = rows
    rows, strategy, report = _interpret_import_rows(rows_raw, mode=mode)
    if not rows:
        return JsonResponse(
            {
                "success": False,
                "error": "Não foi possível estruturar dados úteis da planilha com o modo escolhido.",
                "strategy": strategy,
                "report": report,
            },
            status=400,
        )
    if mode == "auto" and strategy == "fallback_bruto":
        binary_ratio = _binary_ratio_rows(rows, max_cells=5000)
        if binary_ratio >= 0.75:
            return JsonResponse(
                {
                    "success": False,
                    "error": (
                        "A aba selecionada parece uma matriz binária (0/1) sem cabeçalhos descritivos. "
                        "Informe a aba de dados da planilha ou use modo avançado."
                    ),
                    "strategy": strategy,
                    "report": report,
                },
                status=400,
            )

    total_rows = len(rows)
    total_cols = max((len(r) for r in rows), default=0)
    total_cells = total_rows * total_cols
    if total_rows > PO_MAX_IMPORT_ROWS or total_cols > PO_MAX_IMPORT_COLS or total_cells > PO_MAX_IMPORT_CELLS:
        return JsonResponse(
            {
                "success": False,
                "error": (
                    f"Limites: até {PO_MAX_IMPORT_ROWS} linhas, {PO_MAX_IMPORT_COLS} colunas e "
                    f"{PO_MAX_IMPORT_CELLS} células."
                ),
            },
            status=400,
        )

    interpretation_meta = _build_interpretation_metadata(
        raw_rows=rows_raw,
        interpreted_rows=rows,
        strategy=strategy,
        report=report,
        read_diag=read_diag if isinstance(read_diag, dict) else {},
    )

    return JsonResponse(
        {
            "success": True,
            "rows": rows,
            "sheet": sheet_lida,
            "strategy": strategy,
            "report": report,
            "read": read_diag,
            "interpretation_meta": interpretation_meta,
            "meta": {"rows": total_rows, "cols": total_cols, "cells": total_cells},
        }
    )


@login_required
@require_group(GRUPOS.FERRAMENTA_OPERACIONAL)
@require_http_methods(["GET"])
def api_exportar_matriz_excel(request, ambiente_id: int):
    _, obra = _resolver_obra(request)
    ambiente = get_object_or_404(AmbienteOperacional, id=ambiente_id, ativo=True)
    if not obra or ambiente.obra_id != obra.id:
        return JsonResponse({"success": False, "error": "Ambiente não pertence à obra ativa."}, status=403)

    versao = ambiente.versoes.filter(estado=VersaoEstado.DRAFT).order_by("-numero").first()
    if not versao:
        versao = ambiente.versoes.filter(estado=VersaoEstado.PUBLISHED).order_by("-numero").first()
    if not versao:
        return JsonResponse({"success": False, "error": "Ambiente sem versão disponível para exportação."}, status=404)

    rows = _extrair_primeira_matriz_rows(versao.layout if isinstance(versao.layout, dict) else {})
    if not rows:
        return JsonResponse({"success": False, "error": "Nenhuma matriz encontrada para exportação."}, status=400)

    wb = Workbook()
    ws = wb.active
    ws.title = "Mapa de Controle"
    for row in rows:
        ws.append([str(cell or "") for cell in row])

    for col_idx in range(1, min(60, len(rows[0])) + 1):
        letter = ws.cell(row=1, column=col_idx).column_letter
        max_len = 0
        for row_idx in range(1, min(250, len(rows)) + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            max_len = max(max_len, len(str(val)) if val is not None else 0)
        ws.column_dimensions[letter].width = min(28, max(8, max_len + 2))

    stamp = timezone.localtime().strftime("%Y%m%d_%H%M")
    safe_name = slugify(ambiente.nome) or f"ambiente_{ambiente.id}"
    filename = f"{safe_name}_mapa_controle_{stamp}.xlsx"
    payload = BytesIO()
    wb.save(payload)
    payload.seek(0)

    response = HttpResponse(
        payload.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response

