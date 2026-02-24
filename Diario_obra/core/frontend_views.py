"""
Views para frontend do LPlan - Templates Django com HTMX/Alpine.js
"""
# #region agent log
import json
def _dbg(loc, msg, data, hypothesis_id=None):
    try:
        from django.conf import settings
        log_path = getattr(settings, "BASE_DIR", __import__("pathlib").Path(__file__).resolve().parent.parent.parent) / "debug.log"
        payload = {"location": loc, "message": msg, "data": data, "timestamp": __import__("time").time() * 1000}
        if hypothesis_id:
            payload["hypothesisId"] = hypothesis_id
        with open(str(log_path), "a", encoding="utf-8") as f:
            f.write(json.dumps(payload, ensure_ascii=False) + "\n")
    except Exception:
        pass
# #endregion
from functools import wraps
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse, Http404
from django.views.decorators.http import require_http_methods
from django.db.models import Q, Count, Avg, Sum
from django.utils import timezone
from datetime import datetime, timedelta
from .models import (
    Project,
    ProjectMember,
    ProjectOwner,
    ConstructionDiary,
    DiaryComment,
    DiaryImage,
    DailyWorkLog,
    Labor,
    Equipment,
    DiaryStatus,
    Activity,
    ActivityStatus,
)
from django.core.exceptions import PermissionDenied, ValidationError
from accounts.groups import GRUPOS
# PDFGenerator será importado apenas quando necessário (lazy import)
# Suporta WeasyPrint (preferencial) e xhtml2pdf (fallback para Windows)
PDFGenerator = None
WEASYPRINT_AVAILABLE = False
XHTML2PDF_AVAILABLE = False

# Mapeamento obra → contratante para autopreencher o formulário do diário.
# Chave: substring normalizada (lower) do nome ou código da obra.
OBRA_CONTRATANTE_MAP = {
    'entreaguas': 'Incorporadora Adamo',
    'okena': 'JP Empreendimentos',
    'marghot': 'Antonina Hotéis',
    'sunrise': 'Rpontes',
}


def get_contractante_for_project(project):
    """Retorna o nome do contratante: para obras mapeadas usa sempre o mapeamento; senão usa project.client_name."""
    if not project:
        return ''
    name = (getattr(project, 'name', None) or '').strip().lower()
    code = (getattr(project, 'code', None) or '').strip().lower()
    for key, contratante in OBRA_CONTRATANTE_MAP.items():
        if key in name or key in code:
            return contratante
    if getattr(project, 'client_name', None) and (project.client_name or '').strip():
        return (project.client_name or '').strip()
    return ''


def login_view(request):
    """View de login."""
    if request.user.is_authenticated:
        # Sempre redireciona para seleção de sistema (não redireciona automaticamente)
        return redirect('select-system')
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)
            # Limpa obra selecionada anterior (se houver) para forçar nova seleção
            if 'selected_project_id' in request.session:
                del request.session['selected_project_id']
            if 'selected_project_name' in request.session:
                del request.session['selected_project_name']
            if 'selected_project_code' in request.session:
                del request.session['selected_project_code']
            return redirect('select-system')
        else:
            return render(request, 'core/login.html', {'error': 'Credenciais inválidas'})
    
    return render(request, 'core/login.html')


@login_required
def logout_view(request):
    """View de logout."""
    # Limpa a obra selecionada (todos os campos para não vazar dados entre sessões)
    for key in ('selected_project_id', 'selected_project_name', 'selected_project_code'):
        request.session.pop(key, None)
    logout(request)
    return redirect('login')


@login_required
def select_system_view(request):
    """View para seleção de sistema após login."""
    user = request.user
    user_groups = set(user.groups.values_list('name', flat=True))
    
    # Determinar acesso por sistema baseado nos grupos
    has_diario = user.is_superuser or user.is_staff or GRUPOS.GERENTES in user_groups
    has_gestao = user.is_superuser or user.is_staff or bool(
        user_groups & {GRUPOS.ADMINISTRADOR, GRUPOS.RESPONSAVEL_EMPRESA, GRUPOS.APROVADOR, GRUPOS.SOLICITANTE}
    )
    has_mapa = user.is_superuser or user.is_staff or GRUPOS.ENGENHARIA in user_groups
    has_central = user.is_superuser or user.is_staff
    # Dono da obra: se só tem acesso ao portal cliente, redireciona direto
    if not (has_diario or has_gestao or has_mapa or has_central) and _is_work_owner(user):
        return redirect('client-diary-list')
    context = {
        'has_diario': has_diario,
        'has_gestao': has_gestao,
        'has_mapa': has_mapa,
        'has_admin': user.is_superuser or user.is_staff,
        'has_central': has_central,
    }
    return render(request, 'core/select_system.html', context)


@login_required
def central_hub_view(request):
    """Redireciona para o Painel do sistema (hub unificado). Staff/superuser sempre usam o Painel."""
    if not (request.user.is_staff or request.user.is_superuser):
        from django.core.exceptions import PermissionDenied
        raise PermissionDenied('Acesso restrito.')
    from django.shortcuts import redirect
    return redirect('accounts:admin_central')


def _is_work_owner(user):
    """True se o usuário é dono de alguma obra (acesso restrito só às suas obras)."""
    from core.models import ProjectOwner
    return ProjectOwner.objects.filter(user=user).exists()


def _get_projects_for_user(request):
    """Obras que o usuário pode acessar no Diário: staff/superuser vê todas; donos só as que possuem; demais só as vinculadas."""
    from core.models import ProjectOwner
    if request.user.is_staff or request.user.is_superuser:
        return Project.objects.filter(is_active=True).order_by('-created_at')
    # Donos da obra: só veem as obras das quais são donos
    owner_project_ids = list(ProjectOwner.objects.filter(user=request.user).values_list('project_id', flat=True))
    if owner_project_ids:
        return Project.objects.filter(pk__in=owner_project_ids, is_active=True).order_by('-created_at')
    project_ids = list(ProjectMember.objects.filter(user=request.user).values_list('project_id', flat=True))
    # Se não tem vínculo no Diário mas tem obras no GestControll, sincronizar (lista única)
    if not project_ids:
        try:
            from gestao_aprovacao.models import Obra, WorkOrderPermission
            for project_id in Obra.objects.filter(
                permissoes__usuario=request.user,
                permissoes__ativo=True,
            ).exclude(project__isnull=True).values_list('project_id', flat=True).distinct():
                ProjectMember.objects.get_or_create(user=request.user, project_id=project_id)
            project_ids = list(ProjectMember.objects.filter(user=request.user).values_list('project_id', flat=True))
        except Exception:
            pass
    return Project.objects.filter(pk__in=project_ids, is_active=True).order_by('-created_at')


def _user_can_access_project(user, project):
    """Verifica se o usuário pode acessar a obra (dono, vinculado ou staff/superuser)."""
    if user.is_staff or user.is_superuser:
        return True
    from core.models import ProjectOwner
    if ProjectOwner.objects.filter(user=user, project=project).exists():
        return True
    return ProjectMember.objects.filter(user=user, project=project).exists()


@login_required
def select_project_view(request):
    """View para seleção de obra após login. Só aparecem obras às quais o usuário está vinculado (staff vê todas)."""
    projects = _get_projects_for_user(request)

    if request.method == 'POST':
        project_id = request.POST.get('project_id')
        if project_id:
            try:
                project = Project.objects.get(pk=project_id, is_active=True)
                if not _user_can_access_project(request.user, project):
                    messages.error(request, 'Você não está vinculado a esta obra.')
                    return render(request, 'core/select_project.html', {
                        'projects': projects,
                        'selected_project_id': request.session.get('selected_project_id'),
                    })
                request.session['selected_project_id'] = project.id
                request.session['selected_project_name'] = project.name
                request.session['selected_project_code'] = project.code
                return redirect('dashboard')
            except (Project.DoesNotExist, ValueError, TypeError):
                return render(request, 'core/select_project.html', {
                    'error': 'Obra não encontrada ou inativa.',
                    'projects': projects,
                    'selected_project_id': request.session.get('selected_project_id'),
                })

    selected_project_id = request.session.get('selected_project_id')
    return render(request, 'core/select_project.html', {
        'projects': projects,
        'selected_project_id': selected_project_id,
    })


def get_selected_project(request):
    """Helper function para obter a obra selecionada da sessão."""
    project_id = request.session.get('selected_project_id')
    if project_id:
        try:
            return Project.objects.get(pk=project_id, is_active=True)
        except Project.DoesNotExist:
            # Limpa sessão se obra não existe mais
            for key in ('selected_project_id', 'selected_project_name', 'selected_project_code'):
                request.session.pop(key, None)
    return None


def project_required(view_func):
    """Decorator para garantir que uma obra foi selecionada e que o usuário ainda tem acesso a ela."""
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('login')
        if 'selected_project_id' not in request.session:
            return redirect('select-project')
        project = get_selected_project(request)
        if not project:
            return redirect('select-project')
        if not _user_can_access_project(request.user, project):
            for key in ('selected_project_id', 'selected_project_name', 'selected_project_code'):
                request.session.pop(key, None)
            messages.warning(request, 'Você não está mais vinculado a essa obra. Selecione outra.')
            return redirect('select-project')
        return view_func(request, *args, **kwargs)
    return wrapper


@login_required
@project_required
def dashboard_view(request):
    """View do dashboard com KPIs e calendário."""
    project = get_selected_project(request)
    
    # Debug: Log do projeto selecionado
    import logging
    logger = logging.getLogger(__name__)
    logger.debug(f'Dashboard - Projeto selecionado: {project.name} (ID: {project.id})')
    
    # KPIs filtrados pela obra selecionada
    total_diaries = ConstructionDiary.objects.filter(project=project).count()
    # Relatórios em edição (status PREENCHENDO - raro, mas mantido para compatibilidade)
    pending_reports = ConstructionDiary.objects.filter(
        project=project,
        status=DiaryStatus.PREENCHENDO
    ).count()
    approved_reports = ConstructionDiary.objects.filter(
        project=project,
        status=DiaryStatus.APROVADO
    ).count()
    
    # Clima médio (últimos 7 dias) - filtrado pela obra
    last_week = timezone.now().date() - timedelta(days=7)
    recent_diaries = ConstructionDiary.objects.filter(
        project=project,
        date__gte=last_week
    )
    
    # Calcula clima médio real
    weather_counts = {}
    for diary in recent_diaries:
        if diary.weather_conditions:
            weather_lower = diary.weather_conditions.lower()
            if 'sol' in weather_lower:
                weather_counts['sol'] = weather_counts.get('sol', 0) + 1
            elif 'nuvem' in weather_lower or 'nublado' in weather_lower:
                weather_counts['nublado'] = weather_counts.get('nublado', 0) + 1
            elif 'chuva' in weather_lower:
                weather_counts['chuva'] = weather_counts.get('chuva', 0) + 1
    
    if weather_counts:
        avg_weather = max(weather_counts, key=weather_counts.get).title()
    else:
        avg_weather = "N/A"
    
    # Efetivo total (conta todos os funcionários ativos)
    total_workers = Labor.objects.filter(is_active=True).count()
    
    # Estatísticas adicionais
    total_work_hours = ConstructionDiary.objects.filter(
        project=project,
        work_hours__isnull=False
    ).aggregate(total=Sum('work_hours'))['total'] or 0
    
    # Média de horas trabalhadas
    avg_work_hours = ConstructionDiary.objects.filter(
        project=project,
        work_hours__isnull=False
    ).aggregate(avg=Avg('work_hours'))['avg']
    
    # KPIs adicionais baseados no diariodeobra.app
    from core.models import DiaryImage, DiaryVideo, DiaryAttachment, Activity, DailyWorkLog, DiaryOccurrence
    
    # Total de fotos
    total_photos = DiaryImage.objects.filter(
        diary__project=project
    ).count()
    
    # Total de atividades
    total_activities = Activity.objects.filter(project=project).count()
    
    # Total de ocorrências (usando modelo DiaryOccurrence)
    # Usa select_related para otimizar a query
    total_occurrences = DiaryOccurrence.objects.filter(
        diary__project=project
    ).select_related('diary', 'diary__project').count()
    
    # Debug: Verifica ocorrências
    logger.debug(f'Dashboard - Projeto: {project.name} (ID: {project.id})')
    logger.debug(f'Dashboard - Total de ocorrências: {total_occurrences}')
    all_occurrences_debug = DiaryOccurrence.objects.filter(diary__project=project).select_related('diary', 'diary__project')
    logger.debug(f'Dashboard - Ocorrências encontradas: {[occ.id for occ in all_occurrences_debug]}')
    
    # Verifica se há ocorrências em outros projetos (para debug)
    all_occurrences_all_projects = DiaryOccurrence.objects.all().select_related('diary', 'diary__project')
    if all_occurrences_all_projects.exists():
        logger.debug(f'Dashboard - Total de ocorrências em TODOS os projetos: {all_occurrences_all_projects.count()}')
        for occ in all_occurrences_all_projects[:3]:
            logger.debug(f'Dashboard - Ocorrência ID={occ.id} está no projeto: {occ.diary.project.name} (ID: {occ.diary.project.id})')
    
    # Total de comentários (placeholder - pode ser implementado depois)
    # Por enquanto, conta notas nos diários como "comentários"
    total_comments = ConstructionDiary.objects.filter(
        project=project
    ).exclude(general_notes='').count()
    
    # Total de vídeos
    # Usa select_related para otimizar a query
    total_videos = DiaryVideo.objects.filter(
        diary__project=project
    ).select_related('diary', 'diary__project').count()
    
    # Debug: Verifica vídeos
    logger.debug(f'Dashboard - Total de vídeos: {total_videos}')
    all_videos_debug = DiaryVideo.objects.filter(diary__project=project).select_related('diary', 'diary__project')
    logger.debug(f'Dashboard - Vídeos encontrados: {[v.id for v in all_videos_debug]}')
    
    # Verifica se há vídeos em outros projetos (para debug)
    all_videos_all_projects = DiaryVideo.objects.all().select_related('diary', 'diary__project')
    if all_videos_all_projects.exists():
        logger.debug(f'Dashboard - Total de vídeos em TODOS os projetos: {all_videos_all_projects.count()}')
        for vid in all_videos_all_projects[:3]:
            logger.debug(f'Dashboard - Vídeo ID={vid.id} está no projeto: {vid.diary.project.name} (ID: {vid.diary.project.id})')
    
    # Total de anexos
    total_attachments = DiaryAttachment.objects.filter(
        diary__project=project
    ).count()
    
    # Relatórios recentes (últimos 7 para cards; até 25 para modo tabela)
    recent_reports = ConstructionDiary.objects.filter(
        project=project
    ).select_related('project').prefetch_related(
        'images', 'videos', 'occurrences'
    ).order_by('-date', '-created_at')[:7]
    recent_reports_table = ConstructionDiary.objects.filter(
        project=project
    ).select_related('project').prefetch_related(
        'images', 'videos', 'occurrences'
    ).order_by('-date', '-created_at')[:25]
    
    # Fotos recentes (últimas 9)
    recent_photos = DiaryImage.objects.filter(
        diary__project=project
    ).select_related('diary').order_by('-uploaded_at')[:9]
    
    # Vídeos recentes (últimos 6)
    recent_videos = DiaryVideo.objects.filter(
        diary__project=project
    ).select_related('diary').order_by('-uploaded_at')[:6]
    
    # Debug: Verifica vídeos recentes
    logger.debug(f'Dashboard - Vídeos recentes: {len(recent_videos)}')
    logger.debug(f'Dashboard - IDs dos vídeos recentes: {[v.id for v in recent_videos]}')
    
    # Cálculos para Informações da Obra
    project_days_elapsed = 0
    project_days_total = 0
    project_days_remaining = 0
    project_progress_percent = 0
    
    if project.start_date:
        from datetime import date
        today = date.today()
        project_days_elapsed = (today - project.start_date).days
        
        if project.end_date:
            project_days_total = (project.end_date - project.start_date).days
            project_days_remaining = (project.end_date - today).days
            
            if project_days_total > 0:
                project_progress_percent = min(100, max(0, (project_days_elapsed / project_days_total) * 100))
    
    # Debug: Verifica se recent_videos é um QuerySet ou lista
    logger.debug(f'Dashboard - Tipo de recent_videos: {type(recent_videos)}')
    if hasattr(recent_videos, '__iter__'):
        logger.debug(f'Dashboard - recent_videos é iterável, length: {len(list(recent_videos))}')
    
    from django.conf import settings
    context = {
        'project': project,
        'total_diaries': total_diaries,
        'pending_reports': pending_reports,
        'approved_reports': approved_reports,
        'avg_weather': avg_weather,
        'total_workers': total_workers,
        'total_work_hours': total_work_hours,
        'avg_work_hours': round(avg_work_hours, 1) if avg_work_hours else None,
        'total_photos': total_photos,
        'total_activities': total_activities,
        'total_occurrences': total_occurrences,
        'total_comments': total_comments,
        'total_videos': total_videos,
        'total_attachments': total_attachments,
        'recent_reports': recent_reports,
        'recent_reports_table': recent_reports_table,
        'recent_photos': recent_photos,
        'recent_videos': recent_videos,
        # Informações da Obra
        'project_days_elapsed': project_days_elapsed,
        'project_days_total': project_days_total,
        'project_days_remaining': project_days_remaining,
        'project_progress_percent': round(project_progress_percent, 1),
        # Para sidebar
        'total_reports_count': total_diaries,
        'total_photos_count': total_photos,
        'total_videos_count': total_videos,
        'total_activities_count': total_activities,
        'total_occurrences_count': total_occurrences,
        'total_comments_count': total_comments,
        'total_attachments_count': total_attachments,
        # Debug
        'DEBUG': settings.DEBUG,
    }
    
    return render(request, 'core/dashboard.html', context)


@login_required
@project_required
def calendar_events_view(request):
    """API endpoint para eventos do FullCalendar."""
    from datetime import date as date_type
    
    project = get_selected_project(request)
    start = request.GET.get('start')
    end = request.GET.get('end')
    
    try:
        start_date = datetime.fromisoformat(start.replace('Z', '+00:00'))
        end_date = datetime.fromisoformat(end.replace('Z', '+00:00'))
    except (ValueError, AttributeError, TypeError):
        # start/end ausentes ou inválidos: usa período padrão (datetime para .date() abaixo)
        start_date = timezone.now()
        end_date = start_date + timedelta(days=30)
    
    # Ajusta o período para considerar o período da obra
    view_start = start_date.date()
    view_end = end_date.date()
    
    # Se a obra tem datas definidas, limita ao período da obra
    if project.start_date and project.end_date:
        # Usa a interseção entre o período da obra e o período visualizado
        view_start = max(view_start, project.start_date)
        view_end = min(view_end, project.end_date)
    
    # Busca todos os diários no período
    diaries = ConstructionDiary.objects.filter(
        project=project,
        date__gte=view_start,
        date__lte=view_end
    ).select_related('project')
    
    # Cria um conjunto de datas que já têm relatórios
    dates_with_diaries = set(diary.date for diary in diaries)
    
    events = []
    
    # Adiciona eventos para dias com relatórios
    for diary in diaries:
        # Determina cor baseada no status
        if diary.status == DiaryStatus.APROVADO:
            color = '#10b981'  # Verde - Preenchido/Finalizado
            title_status = 'Preenchido'
        elif diary.status == DiaryStatus.SALVAMENTO_PARCIAL:
            color = '#f59e0b'  # Âmbar - Salvamento parcial (rascunho)
            title_status = 'Salvamento Parcial'
        elif diary.status == DiaryStatus.PREENCHENDO:
            # Preenchendo = relatório existe mas ainda está sendo preenchido
            color = '#10b981'  # Verde - Preenchido
            title_status = 'Preenchido'
        else:
            # Status desconhecido ou inválido - trata como preenchido mas com status indefinido
            color = '#6b7280'  # Cinza
            title_status = 'Indefinido'
        
        # Título completo e curto (para preview compacto no calendário)
        if diary.report_number:
            title = f"RDO #{diary.report_number} - {title_status}"
            short_title = f"RDO #{diary.report_number}"
        else:
            creator_name = diary.created_by.get_full_name() or diary.created_by.username
            title = f"{creator_name[:15]}... - {title_status}"
            short_title = "RDO"
        
        events.append({
            'id': diary.id,
            'title': title,
            'start': diary.date.isoformat(),
            'color': color,
            'display': 'block',
            'extendedProps': {
                'status': title_status,
                'diary_id': diary.id,
                'has_diary': True,
                'short_title': short_title,
            },
        })
    
    # Adiciona eventos para dias sem relatórios (dias faltantes)
    # Só mostra dias faltantes se a obra tem período definido
    if project.start_date and project.end_date:
        today = timezone.now().date()
        current_date = view_start
        
        while current_date <= view_end:
            # Verifica se é dia útil (segunda a sexta) - pode ser configurável depois
            # Por enquanto, mostra todos os dias dentro do período da obra
            is_weekday = current_date.weekday() < 5  # 0-4 = segunda a sexta
            
            # Se não tem relatório neste dia e está dentro do período da obra
            # Só exibe evento para hoje ou dias passados (Falta/Atraso). Dias futuros ficam em branco.
            if current_date not in dates_with_diaries and current_date <= today:
                color = '#dc2626'  # Vermelho escuro para borda
                title_status = 'Atraso'
                title = f"Falta relatório - {title_status}"
                short_title = 'Falta'
                events.append({
                    'id': f'missing_{current_date.isoformat()}',
                    'title': title,
                    'start': current_date.isoformat(),
                    'color': color,
                    'display': 'block',
                    'extendedProps': {
                        'status': title_status,
                        'diary_id': None,
                        'has_diary': False,
                        'missing': True,
                        'short_title': short_title,
                    },
                })
            
            current_date += timedelta(days=1)
    
    return JsonResponse(events, safe=False)


@login_required
@project_required
def report_list_view(request):
    """View de listagem de relatórios com filtros HTMX."""
    project = get_selected_project(request)
    diaries = ConstructionDiary.objects.filter(project=project).select_related('project').all()
    
    # Filtros
    search = request.GET.get('search')
    if search:
        diaries = diaries.filter(
            Q(project__code__icontains=search) |
            Q(project__name__icontains=search) |
            Q(general_notes__icontains=search)
        )
    
    date_start = request.GET.get('date_start')
    if date_start:
        try:
            diaries = diaries.filter(date__gte=date_start)
        except ValueError:
            pass
    
    date_end = request.GET.get('date_end')
    if date_end:
        try:
            diaries = diaries.filter(date__lte=date_end)
        except ValueError:
            pass
    
    status = request.GET.get('status')
    if status:
        diaries = diaries.filter(status=status)
    
    # Ordenação
    diaries = diaries.order_by('-date', '-created_at')
    
    # Último relatório para o modal (qualquer status)
    last_diary = ConstructionDiary.objects.filter(
        project=project
    ).order_by('-date', '-created_at').first()
    
    # Projetos que o usuário pode acessar (para o select do modal "Adicionar relatório")
    all_projects = _get_projects_for_user(request)
    
    context = {
        'diaries': diaries,
        'last_diary': last_diary,
        'user': request.user,  # Adiciona user ao contexto para can_be_edited_by
        'project': project,  # Adiciona projeto ao contexto para o modal
        'all_projects': all_projects,  # Projetos acessíveis para o select do modal
    }
    
    # Se for requisição HTMX, retorna apenas o conteúdo
    if request.headers.get('HX-Request'):
        return render(request, 'core/report_list_partial.html', context)
    
    return render(request, 'core/report_list.html', context)


@login_required
@project_required
def diary_detail_view(request, pk):
    """View de detalhe do diário."""
    from collections import defaultdict
    from core.models import DiaryView, DiaryEditLog
    
    # SEMPRE retorna HTML para esta view (é uma view frontend)
    # Ignora qualquer header Accept que possa fazer o DRF interceptar
    # Se vier de /api/diaries/, redireciona para /diaries/
    if request.path.startswith('/api/diaries/'):
        from django.http import HttpResponseRedirect
        return HttpResponseRedirect(f'/diaries/{pk}/')
    
    # Força Content-Type HTML explícito para evitar que DRF intercepte
    # Remove qualquer header Accept que possa fazer o DRF pensar que é JSON
    # SEMPRE força HTML para requisições do frontend
    request.META['HTTP_ACCEPT'] = 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8'
    if 'Accept' in request.META:
        request.META['Accept'] = 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8'
    
    diary = get_object_or_404(
        ConstructionDiary.objects.select_related('project', 'created_by', 'reviewed_by')
        .prefetch_related(
            'images', 'videos',
            'work_logs__activity', 'work_logs__resources_labor', 'work_logs__resources_equipment',
            'occurrences', 'occurrences__tags',
            'owner_comments__author',
        ),
        pk=pk
    )
    # #region agent log
    _dbg("diary_detail_view:loaded", "diary from DB for detail", {
        "diary_pk": diary.pk,
        "occurrences_count": diary.occurrences.count(),
        "inspections_preview": (getattr(diary, "inspections", None) or "")[:80],
        "dds_preview": (getattr(diary, "dds", None) or "")[:80],
    }, "H5")
    # #endregion
    project = get_selected_project(request)
    # Se há projeto na sessão, só permite ver diário desse projeto
    if project is not None and diary.project_id != project.id:
        raise Http404('Relatório não encontrado.')
    # Se não há projeto na sessão (ex.: link direto), só usa o projeto do diário se o usuário tiver acesso
    if project is None:
        if not _user_can_access_project(request.user, diary.project):
            raise Http404('Relatório não encontrado.')
        project = diary.project
        request.session['selected_project_id'] = project.id
        request.session['selected_project_name'] = project.name
        request.session['selected_project_code'] = getattr(project, 'code', '')
    
    # Registra visualização
    DiaryView.objects.create(
        diary=diary,
        viewed_by=request.user,
        ip_address=request.META.get('REMOTE_ADDR')
    )
    
    # Conta visualizações e edições
    view_count = DiaryView.objects.filter(diary=diary).count()
    edit_count = DiaryEditLog.objects.filter(diary=diary).count()
    
    # Processa condições climáticas - melhorado para usar campos específicos
    weather_morning = None
    weather_afternoon = None
    weather_night = None
    
    # Processa clima da manhã usando campos específicos
    if diary.weather_morning_condition:
        condition = diary.weather_morning_condition
        workable = diary.weather_morning_workable
        
        if condition == 'B':  # Bom
            if 'sol' in (diary.weather_conditions or '').lower():
                weather_morning = {'icon': 'sun', 'label': 'Claro', 'condition': 'Trabalhável' if workable == 'T' else 'Não Trabalhável', 'color': 'yellow'}
            else:
                weather_morning = {'icon': 'cloud-sun', 'label': 'Bom', 'condition': 'Trabalhável' if workable == 'T' else 'Não Trabalhável', 'color': 'yellow'}
        elif condition == 'R':  # Ruim
            weather_morning = {'icon': 'cloud-rain', 'label': 'Ruim', 'condition': 'Não Trabalhável', 'color': 'blue'}
        else:
            # Fallback para processamento antigo
            if diary.weather_conditions:
                parts = diary.weather_conditions.split('|')
                for part in parts:
                    part = part.strip()
                    if 'manhã' in part.lower() or 'manha' in part.lower():
                        if 'sol' in part.lower():
                            weather_morning = {'icon': 'sun', 'label': 'Claro', 'condition': 'Praticável', 'color': 'yellow'}
                        elif 'nuvem' in part.lower() or 'nublado' in part.lower():
                            weather_morning = {'icon': 'cloud', 'label': 'Nublado', 'condition': 'Praticável', 'color': 'gray'}
                        elif 'chuva' in part.lower():
                            weather_morning = {'icon': 'cloud-rain', 'label': 'Chuvoso', 'condition': 'Impraticável', 'color': 'blue'}
    
    # Processa clima da tarde usando campos específicos
    if diary.weather_afternoon_condition:
        condition = diary.weather_afternoon_condition
        workable = diary.weather_afternoon_workable
        
        if condition == 'B':  # Bom
            if 'sol' in (diary.weather_conditions or '').lower():
                weather_afternoon = {'icon': 'sun', 'label': 'Claro', 'condition': 'Trabalhável' if workable == 'T' else 'Não Trabalhável', 'color': 'yellow'}
            else:
                weather_afternoon = {'icon': 'cloud-sun', 'label': 'Bom', 'condition': 'Trabalhável' if workable == 'T' else 'Não Trabalhável', 'color': 'yellow'}
        elif condition == 'R':  # Ruim
            weather_afternoon = {'icon': 'cloud-rain', 'label': 'Ruim', 'condition': 'Não Trabalhável', 'color': 'blue'}
        else:
            # Fallback para processamento antigo
            if diary.weather_conditions:
                parts = diary.weather_conditions.split('|')
                for part in parts:
                    part = part.strip()
                    if 'tarde' in part.lower():
                        if 'sol' in part.lower():
                            weather_afternoon = {'icon': 'sun', 'label': 'Claro', 'condition': 'Praticável', 'color': 'yellow'}
                        elif 'nuvem' in part.lower() or 'nublado' in part.lower():
                            weather_afternoon = {'icon': 'cloud', 'label': 'Nublado', 'condition': 'Praticável', 'color': 'gray'}
                        elif 'chuva' in part.lower():
                            weather_afternoon = {'icon': 'cloud-rain', 'label': 'Chuvoso', 'condition': 'Impraticável', 'color': 'blue'}
    
    # Processa clima da noite se habilitado
    if diary.weather_night_enabled and diary.weather_night_type:
        night_type = diary.weather_night_type
        night_workable = diary.weather_night_workable
        
        if night_type == 'C':  # Claro
            weather_night = {'icon': 'moon', 'label': 'Claro', 'condition': 'Praticável' if night_workable == 'P' else 'Impraticável', 'color': 'yellow'}
        elif night_type == 'N':  # Nublado
            weather_night = {'icon': 'cloud-moon', 'label': 'Nublado', 'condition': 'Praticável' if night_workable == 'P' else 'Impraticável', 'color': 'gray'}
        elif night_type == 'CH':  # Chuvoso
            weather_night = {'icon': 'cloud-moon-rain', 'label': 'Chuvoso', 'condition': 'Impraticável', 'color': 'blue'}
    
    # Agrupa mão de obra por tipo
    labor_by_type = {
        'Direto': {},
        'Indireto': {},
        'Terceiros': {},
    }
    
    # Coleta mão de obra de todos os work_logs
    for work_log in diary.work_logs.all():
        for labor in work_log.resources_labor.all():
            labor_type = labor.labor_type
            if labor_type == 'D':
                if labor.name not in labor_by_type['Direto']:
                    labor_by_type['Direto'][labor.name] = 0
                labor_by_type['Direto'][labor.name] += 1
            elif labor_type == 'I':
                if labor.name not in labor_by_type['Indireto']:
                    labor_by_type['Indireto'][labor.name] = 0
                labor_by_type['Indireto'][labor.name] += 1
            elif labor_type == 'T':
                if labor.name not in labor_by_type['Terceiros']:
                    labor_by_type['Terceiros'][labor.name] = 0
                labor_by_type['Terceiros'][labor.name] += 1
    
    # Calcula métricas do projeto
    project = diary.project
    project_days_total = (project.end_date - project.start_date).days if project.end_date and project.start_date else 0
    project_days_elapsed = (diary.date - project.start_date).days if project.end_date and project.start_date else 0
    project_days_remaining = (project.end_date - diary.date).days if project.end_date and project.start_date else 0
    
    # Nome do dia da semana
    from datetime import datetime
    weekday_names = ['Segunda-Feira', 'Terça-Feira', 'Quarta-Feira', 'Quinta-Feira', 'Sexta-Feira', 'Sábado', 'Domingo']
    weekday_name = weekday_names[diary.date.weekday()] if diary.date else ''
    
    # Equipamentos
    equipment_list = []
    for work_log in diary.work_logs.all():
        for equipment in work_log.resources_equipment.all():
            if equipment.name not in [e['name'] for e in equipment_list]:
                equipment_list.append({
                    'name': equipment.name,
                    'code': equipment.code,
                })

    # Mão de obra por categorias (DiaryLaborEntry) - preferência sobre work_log.resources_labor
    labor_entries_by_category = None
    try:
        from .models import DiaryLaborEntry
        entries = DiaryLaborEntry.objects.filter(diary=diary).select_related('cargo', 'cargo__category').order_by('cargo__category__order', 'company', 'cargo__name')
        if entries.exists():
            labor_entries_by_category = {'indireta': [], 'direta': [], 'terceirizada': {}}
            for e in entries:
                slug = e.cargo.category.slug
                item = {'cargo_name': e.cargo.name, 'quantity': e.quantity}
                if slug == 'terceirizada':
                    company = e.company or '(Sem empresa)'
                    if company not in labor_entries_by_category['terceirizada']:
                        labor_entries_by_category['terceirizada'][company] = []
                    labor_entries_by_category['terceirizada'][company].append(item)
                elif slug in labor_entries_by_category:
                    labor_entries_by_category[slug].append(item)
            labor_entries_by_category['terceirizada'] = [{'company': k, 'items': v} for k, v in labor_entries_by_category['terceirizada'].items()]
    except Exception:
        pass
    
    # Assinaturas
    signatures = diary.signatures.all().select_related('signer')
    signature_inspection = signatures.filter(signature_type='inspection').first()
    signature_production = signatures.filter(signature_type='production').first()
    
    context = {
        'diary': diary,
        'user': request.user,
        'weather_morning': weather_morning,
        'weather_afternoon': weather_afternoon,
        'weather_night': weather_night,
        'labor_by_type': labor_by_type,
        'project_days_total': project_days_total,
        'project_days_elapsed': project_days_elapsed,
        'project_days_remaining': project_days_remaining,
        'total_indirect_labor': sum(labor_by_type['Indireto'].values()),
        'total_direct_labor': sum(labor_by_type['Direto'].values()),
        'total_third_party_labor': sum(labor_by_type['Terceiros'].values()),
        'equipment_list': equipment_list,
        'labor_entries_by_category': labor_entries_by_category,
        'weekday_name': weekday_name,
        'view_count': view_count,
        'edit_count': edit_count,
        'edit_logs': DiaryEditLog.objects.filter(diary=diary).select_related('edited_by')[:10],
        'signature_inspection': signature_inspection,
        'signature_production': signature_production,
        'owner_comments': list(diary.owner_comments.select_related('author').order_by('created_at')),
        'can_add_owner_comment': diary.is_approved() and (request.user.is_staff or request.user.groups.filter(name=GRUPOS.GERENTES).exists()),
    }
    
    # Força retorno HTML explícito (evita conflito com API REST)
    # Remove qualquer header que possa fazer o DRF interceptar
    response = render(request, 'core/diary_detail.html', context)
    response['Content-Type'] = 'text/html; charset=utf-8'
    # Remove headers que podem fazer o DRF pensar que é JSON
    if 'Content-Type' in response:
        response['Content-Type'] = 'text/html; charset=utf-8'
    # Garante que não seja tratado como API
    response['X-Content-Type-Options'] = 'nosniff'
    return response


# ==================== PORTAL DO DONO DA OBRA (CLIENTE) ====================

def _client_can_access_diary(user, diary):
    """True se o usuário é dono da obra do diário (pode acessar a página cliente)."""
    return ProjectOwner.objects.filter(project=diary.project, user=user).exists()


def _client_can_comment(diary):
    """True se ainda está na janela de 24h para enviar comentários (a partir de sent_to_owner_at)."""
    if not diary.sent_to_owner_at:
        return False
    from datetime import timedelta
    deadline = diary.sent_to_owner_at + timedelta(hours=24)
    return timezone.now() <= deadline


@login_required
def client_diary_list_view(request):
    """Lista de diários disponíveis para o dono da obra (só obras que ele possui)."""
    if not _is_work_owner(request.user):
        raise Http404("Acesso restrito.")
    projects = _get_projects_for_user(request)
    # Diários aprovados já enviados ao dono (com sent_to_owner_at)
    diaries_by_project = []
    for project in projects:
        diaries = (
            ConstructionDiary.objects.filter(
                project=project,
                status=DiaryStatus.APROVADO,
                sent_to_owner_at__isnull=False,
            )
            .order_by('-date')[:10]
        )
        if diaries:
            diaries_by_project.append({'project': project, 'diaries': list(diaries)})
    context = {
        'diaries_by_project': diaries_by_project,
        'user': request.user,
    }
    return render(request, 'core/client_diary_list.html', context)


@login_required
def client_diary_detail_view(request, pk):
    """Visualização do diário pelo dono da obra: leitura + comentários (24h)."""
    diary = get_object_or_404(
        ConstructionDiary.objects.select_related('project', 'created_by', 'reviewed_by').prefetch_related(
            'images', 'videos', 'work_logs__activity', 'work_logs__resources_labor', 'work_logs__resources_equipment',
            'owner_comments__author',
        ),
        pk=pk,
    )
    if not _client_can_access_diary(request.user, diary):
        raise Http404("Você não tem acesso a este diário.")
    if diary.status != DiaryStatus.APROVADO:
        raise Http404("Diário não disponível para visualização.")
    comments = list(diary.owner_comments.select_related('author').order_by('created_at'))
    can_comment = _client_can_comment(diary)
    comment_deadline = None
    if diary.sent_to_owner_at:
        from datetime import timedelta
        comment_deadline = diary.sent_to_owner_at + timedelta(hours=24)
    # Contexto mínimo para exibir o diário (reutiliza lógica de clima/labor se necessário)
    labor_by_type = {'Direto': {}, 'Indireto': {}, 'Terceiros': {}}
    for work_log in diary.work_logs.all():
        for labor in work_log.resources_labor.all():
            labor_type = labor.labor_type
            if labor_type == 'D':
                labor_by_type['Direto'][labor.name] = labor_by_type['Direto'].get(labor.name, 0) + 1
            elif labor_type == 'I':
                labor_by_type['Indireto'][labor.name] = labor_by_type['Indireto'].get(labor.name, 0) + 1
            elif labor_type == 'T':
                labor_by_type['Terceiros'][labor.name] = labor_by_type['Terceiros'].get(labor.name, 0) + 1
    equipment_list = []
    for work_log in diary.work_logs.all():
        for equipment in work_log.resources_equipment.all():
            if equipment.name not in [e['name'] for e in equipment_list]:
                equipment_list.append({'name': equipment.name, 'code': equipment.code})
    project = diary.project
    project_days_total = (project.end_date - project.start_date).days if project.end_date and project.start_date else 0
    project_days_elapsed = (diary.date - project.start_date).days if project.end_date and project.start_date else 0
    project_days_remaining = (project.end_date - diary.date).days if project.end_date and project.start_date else 0
    from datetime import datetime as dt
    weekday_names = ['Segunda-Feira', 'Terça-Feira', 'Quarta-Feira', 'Quinta-Feira', 'Sexta-Feira', 'Sábado', 'Domingo']
    weekday_name = weekday_names[diary.date.weekday()] if diary.date else ''
    context = {
        'diary': diary,
        'project': project,
        'comments': comments,
        'can_comment': can_comment,
        'comment_deadline': comment_deadline,
        'labor_by_type': labor_by_type,
        'equipment_list': equipment_list,
        'project_days_total': project_days_total,
        'project_days_elapsed': project_days_elapsed,
        'project_days_remaining': project_days_remaining,
        'weekday_name': weekday_name,
        'total_indirect_labor': sum(labor_by_type['Indireto'].values()),
        'total_direct_labor': sum(labor_by_type['Direto'].values()),
        'total_third_party_labor': sum(labor_by_type['Terceiros'].values()),
    }
    return render(request, 'core/client_diary_detail.html', context)


@login_required
@project_required
@require_http_methods(["POST"])
def diary_add_owner_comment_view(request, pk):
    """POST: adiciona comentário da LPLAN no diário (staff/gerentes). Diário deve estar aprovado."""
    diary = get_object_or_404(ConstructionDiary.objects.select_related('project'), pk=pk)
    if diary.project_id != get_selected_project(request).id:
        raise Http404()
    if not diary.is_approved():
        messages.error(request, "Só é possível comentar em diários aprovados.")
        return redirect('diary-detail', pk=pk)
    if not (request.user.is_staff or request.user.groups.filter(name=GRUPOS.GERENTES).exists()):
        raise PermissionDenied("Sem permissão para comentar.")
    text = (request.POST.get('text') or '').strip()
    if not text:
        messages.error(request, "Escreva um comentário.")
        return redirect('diary-detail', pk=pk)
    DiaryComment.objects.create(diary=diary, author=request.user, text=text)
    messages.success(request, "Comentário enviado. O dono da obra poderá ver na página de visualização.")
    return redirect('diary-detail', pk=pk)


@login_required
@require_http_methods(["POST"])
def client_diary_add_comment_view(request, pk):
    """POST: adiciona comentário ao diário (apenas dono da obra, dentro da janela de 24h)."""
    diary = get_object_or_404(ConstructionDiary.objects.select_related('project'), pk=pk)
    if not _client_can_access_diary(request.user, diary):
        raise PermissionDenied("Você não tem acesso a este diário.")
    if not _client_can_comment(diary):
        messages.error(request, "O prazo de 24 horas para enviar comentários foi encerrado.")
        return redirect('client-diary-detail', pk=pk)
    text = (request.POST.get('text') or '').strip()
    if not text:
        messages.error(request, "Escreva um comentário.")
        return redirect('client-diary-detail', pk=pk)
    DiaryComment.objects.create(diary=diary, author=request.user, text=text)
    messages.success(request, "Comentário enviado.")
    return redirect('client-diary-detail', pk=pk)


# ==================== FILTROS DE BUSCA ====================

@login_required
@project_required
def filter_photos_view(request):
    """View para filtro de fotos."""
    project = get_selected_project(request)
    
    # Busca todas as fotos do projeto
    photos = DiaryImage.objects.filter(
        diary__project=project,
        is_approved_for_report=True
    ).select_related('diary').order_by('-diary__date', '-uploaded_at')
    
    # Filtros
    search = request.GET.get('search', '')
    date_start = request.GET.get('date_start')
    date_end = request.GET.get('date_end')
    diary_id = request.GET.get('diary_id')
    
    if search:
        photos = photos.filter(
            Q(caption__icontains=search) |
            Q(diary__project__name__icontains=search) |
            Q(diary__project__code__icontains=search)
        )
    
    if date_start:
        try:
            photos = photos.filter(diary__date__gte=date_start)
        except ValueError:
            pass
    
    if date_end:
        try:
            photos = photos.filter(diary__date__lte=date_end)
        except ValueError:
            pass
    
    if diary_id:
        photos = photos.filter(diary_id=diary_id)
    
    # Paginação
    from django.core.paginator import Paginator
    paginator = Paginator(photos, 24)  # 24 fotos por página (grid 4x6)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    # Estatísticas
    total_photos = photos.count()
    photos_by_date = photos.values('diary__date').annotate(count=Count('id')).order_by('-diary__date')[:10]
    
    context = {
        'photos': page_obj,
        'total_photos': total_photos,
        'photos_by_date': photos_by_date,
        'search': search,
        'date_start': date_start,
        'date_end': date_end,
        'diary_id': diary_id,
    }
    
    return render(request, 'core/filters/photos.html', context)


@login_required
@project_required
def filter_videos_view(request):
    """View para filtro de vídeos."""
    project = get_selected_project(request)
    from core.models import DiaryVideo
    
    # Busca todos os vídeos do projeto
    videos = DiaryVideo.objects.filter(
        diary__project=project
    ).select_related('diary').order_by('-diary__date', '-uploaded_at')
    
    # Filtros
    search = request.GET.get('search', '')
    date_start = request.GET.get('date_start')
    date_end = request.GET.get('date_end')
    diary_id = request.GET.get('diary_id')
    
    if search:
        videos = videos.filter(
            Q(caption__icontains=search) |
            Q(diary__project__name__icontains=search) |
            Q(diary__project__code__icontains=search)
        )
    
    if date_start:
        try:
            videos = videos.filter(diary__date__gte=date_start)
        except ValueError:
            pass
    
    if date_end:
        try:
            videos = videos.filter(diary__date__lte=date_end)
        except ValueError:
            pass
    
    if diary_id:
        videos = videos.filter(diary_id=diary_id)
    
    # Paginação
    from django.core.paginator import Paginator
    paginator = Paginator(videos, 24)  # 24 vídeos por página
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    # Estatísticas
    total_videos = videos.count()
    videos_by_date = videos.values('diary__date').annotate(count=Count('id')).order_by('-diary__date')[:10]
    
    context = {
        'videos': page_obj,
        'total_videos': total_videos,
        'videos_by_date': videos_by_date,
        'search': search,
        'date_start': date_start,
        'date_end': date_end,
        'diary_id': diary_id,
    }
    
    return render(request, 'core/filters/videos.html', context)


@login_required
@project_required
def filter_activities_view(request):
    """View para filtro de atividades."""
    project = get_selected_project(request)
    
    # Busca todas as atividades do projeto
    activities = Activity.objects.filter(project=project).order_by('path')
    
    # Filtros
    search = request.GET.get('search', '')
    status_filter = request.GET.get('status')
    progress_min = request.GET.get('progress_min')
    progress_max = request.GET.get('progress_max')
    
    if search:
        activities = activities.filter(
            Q(name__icontains=search) |
            Q(code__icontains=search) |
            Q(description__icontains=search)
        )
    
    if status_filter:
        activities = activities.filter(status=status_filter)
    
    # Filtro por progresso (se implementado)
    # TODO: Implementar filtro por progresso quando o campo estiver disponível
    
    # Estatísticas
    total_activities = activities.count()
    activities_by_status = activities.values('status').annotate(count=Count('id'))
    
    context = {
        'activities': activities,
        'total_activities': total_activities,
        'activities_by_status': activities_by_status,
        'search': search,
        'status_filter': status_filter,
        'progress_min': progress_min,
        'progress_max': progress_max,
    }
    
    return render(request, 'core/filters/activities.html', context)


@login_required
@project_required
def filter_occurrences_view(request):
    """View para filtro de ocorrências (modelo DiaryOccurrence + campo incidents)."""
    from .models import DiaryOccurrence
    from django.core.paginator import Paginator

    project = get_selected_project(request)

    # Ocorrências do formulário (DiaryOccurrence) – mesma fonte que o dashboard
    occurrences = DiaryOccurrence.objects.filter(
        diary__project=project
    ).select_related('diary', 'diary__project').prefetch_related('tags').order_by('-diary__date', '-created_at')

    # Filtros
    search = request.GET.get('search', '')
    date_start = request.GET.get('date_start')
    date_end = request.GET.get('date_end')

    if search:
        occurrences = occurrences.filter(description__icontains=search)
    if date_start:
        try:
            occurrences = occurrences.filter(diary__date__gte=date_start)
        except ValueError:
            pass
    if date_end:
        try:
            occurrences = occurrences.filter(diary__date__lte=date_end)
        except ValueError:
            pass

    # Estatísticas (total e por data, com os mesmos filtros)
    total_occurrences = occurrences.count()
    occurrences_by_date = list(
        occurrences.values('diary__date')
        .annotate(count=Count('id'))
        .order_by('-diary__date')[:10]
    )

    # Paginação
    paginator = Paginator(occurrences, 20)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    context = {
        'occurrences': page_obj,
        'total_occurrences': total_occurrences,
        'occurrences_by_date': occurrences_by_date,
        'search': search,
        'date_start': date_start,
        'date_end': date_end,
    }

    return render(request, 'core/filters/occurrences.html', context)


@login_required
@project_required
def filter_comments_view(request):
    """View para filtro de comentários."""
    project = get_selected_project(request)
    
    # Busca comentários (general_notes) de todos os diários do projeto
    diaries = ConstructionDiary.objects.filter(
        project=project
    ).exclude(general_notes='').exclude(general_notes__isnull=True)
    
    # Filtros
    search = request.GET.get('search', '')
    date_start = request.GET.get('date_start')
    date_end = request.GET.get('date_end')
    
    if search:
        diaries = diaries.filter(general_notes__icontains=search)
    
    if date_start:
        try:
            diaries = diaries.filter(date__gte=date_start)
        except ValueError:
            pass
    
    if date_end:
        try:
            diaries = diaries.filter(date__lte=date_end)
        except ValueError:
            pass
    
    # Paginação
    from django.core.paginator import Paginator
    paginator = Paginator(diaries, 20)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    # Estatísticas
    total_comments = diaries.count()
    
    context = {
        'diaries': page_obj,
        'total_comments': total_comments,
        'search': search,
        'date_start': date_start,
        'date_end': date_end,
    }
    
    return render(request, 'core/filters/comments.html', context)


@login_required
@project_required
def filter_attachments_view(request):
    """View para filtro de anexos."""
    from .models import DiaryAttachment
    
    project = get_selected_project(request)
    
    # Busca anexos
    attachments = DiaryAttachment.objects.filter(
        diary__project=project
    ).select_related('diary').order_by('-diary__date', '-uploaded_at')
    
    # Filtros
    search = request.GET.get('search', '')
    date_start = request.GET.get('date_start')
    date_end = request.GET.get('date_end')
    
    if search:
        attachments = attachments.filter(
            Q(name__icontains=search) |
            Q(description__icontains=search) |
            Q(diary__project__name__icontains=search)
        )
    
    if date_start:
        try:
            attachments = attachments.filter(diary__date__gte=date_start)
        except ValueError:
            pass
    
    if date_end:
        try:
            attachments = attachments.filter(diary__date__lte=date_end)
        except ValueError:
            pass
    
    # Paginação
    from django.core.paginator import Paginator
    paginator = Paginator(attachments, 20)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    context = {
        'attachments': page_obj,
        'total_attachments': attachments.count(),
        'search': search,
        'date_start': date_start,
        'date_end': date_end,
    }
    
    return render(request, 'core/filters/attachments.html', context)


@login_required
@project_required
def weather_conditions_view(request):
    """View para análise de condições climáticas."""
    project = get_selected_project(request)
    
    # Busca todos os diários com condições climáticas
    diaries = ConstructionDiary.objects.filter(
        project=project
    ).exclude(weather_conditions='').exclude(weather_conditions__isnull=True).order_by('-date')
    
    # Filtros
    date_start = request.GET.get('date_start')
    date_end = request.GET.get('date_end')
    
    if date_start:
        try:
            diaries = diaries.filter(date__gte=date_start)
        except ValueError:
            pass
    
    if date_end:
        try:
            diaries = diaries.filter(date__lte=date_end)
        except ValueError:
            pass
    
    # Análise de condições climáticas
    weather_stats = {
        'sunny_days': 0,
        'cloudy_days': 0,
        'rainy_days': 0,
        'total_days': diaries.count(),
    }
    
    for diary in diaries:
        weather = diary.weather_conditions.lower()
        if 'sol' in weather or 'claro' in weather:
            weather_stats['sunny_days'] += 1
        elif 'nuvem' in weather or 'nublado' in weather:
            weather_stats['cloudy_days'] += 1
        elif 'chuva' in weather or 'chuvoso' in weather:
            weather_stats['rainy_days'] += 1
    
    # Estatísticas de chuva
    rain_diaries = diaries.exclude(rain_occurrence='').exclude(rain_occurrence__isnull=True)
    rain_stats = {
        'weak': rain_diaries.filter(rain_occurrence='F').count(),
        'medium': rain_diaries.filter(rain_occurrence='M').count(),
        'strong': rain_diaries.filter(rain_occurrence='S').count(),
        'total': rain_diaries.count(),
    }
    
    context = {
        'diaries': diaries[:50],  # Limita a 50 para performance
        'weather_stats': weather_stats,
        'rain_stats': rain_stats,
        'date_start': date_start,
        'date_end': date_end,
    }
    
    return render(request, 'core/filters/weather_conditions.html', context)


@login_required
@project_required
def labor_histogram_view(request):
    """View para histograma de mão de obra."""
    project = get_selected_project(request)
    
    # Busca todos os work_logs do projeto
    work_logs = DailyWorkLog.objects.filter(
        diary__project=project
    ).select_related('diary', 'activity').prefetch_related('resources_labor')
    
    # Filtros
    date_start = request.GET.get('date_start')
    date_end = request.GET.get('date_end')
    activity_id = request.GET.get('activity_id')
    
    if date_start:
        try:
            work_logs = work_logs.filter(diary__date__gte=date_start)
        except ValueError:
            pass
    
    if date_end:
        try:
            work_logs = work_logs.filter(diary__date__lte=date_end)
        except ValueError:
            pass
    
    if activity_id:
        work_logs = work_logs.filter(activity_id=activity_id)
    
    # Agrupa mão de obra por tipo e nome
    labor_stats = {
        'Direto': {},
        'Indireto': {},
        'Terceiros': {},
    }
    
    for work_log in work_logs:
        for labor in work_log.resources_labor.all():
            labor_type = 'Direto' if labor.labor_type == 'D' else ('Indireto' if labor.labor_type == 'I' else 'Terceiros')
            if labor.name not in labor_stats[labor_type]:
                labor_stats[labor_type][labor.name] = 0
            labor_stats[labor_type][labor.name] += 1
    
    # Estatísticas por data
    labor_by_date = {}
    for work_log in work_logs:
        date_key = work_log.diary.date.isoformat()
        if date_key not in labor_by_date:
            labor_by_date[date_key] = {'Direto': 0, 'Indireto': 0, 'Terceiros': 0}
        
        for labor in work_log.resources_labor.all():
            labor_type = 'Direto' if labor.labor_type == 'D' else ('Indireto' if labor.labor_type == 'I' else 'Terceiros')
            labor_by_date[date_key][labor_type] += 1
    
    context = {
        'labor_stats': labor_stats,
        'labor_by_date': labor_by_date,
        'date_start': date_start,
        'date_end': date_end,
        'activity_id': activity_id,
        'activities': Activity.objects.filter(project=project).order_by('name'),
    }
    
    return render(request, 'core/filters/labor_histogram.html', context)


@login_required
@project_required
def equipment_histogram_view(request):
    """View para histograma de equipamentos."""
    project = get_selected_project(request)
    
    # Busca todos os work_logs do projeto
    work_logs = DailyWorkLog.objects.filter(
        diary__project=project
    ).select_related('diary', 'activity').prefetch_related('resources_equipment')
    
    # Filtros
    date_start = request.GET.get('date_start')
    date_end = request.GET.get('date_end')
    activity_id = request.GET.get('activity_id')
    
    if date_start:
        try:
            work_logs = work_logs.filter(diary__date__gte=date_start)
        except ValueError:
            pass
    
    if date_end:
        try:
            work_logs = work_logs.filter(diary__date__lte=date_end)
        except ValueError:
            pass
    
    if activity_id:
        work_logs = work_logs.filter(activity_id=activity_id)
    
    # Agrupa equipamentos por nome
    equipment_stats = {}
    
    for work_log in work_logs:
        for equipment in work_log.resources_equipment.all():
            if equipment.name not in equipment_stats:
                equipment_stats[equipment.name] = 0
            equipment_stats[equipment.name] += 1
    
    # Estatísticas por data
    equipment_by_date = {}
    for work_log in work_logs:
        date_key = work_log.diary.date.isoformat()
        if date_key not in equipment_by_date:
            equipment_by_date[date_key] = {}
        
        for equipment in work_log.resources_equipment.all():
            if equipment.name not in equipment_by_date[date_key]:
                equipment_by_date[date_key][equipment.name] = 0
            equipment_by_date[date_key][equipment.name] += 1
    
    context = {
        'equipment_stats': equipment_stats,
        'equipment_by_date': equipment_by_date,
        'date_start': date_start,
        'date_end': date_end,
        'activity_id': activity_id,
        'activities': Activity.objects.filter(project=project).order_by('name'),
    }
    
    return render(request, 'core/filters/equipment_histogram.html', context)


@login_required
@project_required
def diary_form_view(request, pk=None):
    """View de formulário de diário de obra."""
    import logging
    logger = logging.getLogger(__name__)
    from .forms import (
        ConstructionDiaryForm,
        DiaryImageFormSet,
        DailyWorkLogFormSet,
        DiaryOccurrenceFormSet,
    )
    from .services import ProgressService
    
    project = get_selected_project(request)
    
    if pk:
        # Otimiza queries com select_related e prefetch_related
        diary = get_object_or_404(
            ConstructionDiary.objects.select_related('project', 'created_by', 'reviewed_by')
            .prefetch_related('images', 'videos', 'attachments', 'work_logs__activity', 'occurrences__tags'),
            pk=pk, 
            project=project
        )
        if not diary.can_be_edited_by(request.user):
            messages.warning(request, 'Este diário não pode ser editado no momento.')
            return redirect('diary-detail', pk=pk)
    else:
        diary = None
    
    copy_source_diary = None
    copy_opts_list = []
    copy_from_id = None
    copy_options_raw = ''
    
    if request.method == 'POST':
        # Verifica permissão de edição antes de processar (se for edição)
        if diary and not diary.can_be_edited_by(request.user):
            messages.error(request, 'Você não tem permissão para editar este diário.')
            return redirect('diary-detail', pk=pk)
        
        # Valida que projeto existe
        if not project:
            messages.error(request, 'Nenhum projeto selecionado. Selecione um projeto primeiro.')
            return redirect('select-project')
        
        form = ConstructionDiaryForm(request.POST, instance=diary, user=request.user, project=project)
        # Debug: Verifica arquivos recebidos
        import logging
        logger = logging.getLogger(__name__)
        logger.info(f"=== PROCESSANDO FORMULÁRIO DE DIÁRIO ===")
        logger.info(f"POST keys: {list(request.POST.keys())}")
        logger.info(f"FILES keys: {list(request.FILES.keys())}")
        # Log específico para debug de atividades e ocorrências
        post_work_logs = [k for k in request.POST.keys() if 'work_log' in k.lower()]
        post_ocorrencias = [k for k in request.POST.keys() if 'ocorrencia' in k.lower() or (k.startswith('occurrence') and 'TOTAL' in k)]
        logger.warning(f"[DIARY_DEBUG] POST work_logs keys ({len(post_work_logs)}): {post_work_logs[:30]}{'...' if len(post_work_logs) > 30 else ''}")
        logger.warning(f"[DIARY_DEBUG] POST ocorrencias keys ({len(post_ocorrencias)}): {post_ocorrencias[:30]}{'...' if len(post_ocorrencias) > 30 else ''}")
        logger.warning(f"[DIARY_DEBUG] work_logs-TOTAL_FORMS={request.POST.get('work_logs-TOTAL_FORMS', 'N/A')} | ocorrencias-TOTAL_FORMS={request.POST.get('ocorrencias-TOTAL_FORMS', 'N/A')}")
        if request.FILES:
            for key, file in request.FILES.items():
                logger.info(f"  Arquivo: {key} -> {file.name} ({file.size} bytes)")
        
        # IMPORTANTE: Preserva request.FILES antes de qualquer processamento
        # request.FILES pode ser consumido apenas uma vez, então precisamos preservá-lo
        files_dict = {}
        if request.FILES:
            # Cria uma cópia dos arquivos para poder reutilizar
            for key, file_obj in request.FILES.items():
                # Cria uma cópia do arquivo em memória
                from django.core.files.uploadedfile import InMemoryUploadedFile
                from io import BytesIO
                
                # Lê o conteúdo do arquivo
                file_obj.seek(0)  # Garante que está no início
                file_content = file_obj.read()
                file_obj.seek(0)  # Volta ao início para uso posterior
                
                # Cria uma nova instância do arquivo
                files_dict[key] = InMemoryUploadedFile(
                    BytesIO(file_content),
                    None,
                    file_obj.name,
                    file_obj.content_type,
                    file_obj.size,
                    file_obj.charset
                )
        
        # Cria formsets iniciais (antes de salvar o diário)
        image_formset = DiaryImageFormSet(
            request.POST,
            request.FILES,
            instance=diary if diary else None
        )
        
        logger.info(f"Formset criado. Total de forms: {image_formset.total_form_count()}")
        logger.info(f"INITIAL_FORMS: {image_formset.initial_form_count()}")
        logger.info(f"TOTAL_FORMS do POST: {request.POST.get('diaryimage_set-TOTAL_FORMS', 'N/A')}")
        logger.info(f"Arquivos em request.FILES: {list(request.FILES.keys())}")
        
        worklog_formset = DailyWorkLogFormSet(
            request.POST,
            instance=diary if diary else None,
            form_kwargs={'diary': diary if diary else None},
            prefix='work_logs'
        )
        # POST para ocorrências: aceita tanto 'ocorrencias-*' quanto 'occurrences-*' (normaliza para ocorrencias)
        from copy import deepcopy
        _post_occ = deepcopy(request.POST)
        if hasattr(_post_occ, '_mutable'):
            _post_occ._mutable = True
        if 'occurrences-TOTAL_FORMS' in request.POST and 'ocorrencias-TOTAL_FORMS' not in _post_occ:
            for key, value in request.POST.items():
                if key.startswith('occurrences-'):
                    _post_occ[key.replace('occurrences-', 'ocorrencias-', 1)] = value
        occurrence_formset = DiaryOccurrenceFormSet(
            _post_occ,
            instance=diary if diary else None,
            prefix='ocorrencias'
        )
        # #region agent log
        post_keys_occ = [k for k in request.POST.keys() if 'occurrence' in k.lower() or 'ocorrencia' in k.lower()]
        post_keys_ae = [k for k in request.POST.keys() if k in ('inspections', 'dds')]
        _dbg("frontend_views:after_occurrence_formset", "POST keys occurrence/ocorrencia and inspections/dds", {
            "post_occurrence_keys": post_keys_occ,
            "post_inspections_dds": post_keys_ae,
            "ocorrencias_total": request.POST.get("ocorrencias-TOTAL_FORMS"),
            "occurrences_total": request.POST.get("occurrences-TOTAL_FORMS"),
            "occurrence_formset_len": len(occurrence_formset.forms),
            "occurrence_formset_valid": occurrence_formset.is_valid(),
        }, "H1")
        # #endregion
        logger.warning(f"[DIARY_DEBUG] Formsets iniciais (antes de form.is_valid): worklog forms={worklog_formset.total_form_count()}, occurrence forms={occurrence_formset.total_form_count()}; worklog_valid={worklog_formset.is_valid()}, occurrence_valid={occurrence_formset.is_valid()}")
        if not worklog_formset.is_valid():
            for i, f in enumerate(worklog_formset.forms):
                if f.errors:
                    logger.warning(f"[DIARY_DEBUG] Worklog form {i} erros: {f.errors}")
            if worklog_formset.non_form_errors():
                logger.warning(f"[DIARY_DEBUG] Worklog non_form_errors: {worklog_formset.non_form_errors()}")
        if not occurrence_formset.is_valid():
            for i, f in enumerate(occurrence_formset.forms):
                if f.errors:
                    logger.warning(f"[DIARY_DEBUG] Ocorrência form {i} erros: {f.errors}")
            if occurrence_formset.non_form_errors():
                logger.warning(f"[DIARY_DEBUG] Ocorrência non_form_errors: {occurrence_formset.non_form_errors()}")

        # Valida o form primeiro
        import logging
        logger = logging.getLogger(__name__)
        
        if form.is_valid():
            logger.info(f"Form principal válido. Salvando diário...")
            # Se o form for válido, salva o diário primeiro (mesmo que seja None)
            # Isso é necessário para que o formset tenha uma instância válida
            diary = form.save(commit=False)
            # #region agent log
            _dbg("frontend_views:after_form_save_commit_false", "diary inspections/dds from form", {
                "diary_pk": getattr(diary, "pk", None),
                "inspections_preview": (getattr(diary, "inspections", None) or "")[:100],
                "dds_preview": (getattr(diary, "dds", None) or "")[:100],
            }, "H3")
            # #endregion

            # Valida que diary foi criado corretamente (não deve ser None)
            if diary is None:
                logger.error("Form.save(commit=False) retornou None! Isso não deveria acontecer.")
                messages.error(request, 'Erro ao processar formulário. Verifique os dados e tente novamente.')
                # Retorna para o formulário com erros
                form = ConstructionDiaryForm(request.POST, instance=diary, user=request.user, project=project)
                if diary and diary.pk:
                    image_formset = DiaryImageFormSet(instance=diary)
                    worklog_formset = DailyWorkLogFormSet(instance=diary, form_kwargs={'diary': diary}, prefix='work_logs')
                    occurrence_formset = DiaryOccurrenceFormSet(instance=diary, prefix='ocorrencias')
                else:
                    image_formset = DiaryImageFormSet()
                    worklog_formset = DailyWorkLogFormSet(form_kwargs={'diary': None}, prefix='work_logs')
                    occurrence_formset = DiaryOccurrenceFormSet(prefix='ocorrencias')
                from .models import OccurrenceTag
                try:
                    occurrence_tags = OccurrenceTag.objects.filter(is_active=True)
                except Exception:
                    occurrence_tags = []
                context = {
                    'diary': diary if diary and diary.pk else None,
                    'form': form,
                    'image_formset': image_formset,
                    'worklog_formset': worklog_formset,
                    'occurrence_formset': occurrence_formset,
                    'occurrence_tags': occurrence_tags,
                    'project': project,
                    'next_report_number': None,
                    'initial_contractante': request.POST.get('project_client_name', get_contractante_for_project(project)),
                }
                return render(request, 'core/daily_log_form.html', context)
            
            is_new = not diary.pk if diary else True
            logger.info(f"Diário {'novo' if is_new else 'existente'} (pk={diary.pk if diary else None})")
            
            # IMPORTANTE: Prepara o diário mas NÃO salva ainda
            # O diário será salvo dentro da transação atomic() após validar os formsets
            if diary:
                is_partial_save = (
                    request.POST.get('partial_save') == '1' or
                    request.POST.get('as_partial_checkbox') == '1'
                )
                if not diary.pk:
                    diary.created_by = request.user
                    diary.project = project  # Associa à obra selecionada
                elif getattr(diary, 'created_by_id', None) is None:
                    # Diário já existe mas sem criador (legado): define para permitir edição
                    diary.created_by = request.user
                if is_partial_save:
                    diary.status = DiaryStatus.SALVAMENTO_PARCIAL
                    logger.info("Salvamento parcial: status definido como SALVAMENTO_PARCIAL")
                else:
                    # Salvar diário = aprovado (sem fluxo de revisar/aprovar). Envio ao dono da obra após salvar.
                    diary.status = DiaryStatus.APROVADO
                    diary.approved_at = timezone.now()
                    diary.sent_to_owner_at = timezone.now()
                    diary.reviewed_by = request.user
                    logger.info("Salvar diário: status definido como APROVADO (envio ao dono após commit)")
            
            # Recria os formsets ANTES de salvar o diário
            # Isso permite validar os formsets antes de criar o diário no banco
            # Usa a cópia dos arquivos preservada
            from django.utils.datastructures import MultiValueDict
            from copy import deepcopy
            
            # Normaliza os dados do POST para usar o prefixo correto do formset
            # O formset espera 'work_logs' mas os testes podem enviar 'dailyworklog_set'
            normalized_post = deepcopy(request.POST)
            if hasattr(normalized_post, '_mutable'):
                normalized_post._mutable = True
            
            # Converte dailyworklog_set para work_logs se necessário
            worklog_prefix_old = 'dailyworklog_set'
            worklog_prefix_new = 'work_logs'
            
            # Verifica se há dados com o prefixo antigo (dailyworklog_set); só normaliza se
            # o front ainda enviar esse prefixo (ex.: cache). O formset usa prefix='work_logs',
            # então o esperado é work_logs-* no POST; não sobrescrever work_logs com valores antigos.
            if f'{worklog_prefix_old}-TOTAL_FORMS' in request.POST and f'{worklog_prefix_new}-TOTAL_FORMS' not in request.POST:
                total_forms = request.POST.get(f'{worklog_prefix_old}-TOTAL_FORMS', '0')
                initial_forms = request.POST.get(f'{worklog_prefix_old}-INITIAL_FORMS', '0')
                normalized_post[f'{worklog_prefix_new}-TOTAL_FORMS'] = total_forms
                normalized_post[f'{worklog_prefix_new}-INITIAL_FORMS'] = initial_forms
                normalized_post[f'{worklog_prefix_new}-MIN_NUM_FORMS'] = request.POST.get(f'{worklog_prefix_old}-MIN_NUM_FORMS', '0')
                normalized_post[f'{worklog_prefix_new}-MAX_NUM_FORMS'] = request.POST.get(f'{worklog_prefix_old}-MAX_NUM_FORMS', '1000')
                for key, value in request.POST.items():
                    if key.startswith(f'{worklog_prefix_old}-'):
                        new_key = key.replace(f'{worklog_prefix_old}-', f'{worklog_prefix_new}-', 1)
                        normalized_post[new_key] = value
            
            # Normaliza occurrences -> ocorrencias (formset usa prefix 'ocorrencias')
            if 'occurrences-TOTAL_FORMS' in request.POST and 'ocorrencias-TOTAL_FORMS' not in normalized_post:
                for key, value in request.POST.items():
                    if key.startswith('occurrences-'):
                        new_key = key.replace('occurrences-', 'ocorrencias-', 1)
                        normalized_post[new_key] = value
            
            # Garante TOTAL_FORMS mínimo se o POST tiver dados de atividades/ocorrências (evita perder dados)
            wl_total = int(normalized_post.get('work_logs-TOTAL_FORMS', '0'))
            occ_total = int(normalized_post.get('ocorrencias-TOTAL_FORMS', '0'))
            for key in request.POST:
                if key.startswith('work_logs-'):
                    parts = key.split('-')
                    if len(parts) >= 2 and parts[1].isdigit():
                        idx = int(parts[1])
                        if idx + 1 > wl_total:
                            wl_total = idx + 1
                if key.startswith('ocorrencias-'):
                    parts = key.split('-')
                    if len(parts) >= 2 and parts[1].isdigit():
                        idx = int(parts[1])
                        if idx + 1 > occ_total:
                            occ_total = idx + 1
            if wl_total > int(normalized_post.get('work_logs-TOTAL_FORMS', '0')):
                normalized_post['work_logs-TOTAL_FORMS'] = str(wl_total)
            if occ_total > int(normalized_post.get('ocorrencias-TOTAL_FORMS', '0')):
                normalized_post['ocorrencias-TOTAL_FORMS'] = str(occ_total)
            
            # Cria um novo MultiValueDict com os arquivos preservados
            preserved_files = MultiValueDict()
            for key, file_obj in files_dict.items():
                preserved_files.appendlist(key, file_obj)
            
            image_formset = DiaryImageFormSet(
                normalized_post,
                preserved_files,
                instance=diary
            )
            worklog_formset = DailyWorkLogFormSet(
                normalized_post,
                instance=diary,
                form_kwargs={'diary': diary},
                prefix='work_logs'
            )
            occurrence_formset = DiaryOccurrenceFormSet(
                normalized_post,
                instance=diary,
                prefix='ocorrencias'
            )
            
            # IMPORTANTE: Formsets inline precisam de instância salva (com PK)
            # Vamos salvar o diário dentro da transação primeiro, depois validar os formsets
            # Se os formsets críticos falharem, lançamos exceção para fazer rollback
            from django.db import transaction
            from core.models import DiarySignature
            
            try:
                with transaction.atomic():
                    # Inicializa variáveis de validação (serão atualizadas dentro do bloco condicional)
                    image_valid = False
                    worklog_valid = False
                    occurrence_valid = False
                    
                    # 0. SALVA O DIÁRIO PRIMEIRO (dentro da transação para permitir rollback se necessário)
                    # IMPORTANTE: Persiste sempre (novo ou edição) para gravar dados do form e status (ex.: Salvamento Parcial)
                    if diary:
                        try:
                            # #region agent log
                            _dbg("frontend_views:before_diary_save", "diary right before save", {
                                "diary_pk": diary.pk,
                                "inspections_preview": (getattr(diary, "inspections", None) or "")[:80],
                                "dds_preview": (getattr(diary, "dds", None) or "")[:80],
                            }, "H4")
                            # #endregion
                            diary.save()
                            logger.info(f"Diário salvo (id={diary.pk}) para validação de formsets e processamento de dados.")
                        except Exception as e:
                            logger.error(f"ERRO ao salvar diário: {e}", exc_info=True)
                            messages.error(request, f'Erro ao salvar diário: {str(e)}')
                            raise  # Re-raise para fazer rollback da transação
                    
                    # Recria os formsets com a instância salva (agora tem PK)
                    if diary and diary.pk:
                        image_formset = DiaryImageFormSet(
                            normalized_post,
                            preserved_files,
                            instance=diary
                        )
                        worklog_formset = DailyWorkLogFormSet(
                            normalized_post,
                            instance=diary,
                            form_kwargs={'diary': diary},
                            prefix='work_logs'
                        )
                        occurrence_formset = DiaryOccurrenceFormSet(
                            normalized_post,
                            instance=diary,
                            prefix='ocorrencias'
                        )

                        # Re-valida os formsets agora que o diário tem PK
                        total_image_forms = int(normalized_post.get('diaryimage_set-TOTAL_FORMS', '0'))
                        total_worklog_forms = int(normalized_post.get('work_logs-TOTAL_FORMS', '0'))
                        total_occurrence_forms = int(normalized_post.get('ocorrencias-TOTAL_FORMS', '0'))
                        logger.warning(f"[DIARY_DEBUG] Dentro da transação (diary.pk={diary.pk}): total_worklog_forms={total_worklog_forms}, total_occurrence_forms={total_occurrence_forms}; normalized_post work_logs-TOTAL_FORMS={normalized_post.get('work_logs-TOTAL_FORMS')}, ocorrencias-TOTAL_FORMS={normalized_post.get('ocorrencias-TOTAL_FORMS')}")
                        
                        if total_image_forms == 0:
                            image_valid_final = True
                        else:
                            has_image_data = False
                            for i in range(total_image_forms):
                                image_key = f'diaryimage_set-{i}-image'
                                caption_key = f'diaryimage_set-{i}-caption'
                                id_key = f'diaryimage_set-{i}-id'
                                if (image_key in preserved_files and preserved_files.get(image_key)) or \
                                   (caption_key in normalized_post and normalized_post.get(caption_key, '').strip()) or \
                                   (id_key in normalized_post and normalized_post.get(id_key, '').strip()):
                                    has_image_data = True
                                    break
                            image_valid_final = True if not has_image_data else image_formset.is_valid()
                        
                        if total_worklog_forms == 0:
                            worklog_valid_final = True
                        else:
                            has_worklog_data = False
                            for i in range(total_worklog_forms):
                                # Verifica múltiplos campos para detectar se há dados
                                activity_key = f'work_logs-{i}-activity_description'
                                location_key = f'work_logs-{i}-location'
                                notes_key = f'work_logs-{i}-notes'
                                percentage_key = f'work_logs-{i}-percentage_executed_today'
                                progress_key = f'work_logs-{i}-accumulated_progress_snapshot'
                                id_key = f'work_logs-{i}-id'
                                delete_key = f'work_logs-{i}-DELETE'
                                
                                # Verifica se há dados (ignora se estiver marcado para deletar)
                                is_deleted = normalized_post.get(delete_key, '').strip() == 'on'
                                
                                # IMPORTANTE: activity_description é obrigatório, então só considera como tendo dados
                                # se o campo obrigatório estiver preenchido OU se for uma edição (tem ID)
                                activity_description = normalized_post.get(activity_key, '').strip() if activity_key in normalized_post else ''
                                has_id = id_key in normalized_post and normalized_post.get(id_key, '').strip()
                                
                                # Considera como tendo dados apenas se:
                                # 1. Tem ID (é uma edição) OU
                                # 2. Tem activity_description preenchido (campo obrigatório)
                                if not is_deleted and (has_id or activity_description):
                                    has_worklog_data = True
                                    break
                            
                            # Se não há dados válidos, considera válido (formset vazio)
                            # Se há dados, valida o formset
                            if not has_worklog_data:
                                worklog_valid_final = True
                            else:
                                worklog_valid_final = worklog_formset.is_valid()
                                # Log detalhado se falhar
                                if not worklog_valid_final:
                                    logger.warning(f"Formset de worklogs inválido. Erros: {worklog_formset.errors}")
                                    for i, form in enumerate(worklog_formset.forms):
                                        if form.errors:
                                            logger.warning(f"  Form {i} erros: {form.errors}")
                                    if worklog_formset.non_form_errors():
                                        logger.warning(f"  Erros não-form: {worklog_formset.non_form_errors()}")
                        
                        if total_occurrence_forms == 0:
                            occurrence_valid_final = True
                        else:
                            has_occurrence_data = False
                            for i in range(total_occurrence_forms):
                                delete_key = f'ocorrencias-{i}-DELETE'
                                is_deleted = normalized_post.get(delete_key, '').strip() == 'on'
                                if is_deleted:
                                    continue
                                description_key = f'ocorrencias-{i}-description'
                                id_key = f'ocorrencias-{i}-id'
                                if (description_key in normalized_post and normalized_post.get(description_key, '').strip()) or \
                                   (id_key in normalized_post and normalized_post.get(id_key, '').strip()):
                                    has_occurrence_data = True
                                    break
                            occurrence_valid_final = True if not has_occurrence_data else occurrence_formset.is_valid()
                            # #region agent log
                            if not occurrence_valid_final and has_occurrence_data:
                                errs = [(i, f.errors) for i, f in enumerate(occurrence_formset.forms) if f.errors]
                                _dbg("frontend_views:occurrence_formset_invalid", "formset invalid with data", {"errors": errs, "non_form_errors": occurrence_formset.non_form_errors()}, "H2")
                            # #endregion

                        logger.info(f"Re-validação dos formsets (com PK): imagens={image_valid_final}, worklogs={worklog_valid_final}, ocorrências={occurrence_valid_final}")
                        logger.warning(f"[DIARY_DEBUG] Validação final: worklog_valid_final={worklog_valid_final}, occurrence_valid_final={occurrence_valid_final}; worklog_formset.forms={len(worklog_formset.forms)}, occurrence_formset.forms={len(occurrence_formset.forms)}")
                        if not worklog_valid_final and total_worklog_forms > 0:
                            for i, f in enumerate(worklog_formset.forms):
                                if f.errors:
                                    logger.warning(f"[DIARY_DEBUG] [FINAL] Worklog form {i} erros: {f.errors}")
                            if worklog_formset.non_form_errors():
                                logger.warning(f"[DIARY_DEBUG] [FINAL] Worklog non_form_errors: {worklog_formset.non_form_errors()}")
                        if not occurrence_valid_final and total_occurrence_forms > 0:
                            for i, f in enumerate(occurrence_formset.forms):
                                if f.errors:
                                    logger.warning(f"[DIARY_DEBUG] [FINAL] Ocorrência form {i} erros: {f.errors}")
                            if occurrence_formset.non_form_errors():
                                logger.warning(f"[DIARY_DEBUG] [FINAL] Ocorrência non_form_errors: {occurrence_formset.non_form_errors()}")
                        
                        # Se o formset de worklogs é crítico e falhou com dados, coleta erros antes de fazer rollback
                        if not worklog_valid_final and total_worklog_forms > 0:
                            logger.error("Formset de worklogs falhou e há dados. Coletando erros antes do rollback.")
                            # Coleta erros específicos do formset para mostrar ao usuário
                            worklog_errors = []
                            for i, form_obj in enumerate(worklog_formset.forms):
                                if form_obj.errors:
                                    for field, errors in form_obj.errors.items():
                                        for error in errors:
                                            worklog_errors.append(f'Atividade {i+1} - {field}: {error}')
                            if worklog_formset.non_form_errors():
                                for error in worklog_formset.non_form_errors():
                                    worklog_errors.append(f'Erro geral: {error}')
                            
                            error_message = "Erro ao processar atividades. "
                            if worklog_errors:
                                error_message += "Detalhes: " + "; ".join(worklog_errors[:5])  # Limita a 5 erros
                                if len(worklog_errors) > 5:
                                    error_message += f" (e mais {len(worklog_errors) - 5} erro(s))"
                            else:
                                error_message += "Por favor, verifique os dados das atividades e tente novamente."
                            
                            logger.error(f"Erros do formset de worklogs: {worklog_errors}")
                            raise ValueError(error_message)
                        
                        # Atualiza as variáveis de validação para usar os valores finais
                        image_valid = image_valid_final
                        worklog_valid = worklog_valid_final
                        occurrence_valid = occurrence_valid_final
                    else:
                        # Se o diário não existe ou não tem PK, considera formsets vazios (válidos)
                        image_valid = True
                        worklog_valid = True
                        occurrence_valid = True
                        logger.warning("Diário não existe ou não tem PK, considerando formsets vazios como válidos")
                    
                    # Se for edição de diário existente, registra no log
                    if diary and diary.pk and not is_new:
                        # Registra edição
                        try:
                            from core.models import DiaryEditLog
                            DiaryEditLog.objects.create(
                                diary=diary,
                                edited_by=request.user,
                                notes=f"Diário editado via formulário web"
                            )
                        except Exception as e:
                            logger.warning(f"Erro ao criar DiaryEditLog: {e}")
                    
                    # 1. PROCESSAMENTO DE FOTOS
                    from core.models import DiaryImage
                    saved_images = []
                    manually_saved_images = []
                    
                    # Se formset é válido, usa ele para salvar (evita duplicação)
                    if image_valid:
                        saved_images = image_formset.save()
                        logger.info(f"Imagens salvas pelo formset: {len(saved_images)} imagens")
                        
                        # Coleta IDs das imagens salvas pelo formset para evitar processamento manual duplicado
                        formset_saved_ids = {img.id for img in saved_images if img.id}
                        formset_processed_indices = set()
                        
                        # Identifica quais índices do formset foram processados
                        for form in image_formset.forms:
                            if form.instance.pk:
                                # Tenta identificar o índice do form
                                prefix = form.prefix
                                if prefix and prefix.startswith('diaryimage_set-'):
                                    try:
                                        idx = int(prefix.replace('diaryimage_set-', '').split('-')[0])
                                        formset_processed_indices.add(idx)
                                    except (ValueError, IndexError):
                                        pass
                    else:
                        saved_images = []
                        formset_saved_ids = set()
                        formset_processed_indices = set()
                        logger.warning("Formset de imagens inválido, processando manualmente...")
                    
                    # Processa fotos manualmente APENAS se:
                    # 1. Formset falhou (image_valid = False), OU
                    # 2. Há fotos que não foram processadas pelo formset (fotos adicionadas dinamicamente)
                    total_forms = int(normalized_post.get('diaryimage_set-TOTAL_FORMS', '0'))
                    logger.info(f"TOTAL_FORMS do formset de imagens: {total_forms}")
                    
                    # Processa manualmente apenas se formset falhou ou para fotos não processadas
                    if not image_valid or total_forms > len(formset_processed_indices):
                        processed_image_indices = set()
                        for i in range(total_forms):
                            # Se formset processou este índice e foi válido, pula
                            if image_valid and i in formset_processed_indices:
                                continue
                            
                            if i in processed_image_indices:
                                continue
                            
                            image_key = f'diaryimage_set-{i}-image'
                            id_key = f'diaryimage_set-{i}-id'
                            delete_key = f'diaryimage_set-{i}-DELETE'
                            caption_key = f'diaryimage_set-{i}-caption'
                            approved_key = f'diaryimage_set-{i}-is_approved_for_report'
                            
                            if delete_key in request.POST and request.POST[delete_key] == 'on':
                                if id_key in request.POST:
                                    image_id = request.POST.get(id_key, '').strip()
                                    if image_id and image_id not in formset_saved_ids:
                                        try:
                                            existing_image = DiaryImage.objects.get(id=image_id, diary=diary)
                                            existing_image.delete()
                                            logger.info(f"Imagem deletada manualmente: ID={image_id}")
                                        except DiaryImage.DoesNotExist:
                                            pass
                                continue
                            
                            # Processa apenas se há arquivo E não foi processado pelo formset
                            if image_key in preserved_files:
                                image_file = preserved_files[image_key]
                                
                                # Valida arquivo antes de processar
                                try:
                                    from .utils.file_validators import validate_image_file
                                    validate_image_file(image_file)
                                except ValidationError as e:
                                    logger.error(f"Erro de validação na imagem {image_key}: {e}")
                                    messages.error(request, f'Erro ao processar imagem: {e}')
                                    continue
                                
                                image_id = request.POST.get(id_key, '').strip()
                                
                                # Se tem ID e já foi salvo pelo formset, pula (evita duplicação)
                                if image_id and image_id in formset_saved_ids:
                                    logger.info(f"Imagem {image_id} já processada pelo formset, pulando processamento manual")
                                    continue
                                
                                # Se tem ID mas não foi salvo pelo formset, atualiza
                                if image_id:
                                    try:
                                        existing = DiaryImage.objects.get(id=image_id, diary=diary)
                                        existing.image = image_file
                                        if caption_key in request.POST:
                                            existing.caption = request.POST[caption_key].strip()
                                        if approved_key in request.POST:
                                            existing.is_approved_for_report = (request.POST[approved_key] == 'on')
                                        existing.save()
                                        manually_saved_images.append(existing)
                                        logger.info(f"Imagem atualizada manualmente: ID={existing.id}")
                                    except DiaryImage.DoesNotExist:
                                        # Cria nova se não encontrou (legenda obrigatória)
                                        caption = request.POST.get(caption_key, '').strip()
                                        if not caption:
                                            continue
                                        is_approved = approved_key in request.POST and request.POST[approved_key] == 'on'
                                        new_image = DiaryImage.objects.create(
                                            diary=diary,
                                            image=image_file,
                                            caption=caption,
                                            is_approved_for_report=is_approved
                                        )
                                        manually_saved_images.append(new_image)
                                        logger.info(f"Imagem nova criada manualmente: ID={new_image.id}")
                                else:
                                    # Nova imagem sem ID - verifica se não foi salva pelo formset
                                    already_saved = False
                                    if saved_images:
                                        # Compara pelo tamanho e nome do arquivo
                                        for saved_img in saved_images:
                                            if (hasattr(saved_img, 'image') and saved_img.image and 
                                                saved_img.image.size == image_file.size):
                                                already_saved = True
                                                break
                                    
                                    if not already_saved:
                                        caption = request.POST.get(caption_key, '').strip()
                                        if not caption:
                                            continue
                                        is_approved = approved_key in request.POST and request.POST[approved_key] == 'on'
                                        new_image = DiaryImage.objects.create(
                                            diary=diary,
                                            image=image_file,
                                            caption=caption,
                                            is_approved_for_report=is_approved
                                        )
                                        manually_saved_images.append(new_image)
                                        logger.info(f"Imagem nova salva manualmente: ID={new_image.id}")
                                
                                processed_image_indices.add(i)
                    
                    all_saved_images = list(saved_images) + manually_saved_images
                    logger.info(f"Total de imagens salvas: {len(all_saved_images)} (formset: {len(saved_images)}, manual: {len(manually_saved_images)})")
                    
                    # 2. PROCESSAMENTO DE VÍDEOS (sempre processa)
                    from core.models import DiaryVideo
                    saved_videos = []
                    
                    if diary and diary.pk:
                        for video in diary.videos.all():
                            video_id = str(video.id)
                            delete_key = f'video_delete_{video_id}'
                            video_file_key = f'video_{video_id}'
                            caption_key = f'video_caption_{video_id}'
                            
                            if delete_key in request.POST:
                                logger.info(f"Deletando vídeo ID={video_id}")
                                video.delete()
                                continue
                            
                            if caption_key in request.POST:
                                video.caption = request.POST[caption_key].strip()
                                video.save()
                                if video not in saved_videos:
                                    saved_videos.append(video)
                            
                            if video_file_key in preserved_files:
                                logger.info(f"Atualizando vídeo ID={video_id}")
                                video_file = preserved_files[video_file_key]
                                # Valida arquivo antes de processar
                                try:
                                    from .utils.file_validators import validate_video_file
                                    validate_video_file(video_file)
                                except ValidationError as e:
                                    logger.error(f"Erro de validação no vídeo {video_id}: {e}")
                                    messages.error(request, f'Erro ao processar vídeo: {e}')
                                    continue
                                video.video = video_file
                                if caption_key in request.POST:
                                    video.caption = request.POST[caption_key].strip()
                                video.save()
                                if video not in saved_videos:
                                    saved_videos.append(video)
                    
                    processed_video_indices = set()
                    for key in preserved_files.keys():
                        if key.startswith('video_new_'):
                            try:
                                index = int(key.replace('video_new_', ''))
                                if index in processed_video_indices:
                                    continue
                                processed_video_indices.add(index)
                                
                                video_file = preserved_files[key]
                                # Valida arquivo antes de processar
                                try:
                                    from .utils.file_validators import validate_video_file
                                    validate_video_file(video_file)
                                except ValidationError as e:
                                    logger.error(f"Erro de validação no vídeo novo {key}: {e}")
                                    messages.error(request, f'Erro ao processar vídeo: {e}')
                                    continue
                                
                                caption_key = f'video_caption_new_{index}'
                                caption = request.POST.get(caption_key, '').strip()
                                if not caption:
                                    continue
                                video = DiaryVideo.objects.create(
                                    diary=diary,
                                    video=video_file,
                                    caption=caption,
                                    is_approved_for_report=True
                                )
                                saved_videos.append(video)
                                logger.info(f"Vídeo novo salvo: ID={video.id}")
                            except (ValueError, IndexError) as e:
                                logger.warning(f"Erro ao processar vídeo {key}: {e}")
                    
                    logger.info(f"Vídeos salvos: {len(saved_videos)} vídeos")
                    
                    # 3. PROCESSAMENTO DE ANEXOS (sempre processa)
                    from core.models import DiaryAttachment
                    saved_attachments = []
                    
                    if diary and diary.pk:
                        kept_attachment_ids = set()
                        for key in request.POST.keys():
                            if key.startswith('attachment_name_') and not key.startswith('attachment_name_new_'):
                                try:
                                    attachment_id = int(key.replace('attachment_name_', ''))
                                    kept_attachment_ids.add(attachment_id)
                                except ValueError:
                                    pass
                        
                        for key in preserved_files.keys():
                            if key.startswith('attachment_') and not key.startswith('attachment_new_'):
                                try:
                                    attachment_id = int(key.replace('attachment_', ''))
                                    kept_attachment_ids.add(attachment_id)
                                except ValueError:
                                    pass
                        
                        for attachment in list(diary.attachments.all()):
                            if attachment.id not in kept_attachment_ids:
                                logger.info(f"Deletando anexo ID={attachment.id}")
                                attachment.delete()
                                continue
                            
                            attachment_id = str(attachment.id)
                            file_key = f'attachment_{attachment_id}'
                            name_key = f'attachment_name_{attachment_id}'
                            
                            if name_key in request.POST:
                                attachment.name = request.POST[name_key].strip()
                                attachment.save()
                                if attachment not in saved_attachments:
                                    saved_attachments.append(attachment)
                            
                            if file_key in preserved_files:
                                logger.info(f"Atualizando anexo ID={attachment_id}")
                                attachment_file = preserved_files[file_key]
                                # Valida arquivo antes de processar
                                try:
                                    from .utils.file_validators import validate_attachment_file
                                    validate_attachment_file(attachment_file)
                                except ValidationError as e:
                                    logger.error(f"Erro de validação no anexo {attachment_id}: {e}")
                                    messages.error(request, f'Erro ao processar anexo: {e}')
                                    continue
                                attachment.file = attachment_file
                                if name_key in request.POST:
                                    attachment.name = request.POST[name_key].strip()
                                attachment.save()
                                if attachment not in saved_attachments:
                                    saved_attachments.append(attachment)
                    
                    processed_attachment_indices = set()
                    for key in preserved_files.keys():
                        if key.startswith('attachment_new_'):
                            try:
                                index = int(key.replace('attachment_new_', ''))
                                if index in processed_attachment_indices:
                                    continue
                                processed_attachment_indices.add(index)
                                
                                attachment_file = preserved_files[key]
                                # Valida arquivo antes de processar
                                try:
                                    from .utils.file_validators import validate_attachment_file
                                    validate_attachment_file(attachment_file)
                                except ValidationError as e:
                                    logger.error(f"Erro de validação no anexo novo {key}: {e}")
                                    messages.error(request, f'Erro ao processar anexo: {e}')
                                    continue
                                
                                name_key = f'attachment_name_new_{index}'
                                name = request.POST.get(name_key, '').strip() or attachment_file.name
                                
                                attachment = DiaryAttachment.objects.create(
                                    diary=diary,
                                    file=attachment_file,
                                    name=name
                                )
                                saved_attachments.append(attachment)
                                logger.info(f"Anexo novo salvo: ID={attachment.id}")
                            except (ValueError, IndexError) as e:
                                logger.warning(f"Erro ao processar anexo {key}: {e}")
                    
                    logger.info(f"Anexos salvos: {len(saved_attachments)} anexos")
                    
                    # 4. PROCESSAMENTO DE MÃO DE OBRA E EQUIPAMENTOS (sempre processa)
                    import json
                    from core.models import Labor, Equipment, DiaryLaborEntry, LaborCargo
                    
                    # Novo sistema: mão de obra por categorias/cargos (diary_labor_data)
                    diary_labor_json = request.POST.get('diary_labor_data', '')
                    if diary_labor_json:
                        try:
                            diary_labor_data = json.loads(diary_labor_json) if diary_labor_json else []
                            DiaryLaborEntry.objects.filter(diary=diary).delete()
                            for item in diary_labor_data:
                                cargo_id = item.get('cargo_id')
                                quantity = max(1, int(item.get('quantity', 1)))
                                company = (item.get('company') or '').strip()
                                if not cargo_id or not LaborCargo.objects.filter(pk=cargo_id).exists():
                                    continue
                                DiaryLaborEntry.objects.create(
                                    diary=diary,
                                    cargo_id=cargo_id,
                                    quantity=quantity,
                                    company=company
                                )
                            logger.info(f"DiaryLaborEntry: {len(diary_labor_data)} itens salvos")
                            request._labor_objects = []
                        except (json.JSONDecodeError, ValueError, TypeError) as e:
                            logger.warning(f"Erro ao processar diary_labor_data: {e}")
                            request._labor_objects = []
                    else:
                        # Legado: labor_data (nome/role/quantidade)
                        labor_data_json = request.POST.get('labor_data', '[]')
                        try:
                            labor_data = json.loads(labor_data_json) if labor_data_json else []
                            logger.info(f"Dados de mão de obra recebidos: {len(labor_data)} itens")
                            
                            labor_objects = []
                            for labor_item in labor_data:
                                labor_name = labor_item.get('name', '').strip()
                                labor_role = labor_item.get('role', '')
                                labor_quantity = int(labor_item.get('quantity', 1))
                                
                                if not labor_name or not labor_role:
                                    continue
                                
                                labor, created = Labor.objects.get_or_create(
                                    name=labor_name,
                                    role=labor_role,
                                    defaults={
                                        'is_active': True
                                    }
                                )
                                
                                for _ in range(labor_quantity):
                                    labor_objects.append(labor)
                                
                                logger.info(f"Labor processado: {labor.name} ({labor.get_role_display()}) x{labor_quantity}")
                            
                            request._labor_objects = labor_objects
                        except (json.JSONDecodeError, ValueError) as e:
                            logger.warning(f"Erro ao processar dados de mão de obra: {e}")
                            request._labor_objects = []
                    
                    equipment_data_json = request.POST.get('equipment_data', '[]')
                    try:
                        equipment_data = json.loads(equipment_data_json) if equipment_data_json else []
                        logger.info(f"Dados de equipamentos recebidos: {len(equipment_data)} itens")
                        logger.warning("[DIARY_DEBUG] equipment_data no POST: len=%s, payload (300 chars)=%s", len(equipment_data), (equipment_data_json or '')[:300])
                        
                        equipment_items = []  # lista de (equipment, quantity) para through
                        for equipment_item in equipment_data:
                            equipment_name = equipment_item.get('name', '').strip()
                            equipment_quantity = int(equipment_item.get('quantity', 1))
                            
                            if not equipment_name:
                                continue
                            
                            equipment_code = f"EQ-{equipment_name.upper().replace(' ', '-')[:20]}"
                            
                            equipment, created = Equipment.objects.get_or_create(
                                code=equipment_code,
                                defaults={
                                    'name': equipment_name,
                                    'equipment_type': equipment_name,
                                    'is_active': True
                                }
                            )
                            
                            equipment_items.append((equipment, equipment_quantity))
                            logger.info(f"Equipment processado: {equipment.name} x{equipment_quantity}")
                        
                        request._equipment_items = equipment_items
                    except (json.JSONDecodeError, ValueError) as e:
                        logger.warning(f"Erro ao processar dados de equipamentos: {e}")
                        request._equipment_items = []
                    
                    # 5. PROCESSAMENTO DE WORKLOGS (prioridade: JSON > formset)
                    saved_worklogs = []
                    work_logs_json_str = (request.POST.get('work_logs_json') or '').strip()
                    occurrences_json_str = (request.POST.get('occurrences_json') or '').strip()
                    # Só usa JSON quando o payload tem conteúdo; [] ou vazio evita apagar dados por engano
                    use_worklogs_json = 'work_logs_json' in request.POST and work_logs_json_str not in ('', '[]')
                    use_occurrences_json = 'occurrences_json' in request.POST and occurrences_json_str not in ('', '[]')
                    logger.warning(
                        "[DIARY_DEBUG] Antes de salvar: worklog_valid=%s, occurrence_valid=%s; "
                        "use_worklogs_json=%s (len=%s), use_occurrences_json=%s (len=%s); request.user.pk=%s",
                        worklog_valid, occurrence_valid,
                        use_worklogs_json, len(work_logs_json_str),
                        use_occurrences_json, len(occurrences_json_str),
                        getattr(request.user, 'pk', None),
                    )
                    logger.warning(
                        "[DIARY_DEBUG] POST keys: work_logs_json=%s, occurrences_json=%s (se use_*=False com payload, checar nome da chave/encoding)",
                        "work_logs_json" in request.POST,
                        "occurrences_json" in request.POST,
                    )
                    if use_worklogs_json:
                        from core.diary_json_services import create_worklogs_from_json
                        create_worklogs_from_json(diary, project, work_logs_json_str)
                        saved_worklogs = list(diary.work_logs.all())
                        logger.info(f"Worklogs criados a partir de JSON; total no diário: {len(saved_worklogs)}")
                    elif worklog_valid:
                        saved_worklogs = worklog_formset.save()
                        logger.info(f"Worklogs salvos pelo formset: {len(saved_worklogs)} worklogs")
                        logger.warning(f"[DIARY_DEBUG] Worklogs salvos: {len(saved_worklogs)} itens")
                    else:
                        logger.warning("Formset de worklogs inválido, pulando processamento")
                    
                    # Se não há worklogs mas há mão de obra ou equipamentos, cria um worklog padrão
                    # IMPORTANTE: Isso deve acontecer mesmo se o formset falhou
                    if (not saved_worklogs and 
                        ((hasattr(request, '_labor_objects') and request._labor_objects) or 
                         (hasattr(request, '_equipment_items') and request._equipment_items))):
                        from core.models import DailyWorkLog, Activity
                        from django.db import IntegrityError
                        
                        # Valida que project existe
                        if not project:
                            logger.error("Tentativa de criar worklog padrão sem projeto")
                            raise ValueError("Projeto é obrigatório para criar worklog padrão")
                        
                        # Cria uma Activity genérica para o worklog padrão
                        # Treebeard requer usar add_root() para criar nós raiz
                        with transaction.atomic():
                            # Tenta buscar Activity existente primeiro
                            try:
                                default_activity = Activity.objects.get(
                                    project=project,
                                    name='Registro Geral de Mão de Obra e Equipamentos'
                                )
                                created = False
                            except Activity.DoesNotExist:
                                # Se não existe, cria como raiz usando add_root()
                                from decimal import Decimal
                                default_activity = Activity.add_root(
                                    project=project,
                                    name='Registro Geral de Mão de Obra e Equipamentos',
                                    code='GEN-MAO-OBRA-EQUIP',
                                    description='Atividade genérica para registro de mão de obra e equipamentos sem atividade específica',
                                    weight=Decimal('0.00'),
                                    status=ActivityStatus.NOT_STARTED
                                )
                                created = True
                                logger.info(f"Activity padrão criada como raiz: {default_activity.code}")
                            
                            # Verifica se já existe worklog com esta activity e diary (unique_together)
                            # Usa get_or_create para evitar race condition
                            try:
                                default_worklog, created_worklog = DailyWorkLog.objects.get_or_create(
                                    diary=diary,
                                    activity=default_activity,
                                    defaults={
                                        'location': 'Geral',
                                        'notes': 'Registro de mão de obra e equipamentos do dia',
                                        'percentage_executed_today': 0,
                                        'accumulated_progress_snapshot': 0
                                    }
                                )
                                if created_worklog:
                                    logger.info(f"Worklog padrão criado (ID={default_worklog.id}) com Activity {default_activity.id}")
                                else:
                                    logger.info(f"Worklog padrão já existe (ID={default_worklog.id}), reutilizando")
                            except IntegrityError as e:
                                # Se ainda assim houver erro de integridade, tenta buscar novamente
                                logger.warning(f"Erro de integridade ao criar worklog padrão: {e}. Tentando buscar existente...")
                                try:
                                    default_worklog = DailyWorkLog.objects.get(
                                        diary=diary,
                                        activity=default_activity
                                    )
                                    logger.info(f"Worklog padrão encontrado após erro (ID={default_worklog.id})")
                                except DailyWorkLog.DoesNotExist:
                                    logger.error(f"Não foi possível criar ou encontrar worklog padrão: {e}")
                                    raise
                        
                        saved_worklogs = [default_worklog]
                    
                    # Associa mão de obra e equipamentos aos worklogs (sempre, mesmo se formset falhou)
                    if hasattr(request, '_labor_objects') and request._labor_objects and saved_worklogs:
                        for worklog in saved_worklogs:
                            if worklog.pk:
                                # Remove duplicados antes de adicionar
                                existing_labor_ids = set(worklog.resources_labor.values_list('id', flat=True))
                                new_labor_objects = [lab for lab in request._labor_objects if lab.id not in existing_labor_ids]
                                if new_labor_objects:
                                    worklog.resources_labor.add(*new_labor_objects)
                                    logger.info(f"Mão de obra associada ao worklog {worklog.id}: {len(new_labor_objects)} novos itens (total: {len(request._labor_objects)})")
                                else:
                                    logger.info(f"Mão de obra já associada ao worklog {worklog.id}, nenhum item novo")
                    
                    if hasattr(request, '_equipment_items') and request._equipment_items is not None and saved_worklogs:
                        from .models import DailyWorkLogEquipment
                        for worklog in saved_worklogs:
                            if worklog.pk:
                                # Payload é a fonte da verdade: substitui equipamentos do worklog via through (com quantidade)
                                DailyWorkLogEquipment.objects.filter(work_log=worklog).delete()
                                if request._equipment_items:
                                    for equipment, qty in request._equipment_items:
                                        DailyWorkLogEquipment.objects.create(
                                            work_log=worklog,
                                            equipment=equipment,
                                            quantity=max(1, int(qty)),
                                        )
                                    logger.info(f"Equipamentos associados ao worklog {worklog.id}: {len(request._equipment_items)} itens (com quantidade)")
                    
                    # Recalcula progresso quando worklogs foram salvos (JSON ou formset)
                    if use_worklogs_json and saved_worklogs:
                        try:
                            from .services import ProgressService
                            for wl in saved_worklogs:
                                if wl.pk and wl.activity_id:
                                    try:
                                        ProgressService.calculate_rollup_progress(wl.activity_id)
                                    except Exception as e:
                                        logger.warning(f"Erro ao recalcular progresso da atividade {wl.activity_id}: {e}", exc_info=True)
                        except Exception as e:
                            logger.warning(f"Erro ao recalcular progresso (JSON): {e}", exc_info=True)
                    elif worklog_valid:
                        for form_obj in worklog_formset.forms:
                            if form_obj.instance.pk and form_obj.instance.activity:
                                try:
                                    from .services import ProgressService
                                    ProgressService.calculate_rollup_progress(form_obj.instance.activity_id)
                                except Exception as e:
                                    logger.warning(f"Erro ao recalcular progresso da atividade {form_obj.instance.activity_id}: {e}", exc_info=True)
                    
                    # 6. PROCESSAMENTO DE OCORRÊNCIAS (prioridade: JSON > formset)
                    # #region agent log
                    _dbg("frontend_views:before_occurrence_save", "occurrence_valid and formset state", {
                        "occurrence_valid": occurrence_valid,
                        "occurrence_formset_forms_len": len(occurrence_formset.forms),
                        "use_occurrences_json": use_occurrences_json,
                    }, "H2")
                    # #endregion
                    if use_occurrences_json:
                        from core.diary_json_services import create_occurrences_from_json
                        created_occurrences = create_occurrences_from_json(diary, occurrences_json_str, request.user)
                        logger.info("Ocorrências criadas a partir de JSON: %s itens (diary_id=%s)", len(created_occurrences), diary.pk)
                    elif occurrence_valid:
                        occurrences = occurrence_formset.save(commit=False)
                        # #region agent log
                        _dbg("frontend_views:after_occurrence_save", "occurrences saved", {"count": len(occurrences)}, "H2")
                        # #endregion
                        for occurrence in occurrences:
                            if not occurrence.pk:
                                occurrence.created_by = request.user
                            occurrence.save()
                        occurrence_formset.save_m2m()
                        logger.info(f"Ocorrências salvas: {len(occurrences)} ocorrências")
                        logger.warning(f"[DIARY_DEBUG] Ocorrências salvas: {len(occurrences)} itens")
                    else:
                        logger.warning("Formset de ocorrências inválido, pulando processamento")
                    
                    # 7. ATUALIZA INFORMAÇÕES DO PROJETO
                    if project:
                        project_updated = False
                        if 'project_name' in request.POST and request.POST['project_name']:
                            project.name = request.POST['project_name']
                            project_updated = True
                        if 'project_client_name' in request.POST:
                            project.client_name = request.POST['project_client_name']
                            project_updated = True
                        if 'project_address' in request.POST:
                            project.address = request.POST['project_address']
                            project_updated = True
                        if 'project_responsible' in request.POST:
                            project.responsible = request.POST['project_responsible']
                            project_updated = True
                        
                        if project_updated:
                            project.save()
                    
                    # 8. SALVA ASSINATURAS (assinatura obrigatória, exceto em Salvamento Parcial)
                    signature_inspection = request.POST.get('signature_inspection')
                    signature_production = request.POST.get('signature_production')
                    is_partial_save = (
                        request.POST.get('partial_save') == '1' or
                        request.POST.get('as_partial_checkbox') == '1'
                    )
                    
                    if not is_partial_save and (not signature_inspection or not signature_inspection.strip()):
                        raise ValueError("A assinatura do responsável pelo preenchimento é obrigatória. Preencha a seção Assinaturas.")
                    
                    if signature_inspection:
                        DiarySignature.objects.update_or_create(
                            diary=diary,
                            signature_type='inspection',
                            defaults={
                                'signer': request.user,
                                'signature_data': signature_inspection
                            }
                        )
                    
                    if signature_production:
                        DiarySignature.objects.update_or_create(
                            diary=diary,
                            signature_type='production',
                            defaults={
                                'signer': request.user,
                                'signature_data': signature_production
                            }
                        )
                    
            except ValueError as e:
                # Erro específico de validação de formset - rollback automático
                logger.error(f"Erro de validação: {e}", exc_info=True)
                # Usa a mensagem detalhada da exceção (já contém os erros específicos)
                messages.error(request, str(e))
                # A transação fará rollback automaticamente, então o diário não será salvo
                # Recria o diário sem PK para o formulário
                if diary and diary.pk:
                    try:
                        # Tenta refresh do banco, mas se não existir (rollback), recria instância
                        diary.refresh_from_db()
                        # Se após refresh ainda tem PK, significa que era um diário existente
                        # Se não tem PK, foi revertido pelo rollback
                        if not diary.pk:
                            # Recria a instância do form
                            form_temp = ConstructionDiaryForm(request.POST, user=request.user, project=project)
                            if form_temp.is_valid():
                                diary = form_temp.save(commit=False)
                            else:
                                diary = None
                    except Exception:
                        # Se não existe mais no banco (rollback de diário novo), recria instância
                        try:
                            form_temp = ConstructionDiaryForm(request.POST, user=request.user, project=project)
                            if form_temp.is_valid():
                                diary = form_temp.save(commit=False)
                            else:
                                diary = None
                        except Exception:
                            diary = None
                
                # Retorna para o formulário com erros
                form = ConstructionDiaryForm(request.POST, instance=diary, user=request.user, project=project)
                files_for_formset = preserved_files if 'preserved_files' in locals() else request.FILES
                # Garante que normalized_post está definido (pode não estar se exceção ocorreu antes)
                if 'normalized_post' not in locals():
                    from copy import deepcopy
                    normalized_post = deepcopy(request.POST)
                    if hasattr(normalized_post, '_mutable'):
                        normalized_post._mutable = True
                    # Normaliza prefixo de worklogs só se o POST veio com dailyworklog_set (não sobrescrever work_logs)
                    if 'dailyworklog_set-TOTAL_FORMS' in request.POST and 'work_logs-TOTAL_FORMS' not in request.POST:
                        normalized_post['work_logs-TOTAL_FORMS'] = request.POST.get('dailyworklog_set-TOTAL_FORMS', '0')
                        normalized_post['work_logs-INITIAL_FORMS'] = request.POST.get('dailyworklog_set-INITIAL_FORMS', '0')
                        for key, value in request.POST.items():
                            if key.startswith('dailyworklog_set-'):
                                new_key = key.replace('dailyworklog_set-', 'work_logs-', 1)
                                normalized_post[new_key] = value
                    # Normaliza occurrences -> ocorrencias
                    if 'occurrences-TOTAL_FORMS' in request.POST and 'ocorrencias-TOTAL_FORMS' not in normalized_post:
                        for key, value in request.POST.items():
                            if key.startswith('occurrences-'):
                                normalized_post[key.replace('occurrences-', 'ocorrencias-', 1)] = value
                image_formset = DiaryImageFormSet(request.POST, files_for_formset, instance=diary)
                worklog_formset = DailyWorkLogFormSet(normalized_post, instance=diary, form_kwargs={'diary': diary}, prefix='work_logs')
                occurrence_formset = DiaryOccurrenceFormSet(normalized_post, instance=diary, prefix='ocorrencias')
                from .models import OccurrenceTag
                try:
                    occurrence_tags = OccurrenceTag.objects.filter(is_active=True)
                except Exception:
                    occurrence_tags = []
                context = {
                    'diary': diary if diary and diary.pk else None,
                    'form': form,
                    'image_formset': image_formset,
                    'worklog_formset': worklog_formset,
                    'occurrence_formset': occurrence_formset,
                    'occurrence_tags': occurrence_tags,
                    'project': project,
                    'next_report_number': None,
                    'initial_contractante': request.POST.get('project_client_name', get_contractante_for_project(project)),
                }
                return render(request, 'core/daily_log_form.html', context)
            except Exception as e:
                logger.error(f"ERRO durante processamento: {e}", exc_info=True)
                logger.warning(f"[DIARY_DEBUG] Exceção durante salvamento do diário: {type(e).__name__}: {e}")
                messages.error(request, f'Erro ao processar dados: {str(e)}')
                # A transação fará rollback automaticamente, então o diário não será salvo
                # Se for um diário novo que foi criado na transação, será revertido
                # Se for edição, as mudanças serão revertidas
                
                # Recria o diário sem PK para o formulário (após rollback)
                if diary and diary.pk:
                    try:
                        diary.refresh_from_db()
                    except Exception:
                        # Se não existe mais no banco (rollback), recria instância
                        diary = None
                
                # Retorna para o formulário com erros
                form = ConstructionDiaryForm(request.POST, instance=diary, user=request.user, project=project)
                files_for_formset = preserved_files if 'preserved_files' in locals() else request.FILES
                # Garante que normalized_post está definido (pode não estar se exceção ocorreu antes)
                if 'normalized_post' not in locals():
                    from copy import deepcopy
                    normalized_post = deepcopy(request.POST)
                    if hasattr(normalized_post, '_mutable'):
                        normalized_post._mutable = True
                    # Normaliza prefixo de worklogs só se o POST veio com dailyworklog_set (não sobrescrever work_logs)
                    if 'dailyworklog_set-TOTAL_FORMS' in request.POST and 'work_logs-TOTAL_FORMS' not in request.POST:
                        normalized_post['work_logs-TOTAL_FORMS'] = request.POST.get('dailyworklog_set-TOTAL_FORMS', '0')
                        normalized_post['work_logs-INITIAL_FORMS'] = request.POST.get('dailyworklog_set-INITIAL_FORMS', '0')
                        for key, value in request.POST.items():
                            if key.startswith('dailyworklog_set-'):
                                new_key = key.replace('dailyworklog_set-', 'work_logs-', 1)
                                normalized_post[new_key] = value
                    # Normaliza occurrences -> ocorrencias
                    if 'occurrences-TOTAL_FORMS' in request.POST and 'ocorrencias-TOTAL_FORMS' not in normalized_post:
                        for key, value in request.POST.items():
                            if key.startswith('occurrences-'):
                                normalized_post[key.replace('occurrences-', 'ocorrencias-', 1)] = value
                image_formset = DiaryImageFormSet(request.POST, files_for_formset, instance=diary)
                worklog_formset = DailyWorkLogFormSet(normalized_post, instance=diary, form_kwargs={'diary': diary}, prefix='work_logs')
                occurrence_formset = DiaryOccurrenceFormSet(normalized_post, instance=diary, prefix='ocorrencias')
                from .models import OccurrenceTag
                try:
                    occurrence_tags = OccurrenceTag.objects.filter(is_active=True)
                except Exception:
                    occurrence_tags = []
                context = {
                    'diary': diary if diary and diary.pk else None,
                    'form': form,
                    'image_formset': image_formset,
                    'worklog_formset': worklog_formset,
                    'occurrence_formset': occurrence_formset,
                    'occurrence_tags': occurrence_tags,
                    'project': project,
                    'next_report_number': None,
                    'initial_contractante': request.POST.get('project_client_name', get_contractante_for_project(project)),
                }
                return render(request, 'core/daily_log_form.html', context)
            
            # Mensagem de sucesso ou aviso
            # Se chegou aqui, a transação foi commitada com sucesso
            if image_valid and worklog_valid and occurrence_valid:
                from django.urls import reverse
                # Salvar diário (não rascunho) = diário aprovado → enviar e-mail ao dono da obra
                if not is_partial_save and diary and diary.status == DiaryStatus.APROVADO:
                    try:
                        from .diary_email import send_diary_to_owners
                        send_diary_to_owners(diary)
                    except Exception as e:
                        logger.exception("Erro ao enviar diário aos donos da obra: %s", e)
                if is_partial_save:
                    messages.success(request, 'Diário salvo parcialmente. Você pode continuar o preenchimento depois.')
                    return redirect('report-list')
                if is_new:
                    messages.success(request, f'Diário criado com sucesso! Relatório #{diary.report_number or "em processamento"}')
                    return redirect('report-list')
                else:
                    messages.success(request, f'Diário atualizado com sucesso!')
                    return redirect(reverse('diary-detail', kwargs={'pk': diary.pk}))
            else:
                # Alguns formsets falharam, mas dados foram salvos na transação: re-renderiza com erros
                error_parts = []
                if not image_valid:
                    error_parts.append('fotos')
                if not worklog_valid:
                    error_parts.append('atividades')
                if not occurrence_valid:
                    error_parts.append('ocorrências')
                messages.warning(request, f'Diário salvo, mas alguns dados não puderam ser processados: {", ".join(error_parts)}. Verifique os erros abaixo.')
                form = ConstructionDiaryForm(request.POST, instance=diary, user=request.user, project=project)
                files_for_retry = request.FILES
                if locals().get('preserved_files'):
                    files_for_retry = preserved_files
                image_formset = DiaryImageFormSet(normalized_post, files_for_retry, instance=diary)
                worklog_formset = DailyWorkLogFormSet(normalized_post, instance=diary, form_kwargs={'diary': diary}, prefix='work_logs')
                occurrence_formset = DiaryOccurrenceFormSet(normalized_post, instance=diary, prefix='ocorrencias')
                from .models import OccurrenceTag
                try:
                    occurrence_tags = OccurrenceTag.objects.filter(is_active=True)
                except Exception:
                    occurrence_tags = []
                context = {
                    'diary': diary,
                    'form': form,
                    'image_formset': image_formset,
                    'worklog_formset': worklog_formset,
                    'occurrence_formset': occurrence_formset,
                    'occurrence_tags': occurrence_tags,
                    'project': project,
                    'next_report_number': None,
                    'initial_contractante': request.POST.get('project_client_name', get_contractante_for_project(project)),
                }
                return render(request, 'core/daily_log_form.html', context)
        else:
            # Form principal inválido - coleta erros
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Form principal INVÁLIDO. Erros: {form.errors}")
            logger.warning(f"[DIARY_DEBUG] Form principal inválido; form.errors={dict(form.errors)}")
            
            errors = []
            error_details = []
            
            if not form.is_valid():
                errors.append('Erros no formulário principal')
                for field, field_errors in form.errors.items():
                    # Remove mensagens duplicadas de unique_together do Django
                    unique_error_msg = "Diário de Obra com este Projeto e Data já existe."
                    filtered_errors = [e for e in field_errors if unique_error_msg not in str(e)]
                    if filtered_errors:
                        error_details.append(f'{field}: {", ".join(filtered_errors)}')
                        logger.error(f"Erro no campo {field}: {', '.join(filtered_errors)}")
            
            if not image_formset.is_valid():
                errors.append('Erros nas imagens')
                # Adiciona detalhes dos erros de imagem
                for i, form_obj in enumerate(image_formset.forms):
                    if form_obj.errors:
                        error_details.append(f'Foto {i+1}: {", ".join([", ".join(errors) for errors in form_obj.errors.values()])}')
                # Adiciona erros não-form (erros do formset)
                if image_formset.non_form_errors():
                    error_details.append(f'Erros gerais nas fotos: {", ".join(image_formset.non_form_errors())}')
            
            if not worklog_formset.is_valid():
                errors.append('Erros nas atividades')
                for i, form_obj in enumerate(worklog_formset.forms):
                    if form_obj.errors:
                        error_details.append(f'Atividade {i+1}: {", ".join([", ".join(errors) for errors in form_obj.errors.values()])}')
            
            if not occurrence_formset.is_valid():
                errors.append('Erros nas ocorrências')
                for i, form_obj in enumerate(occurrence_formset.forms):
                    if form_obj.errors:
                        error_details.append(f'Ocorrência {i+1}: {", ".join([", ".join(errors) for errors in form_obj.errors.values()])}')
            
            if errors:
                error_message = 'Por favor, corrija os erros abaixo: ' + ', '.join(errors)
                if error_details:
                    error_message += '\n\nDetalhes:\n' + '\n'.join(error_details[:10])  # Limita a 10 detalhes
                messages.error(request, error_message)
            # Re-renderiza o formulário com os dados do POST e os erros (evita perda de dados)
            from django.utils.datastructures import MultiValueDict
            from copy import deepcopy
            normalized_post = deepcopy(request.POST)
            if hasattr(normalized_post, '_mutable'):
                normalized_post._mutable = True
            if 'dailyworklog_set-TOTAL_FORMS' in request.POST and 'work_logs-TOTAL_FORMS' not in request.POST:
                normalized_post['work_logs-TOTAL_FORMS'] = request.POST.get('dailyworklog_set-TOTAL_FORMS', '0')
                normalized_post['work_logs-INITIAL_FORMS'] = request.POST.get('dailyworklog_set-INITIAL_FORMS', '0')
                for key, value in request.POST.items():
                    if key.startswith('dailyworklog_set-'):
                        normalized_post[key.replace('dailyworklog_set-', 'work_logs-', 1)] = value
            if 'occurrences-TOTAL_FORMS' in request.POST and 'ocorrencias-TOTAL_FORMS' not in normalized_post:
                for key, value in request.POST.items():
                    if key.startswith('occurrences-'):
                        normalized_post[key.replace('occurrences-', 'ocorrencias-', 1)] = value
            files_for_formset = request.FILES
            if locals().get('files_dict'):
                files_for_formset = MultiValueDict()
                for k, v in files_dict.items():
                    files_for_formset.appendlist(k, v)
            image_formset = DiaryImageFormSet(request.POST, files_for_formset, instance=diary)
            worklog_formset = DailyWorkLogFormSet(normalized_post, instance=diary, form_kwargs={'diary': diary}, prefix='work_logs')
            occurrence_formset = DiaryOccurrenceFormSet(normalized_post, instance=diary, prefix='ocorrencias')
            from .models import OccurrenceTag
            try:
                occurrence_tags = OccurrenceTag.objects.filter(is_active=True)
            except Exception:
                occurrence_tags = []
            context = {
                'diary': diary if diary and getattr(diary, 'pk', None) else None,
                'form': form,
                'image_formset': image_formset,
                'worklog_formset': worklog_formset,
                'occurrence_formset': occurrence_formset,
                'occurrence_tags': occurrence_tags,
                'project': project,
                'next_report_number': None,
                'initial_contractante': request.POST.get('project_client_name', get_contractante_for_project(project)),
            }
            return render(request, 'core/daily_log_form.html', context)
    else:
        # GET request - mostra formulário
        # Se vier da lista com projeto/data (modal "Adicionar relatório"), troca a obra na sessão e redireciona
        get_project_id = request.GET.get('project')
        if get_project_id and not pk:
            try:
                proj = Project.objects.get(pk=get_project_id, is_active=True)
                if _user_can_access_project(request.user, proj):
                    request.session['selected_project_id'] = proj.id
                    request.session['selected_project_name'] = proj.name
                    request.session['selected_project_code'] = getattr(proj, 'code', '') or ''
                    get_date = (request.GET.get('date') or '').strip()
                    if get_date:
                        from django.urls import reverse
                        from urllib.parse import urlencode
                        return redirect(reverse('diary-new') + '?' + urlencode({'date': get_date}))
                    return redirect('diary-new')
            except (Project.DoesNotExist, ValueError, TypeError):
                pass  # Mantém projeto atual e ignora parâmetro inválido
        # Sincroniza contratante da obra: se o mapeamento define um valor, grava no projeto (corrige vazio ou valor antigo)
        if project:
            default_cli = get_contractante_for_project(project)
            if default_cli and (getattr(project, 'client_name', None) or '').strip() != default_cli:
                project.client_name = default_cli
                project.save(update_fields=['client_name'])
        # Verifica se há data passada via GET
        initial_date = request.GET.get('date')
        form = ConstructionDiaryForm(instance=diary, user=request.user, project=project)
        
        # Se houver data no GET e for um novo diário, preenche o campo
        if initial_date and not diary:
            try:
                from datetime import datetime
                # Tenta converter a data do formato dd/mm/yyyy para yyyy-mm-dd
                if '/' in initial_date:
                    date_obj = datetime.strptime(initial_date, '%d/%m/%Y').date()
                else:
                    date_obj = datetime.strptime(initial_date, '%Y-%m-%d').date()
                form.initial['date'] = date_obj
            except (ValueError, TypeError):
                pass  # Se não conseguir converter, deixa o formulário usar o padrão
        
        if diary and diary.pk:
            image_formset = DiaryImageFormSet(instance=diary)
            worklog_formset = DailyWorkLogFormSet(instance=diary, form_kwargs={'diary': diary}, prefix='work_logs')
            occurrence_formset = DiaryOccurrenceFormSet(instance=diary, prefix='ocorrencias')
        else:
            image_formset = DiaryImageFormSet()
            worklog_formset = DailyWorkLogFormSet(form_kwargs={'diary': None}, prefix='work_logs')
            occurrence_formset = DiaryOccurrenceFormSet(prefix='ocorrencias')
        
        # Copiar de relatório anterior: copy_from=<pk>&copy=climate,labor,equipment,activities,ocorrencias,all
        copy_from_id = request.GET.get('copy_from')
        copy_options_raw = request.GET.get('copy', '') or ''
        copy_source_diary = None
        copy_opts_list = []
        if project and copy_from_id and copy_options_raw:
            copy_opts = [x.strip().lower() for x in copy_options_raw.split(',') if x.strip()]
            if 'all' in copy_opts:
                copy_opts = ['climate', 'labor', 'equipment', 'activities', 'ocorrencias']
            try:
                from .models import ConstructionDiary as CD
                src = CD.objects.filter(project=project, pk=copy_from_id).select_related('project').prefetch_related(
                    'work_logs__activity', 'work_logs__resources_equipment', 'occurrences__tags'
                ).first()
                if src and (not diary or src.pk != diary.pk):
                    copy_source_diary = src
                    copy_opts_list = copy_opts
                    # Form initial a partir do relatório fonte
                    if any(o in copy_opts for o in ('climate',)):
                        climate_fields = ('weather_conditions', 'weather_morning_condition', 'weather_morning_workable',
                                         'weather_afternoon_condition', 'weather_afternoon_workable',
                                         'weather_night_enabled', 'weather_night_type', 'weather_night_workable',
                                         'pluviometric_index', 'rain_occurrence', 'rain_observations')
                        for f in climate_fields:
                            if f in form.fields and hasattr(src, f):
                                form.initial[f] = getattr(src, f)
                    # Atividades e ocorrências: preencher formsets iniciais (só para novo diário)
                    if not diary and 'activities' in copy_opts and src.work_logs.exists():
                        worklog_initial = []
                        for wl in src.work_logs.prefetch_related('activity').all():
                            worklog_initial.append({
                                'location': wl.location or '',
                                'work_stage': getattr(wl, 'work_stage', 'AN') or 'AN',
                                'percentage_executed_today': wl.percentage_executed_today,
                                'accumulated_progress_snapshot': wl.accumulated_progress_snapshot,
                                'notes': wl.notes or '',
                                'activity_description': wl.activity.name if wl.activity else '',
                            })
                        if worklog_initial:
                            worklog_formset = DailyWorkLogFormSet(initial=worklog_initial, form_kwargs={'diary': None}, prefix='work_logs')
                    if not diary and 'ocorrencias' in copy_opts and src.occurrences.exists():
                        occ_initial = []
                        for o in src.occurrences.prefetch_related('tags').all():
                            occ_initial.append({
                                'description': o.description or '',
                                'tags': list(o.tags.values_list('pk', flat=True)),
                            })
                        if occ_initial:
                            occurrence_formset = DiaryOccurrenceFormSet(initial=occ_initial, prefix='ocorrencias')
            except Exception:
                copy_source_diary = None
    
    # Prepara dados para o template
    from .models import OccurrenceTag, LaborCategory
    try:
        occurrence_tags = OccurrenceTag.objects.filter(is_active=True)
    except Exception:
        # Se as tabelas ainda não existirem (migration não aplicada)
        occurrence_tags = []

    # Categorias e cargos para seleção de mão de obra por blocos
    try:
        labor_categories = LaborCategory.objects.prefetch_related('cargos').order_by('order')
        labor_terceirizada_cargos = []
        for cat in labor_categories:
            if cat.slug == 'terceirizada':
                labor_terceirizada_cargos = [{'id': c.id, 'name': c.name} for c in cat.cargos.all()]
                break
    except Exception:
        labor_categories = []
        labor_terceirizada_cargos = []

    # Categorias e equipamentos padrão para seleção no diário
    try:
        from .models import EquipmentCategory
        equipment_categories = EquipmentCategory.objects.prefetch_related('items').order_by('order')
    except Exception:
        equipment_categories = []

    # Registros de mão de obra já salvos (para edição ou cópia de relatório anterior)
    existing_diary_labor = []
    labor_source = (copy_source_diary if copy_source_diary and 'labor' in copy_opts_list else None) or (diary if diary and diary.pk else None)
    if labor_source:
        try:
            from .models import DiaryLaborEntry
            for e in DiaryLaborEntry.objects.filter(diary=labor_source).select_related('cargo'):
                existing_diary_labor.append({
                    'cargo_id': e.cargo_id,
                    'quantity': e.quantity,
                    'company': e.company or '',
                })
        except Exception as e:
            logger.warning("[DIARY_DEBUG] Erro ao montar existing_diary_labor (cópia): %s", e, exc_info=True)
    
    # Equipamentos já salvos no diário (para edição ou cópia) – agregados por nome e quantity
    existing_diary_equipment = []
    equipment_source = (copy_source_diary if copy_source_diary and 'equipment' in copy_opts_list else None) or (diary if diary and diary.pk else None)
    if equipment_source:
        try:
            from collections import defaultdict
            agg = defaultdict(int)
            equipment_ids = {}
            # Usar through (DailyWorkLogEquipment) para respeitar quantidade; fallback para M2M antigo
            from .models import DailyWorkLogEquipment
            through_rows = DailyWorkLogEquipment.objects.filter(
                work_log__diary=equipment_source
            ).select_related('equipment')
            if through_rows.exists():
                for row in through_rows:
                    name = row.equipment.name
                    agg[name] += row.quantity
                    equipment_ids[name] = row.equipment_id
            else:
                # Fallback: diários antigos sem through (só M2M) contam 1 por ocorrência
                for wl in equipment_source.work_logs.prefetch_related('resources_equipment').all():
                    for eq in wl.resources_equipment.all():
                        agg[eq.name] += 1
                        equipment_ids[eq.name] = eq.id
            for name, qty in agg.items():
                existing_diary_equipment.append({
                    'name': name,
                    'quantity': qty,
                    'equipment_id': equipment_ids.get(name),
                })
            logger.info("[DIARY_DEBUG] existing_diary_equipment montado: %s itens (copy_source=%s)", len(existing_diary_equipment), bool(copy_source_diary and 'equipment' in copy_opts_list))
        except Exception as e:
            logger.warning("[DIARY_DEBUG] Erro ao montar existing_diary_equipment (cópia): %s", e, exc_info=True)
    
    # Calcula próximo número do relatório se for novo diário (otimizado)
    next_report_number = None
    if not diary or not diary.pk:
        last_diary_for_number = ConstructionDiary.objects.filter(
            project=project
        ).only('report_number').order_by('-report_number').first()
        
        if last_diary_for_number and last_diary_for_number.report_number:
            next_report_number = last_diary_for_number.report_number + 1
        else:
            next_report_number = 1
    
    # Último relatório do projeto (para "Copiar dados de relatório anterior" — sempre só o último, escala sem dropdown)
    last_diary_for_copy = None
    if project:
        qs = ConstructionDiary.objects.filter(project=project).order_by('-date', '-report_number')
        if diary and diary.pk:
            qs = qs.exclude(pk=diary.pk)
        d = qs.only('pk', 'date', 'report_number').first()
        if d:
            try:
                date_str = d.date.strftime('%d/%m/%Y') if d.date else ''
            except Exception:
                date_str = str(d.date) if d.date else ''
            last_diary_for_copy = {
                'id': d.pk,
                'date': date_str,
                'report_number': d.report_number or '-',
            }

    context = {
        'diary': diary if diary and diary.pk else None,  # Só passa se já estiver salvo
        'form': form,
        'image_formset': image_formset,
        'worklog_formset': worklog_formset,
        'occurrence_formset': occurrence_formset,
        'occurrence_tags': occurrence_tags,
        'labor_categories': labor_categories,
        'labor_terceirizada_cargos': labor_terceirizada_cargos,
        'existing_diary_labor': existing_diary_labor,
        'existing_diary_equipment': existing_diary_equipment,
        'equipment_categories': equipment_categories,
        'project': project,  # Adiciona projeto ao contexto
        'next_report_number': next_report_number,  # Próximo número do relatório
        'initial_contractante': get_contractante_for_project(project),
        'last_diary_for_copy': last_diary_for_copy,
        'copy_from_id': copy_from_id,
        'copy_options': copy_options_raw if copy_source_diary else '',
        'copy_source_diary': copy_source_diary,
    }
    
    return render(request, 'core/daily_log_form.html', context)


@login_required
@project_required
def diary_pdf_view(request, pk, pdf_type='normal'):
    """
    View para gerar e retornar PDF do diário.
    pdf_type: 'normal', 'detailed', 'no_photos'
    """
    # Importação lazy do PDFGenerator
    global PDFGenerator, WEASYPRINT_AVAILABLE, XHTML2PDF_AVAILABLE
    
    if PDFGenerator is None:
        try:
            from .utils.pdf_generator import (
                PDFGenerator as PDFGen, 
                WEASYPRINT_AVAILABLE as WP_AVAILABLE,
                XHTML2PDF_AVAILABLE as XP_AVAILABLE
            )
            PDFGenerator = PDFGen
            WEASYPRINT_AVAILABLE = WP_AVAILABLE
            XHTML2PDF_AVAILABLE = XP_AVAILABLE
        except Exception as e:
            # Qualquer falha (Cairo, dependências do xhtml2pdf, etc.) → mensagem amigável
            WEASYPRINT_AVAILABLE = False
            XHTML2PDF_AVAILABLE = False
            messages.error(
                request,
                "Geração de PDF temporariamente indisponível (dependências do sistema). "
                "No Windows, instale: pip install xhtml2pdf e reinicie o servidor."
            )
            return redirect('diary-detail', pk=pk)
    
    project = get_selected_project(request)
    diary = get_object_or_404(ConstructionDiary, pk=pk, project=project)
    
    # Verifica se pelo menos uma biblioteca está disponível
    if not WEASYPRINT_AVAILABLE and not XHTML2PDF_AVAILABLE:
        messages.error(
            request,
            "Geração de PDF indisponível. No Windows: 1) pip install xhtml2pdf  2) Reinicie o servidor. "
            "Se ainda falhar (erro de Cairo), instale o GTK3 Runtime: https://github.com/tschoonj/GTK-for-Windows-Runtime-Environment-Installer/releases"
        )
        return redirect('diary-detail', pk=pk)
    
    try:
        # Gera PDF baseado no tipo
        pdf_bytes = PDFGenerator.generate_diary_pdf(diary.id, pdf_type=pdf_type)
        
        if pdf_bytes:
            response = HttpResponse(
                pdf_bytes.getvalue(),
                content_type='application/pdf'
            )
            type_suffix = {
                'normal': '',
                'detailed': '_detalhado',
                'no_photos': '_sem_fotos'
            }.get(pdf_type, '')
            filename = f"diario_{diary.project.code}_{diary.date.strftime('%Y%m%d')}{type_suffix}.pdf"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
        else:
            messages.error(request, "Erro ao gerar PDF. Tente novamente.")
            return redirect('diary-detail', pk=pk)
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Erro ao gerar PDF do diário {diary.id}: {str(e)}", exc_info=True)
        messages.error(request, f"Erro ao gerar PDF: {str(e)}")
        return redirect('diary-detail', pk=pk)


@login_required
@project_required
def diary_excel_view(request, pk):
    """View para gerar e retornar Excel do diário."""
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    
    project = get_selected_project(request)
    diary = get_object_or_404(ConstructionDiary, pk=pk, project=project)
    
    # Cria workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"Diario {diary.date.strftime('%d-%m-%Y')}"  # Remove / para evitar erro no Excel
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=14)
    
    # Título
    ws['A1'] = f"Relatório Diário de Obra - {diary.project.name}"
    ws['A1'].font = title_font
    ws.merge_cells('A1:D1')
    
    row = 3
    
    # Informações do Relatório
    ws[f'A{row}'] = 'N° do Relatório:'
    ws[f'B{row}'] = diary.report_number or '-'
    row += 1
    ws[f'A{row}'] = 'Data:'
    ws[f'B{row}'] = diary.date.strftime('%d/%m/%Y')
    row += 1
    ws[f'A{row}'] = 'Obra:'
    ws[f'B{row}'] = diary.project.name
    row += 2
    
    # Condições Climáticas
    ws[f'A{row}'] = 'Condições Climáticas'
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws.merge_cells(f'A{row}:D{row}')
    row += 1
    ws[f'A{row}'] = 'Manhã:'
    ws[f'B{row}'] = diary.weather_conditions or '-'
    row += 1
    ws[f'A{row}'] = 'Tarde:'
    ws[f'B{row}'] = diary.weather_conditions or '-'
    row += 2
    
    # Atividades
    ws[f'A{row}'] = 'Atividades'
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws.merge_cells(f'A{row}:D{row}')
    row += 1
    ws[f'A{row}'] = 'Código'
    ws[f'B{row}'] = 'Descrição'
    ws[f'C{row}'] = 'Progresso (%)'
    ws[f'D{row}'] = 'Local'
    for cell in [ws[f'A{row}'], ws[f'B{row}'], ws[f'C{row}'], ws[f'D{row}']]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    row += 1
    
    for work_log in diary.work_logs.all():
        ws[f'A{row}'] = work_log.activity.code
        ws[f'B{row}'] = work_log.activity.name
        ws[f'C{row}'] = float(work_log.percentage_executed_today)
        ws[f'D{row}'] = work_log.location or '-'
        row += 1
    
    # Ajusta largura das colunas
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 25
    
    # Resposta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"diario_{diary.project.code}_{diary.date.strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def profile_view(request):
    """View de perfil do usuário."""
    from django.contrib import messages
    from .forms import ProfileEditForm
    
    user = request.user
    
    # Estatísticas do usuário
    from .models import ConstructionDiary, DiaryImage
    
    project = get_selected_project(request)
    user_diaries_count = 0
    user_photos_count = 0
    
    if project:
        user_diaries_count = ConstructionDiary.objects.filter(
            project=project,
            created_by=user
        ).count()
        
        user_photos_count = DiaryImage.objects.filter(
            diary__project=project,
            diary__created_by=user
        ).count()
    
    # Processa o formulário se for POST
    if request.method == 'POST':
        form = ProfileEditForm(request.POST, user=user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Perfil atualizado com sucesso!')
            # Se a senha foi alterada, faz logout e redireciona para login
            if form.cleaned_data.get('new_password'):
                from django.contrib.auth import logout
                logout(request)
                messages.info(request, 'Sua senha foi alterada. Por favor, faça login novamente.')
                return redirect('login')
            return redirect('profile')
        else:
            messages.error(request, 'Por favor, corrija os erros abaixo.')
    else:
        form = ProfileEditForm(user=user)
    
    context = {
        'user': user,
        'user_diaries_count': user_diaries_count,
        'user_photos_count': user_photos_count,
        'form': form,
    }
    
    return render(request, 'core/profile.html', context)


@login_required
@project_required
def activity_form_view(request, project_id, pk=None, parent_id=None):
    """View para criar/editar atividades na EAP."""
    from django.contrib import messages
    from .forms import ActivityForm
    
    project = get_selected_project(request)
    
    # Verifica se o projeto da URL corresponde ao selecionado
    if project.id != project_id:
        from django.core.exceptions import PermissionDenied
        raise PermissionDenied("Projeto não corresponde ao selecionado.")
    
    if pk:
        activity = get_object_or_404(Activity, pk=pk, project=project)
        form_title = "Editar Atividade"
        parent = activity.get_parent() if not activity.is_root() else None
    else:
        activity = None
        form_title = "Nova Atividade"
        if parent_id:
            parent = get_object_or_404(Activity, pk=parent_id, project=project)
        else:
            parent = None
    
    if request.method == 'POST':
        form = ActivityForm(request.POST, instance=activity)
        if form.is_valid():
            activity = form.save(commit=False)
            activity.project = project
            
            if not activity.pk:  # Nova atividade
                if parent:
                    activity = parent.add_child(instance=activity)
                else:
                    # Cria como raiz
                    activity = Activity.add_root(instance=activity)
            else:
                activity.save()
            
            messages.success(request, f'Atividade "{activity.name}" foi salva com sucesso!')
            return redirect('project-activities-tree', project_id=project.id)
        else:
            messages.error(request, 'Por favor, corrija os erros abaixo.')
    else:
        form = ActivityForm(instance=activity)
    
    # Lista de atividades para seleção de pai (se edição)
    available_parents = Activity.objects.filter(project=project)
    if activity:
        # Exclui a própria atividade e seus descendentes
        available_parents = available_parents.exclude(
            path__startswith=activity.path
        )
    
    context = {
        'form': form,
        'activity': activity,
        'project': project,
        'parent': parent,
        'form_title': form_title,
        'available_parents': available_parents,
    }
    
    return render(request, 'core/activity_form.html', context)


@login_required
@project_required
def activity_delete_view(request, project_id, pk):
    """View para deletar atividade."""
    from django.contrib import messages
    
    project = get_selected_project(request)
    activity = get_object_or_404(Activity, pk=pk, project=project)
    
    if request.method == 'POST':
        # Verifica se a atividade tem filhos
        if activity.get_children().exists():
            messages.error(request, f'Não é possível deletar a atividade "{activity.name}" pois ela possui atividades filhas.')
            return redirect('project-activities-tree', project_id=project.id)
        
        activity_name = activity.name
        activity.delete()
        messages.success(request, f'Atividade "{activity_name}" foi deletada com sucesso!')
        return redirect('project-activities-tree', project_id=project.id)
    
    context = {
        'activity': activity,
        'project': project,
    }
    
    return render(request, 'core/activity_delete_confirm.html', context)


@login_required
@project_required
def labor_list_view(request):
    """View para listar mão de obra."""
    project = get_selected_project(request)
    
    from .models import Labor
    
    labor_list = Labor.objects.filter(is_active=True).order_by('labor_type', 'name')
    
    # Filtros
    search = request.GET.get('search', '')
    labor_type_filter = request.GET.get('labor_type')
    
    if search:
        labor_list = labor_list.filter(
            Q(name__icontains=search) |
            Q(role_custom__icontains=search) |
            Q(company__icontains=search)
        )
    
    if labor_type_filter:
        labor_list = labor_list.filter(labor_type=labor_type_filter)
    
    # Paginação
    from django.core.paginator import Paginator
    paginator = Paginator(labor_list, 20)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    # Estatísticas
    total_labor = labor_list.count()
    labor_by_type = labor_list.values('labor_type').annotate(count=Count('id'))
    
    context = {
        'labor_list': page_obj,
        'total_labor': total_labor,
        'labor_by_type': labor_by_type,
        'search': search,
        'labor_type_filter': labor_type_filter,
    }
    
    return render(request, 'core/labor_list.html', context)


@login_required
@project_required
def labor_form_view(request, pk=None):
    """View para criar/editar mão de obra."""
    from django.contrib import messages
    from .forms import LaborForm
    
    if pk:
        labor = get_object_or_404(Labor, pk=pk)
        form_title = "Editar Mão de Obra"
    else:
        labor = None
        form_title = "Nova Mão de Obra"
    
    if request.method == 'POST':
        form = LaborForm(request.POST, instance=labor)
        if form.is_valid():
            labor = form.save()
            messages.success(request, f'Mão de obra "{labor.name}" foi salva com sucesso!')
            return redirect('labor-list')
        else:
            messages.error(request, 'Por favor, corrija os erros abaixo.')
    else:
        form = LaborForm(instance=labor)
    
    context = {
        'form': form,
        'labor': labor,
        'form_title': form_title,
    }
    
    return render(request, 'core/labor_form.html', context)


@login_required
@project_required
def equipment_list_view(request):
    """View para listar equipamentos."""
    project = get_selected_project(request)
    
    from .models import Equipment
    
    equipment_list = Equipment.objects.filter(is_active=True).order_by('name')
    
    # Filtros
    search = request.GET.get('search', '')
    
    if search:
        equipment_list = equipment_list.filter(
            Q(name__icontains=search) |
            Q(equipment_type__icontains=search) |
            Q(description__icontains=search)
        )
    
    # Paginação
    from django.core.paginator import Paginator
    paginator = Paginator(equipment_list, 20)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    # Estatísticas
    total_equipment = equipment_list.count()
    
    context = {
        'equipment_list': page_obj,
        'total_equipment': total_equipment,
        'search': search,
    }
    
    return render(request, 'core/equipment_list.html', context)


@login_required
@project_required
def equipment_form_view(request, pk=None):
    """View para criar/editar equipamentos."""
    from django.contrib import messages
    from .forms import EquipmentForm
    
    if pk:
        equipment = get_object_or_404(Equipment, pk=pk)
        form_title = "Editar Equipamento"
    else:
        equipment = None
        form_title = "Novo Equipamento"
    
    if request.method == 'POST':
        form = EquipmentForm(request.POST, instance=equipment)
        if form.is_valid():
            equipment = form.save()
            messages.success(request, f'Equipamento "{equipment.name}" foi salvo com sucesso!')
            return redirect('equipment-list')
        else:
            messages.error(request, 'Por favor, corrija os erros abaixo.')
    else:
        form = EquipmentForm(instance=equipment)
    
    context = {
        'form': form,
        'equipment': equipment,
        'form_title': form_title,
    }
    
    return render(request, 'core/equipment_form.html', context)


@login_required
def notifications_view(request):
    """View para listar notificações do usuário."""
    from .models import Notification
    
    notifications = Notification.objects.filter(
        user=request.user
    ).select_related('related_diary', 'related_diary__project').order_by('-created_at')
    
    # Marca notificações como lidas
    unread_count = notifications.filter(is_read=False).count()
    
    # Filtros
    filter_type = request.GET.get('type')
    filter_read = request.GET.get('read')
    
    if filter_type:
        notifications = notifications.filter(notification_type=filter_type)
    
    if filter_read == 'unread':
        notifications = notifications.filter(is_read=False)
    elif filter_read == 'read':
        notifications = notifications.filter(is_read=True)
    
    # Paginação
    from django.core.paginator import Paginator
    paginator = Paginator(notifications, 20)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    context = {
        'notifications': page_obj,
        'unread_count': unread_count,
        'filter_type': filter_type,
        'filter_read': filter_read,
    }
    
    return render(request, 'core/notifications.html', context)


@login_required
def notification_mark_read_view(request, pk):
    """View para marcar notificação como lida."""
    from django.http import JsonResponse
    from .models import Notification
    
    notification = get_object_or_404(Notification, pk=pk, user=request.user)
    notification.is_read = True
    notification.save()
    
    if request.headers.get('HX-Request'):
        return render(request, 'core/partials/notification_item.html', {'notification': notification})
    
    return JsonResponse({'status': 'success'})


@login_required
def notification_mark_all_read_view(request):
    """View para marcar todas as notificações como lidas."""
    from django.http import JsonResponse
    from django.contrib import messages
    from django.shortcuts import redirect
    from .models import Notification
    
    updated = Notification.objects.filter(
        user=request.user,
        is_read=False
    ).update(is_read=True)
    
    messages.success(request, f'{updated} notificação(ões) marcada(s) como lida(s).')
    
    if request.headers.get('HX-Request'):
        return redirect('notifications')
    
    return redirect('notifications')


@login_required
@project_required
def analytics_view(request):
    """View de análise de dados e estatísticas."""
    project = get_selected_project(request)
    
    from .models import ConstructionDiary, DiaryImage, Activity, DailyWorkLog, Labor, Equipment
    from django.db.models import Count, Avg, Sum, Q
    from datetime import datetime, timedelta
    
    # Estatísticas gerais
    total_diaries = ConstructionDiary.objects.filter(project=project).count()
    total_photos = DiaryImage.objects.filter(diary__project=project).count()
    total_activities = Activity.objects.filter(project=project).count()
    
    # Relatórios por status
    diaries_by_status = ConstructionDiary.objects.filter(
        project=project
    ).values('status').annotate(count=Count('id'))
    
    status_labels = {
        'PR': 'Preenchendo',
        'RV': 'Revisar',
        'AP': 'Aprovado',
    }
    
    status_data = {}
    for item in diaries_by_status:
        status_data[status_labels.get(item['status'], item['status'])] = item['count']
    
    # Relatórios por mês (últimos 6 meses)
    six_months_ago = timezone.now().date() - timedelta(days=180)
    from django.db import connection
    if connection.vendor == 'sqlite':
        diaries_by_month = ConstructionDiary.objects.filter(
            project=project,
            date__gte=six_months_ago
        ).extra(
            select={'month': "strftime('%%Y-%%m', date)"}
        ).values('month').annotate(count=Count('id')).order_by('month')
    else:
        # PostgreSQL ou MySQL
        from django.db.models.functions import TruncMonth
        diaries_by_month = ConstructionDiary.objects.filter(
            project=project,
            date__gte=six_months_ago
        ).annotate(
            month=TruncMonth('date')
        ).values('month').annotate(count=Count('id')).order_by('month')
    
    # Top atividades mais frequentes
    top_activities = DailyWorkLog.objects.filter(
        activity__project=project
    ).values('activity__name').annotate(
        count=Count('id')
    ).order_by('-count')[:10]
    
    # Distribuição de mão de obra
    labor_by_type = Labor.objects.filter(
        work_logs__activity__project=project
    ).values('labor_type').annotate(
        count=Count('id', distinct=True)
    ).order_by('labor_type')
    
    # Total de horas trabalhadas
    total_hours = ConstructionDiary.objects.filter(
        project=project,
        work_hours__isnull=False
    ).aggregate(total=Sum('work_hours'))['total'] or 0
    
    context = {
        'project': project,
        'total_diaries': total_diaries,
        'total_photos': total_photos,
        'total_activities': total_activities,
        'status_data': status_data,
        'diaries_by_month': list(diaries_by_month),
        'top_activities': list(top_activities),
        'labor_by_type': list(labor_by_type),
        'total_hours': total_hours,
    }
    
    return render(request, 'core/analytics.html', context)


@login_required
def project_form_view(request, pk=None):
    """View para criar/editar projetos."""
    from django.contrib import messages
    from .forms import ProjectForm
    from django.core.exceptions import PermissionDenied
    
    # Verifica permissão
    if not (request.user.is_staff or request.user.is_superuser):
        raise PermissionDenied("Você não tem permissão para criar ou editar projetos.")
    
    if pk:
        project = get_object_or_404(Project, pk=pk)
        form_title = "Editar Obra"
    else:
        project = None
        form_title = "Nova Obra"
    
    if request.method == 'POST':
        form = ProjectForm(request.POST, instance=project)
        if form.is_valid():
            project = form.save(commit=False)
            # Nova obra: sempre criar como ativa para aparecer no Diário, Painel e GestControll
            if project.pk is None:
                project.is_active = True
            project.save()
            from core.sync_obras import sync_project_to_gestao_and_mapa
            sync_project_to_gestao_and_mapa(project)
            messages.success(request, f'Obra "{project.name}" foi salva e sincronizada com GestControll e Mapa.')
            return redirect('central_project_list')
        else:
            messages.error(request, 'Por favor, corrija os erros abaixo.')
    else:
        form = ProjectForm(instance=project)
    
    context = {
        'form': form,
        'project': project,
        'form_title': form_title,
    }
    
    return render(request, 'core/project_form.html', context)


def _redirect_anonymous_to_login(view_func):
    """Redireciona anônimos para o login (302). Garante comportamento consistente em testes e produção."""
    @wraps(view_func)
    def _wrapped(request, *args, **kwargs):
        if not getattr(request.user, 'is_authenticated', False):
            from django.conf import settings
            from django.contrib.auth.views import redirect_to_login
            return redirect_to_login(request.get_full_path())
        return view_func(request, *args, **kwargs)
    return _wrapped


@_redirect_anonymous_to_login
@login_required
def project_list_view(request):
    """
    View de listagem de projetos (Central).
    Apenas usuários staff/superuser podem acessar. Gerentes usam select-project.
    """
    from django.db.models import Count
    from django.core.exceptions import PermissionDenied
    
    # Anônimo: sempre redirect (302). Evita 403 quando decorators não redirecionam (ex.: em testes).
    if not getattr(request.user, 'is_authenticated', False):
        from django.contrib.auth.views import redirect_to_login
        return redirect_to_login(request.get_full_path())
    # Apenas staff/superuser; Gerentes sem staff recebem 403
    if not (request.user.is_staff or request.user.is_superuser):
        raise PermissionDenied("Você não tem permissão para acessar esta página.")
    
    # Lista todas as obras (ativas e inativas) para aparecer no Painel; inativas podem ser reativadas ao editar
    projects = Project.objects.annotate(
        diaries_count=Count('diaries', distinct=True),
        activities_count=Count('activities', distinct=True)
    ).order_by('-created_at')
    
    context = {
        'projects': projects,
    }
    
    return render(request, 'core/project_list.html', context)


@login_required
@require_http_methods(["POST"])
def project_delete_view(request, pk):
    """Exclui uma obra (apenas staff/superuser). Remove também a obra no GestControll."""
    from django.contrib import messages
    from django.core.exceptions import PermissionDenied

    if not (request.user.is_staff or request.user.is_superuser):
        raise PermissionDenied("Você não tem permissão para excluir obras.")

    project = get_object_or_404(Project, pk=pk)
    name = project.name

    # Remove a obra correspondente no GestControll para sumir da listagem de lá também
    from gestao_aprovacao.models import Obra as ObraGestao
    from django.db.models import ProtectedError

    obras_gestao = ObraGestao.objects.filter(project=project)
    try:
        obras_gestao.delete()
        removido_gestao = True
    except ProtectedError:
        # Há pedidos vinculados; não é possível excluir a obra no GestControll
        removido_gestao = False

    project.delete()

    # Se a obra excluída era a selecionada na sessão, limpa para evitar referência inválida
    if request.session.get('selected_project_id') == pk:
        for key in ('selected_project_id', 'selected_project_name', 'selected_project_code'):
            request.session.pop(key, None)

    if removido_gestao:
        messages.success(request, f'Obra "{name}" foi excluída do Diário e do GestControll.')
    else:
        messages.warning(
            request,
            f'Obra "{name}" foi excluída do Painel. No GestControll ela permanece porque há pedidos vinculados; exclua por lá se desejar.'
        )
    return redirect('central_project_list')


