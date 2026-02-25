from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, F, Sum, Case, When, Value, IntegerField
from django.http import JsonResponse, HttpResponse
from accounts.decorators import require_group
from accounts.groups import GRUPOS
from mapa_obras.models import Obra, LocalObra
from suprimentos.models import ItemMapa, Insumo, HistoricoAlteracao, RecebimentoObra, AlocacaoRecebimento
from suprimentos.forms import InsumoForm, ItemMapaForm, SiengeImportUploadForm
from collections import defaultdict
from datetime import datetime
from uuid import uuid4
from decimal import Decimal, ROUND_HALF_UP
from io import BytesIO
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def get_obra_da_sessao(request):
    """Obtém a obra da sessão ou a primeira disponível. Só considera obras às quais o usuário está vinculado."""
    from mapa_obras.views import _get_obras_for_user, _user_can_access_obra
    obra_id = request.session.get('obra_id')
    if obra_id:
        try:
            obra = Obra.objects.get(id=obra_id, ativa=True)
            if _user_can_access_obra(request, obra):
                return obra
        except Obra.DoesNotExist:
            pass
        request.session.pop('obra_id', None)
    # Fallback: primeira obra permitida ao usuário
    obras = _get_obras_for_user(request)
    obra = obras.first()
    if obra:
        request.session['obra_id'] = obra.id
    return obra


@login_required
@require_group(GRUPOS.ENGENHARIA)
def mapa_engenharia(request):
    """Mapa editável para Engenharia. Só exibe obras às quais o usuário está vinculado."""
    from mapa_obras.views import _get_obras_for_user, _user_can_access_obra
    obras = _get_obras_for_user(request)
    
    # PRIORIDADE: 1) GET param, 2) Sessão (só aceita obra permitida)
    obra_id = request.GET.get('obra')
    if obra_id:
        try:
            obra = Obra.objects.get(id=int(obra_id), ativa=True)
            if _user_can_access_obra(request, obra):
                request.session['obra_id'] = obra.id
        except (Obra.DoesNotExist, ValueError):
            pass
    else:
        # Usar obra da sessão
        obra_sessao = get_obra_da_sessao(request)
        if obra_sessao:
            obra_id = str(obra_sessao.id)
    
    categoria = request.GET.get('categoria', '')
    local_id = request.GET.get('local', '')
    prioridade = request.GET.get('prioridade', '')
    status_filtro = request.GET.get('status', '')
    search = request.GET.get('search', '')
    
    # SEMPRE filtrar pela obra da sessão (segregação estrita)
    # CORREÇÃO PRIORIDADE 1: Otimizar N+1 queries com prefetch_related e annotate
    if not obra_id:
        # Fallback seguro
        itens = ItemMapa.objects.none()
    else:
        from django.db.models import Prefetch
        from suprimentos.models import AlocacaoRecebimento
        
        # Otimização: usar annotate para calcular quantidade_alocada_local em uma query
        # e prefetch_related para carregar relacionamentos
        itens = ItemMapa.objects.filter(obra_id=obra_id).select_related(
            'obra', 'insumo', 'local_aplicacao'
        ).prefetch_related(
            # Prefetch alocações para evitar N+1 queries ao acessar quantidade_alocada_local
            Prefetch(
                'alocacoes',
                queryset=AlocacaoRecebimento.objects.only('quantidade_alocada'),
                to_attr='alocacoes_cache'
            )
        ).annotate(
            # Calcular quantidade_alocada_local diretamente no banco
            quantidade_alocada_annotated=Sum('alocacoes__quantidade_alocada')
        )
    
    if categoria:
        itens = itens.filter(categoria__icontains=categoria)
    
    if local_id:
        itens = itens.filter(local_aplicacao_id=local_id)
    
    if prioridade:
        itens = itens.filter(prioridade=prioridade)
    
    if search:
        # Busca expandida: nome, código, SC, PC, fornecedor, local, responsável
        itens = itens.filter(
            Q(insumo__descricao__icontains=search) |
            Q(insumo__codigo_sienge__icontains=search) |
            Q(descricao_override__icontains=search) |
            Q(numero_sc__icontains=search) |
            Q(numero_pc__icontains=search) |
            Q(empresa_fornecedora__icontains=search) |
            Q(local_aplicacao__nome__icontains=search) |
            Q(responsavel__icontains=search)
        )
    
    # Filtro por status (deve ser aplicado por último, pois usa propriedades calculadas)
    # IMPORTANTE: Alguns status dependem de propriedades calculadas, então precisamos converter para lista
    if status_filtro:
        # Converter para lista para poder usar propriedades calculadas
        itens_lista = list(itens)
        
        if status_filtro == 'LEVANTAMENTO':
            # Sem SC
            itens_lista = [item for item in itens_lista if not item.numero_sc or item.numero_sc.strip() == '']
        elif status_filtro == 'AGUARDANDO_COMPRA':
            # Tem SC mas não tem PC
            itens_lista = [item for item in itens_lista 
                          if item.numero_sc and item.numero_sc.strip() != '' 
                          and (not item.numero_pc or item.numero_pc.strip() == '')]
        elif status_filtro == 'AGUARDANDO_ENTREGA':
            # Tem PC mas não recebeu nada ou recebeu parcialmente
            itens_lista = [item for item in itens_lista 
                          if item.numero_pc and item.numero_pc.strip() != ''
                          and item.quantidade_recebida_obra < item.quantidade_solicitada_sienge]
        elif status_filtro == 'AGUARDANDO_ALOCACAO':
            # Recebeu mas não alocou nada para este local
            itens_lista = [item for item in itens_lista 
                          if item.quantidade_recebida_obra > 0 
                          and item.quantidade_alocada_local == 0]
        elif status_filtro == 'PARCIAL':
            # Alocou parcialmente
            itens_lista = [item for item in itens_lista 
                          if item.quantidade_alocada_local > 0 
                          and ((item.quantidade_solicitada_sienge > 0 and item.quantidade_alocada_local < item.quantidade_solicitada_sienge) or
                               (item.quantidade_solicitada_sienge == 0 and item.quantidade_planejada > 0 and item.quantidade_alocada_local < item.quantidade_planejada))]
        elif status_filtro == 'ENTREGUE':
            # Totalmente alocado
            itens_lista = [item for item in itens_lista 
                          if ((item.quantidade_solicitada_sienge > 0 and item.quantidade_alocada_local >= item.quantidade_solicitada_sienge) or
                              (item.quantidade_solicitada_sienge == 0 and item.quantidade_planejada > 0 and item.quantidade_alocada_local >= item.quantidade_planejada))]
        elif status_filtro == 'ATRASADO':
            # Prazo vencido e não entregue
            from django.utils import timezone
            hoje = timezone.now().date()
            itens_lista = [item for item in itens_lista 
                          if item.prazo_necessidade 
                          and item.prazo_necessidade < hoje 
                          and ((item.quantidade_solicitada_sienge > 0 and item.quantidade_alocada_local < item.quantidade_solicitada_sienge) or
                               (item.quantidade_solicitada_sienge == 0 and item.quantidade_planejada > 0 and item.quantidade_alocada_local < item.quantidade_planejada))]
        
        # Converter lista filtrada de volta para queryset usando IDs
        # IMPORTANTE: Reaplicar otimizações (annotate/prefetch_related) para manter performance
        if itens_lista:
            ids_filtrados = [item.id for item in itens_lista]
            # Reaplicar otimizações que foram perdidas na conversão para lista
            itens = ItemMapa.objects.filter(id__in=ids_filtrados).select_related(
                'obra', 'insumo', 'local_aplicacao'
            ).prefetch_related(
                Prefetch(
                    'alocacoes',
                    queryset=AlocacaoRecebimento.objects.only('quantidade_alocada'),
                    to_attr='alocacoes_cache'
                )
            ).annotate(
                quantidade_alocada_annotated=Sum('alocacoes__quantidade_alocada')
            )
        else:
            itens = ItemMapa.objects.none()
    
    obra_selecionada = None
    if obra_id:
        obra_selecionada = get_object_or_404(Obra, id=obra_id)
        # IMPORTANTE: Filtrar locais APENAS da obra selecionada (segregação)
        locais = LocalObra.objects.filter(obra_id=obra_id).order_by('tipo', 'nome')
    else:
        locais = LocalObra.objects.none()
    
    # Categorias únicas DESTA OBRA (não de todas)
    if obra_id:
        categorias = ItemMapa.objects.filter(obra_id=obra_id).values_list('categoria', flat=True).distinct().order_by('categoria')
    else:
        categorias = []
    
    # Insumos para o formulário de criação
    insumos = Insumo.objects.filter(ativo=True).order_by('descricao')
    
    # Formulário para criar insumo (para o modal)
    form_insumo = InsumoForm()
    
    # KPIs
    itens_queryset = itens
    kpis = {
        'total': itens_queryset.count(),
        'atrasados': sum(1 for item in itens_queryset if item.is_atrasado),
        'solicitados': itens_queryset.exclude(numero_sc='').count(),
        'em_compra': itens_queryset.exclude(numero_sc='').filter(numero_pc='').count(),
        # Parciais: tem alocação mas não completou (baseado em alocação manual, não recebimento)
        'parciais': sum(1 for item in itens_queryset if 
            item.quantidade_alocada_local > 0 and 
            ((item.quantidade_solicitada_sienge > 0 and item.quantidade_alocada_local < item.quantidade_solicitada_sienge) or
             (item.quantidade_solicitada_sienge == 0 and item.quantidade_planejada > 0 and item.quantidade_alocada_local < item.quantidade_planejada))
        ),
        # Entregues: totalmente alocado (baseado em alocação manual, não recebimento)
        'entregues': sum(1 for item in itens_queryset if 
            ((item.quantidade_solicitada_sienge > 0 and item.quantidade_alocada_local >= item.quantidade_solicitada_sienge) or
             (item.quantidade_solicitada_sienge == 0 and item.quantidade_planejada > 0 and item.quantidade_alocada_local >= item.quantidade_planejada))
        ),
    }
    
    categorias_opcoes = ItemMapa.CATEGORIA_CHOICES
    categorias_opcoes_values = [v for v, _ in categorias_opcoes]
    # Categorias que existem no banco mas não estão mais na lista fechada (legado)
    categorias_legado = [c for c in categorias if c and c not in categorias_opcoes_values]

    context = {
        'obras': obras,
        'obra_selecionada': obra_selecionada,
        'locais': locais,
        'categorias': categorias,
        'categorias_legado': categorias_legado,
        'insumos': insumos,
        'form_insumo': form_insumo,
        # Ordenação: "A CLASSIFICAR" sempre primeiro, depois alfabético
        'itens': itens.annotate(
            ordem_categoria=Case(
                When(categoria='A CLASSIFICAR', then=Value(0)),
                default=Value(1),
                output_field=IntegerField()
            )
        ).order_by('ordem_categoria', 'categoria', 'insumo__descricao'),
        'kpis': kpis,
        'filtros': {
            'obra_id': obra_id,
            'categoria': categoria,
            'local_id': local_id,
            'prioridade': prioridade,
            'status': status_filtro,
            'search': search,
        }
        ,
        'categorias_opcoes': categorias_opcoes,
        'categorias_opcoes_values': categorias_opcoes_values,
    }
    
    return render(request, 'suprimentos/mapa_engenharia.html', context)


@login_required
@require_group(GRUPOS.ENGENHARIA)
def exportar_mapa_excel(request):
    """Exporta o mapa de suprimentos para Excel com formatação."""
    # Aplicar os mesmos filtros da view mapa_engenharia
    obra_id = request.GET.get('obra')
    if not obra_id:
        obra_sessao = get_obra_da_sessao(request)
        if obra_sessao:
            obra_id = str(obra_sessao.id)
    
    categoria = request.GET.get('categoria', '')
    local_id = request.GET.get('local', '')
    prioridade = request.GET.get('prioridade', '')
    search = request.GET.get('search', '')
    
    # Filtrar itens
    # CORREÇÃO: Otimizar N+1 queries na exportação também
    if not obra_id:
        itens = ItemMapa.objects.none()
    else:
        from django.db.models import Prefetch
        from suprimentos.models import AlocacaoRecebimento
        
        itens = ItemMapa.objects.filter(obra_id=obra_id).select_related(
            'obra', 'insumo', 'local_aplicacao'
        ).prefetch_related(
            Prefetch(
                'alocacoes',
                queryset=AlocacaoRecebimento.objects.only('quantidade_alocada'),
                to_attr='alocacoes_cache'
            )
        ).annotate(
            quantidade_alocada_annotated=Sum('alocacoes__quantidade_alocada')
        )
    
    if categoria:
        itens = itens.filter(categoria__icontains=categoria)
    if local_id:
        itens = itens.filter(local_aplicacao_id=local_id)
    if prioridade:
        itens = itens.filter(prioridade=prioridade)
    if search:
        itens = itens.filter(
            Q(insumo__descricao__icontains=search) |
            Q(insumo__codigo_sienge__icontains=search) |
            Q(descricao_override__icontains=search)
        )
    
    # Ordenar como na view
    itens = itens.annotate(
        ordem_categoria=Case(
            When(categoria='A CLASSIFICAR', then=Value(0)),
            default=Value(1),
            output_field=IntegerField()
        )
    ).order_by('ordem_categoria', 'categoria', 'insumo__descricao')
    
    # Converter para lista para evitar problemas de acesso
    itens_lista = list(itens)
    
    # Preparar dados para Excel - EXATAMENTE como aparece na tela
    dados = []
    categoria_anterior = None
    itens_por_categoria = {}
    
    # Agrupar itens por categoria para contar
    for item in itens_lista:
        cat = item.categoria or ''
        if cat not in itens_por_categoria:
            itens_por_categoria[cat] = []
        itens_por_categoria[cat].append(item)
    
    for item in itens_lista:
        # Adicionar linha de categoria se mudou (igual na tela)
        if item.categoria != categoria_anterior:
            categoria_nome = item.categoria or ''
            quantidade_itens = len(itens_por_categoria.get(categoria_nome, []))
            texto_categoria = f"{categoria_nome} ({quantidade_itens} itens)" if categoria_nome else ''
            
            dados.append({
                '1. CATEGORIA': texto_categoria,  # Formato: "CATEGORIA (X itens)"
                '2. CÓDIGO DO INSUMO': '',
                '3. DESCRIÇÃO DO ITEM': '',
                '4. LOCAL': '',
                '5. RESPONSÁVEL': '',
                '6. PRAZO': '',
                '7. QUANTITATIVO': '',
                '8. UND': '',
                '9. Nº SOLICITAÇÃO': '',
                '10. Nº PEDIDO DE COMPRA': '',
                '11. EMPRESA RESPONSÁVEL': '',
                '12. PRAZO RECEBIMENTO': '',
                '13. QUANTIDADE RECEBIDA': '',
                '14. SALDO A SER ENTREGUE': '',
                '15. STATUS': '',
                '16. PRIORIDADE': '',
                '17. OBSERVAÇÃO': '',
                '_is_categoria_header': True,
                '_item': None,
            })
            categoria_anterior = item.categoria
        
        # Função auxiliar para formatar números brasileiros (igual ao filtro format_quantidade)
        def formatar_numero_br(valor):
            """Formata número com separador de milhar (ponto) e vírgula decimal - formato brasileiro"""
            if valor is None or valor == '':
                return '0,00'
            try:
                # Converter para Decimal preservando precisão
                if isinstance(valor, (int, float)):
                    valor_decimal = Decimal(str(valor))
                elif isinstance(valor, Decimal):
                    valor_decimal = valor
                else:
                    valor_decimal = Decimal(str(valor))
                
                # Quantificar para 2 casas decimais
                valor_quantized = valor_decimal.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                valor_str = str(valor_quantized)
                
                # Separar parte inteira e decimal
                if '.' in valor_str:
                    integer_part, decimal_part = valor_str.split('.')
                else:
                    integer_part = valor_str
                    decimal_part = '00'
                
                # Garantir 2 dígitos decimais
                if len(decimal_part) < 2:
                    decimal_part = decimal_part.ljust(2, '0')
                elif len(decimal_part) > 2:
                    decimal_part = decimal_part[:2]
                
                # Adicionar separador de milhar (ponto) na parte inteira
                integer_formatted = ''
                for i, digit in enumerate(reversed(integer_part)):
                    if i > 0 and i % 3 == 0:
                        integer_formatted = '.' + integer_formatted
                    integer_formatted = digit + integer_formatted
                
                # Retornar formato brasileiro: "20.000,00"
                return f"{integer_formatted},{decimal_part}"
            except (ValueError, TypeError, AttributeError):
                return '0,00'
        
        # Formatar quantidade recebida como na tela (alocado/solicitado)
        if item.numero_sc and item.quantidade_solicitada_sienge > 0:
            alocado = formatar_numero_br(item.quantidade_alocada_local or 0)
            solicitado = formatar_numero_br(item.quantidade_solicitada_sienge or 0)
            qtd_recebida_str = f"{alocado}/{solicitado}"
        else:
            alocado = formatar_numero_br(item.quantidade_alocada_local or 0)
            planejado = formatar_numero_br(item.quantidade_planejada or 0)
            qtd_recebida_str = f"{alocado}/{planejado}"
        
        # Formatar saldo como na tela (manter como número para formatação do Excel)
        if item.numero_sc:
            saldo_valor = item.saldo_a_entregar_sienge or Decimal('0.00')
        else:
            if item.saldo_negativo:
                saldo_valor = item.saldo_local_diferenca or Decimal('0.00')
            else:
                saldo_valor = item.saldo_a_alocar_local or Decimal('0.00')
        
        # Formatar quantitativo
        quantitativo_valor = ''
        if item.quantidade_solicitada_sienge > 0:
            quantitativo_valor = formatar_numero_br(item.quantidade_solicitada_sienge)
        
        dados.append({
            '1. CATEGORIA': str(item.categoria or ''),
            '2. CÓDIGO DO INSUMO': str(item.insumo.codigo_sienge or ''),
            '3. DESCRIÇÃO DO ITEM': str(item.descricao_override or item.insumo.descricao or ''),
            '4. LOCAL': str(item.local_aplicacao.nome if item.local_aplicacao else ''),
            '5. RESPONSÁVEL': str(item.responsavel or ''),
            '6. PRAZO': item.prazo_necessidade.strftime('%d/%m/%Y') if item.prazo_necessidade else '',
            '7. QUANTITATIVO': quantitativo_valor,
            '8. UND': str(item.insumo.unidade or 'UND'),
            '9. Nº SOLICITAÇÃO': str(item.numero_sc or ''),
            '10. Nº PEDIDO DE COMPRA': str(item.numero_pc or ''),
            '11. EMPRESA RESPONSÁVEL': str(item.empresa_fornecedora or ''),
            '12. PRAZO RECEBIMENTO': item.prazo_recebimento.strftime('%d/%m/%Y') if item.prazo_recebimento else '',
            '13. QUANTIDADE RECEBIDA': qtd_recebida_str,  # Formato: alocado/solicitado
            '14. SALDO A SER ENTREGUE': formatar_numero_br(saldo_valor),
            '15. STATUS': str(item.status_etapa or ''),
            '16. PRIORIDADE': str(item.get_prioridade_display() if item.prioridade else ''),
            '17. OBSERVAÇÃO': str(item.observacao_eng or ''),
            '_is_categoria_header': False,
            '_item': item,  # Guardar item para formatação de cores
            '_saldo_valor': saldo_valor,  # Guardar valor numérico para formatação
        })
    
    # Se não houver dados, criar DataFrame vazio com colunas
    if not dados:
        dados = [{
            '1. CATEGORIA': '', '2. CÓDIGO DO INSUMO': '', '3. DESCRIÇÃO DO ITEM': '', '4. LOCAL': '',
            '5. RESPONSÁVEL': '', '6. PRAZO': '', '7. QUANTITATIVO': '', '8. UND': '',
            '9. Nº SOLICITAÇÃO': '', '10. Nº PEDIDO DE COMPRA': '', '11. EMPRESA RESPONSÁVEL': '',
            '12. PRAZO RECEBIMENTO': '', '13. QUANTIDADE RECEBIDA': '', '14. SALDO A SER ENTREGUE': '',
            '15. STATUS': '', '16. PRIORIDADE': '', '17. OBSERVAÇÃO': '',
            '_is_categoria_header': False, '_item': None
        }]
    
    # Criar DataFrame (remover colunas auxiliares antes)
    colunas_excel = [
        '1. CATEGORIA', '2. CÓDIGO DO INSUMO', '3. DESCRIÇÃO DO ITEM', '4. LOCAL',
        '5. RESPONSÁVEL', '6. PRAZO', '7. QUANTITATIVO', '8. UND',
        '9. Nº SOLICITAÇÃO', '10. Nº PEDIDO DE COMPRA', '11. EMPRESA RESPONSÁVEL',
        '12. PRAZO RECEBIMENTO', '13. QUANTIDADE RECEBIDA', '14. SALDO A SER ENTREGUE',
        '15. STATUS', '16. PRIORIDADE', '17. OBSERVAÇÃO'
    ]
    
    # Criar DataFrame apenas com as colunas de dados
    dados_limpos = []
    for linha in dados:
        linha_limpa = {col: linha.get(col, '') for col in colunas_excel}
        # Manter referências para formatação
        linha_limpa['_is_categoria_header'] = linha.get('_is_categoria_header', False)
        linha_limpa['_item'] = linha.get('_item')
        dados_limpos.append(linha_limpa)
    
    # Criar DataFrame apenas com colunas de dados (sem auxiliares)
    df = pd.DataFrame([{col: linha.get(col, '') for col in colunas_excel} for linha in dados_limpos])
    
    # Criar arquivo Excel em memória
    output = BytesIO()
    
    try:
        with pd.ExcelWriter(output, engine='openpyxl', mode='w') as writer:
            # Escrever os dados começando na linha 0 (pandas criará cabeçalhos na linha 1)
            df.to_excel(writer, sheet_name='Mapa de Suprimentos', index=False, startrow=0)
            
            # Acessar a planilha para formatação
            workbook = writer.book
            worksheet = writer.sheets['Mapa de Suprimentos']
            
            # Estilos
            header_group_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=11)
            categoria_font = Font(bold=True, size=11)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Inserir linha 1 para cabeçalhos agrupados (antes dos cabeçalhos das colunas que estão na linha 1)
            worksheet.insert_rows(1)
            
            # Linha 1: Cabeçalhos agrupados (REQUISIÇÃO, SOLICITAÇÃO, ENTREGA)
            worksheet.merge_cells('A1:H1')  # REQUISIÇÃO DE PRODUTO/SERVIÇO
            worksheet.merge_cells('I1:K1')  # SOLICITAÇÃO DE COMPRA
            worksheet.merge_cells('L1:N1')  # ENTREGA DE PRODUTO/SERVIÇO
            worksheet.merge_cells('O1:Q1')  # Vazio (17 colunas no total)
            
            cell_req = worksheet['A1']
            cell_req.value = 'REQUISIÇÃO DE PRODUTO/SERVIÇO'
            cell_req.fill = header_group_fill
            cell_req.font = header_font
            cell_req.alignment = Alignment(horizontal='center', vertical='center')
            cell_req.border = border
            
            cell_sol = worksheet['I1']
            cell_sol.value = 'SOLICITAÇÃO DE COMPRA'
            cell_sol.fill = header_group_fill
            cell_sol.font = header_font
            cell_sol.alignment = Alignment(horizontal='center', vertical='center')
            cell_sol.border = border
            
            cell_ent = worksheet['L1']
            cell_ent.value = 'ENTREGA DE PRODUTO/SERVIÇO'
            cell_ent.fill = header_group_fill
            cell_ent.font = header_font
            cell_ent.alignment = Alignment(horizontal='center', vertical='center')
            cell_ent.border = border
            
            # Linha 2: Cabeçalhos das colunas (já criada pelo pandas, agora na linha 2)
            for col_idx, col_name in enumerate(colunas_excel, start=1):
                cell = worksheet.cell(row=2, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = border
            
            # Ajustar largura das colunas (aumentadas para evitar corte)
            column_widths = {
                'A': 25,  # 1. CATEGORIA
                'B': 18,  # 2. CÓDIGO DO INSUMO
                'C': 50,  # 3. DESCRIÇÃO DO ITEM (aumentada para textos longos)
                'D': 25,  # 4. LOCAL
                'E': 20,  # 5. RESPONSÁVEL
                'F': 12,  # 6. PRAZO
                'G': 20,  # 7. QUANTITATIVO
                'H': 10,  # 8. UND
                'I': 18,  # 9. Nº SOLICITAÇÃO
                'J': 18,  # 10. Nº PEDIDO DE COMPRA
                'K': 30,  # 11. EMPRESA RESPONSÁVEL (aumentada)
                'L': 15,  # 12. PRAZO RECEBIMENTO
                'M': 25,  # 13. QUANTIDADE RECEBIDA (aumentada para formato alocado/solicitado)
                'N': 20,  # 14. SALDO A SER ENTREGUE
                'O': 30,  # 15. STATUS (aumentada)
                'P': 15,  # 16. PRIORIDADE
                'Q': 40,  # 17. OBSERVAÇÃO (aumentada)
            }
            
            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width
            
            # Formatar linhas de dados (começando na linha 3, pois linha 1 é agrupador e linha 2 é cabeçalho)
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=3, max_row=worksheet.max_row), start=0):
                if row_idx < len(dados_limpos):
                    linha_dados = dados_limpos[row_idx]
                    item = linha_dados.get('_item')
                    is_categoria_header = linha_dados.get('_is_categoria_header', False)
                    
                    if is_categoria_header:
                        # Linha de categoria: fundo cinza e texto em negrito
                        categoria_fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
                        # Mesclar todas as colunas (A até Q = 17 colunas)
                        row_num = row[0].row
                        worksheet.merge_cells(f'A{row_num}:Q{row_num}')
                        # Aplicar formatação apenas na primeira célula (que agora está mesclada)
                        cell_categoria = row[0]
                        cell_categoria.fill = categoria_fill
                        cell_categoria.font = categoria_font
                        cell_categoria.border = border
                        cell_categoria.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                        # Aplicar bordas nas células mescladas
                        for cell in row:
                            cell.border = border
                    elif item:
                        # Aplicar cor de fundo baseada no status
                        status_fill = None
                        if item.is_atrasado:
                            status_fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
                        elif item.status_css == 'status-verde':
                            status_fill = PatternFill(start_color='E6F7E6', end_color='E6F7E6', fill_type='solid')
                        elif item.status_css == 'status-laranja':
                            status_fill = PatternFill(start_color='FFF4E6', end_color='FFF4E6', fill_type='solid')
                        elif item.status_css == 'status-amarelo':
                            status_fill = PatternFill(start_color='FFFCE6', end_color='FFFCE6', fill_type='solid')
                        elif item.status_css == 'status-vermelho':
                            status_fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
                        elif item.status_css == 'status-azul':
                            status_fill = PatternFill(start_color='E6F0FF', end_color='E6F0FF', fill_type='solid')
                        
                        if status_fill:
                            for cell in row:
                                cell.fill = status_fill
                                cell.border = border
                                cell.alignment = Alignment(vertical='center', horizontal='left', wrap_text=True)
                        else:
                            for cell in row:
                                cell.border = border
                                cell.alignment = Alignment(vertical='center', horizontal='left', wrap_text=True)
                        
                        # Formatação especial para colunas numéricas (centralizar)
                        # Coluna G (7. QUANTITATIVO) - centralizar
                        if len(row) > 6:
                            row[6].alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)
                        # Coluna M (13. QUANTIDADE RECEBIDA) - centralizar
                        if len(row) > 12:
                            row[12].alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)
                        # Coluna N (14. SALDO A SER ENTREGUE) - centralizar
                        if len(row) > 13:
                            row[13].alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)
                    else:
                        # Linha sem item: apenas bordas
                        for cell in row:
                            cell.border = border
                            cell.alignment = Alignment(vertical='center', wrap_text=True)
            
            # Congelar primeira e segunda linha (cabeçalhos)
            worksheet.freeze_panes = 'A3'
            
            # Adicionar informações no rodapé
            if itens_lista:
                obra_nome = itens_lista[0].obra.nome
                data_exportacao = datetime.now().strftime('%d/%m/%Y %H:%M')
                worksheet.cell(row=worksheet.max_row + 2, column=1, value=f'Exportado em: {data_exportacao}')
                worksheet.cell(row=worksheet.max_row, column=1, value=f'Obra: {obra_nome}')
            else:
                obra_nome = 'N/A'
    
    except Exception as e:
        # Em caso de erro, retornar erro HTTP
        from django.http import HttpResponseServerError
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f'Erro ao gerar Excel: {str(e)}', exc_info=True)
        return HttpResponseServerError(f'Erro ao gerar Excel: {str(e)}')
    
    output.seek(0)
    
    # Nome do arquivo
    obra_nome_arquivo = obra_nome.replace(' ', '_') if obra_nome != 'N/A' else 'Mapa'
    nome_arquivo = f'Mapa_Suprimentos_{obra_nome_arquivo}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    # Ler o conteúdo do buffer
    excel_data = output.getvalue()
    output.close()
    
    response = HttpResponse(
        excel_data,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{nome_arquivo}"'
    response['Content-Length'] = len(excel_data)
    
    return response


@login_required
@require_group(GRUPOS.ENGENHARIA)
def criar_item_mapa(request):
    """Cria um novo item no mapa usando Django Form."""
    obra_id = request.GET.get('obra') or request.POST.get('obra')
    
    if request.method == 'POST':
        form = ItemMapaForm(request.POST, obra_id=obra_id)
        
        if form.is_valid():
            try:
                item = form.save(commit=False)
                item.criado_por = request.user
                item.save()
                
                # Se for AJAX, retornar JSON
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': True,
                        'message': 'Item criado com sucesso!',
                        'item_id': item.id
                    })
                
                # Se for POST normal, redirecionar
                messages.success(request, 'Item criado com sucesso!')
                return redirect('mapa:mapa')
            except Exception as e:
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': False,
                        'error': str(e)
                    }, status=500)
                messages.error(request, f'Erro ao criar item: {str(e)}')
        else:
            # Se for AJAX, retornar erros em JSON
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                errors = {}
                for field, error_list in form.errors.items():
                    errors[field] = error_list[0] if error_list else 'Erro de validação'
                
                return JsonResponse({
                    'success': False,
                    'error': 'Erro de validação',
                    'errors': errors
                }, status=400)
    else:
        form = ItemMapaForm(obra_id=obra_id)
    
    # Se for GET ou POST com erro (não AJAX), retornar erro JSON
    # Normalmente isso não acontece pois o modal é usado via AJAX
    if request.headers.get('X-Requested-With') != 'XMLHttpRequest':
        return JsonResponse({
            'success': False,
            'error': 'Esta view deve ser acessada via AJAX'
        }, status=400)


@login_required
@require_group(GRUPOS.ENGENHARIA)
def criar_insumo(request):
    """Cria um novo insumo via formulário Django."""
    if request.method == 'POST':
        form = InsumoForm(request.POST)
        if form.is_valid():
            insumo = form.save(commit=False)
            insumo.ativo = True
            insumo.save()
            
            # Se for requisição AJAX, retornar JSON
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': f'Insumo "{insumo.descricao}" criado com sucesso!',
                    'insumo': {
                        'id': insumo.id,
                        'codigo_sienge': insumo.codigo_sienge,
                        'descricao': insumo.descricao,
                        'unidade': insumo.unidade
                    }
                })
            
            # Se for POST normal, redirecionar com mensagem
            messages.success(request, f'Insumo "{insumo.descricao}" criado com sucesso!')
            return redirect('mapa:mapa')
        else:
            # Se for AJAX, retornar erros em JSON
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'error': 'Erro de validação',
                    'errors': form.errors
                }, status=400)
    else:
        form = InsumoForm()
    
    # Se for GET (não AJAX), retornar erro JSON
    # Normalmente isso não acontece pois o modal é usado via AJAX
    if request.headers.get('X-Requested-With') != 'XMLHttpRequest':
        return JsonResponse({
            'success': False,
            'error': 'Esta view deve ser acessada via AJAX'
        }, status=400)


@login_required
@require_group(GRUPOS.ENGENHARIA)
def criar_levantamento_rapido(request):
    """
    Cria um NOVO ITEM de LEVANTAMENTO (pré-Sienge) de forma prática.
    - Cria um Insumo interno (com código automático) se necessário.
    - Cria o ItemMapa com numero_sc vazio => status 1) LEVANTAMENTO.
    """
    if request.method != 'POST':
        return redirect('mapa:mapa')

    obra_id = request.POST.get('obra') or request.session.get('obra_id')
    if not obra_id:
        messages.error(request, 'Selecione uma obra antes de criar um levantamento.')
        return redirect('mapa:mapa')

    descricao = (request.POST.get('descricao_insumo') or '').strip()
    unidade = (request.POST.get('unidade') or '').strip().upper()
    codigo_manual = (request.POST.get('codigo_insumo') or '').strip()

    if not descricao:
        messages.error(request, 'Informe a descrição do insumo.')
        return redirect(f"{redirect('mapa:mapa').url}?obra={obra_id}")

    # Gerar código interno (não conflita com códigos do Sienge)
    def gerar_codigo_interno():
        return f"SM-LEV-{uuid4().hex[:10].upper()}"

    codigo_final = codigo_manual or gerar_codigo_interno()
    # Garantir unicidade
    while Insumo.objects.filter(codigo_sienge=codigo_final).exists():
        codigo_final = gerar_codigo_interno()

    insumo = Insumo.objects.create(
        codigo_sienge=codigo_final,
        descricao=descricao,
        unidade=unidade,
        ativo=True
    )

    # Campos do item
    categoria = (request.POST.get('categoria') or 'A CLASSIFICAR').strip() or 'A CLASSIFICAR'
    categoria = ' '.join(categoria.split())
    categorias_validas = {v for v, _ in ItemMapa.CATEGORIA_CHOICES}
    if categoria not in categorias_validas:
        messages.error(request, 'Categoria inválida. Selecione uma opção da lista.')
        return redirect(f"{redirect('mapa:mapa').url}?obra={obra_id}")
    prioridade = (request.POST.get('prioridade') or 'MEDIA').strip() or 'MEDIA'
    responsavel = (request.POST.get('responsavel') or '').strip()
    observacao = (request.POST.get('observacao_eng') or '').strip()
    local_id = request.POST.get('local_aplicacao') or None
    prazo_necessidade = request.POST.get('prazo_necessidade') or None
    quantidade_planejada = request.POST.get('quantidade_planejada') or '0'

    try:
        qtd = Decimal(str(quantidade_planejada).replace(',', '.'))
    except Exception:
        qtd = Decimal('0.00')

    item = ItemMapa(
        obra_id=obra_id,
        insumo=insumo,
        categoria=categoria,
        prioridade=prioridade,
        responsavel=responsavel,
        observacao_eng=observacao,
        criado_por=request.user,
        quantidade_planejada=qtd
    )

    if local_id:
        item.local_aplicacao_id = local_id

    if prazo_necessidade:
        try:
            item.prazo_necessidade = datetime.strptime(prazo_necessidade, '%Y-%m-%d').date()
        except Exception:
            item.prazo_necessidade = None

    item.save()

    # Registrar no histórico
    HistoricoAlteracao.registrar(
        obra=item.obra,
        usuario=request.user,
        tipo='CRIACAO',
        descricao=f'Novo Levantamento criado: "{insumo.descricao}" ({item.quantidade_planejada} {insumo.unidade})',
        item_mapa=item,
        campo_alterado='levantamento'
    )

    messages.success(request, f'Levantamento criado: {insumo.descricao}')
    return redirect(f"{redirect('engenharia:mapa').url}?obra={obra_id}")


@login_required
@require_group(GRUPOS.ENGENHARIA)
def importar_sienge_upload(request):
    """
    Tela de upload do arquivo exportado do Sienge para importar RecebimentoObra.
    Usa o management command existente (importar_mapa_controle).
    """
    from django.core.management import call_command
    from io import StringIO
    import tempfile
    import os

    from mapa_obras.views import _get_obras_for_user
    obras = _get_obras_for_user(request)
    form = SiengeImportUploadForm()
    log_output = None
    obra_contexto = get_obra_da_sessao(request)
    import_history = HistoricoAlteracao.objects.filter(
        tipo='IMPORTACAO'
    ).select_related('usuario', 'obra').order_by('-data_hora')[:25]

    if request.method == 'POST':
        form = SiengeImportUploadForm(request.POST, request.FILES)
        if form.is_valid():
            arquivo = form.cleaned_data['arquivo']
            obra_fallback = obra_contexto
            skiprows = 0
            
            # Hash do arquivo para evitar reimportação acidental do MESMO arquivo
            import hashlib
            hasher = hashlib.sha256()
            for chunk in arquivo.chunks():
                hasher.update(chunk)
            file_hash = hasher.hexdigest()
            # Reset do cursor do upload para salvar novamente abaixo
            try:
                arquivo.seek(0)
            except Exception:
                pass
            
            if obra_fallback:
                ja_importado = HistoricoAlteracao.objects.filter(
                    obra=obra_fallback,
                    tipo='IMPORTACAO',
                    valor_anterior=file_hash
                ).order_by('-data_hora').first()
                if ja_importado:
                    messages.warning(
                        request,
                        f'Este mesmo arquivo já foi importado em {ja_importado.data_hora.strftime("%d/%m %H:%M")} '
                        f'por {ja_importado.usuario.username if ja_importado.usuario else "usuário"} (evitando duplicação).'
                    )
                    return render(request, 'suprimentos/importar_sienge.html', {
                        'form': form,
                        'obras': obras,
                        'log_output': None,
                    })

            # Salvar em arquivo temporário
            with tempfile.TemporaryDirectory() as tmpdir:
                original_path = os.path.join(tmpdir, 'upload' + os.path.splitext(arquivo.name)[1].lower())
                with open(original_path, 'wb') as out:
                    for chunk in arquivo.chunks():
                        out.write(chunk)

                path_to_import = original_path
                ext = os.path.splitext(original_path)[1].lower()

                # Se for Excel, converter para CSV separado por ';' (o comando lê CSV)
                if ext in ('.xlsx', '.xls'):
                    try:
                        import pandas as pd
                        # O arquivo do Sienge costuma ter logo/título e o cabeçalho real começa algumas linhas abaixo.
                        # Além disso, alguns arquivos vêm com múltiplas abas ou cabeçalhos repetidos (por página).
                        # Então detectamos (aba + linha do header) automaticamente antes de converter.
                        import unicodedata

                        def norm_xlsx(s: str) -> str:
                            s = '' if s is None else str(s)
                            s = unicodedata.normalize('NFKD', s).encode('ASCII', 'ignore').decode('ASCII')
                            return s.strip().upper()

                        sc_headers = {norm_xlsx(x) for x in ['Nº DA SC', 'N DA SC', 'NUMERO SC', 'NUMERO_DA_SC', 'SC', 'NSC']}
                        obra_headers = {norm_xlsx(x) for x in ['CÓD. OBRA', 'COD OBRA', 'CODIGO OBRA', 'CODIGO_DA_OBRA', 'COD_OBRA', 'OBRA']}
                        insumo_headers = {norm_xlsx(x) for x in ['CÓD. INSUMO', 'COD INSUMO', 'CODIGO INSUMO', 'COD_INSUMO']}

                        def detectar_header_em_raw(raw_df):
                            for i in range(0, min(120, len(raw_df))):
                                row_vals = [norm_xlsx(v) for v in raw_df.iloc[i].tolist()]
                                if any(v in sc_headers for v in row_vals) and any(v in insumo_headers for v in row_vals):
                                    return i
                            # fallback: só achar SC
                            for i in range(0, min(120, len(raw_df))):
                                row_vals = [norm_xlsx(v) for v in raw_df.iloc[i].tolist()]
                                if any(v in sc_headers for v in row_vals):
                                    return i
                            return None

                        # Escolher a melhor aba: aquela com mais linhas válidas (SC + Insumo)
                        best = None
                        best_score = -1
                        best_sheet = None
                        best_header_row = None

                        xls = None
                        df = None
                        try:
                            xls = pd.ExcelFile(original_path)
                            for sheet in xls.sheet_names:
                                raw = pd.read_excel(xls, sheet_name=sheet, header=None, dtype=str)
                                header_row = detectar_header_em_raw(raw)
                                if header_row is None:
                                    continue

                                df_try = raw.iloc[header_row + 1:].copy()
                                df_try.columns = raw.iloc[header_row].tolist()
                                df_try = df_try.dropna(how='all')
                                df_try = df_try.loc[:, [c for c in df_try.columns if str(c).strip() not in ('', 'nan')]]

                                # identificar colunas principais
                                cols_norm = {c: norm_xlsx(c) for c in df_try.columns}
                                sc_col = next((c for c, cn in cols_norm.items() if cn in sc_headers), None)
                                insumo_col = next((c for c, cn in cols_norm.items() if cn in insumo_headers), None)
                                if not sc_col or not insumo_col:
                                    continue

                                # Remover linhas de cabeçalho repetido no meio do arquivo
                                df_try = df_try[~df_try[sc_col].apply(lambda v: norm_xlsx(v) in sc_headers)]

                                # Forward-fill para evitar perder linhas quando Excel deixa SC/Obra/Insumo em branco nas quebras
                                # Também forward-fill no Item se existir
                                item_col = next((c for c, cn in cols_norm.items() if norm_xlsx(c) in {'ITEM', 'N. ITEM', 'N ITEM', 'NUMERO ITEM'}), None)
                                for col_ff in (sc_col, insumo_col, item_col):
                                    if col_ff:
                                        df_try[col_ff] = df_try[col_ff].replace({'': pd.NA, 'nan': pd.NA, None: pd.NA}).ffill()
                                
                                # Garantir que todas as linhas com dados sejam mantidas
                                df_try = df_try.dropna(subset=[sc_col], how='all')

                                score = int(df_try[sc_col].notna().sum())
                                if score > best_score:
                                    best_score = score
                                    best = df_try
                                    best_sheet = sheet
                                    best_header_row = header_row

                            if best is None:
                                raise Exception('Não consegui detectar o cabeçalho do Sienge no Excel (aba/colunas).')

                            df = best.copy()  # Criar cópia para evitar referência ao Excel

                            messages.info(
                                request,
                                f'Excel detectado: aba "{best_sheet}", header na linha {best_header_row + 1}. Linhas lidas: {len(df)}.'
                            )

                            # Garantir que todas as colunas sejam convertidas para string antes de salvar
                            for col in df.columns:
                                df[col] = df[col].astype(str).replace('nan', '', regex=False)
                            
                            csv_path = os.path.join(tmpdir, 'MAPA_CONTROLE.csv')
                            df.to_csv(csv_path, sep=';', index=False, encoding='utf-8-sig')
                            path_to_import = csv_path
                            
                            # Log adicional para debug
                            import logging
                            logger = logging.getLogger(__name__)
                            logger.info(f'Excel convertido: {len(df)} linhas, {len(df.columns)} colunas')
                        finally:
                            # Fechar explicitamente o arquivo Excel para evitar erro de permissão no Windows
                            if 'xls' in locals() and xls is not None:
                                try:
                                    xls.close()
                                except Exception:
                                    pass
                            # Forçar garbage collection para liberar handles de arquivo
                            import gc
                            gc.collect()
                    except ImportError:
                        # Garantir que o arquivo seja fechado mesmo em caso de exceção
                        if 'xls' in locals() and xls is not None:
                            try:
                                xls.close()
                            except Exception:
                                pass
                        messages.error(
                            request,
                            'Para importar Excel (.xlsx) é necessário instalar a dependência "openpyxl". '
                            'Como alternativa, exporte do Sienge como CSV.'
                        )
                        return render(request, 'suprimentos/importar_sienge.html', {
                            'form': form,
                            'obras': obras,
                            'log_output': None,
                        })
                    except Exception as e:
                        # Garantir que o arquivo seja fechado mesmo em caso de exceção
                        if 'xls' in locals() and xls is not None:
                            try:
                                xls.close()
                            except Exception:
                                pass
                        messages.error(request, f'Erro ao ler Excel: {str(e)}')
                        return render(request, 'suprimentos/importar_sienge.html', {
                            'form': form,
                            'obras': obras,
                            'log_output': None,
                        })

                # Auto-detect "skiprows" (quando o arquivo tem linhas antes do cabeçalho)
                import unicodedata

                def norm(s: str) -> str:
                    s = '' if s is None else str(s)
                    s = unicodedata.normalize('NFKD', s).encode('ASCII', 'ignore').decode('ASCII')
                    return s.strip().upper()

                sc_headers = {norm(x) for x in ['Nº DA SC', 'N DA SC', 'NUMERO SC', 'NUMERO_DA_SC', 'SC', 'NSC']}
                obra_headers = {norm(x) for x in ['CÓD. OBRA', 'COD OBRA', 'CODIGO OBRA', 'CODIGO_DA_OBRA', 'COD_OBRA', 'OBRA']}

                def detectar_header_csv(fp: str):
                    encodings = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']
                    for enc in encodings:
                        try:
                            with open(fp, 'r', encoding=enc, errors='replace') as f:
                                lines = []
                                for _ in range(0, 60):
                                    line = f.readline()
                                    if not line:
                                        break
                                    lines.append(line)
                            for idx, line in enumerate(lines):
                                if ';' not in line:
                                    continue
                                tokens = [norm(t) for t in line.split(';')]
                                if any(t in sc_headers for t in tokens):
                                    tem_obra = any(t in obra_headers for t in tokens)
                                    return idx, tem_obra
                        except Exception:
                            continue
                    return 0, False

                skiprows, tem_coluna_obra = detectar_header_csv(path_to_import)

                # Se não tiver coluna de obra no arquivo, usamos a obra do contexto (sessão) automaticamente
                obra_codigo_fallback = obra_fallback.codigo_sienge if (obra_fallback and not tem_coluna_obra) else None

                out = StringIO()
                try:
                    call_command(
                        'importar_mapa_controle',
                        file=path_to_import,
                        obra_codigo=obra_codigo_fallback,
                        skiprows=skiprows,
                        stdout=out
                    )
                    log_output = out.getvalue()

                    # Registrar no histórico (por obra da sessão/fallback)
                    if obra_fallback:
                        HistoricoAlteracao.registrar(
                            obra=obra_fallback,
                            usuario=request.user,
                            tipo='IMPORTACAO',
                            descricao=f'Importação Sienge realizada ({arquivo.name})',
                            campo_alterado='SUCESSO',
                            valor_anterior=file_hash,
                            valor_novo=arquivo.name,
                            ip_address=request.META.get('REMOTE_ADDR')
                        )

                    messages.success(request, 'Importação concluída com sucesso.')
                except Exception as e:
                    log_output = out.getvalue() or None
                    messages.error(request, f'Erro ao importar: {str(e)}')
                    if obra_fallback:
                        HistoricoAlteracao.registrar(
                            obra=obra_fallback,
                            usuario=request.user,
                            tipo='IMPORTACAO',
                            descricao=f'Falha na importação Sienge ({arquivo.name})',
                            campo_alterado='ERRO',
                            valor_anterior=file_hash,
                            valor_novo=str(e)[:500],
                            ip_address=request.META.get('REMOTE_ADDR')
                        )
    # Atualizar histórico após POST
    import_history = HistoricoAlteracao.objects.filter(
        tipo='IMPORTACAO'
    ).select_related('usuario', 'obra').order_by('-data_hora')[:25]

    return render(request, 'suprimentos/importar_sienge.html', {
        'form': form,
        'obras': obras,
        'log_output': log_output,
        'obra_contexto': obra_contexto,
        'import_history': import_history,
    })


@login_required
@require_group(GRUPOS.ENGENHARIA)
def dashboard_2(request):
    """Dashboard 2 - LPLAN Tactical Command - Pipeline Denso por Local. Só exibe obras permitidas ao usuário."""
    from collections import defaultdict
    from decimal import Decimal
    from mapa_obras.views import _get_obras_for_user, _user_can_access_obra
    
    obras = _get_obras_for_user(request)
    
    # PRIORIDADE: 1) GET param, 2) Sessão (só aceita obra permitida)
    obra_id = request.GET.get('obra')
    if obra_id:
        try:
            obra = Obra.objects.get(id=int(obra_id), ativa=True)
            if _user_can_access_obra(request, obra):
                request.session['obra_id'] = obra.id
        except (Obra.DoesNotExist, ValueError):
            obra_id = None
    if not obra_id:
        obra_sessao = get_obra_da_sessao(request)
        if obra_sessao:
            obra_id = str(obra_sessao.id)
    
    obra_selecionada = None
    local_selecionado_id = (request.GET.get('local') or '').strip()
    item_selecionado_id = request.GET.get('item', '')
    
    if obra_id:
        obra_selecionada = get_object_or_404(Obra, id=obra_id)
        # CORREÇÃO: Otimizar N+1 queries no Dashboard 2 também
        from django.db.models import Prefetch
        from suprimentos.models import AlocacaoRecebimento
        
        todos_itens = list(ItemMapa.objects.filter(
            obra_id=obra_id
        ).select_related('insumo', 'local_aplicacao', 'obra').prefetch_related(
            Prefetch(
                'alocacoes',
                queryset=AlocacaoRecebimento.objects.only('quantidade_alocada'),
                to_attr='alocacoes_cache'
            )
        ).annotate(
            quantidade_alocada_annotated=Sum('alocacoes__quantidade_alocada')
        ))
        
        hoje = datetime.now().date()
        itens_ativos = [i for i in todos_itens if not i.nao_aplica]
        
        # ====== NÍVEL 1: AGRUPAR POR LOCAL ======
        # Agrupar itens por local_aplicacao
        itens_por_local = defaultdict(list)
        itens_sem_local = []
        
        for item in itens_ativos:
            if item.local_aplicacao:
                itens_por_local[item.local_aplicacao].append(item)
            else:
                itens_sem_local.append(item)
        
        # Aplicar filtros e busca ANTES de agrupar por local (para filtrar na home)
        busca_query = request.GET.get('busca', '').strip()
        filtro_status = request.GET.get('status', '').strip()
        fornecedor_filtro_top = request.GET.get('fornecedor', '').strip()
        categoria_filtro_top = request.GET.get('categoria', '').strip()
        
        # Função auxiliar para verificar se um item passa nos filtros
        def item_passa_filtros(item):
            # Busca
            if busca_query:
                busca_lower = busca_query.lower()
                if not (busca_lower in (item.insumo.descricao or '').lower() or
                       busca_lower in (item.insumo.codigo_sienge or '').lower() or
                       busca_lower in (item.numero_sc or '').lower() or
                       busca_lower in (item.empresa_fornecedora or '').lower() or
                       busca_lower in (item.categoria or '').lower()):
                    return False
            
            # Filtro por fornecedor
            if fornecedor_filtro_top:
                if (item.empresa_fornecedora or '').strip() != fornecedor_filtro_top:
                    return False
            
            # Filtro por categoria
            if categoria_filtro_top:
                if (item.categoria or '').strip() != categoria_filtro_top:
                    return False
            
            # Filtro de status
            if filtro_status:
                if filtro_status == 'falta_comprar':
                    if not (item.numero_sc and item.quantidade_solicitada_sienge < item.quantidade_planejada):
                        return False
                elif filtro_status == 'a_caminho':
                    if not (item.numero_sc and item.quantidade_recebida_obra < item.quantidade_solicitada_sienge):
                        return False
                elif filtro_status == 'nao_alocado':
                    if not (item.quantidade_alocada_local < item.quantidade_recebida_obra and item.quantidade_recebida_obra > 0):
                        return False
                elif filtro_status == 'com_sc':
                    if not item.numero_sc:
                        return False
                elif filtro_status == 'sem_sc':
                    if item.numero_sc:
                        return False
                elif filtro_status == 'atrasados':
                    if not item.is_atrasado:
                        return False
                elif filtro_status == 'recebido':
                    if not ((item.quantidade_recebida_obra or 0) > 0):
                        return False
            
            return True
        
        # Filtrar itens antes de agrupar
        if busca_query or filtro_status or fornecedor_filtro_top or categoria_filtro_top:
            itens_ativos = [i for i in itens_ativos if item_passa_filtros(i)]
            # Reagrupar após filtros
            itens_por_local = defaultdict(list)
            itens_sem_local = []
            for item in itens_ativos:
                if item.local_aplicacao:
                    itens_por_local[item.local_aplicacao].append(item)
                else:
                    itens_sem_local.append(item)
        
        # Preparar dados dos locais
        locais_data = []
        for local, itens_local in sorted(itens_por_local.items(), key=lambda x: x[0].nome):
            # Calcular estatísticas do local
            total_planejado = sum((i.quantidade_planejada or Decimal('0.00')) for i in itens_local)
            total_recebido = sum((i.quantidade_recebida_obra or Decimal('0.00')) for i in itens_local)
            total_alocado = sum((i.quantidade_alocada_local or Decimal('0.00')) for i in itens_local)
            itens_atrasados_local = sum(1 for i in itens_local if i.is_atrasado)
            
            locais_data.append({
                'local_id': local.id,
                'local_nome': local.nome,
                'total_itens': len(itens_local),
                'total_planejado': float(total_planejado),
                'total_recebido': float(total_recebido),
                'total_alocado': float(total_alocado),
                'itens_atrasados': itens_atrasados_local,
                'itens': itens_local,
            })
        
        # Adicionar itens sem local
        if itens_sem_local:
            total_planejado_sl = sum((i.quantidade_planejada or Decimal('0.00')) for i in itens_sem_local)
            total_recebido_sl = sum((i.quantidade_recebida_obra or Decimal('0.00')) for i in itens_sem_local)
            total_alocado_sl = sum((i.quantidade_alocada_local or Decimal('0.00')) for i in itens_sem_local)
            itens_atrasados_sl = sum(1 for i in itens_sem_local if i.is_atrasado)
            
            locais_data.append({
                'local_id': None,
                'local_nome': 'Sem Local Definido',
                'total_itens': len(itens_sem_local),
                'total_planejado': float(total_planejado_sl),
                'total_recebido': float(total_recebido_sl),
                'total_alocado': float(total_alocado_sl),
                'itens_atrasados': itens_atrasados_sl,
                'itens': itens_sem_local,
            })
        
        # ====== KPIs GLOBAIS DA OBRA ======
        kpi_total_itens = len(itens_ativos)
        kpi_com_sc = sum(1 for i in itens_ativos if i.numero_sc)
        kpi_sem_sc = kpi_total_itens - kpi_com_sc
        kpi_atrasados = sum(1 for i in itens_ativos if i.is_atrasado)
        kpi_total_locais = len(locais_data)
        kpi_recebidos = sum(1 for i in itens_ativos if (i.quantidade_recebida_obra or 0) > 0)
        kpi_alocados = sum(1 for i in itens_ativos if (i.quantidade_alocada_local or 0) > 0)
        
        # Percentuais para progress bars
        kpi_pct_sc = round((kpi_com_sc / kpi_total_itens * 100) if kpi_total_itens > 0 else 0)
        kpi_pct_recebidos = round((kpi_recebidos / kpi_total_itens * 100) if kpi_total_itens > 0 else 0)
        kpi_pct_alocados = round((kpi_alocados / kpi_total_itens * 100) if kpi_total_itens > 0 else 0)
        
        # ====== NÍVEL 2: ITENS DO LOCAL SELECIONADO ======
        itens_local_selecionado = []
        local_selecionado_obj = None
        
        if local_selecionado_id:
            try:
                if local_selecionado_id == 'todos':
                    # Mostrar TODOS os itens da obra
                    itens_local_selecionado = itens_ativos
                    local_selecionado_obj = {'id': 'todos', 'nome': 'Todos os Locais'}
                elif local_selecionado_id == 'sem-local':
                    itens_local_selecionado = itens_sem_local
                    local_selecionado_obj = {'id': None, 'nome': 'Sem Local Definido'}
                else:
                    local_selecionado_obj = LocalObra.objects.get(id=local_selecionado_id, obra=obra_selecionada)
                    itens_local_selecionado = itens_por_local.get(local_selecionado_obj, [])
            except LocalObra.DoesNotExist:
                pass
        else:
            # PADRÃO: Mostrar TODOS os itens quando nenhum local é selecionado
            local_selecionado_id = 'todos'
            itens_local_selecionado = itens_ativos
            local_selecionado_obj = {'id': 'todos', 'nome': 'Todos os Locais'}
        
        # Preparar dados dos itens com matriz de 4 colunas técnicas
        # PRE-FETCH: Carregar TODAS as alocações da obra de uma vez (evita N+1 queries)
        todas_alocacoes = AlocacaoRecebimento.objects.filter(
            obra=obra_selecionada
        ).values('insumo_id', 'item_mapa__numero_sc').annotate(
            total=Sum('quantidade_alocada')
        )
        
        # Indexar alocações por (insumo_id, numero_sc) e por insumo_id (sem SC)
        alocacoes_por_insumo_sc = {}  # (insumo_id, sc) -> total
        alocacoes_por_insumo = {}     # insumo_id -> total (para itens sem SC)
        for aloc in todas_alocacoes:
            insumo_id = aloc['insumo_id']
            sc = aloc['item_mapa__numero_sc'] or ''
            total = aloc['total'] or Decimal('0.00')
            
            # Acumular por insumo (para itens sem SC)
            alocacoes_por_insumo[insumo_id] = alocacoes_por_insumo.get(insumo_id, Decimal('0.00')) + total
            
            # Acumular por (insumo, sc) para itens com SC
            if sc:
                key = (insumo_id, sc)
                alocacoes_por_insumo_sc[key] = alocacoes_por_insumo_sc.get(key, Decimal('0.00')) + total
        
        itens_pipeline = []
        for item in itens_local_selecionado:
            # COLUNA 1: O Plano - Quantidade levantada inicialmente
            qtd_planejada = float(item.quantidade_planejada or Decimal('0.00'))
            
            # COLUNA 2: O Pedido - SC + Quantidade comprada
            numero_sc = item.numero_sc or ''
            qtd_solicitada_sienge = float(item.quantidade_solicitada_sienge or Decimal('0.00'))
            tem_sc = bool(numero_sc)
            
            # COLUNA 3: O Pátio - Saldo total na obra (Sienge MAX)
            qtd_recebida_obra = float(item.quantidade_recebida_obra or Decimal('0.00'))
            
            # COLUNA 4: O Destino - Quantidade Alocada (usando dados pre-fetched)
            if numero_sc:
                total_alocado_insumo = alocacoes_por_insumo_sc.get(
                    (item.insumo_id, numero_sc), Decimal('0.00')
                )
            else:
                total_alocado_insumo = alocacoes_por_insumo.get(
                    item.insumo_id, Decimal('0.00')
                )
            qtd_alocada_total = float(total_alocado_insumo)
            
            # Alocação específica deste item (para mostrar no tooltip)
            qtd_alocada_item = float(item.quantidade_alocada_local or Decimal('0.00'))
            
            # Calcular gaps para alertas visuais
            gap_comprado = qtd_planejada - qtd_solicitada_sienge if tem_sc else qtd_planejada
            gap_recebido = qtd_solicitada_sienge - qtd_recebida_obra if tem_sc else 0.0
            gap_alocado = qtd_recebida_obra - qtd_alocada_total
            
            # Flags de alerta
            alerta_falta_comprar = tem_sc and qtd_solicitada_sienge < qtd_planejada
            alerta_a_caminho = tem_sc and qtd_recebida_obra < qtd_solicitada_sienge
            alerta_nao_alocado = qtd_alocada_total < qtd_recebida_obra and qtd_recebida_obra > 0
            
            # Fornecedor (se tiver)
            fornecedor = item.empresa_fornecedora or ''
            
            # Pipeline progress: 4 etapas (PLANO->PEDIDO->PÁTIO->DESTINO)
            # Cada etapa vale 25%; calculamos o progresso real
            pipeline_steps = 0
            if qtd_planejada > 0:
                pipeline_steps = 1  # Tem plano
                if tem_sc:
                    pipeline_steps = 2  # Tem pedido
                    if qtd_recebida_obra > 0:
                        pipeline_steps = 3  # Recebido no pátio
                        if qtd_alocada_total > 0:
                            pipeline_steps = 4  # Alocado
            pipeline_pct = round(pipeline_steps * 25)
            
            # Status textual para badge
            if pipeline_steps == 0:
                status_label = 'Sem Plano'
                status_class = 'status-empty'
            elif pipeline_steps == 1:
                status_label = 'Levantamento'
                status_class = 'status-plan'
            elif pipeline_steps == 2:
                status_label = 'Comprado'
                status_class = 'status-ordered'
            elif pipeline_steps == 3:
                status_label = 'No Patio'
                status_class = 'status-received'
            else:
                status_label = 'Alocado'
                status_class = 'status-allocated'
            
            itens_pipeline.append({
                'id': item.id,
                'insumo_id': item.insumo.id,
                'insumo_codigo': item.insumo.codigo_sienge or '',
                'insumo_descricao': item.insumo.descricao,
                'insumo_unidade': item.insumo.unidade,
                'categoria': item.categoria or 'A CLASSIFICAR',
                'local_nome': item.local_aplicacao.nome if item.local_aplicacao else 'Sem local',
                'fornecedor': fornecedor,
                # Matriz de 4 colunas
                'coluna_1_plano': qtd_planejada,
                'coluna_2_pedido_sc': numero_sc,
                'coluna_2_pedido_qtd': qtd_solicitada_sienge,
                'coluna_3_patio': qtd_recebida_obra,
                'coluna_4_destino_qtd': qtd_alocada_total,
                'coluna_4_destino_local': item.local_aplicacao.nome if item.local_aplicacao else 'Sem local',
                'coluna_4_destino_item_qtd': qtd_alocada_item,
                # Gaps
                'gap_comprado': gap_comprado,
                'gap_recebido': gap_recebido,
                'gap_alocado': gap_alocado,
                # Alertas
                'alerta_falta_comprar': alerta_falta_comprar,
                'alerta_a_caminho': alerta_a_caminho,
                'alerta_nao_alocado': alerta_nao_alocado,
                # Flag de atrasado
                'is_atrasado': item.is_atrasado,
                # Pipeline progress
                'pipeline_pct': pipeline_pct,
                'pipeline_steps': pipeline_steps,
                'status_label': status_label,
                'status_class': status_class,
            })
        
        # Buscar opções disponíveis de TODOS os itens da obra (não só do local)
        # Isso garante que os filtros sempre tenham opções, mesmo quando não há itens no local
        todos_itens_obra = ItemMapa.objects.filter(obra=obra_selecionada).exclude(nao_aplica=True)
        categorias_disponiveis = sorted(set(cat for cat in todos_itens_obra.values_list('categoria', flat=True) if cat))
        fornecedores_disponiveis = sorted(set(forn for forn in todos_itens_obra.values_list('empresa_fornecedora', flat=True) if forn))
        
        # Resumo por categoria (antes de filtrar): para "Onde estão os cimentos?" no mobile
        _cat = defaultdict(lambda: {'total': 0, 'no_patio': 0, 'a_caminho': 0, 'sem_sc': 0})
        for _item in itens_pipeline:
            c = (_item.get('categoria') or 'A CLASSIFICAR').strip() or 'A CLASSIFICAR'
            _cat[c]['total'] += 1
            q_patio = float(_item.get('coluna_3_patio') or 0)
            q_pedido = float(_item.get('coluna_2_pedido_qtd') or 0)
            tem_sc = bool(_item.get('coluna_2_pedido_sc'))
            if q_patio > 0:
                _cat[c]['no_patio'] += 1
            elif tem_sc and q_patio < q_pedido:
                _cat[c]['a_caminho'] += 1
            elif not tem_sc:
                _cat[c]['sem_sc'] += 1
        categorias_com_contagem = [
            {'nome': k, 'total': v['total'], 'no_patio': v['no_patio'], 'a_caminho': v['a_caminho'], 'sem_sc': v['sem_sc']}
            for k, v in sorted(_cat.items())
        ]
        
        # Filtro por status/alertas nos itens do pipeline
        # (filtro_status e busca_query já foram lidos acima e usados para filtrar itens_ativos;
        #  aqui só aplicamos nos dados formatados do pipeline)
        if filtro_status:
            if filtro_status == 'falta_comprar':
                itens_pipeline = [item for item in itens_pipeline if item.get('alerta_falta_comprar')]
            elif filtro_status == 'a_caminho':
                itens_pipeline = [item for item in itens_pipeline if item.get('alerta_a_caminho')]
            elif filtro_status == 'nao_alocado':
                itens_pipeline = [item for item in itens_pipeline if item.get('alerta_nao_alocado')]
            elif filtro_status == 'com_sc':
                itens_pipeline = [item for item in itens_pipeline if item.get('coluna_2_pedido_sc')]
            elif filtro_status == 'sem_sc':
                itens_pipeline = [item for item in itens_pipeline if not item.get('coluna_2_pedido_sc')]
            elif filtro_status == 'atrasados':
                itens_pipeline = [item for item in itens_pipeline if item.get('is_atrasado')]
            elif filtro_status == 'recebido':
                itens_pipeline = [item for item in itens_pipeline if item.get('coluna_3_patio', 0) > 0]
        
        # Filtro por categoria
        categoria_filtro = request.GET.get('categoria', '').strip()
        if categoria_filtro:
            itens_pipeline = [
                item for item in itens_pipeline
                if (item.get('categoria') or '').strip() == categoria_filtro
            ]
        
        # Filtro por fornecedor
        fornecedor_filtro = request.GET.get('fornecedor', '').strip()
        if fornecedor_filtro:
            itens_pipeline = [item for item in itens_pipeline if item.get('fornecedor') == fornecedor_filtro]
        
        # Totalizadores para o Modo Tabela (rodapé da tabela Nível 2)
        drill_totals = {'plano': 0.0, 'pedido': 0.0, 'patio': 0.0, 'destino': 0.0}
        for _item in itens_pipeline:
            drill_totals['plano'] += float(_item.get('coluna_1_plano') or 0)
            drill_totals['pedido'] += float(_item.get('coluna_2_pedido_qtd') or 0)
            drill_totals['patio'] += float(_item.get('coluna_3_patio') or 0)
            drill_totals['destino'] += float(_item.get('coluna_4_destino_qtd') or 0)
        
        # ====== NÍVEL 3: DETALHES DO ITEM PARA BOTTOM SHEET ======
        item_detalhes = None
        if item_selecionado_id:
            try:
                item_obj = ItemMapa.objects.get(id=item_selecionado_id, obra=obra_selecionada)
                
                # Buscar saldo máximo do Sienge (regra do MAX)
                saldo_maximo = Decimal('0.00')
                if item_obj.numero_sc:
                    recebimentos = RecebimentoObra.objects.filter(
                        obra=obra_selecionada,
                        numero_sc=item_obj.numero_sc,
                        insumo=item_obj.insumo
                    )
                    if recebimentos.exists():
                        saldo_maximo = max(
                            (r.quantidade_recebida or Decimal('0.00')) for r in recebimentos
                        )
                
                # Alocações existentes
                alocacoes = AlocacaoRecebimento.objects.filter(
                    item_mapa=item_obj
                ).select_related('item_mapa__local_aplicacao')
                
                total_alocado = sum((a.quantidade_alocada or Decimal('0.00')) for a in alocacoes)
                saldo_disponivel = saldo_maximo - total_alocado
                qtd_planejada = float(item_obj.quantidade_planejada or Decimal('0.00'))
                tem_sc = bool(item_obj.numero_sc)
                # Status no pipeline (igual à lista)
                pipeline_steps = 0
                if qtd_planejada > 0:
                    pipeline_steps = 1
                    if tem_sc:
                        pipeline_steps = 2
                        if saldo_maximo > 0:
                            pipeline_steps = 3
                            if total_alocado > 0:
                                pipeline_steps = 4
                pipeline_pct = pipeline_steps * 25
                if pipeline_steps == 0:
                    status_label, status_class = 'Sem Plano', 'status-empty'
                elif pipeline_steps == 1:
                    status_label, status_class = 'Levantamento', 'status-plan'
                elif pipeline_steps == 2:
                    status_label, status_class = 'Comprado', 'status-ordered'
                elif pipeline_steps == 3:
                    status_label, status_class = 'No Pátio', 'status-received'
                else:
                    status_label, status_class = 'Alocado', 'status-allocated'
                percentual_recebido = round((float(saldo_maximo) / qtd_planejada * 100) if qtd_planejada > 0 else 0)
                if pipeline_steps == 0:
                    proximo_passo = 'Planejar necessidade'
                elif not tem_sc:
                    proximo_passo = 'Emitir Solicitação de Compra (SC)'
                elif saldo_maximo <= 0:
                    proximo_passo = 'Aguardar recebimento no pátio'
                elif total_alocado <= 0:
                    proximo_passo = 'Alocar ao local (em outra tela)'
                else:
                    proximo_passo = 'Concluído'
                # Mesmo insumo em outros locais (visão consolidada)
                outros_locais = list(
                    ItemMapa.objects.filter(
                        obra=obra_selecionada,
                        insumo=item_obj.insumo
                    ).exclude(id=item_obj.id).select_related('local_aplicacao').values(
                        'id', 'local_aplicacao__nome', 'quantidade_planejada', 'numero_sc'
                    )[:10]
                )
                outros_locais = [
                    {
                        'local_nome': (x['local_aplicacao__nome'] or 'Sem local'),
                        'quantidade_planejada': float(x['quantidade_planejada'] or 0),
                        'tem_sc': bool(x['numero_sc']),
                        'item_id': x['id'],
                    }
                    for x in outros_locais
                ]
                # Outros itens no mesmo local (contexto do bloco)
                outros_itens_mesmo_local = []
                if item_obj.local_aplicacao_id:
                    outros_itens_mesmo_local = list(
                        ItemMapa.objects.filter(
                            obra=obra_selecionada,
                            local_aplicacao_id=item_obj.local_aplicacao_id
                        ).exclude(id=item_obj.id).select_related('insumo')[:5]
                    )
                    outros_itens_mesmo_local = [
                        {
                            'item_id': i.id,
                            'descricao': (i.insumo.descricao or '')[:50],
                            'codigo': (i.insumo.codigo_sienge or ''),
                            'quantidade_planejada': float(i.quantidade_planejada or 0),
                            'unidade': (i.insumo.unidade or ''),
                        }
                        for i in outros_itens_mesmo_local
                    ]
                item_detalhes = {
                    'item_id': item_obj.id,
                    'insumo_id': item_obj.insumo.id,
                    'insumo_codigo': item_obj.insumo.codigo_sienge or '',
                    'insumo_descricao': item_obj.insumo.descricao,
                    'insumo_unidade': item_obj.insumo.unidade,
                    'numero_sc': item_obj.numero_sc or '',
                    'local_nome': item_obj.local_aplicacao.nome if item_obj.local_aplicacao else 'Sem local',
                    'local_id': item_obj.local_aplicacao_id,
                    'categoria': item_obj.categoria or '',
                    'empresa_fornecedora': item_obj.empresa_fornecedora or '',
                    'saldo_maximo': float(saldo_maximo),
                    'total_alocado': float(total_alocado),
                    'saldo_disponivel': float(saldo_disponivel),
                    'quantidade_planejada': qtd_planejada,
                    'quantidade_recebida': float(item_obj.quantidade_recebida_obra or Decimal('0.00')),
                    'quantidade_solicitada': float(item_obj.quantidade_solicitada_sienge or Decimal('0.00')),
                    'status_label': status_label,
                    'status_class': status_class,
                    'pipeline_pct': pipeline_pct,
                    'percentual_recebido': percentual_recebido,
                    'proximo_passo': proximo_passo,
                    'is_atrasado': getattr(item_obj, 'is_atrasado', False),
                    'outros_locais': outros_locais,
                    'outros_itens_mesmo_local': outros_itens_mesmo_local,
                }
            except ItemMapa.DoesNotExist:
                pass
    
    # Buscar opções disponíveis (se não foi definido antes)
    if not obra_id:
        categorias_disponiveis = []
        categorias_com_contagem = []
        fornecedores_disponiveis = []
        categoria_filtro = ''
        fornecedor_filtro = ''
        filtro_status = ''
        busca_query = ''
        itens_pipeline = []
        item_detalhes = None
        local_selecionado_obj = None
        locais_data = []
        kpi_total_itens = 0
        kpi_com_sc = 0
        kpi_sem_sc = 0
        kpi_atrasados = 0
        kpi_total_locais = 0
        kpi_recebidos = 0
        kpi_alocados = 0
        kpi_pct_sc = 0
        kpi_pct_recebidos = 0
        kpi_pct_alocados = 0
        drill_totals = {'plano': 0.0, 'pedido': 0.0, 'patio': 0.0, 'destino': 0.0}
    elif obra_selecionada and not categorias_disponiveis:
        # Garantir que sempre temos listas, mesmo vazias
        todos_itens_obra = ItemMapa.objects.filter(obra=obra_selecionada).exclude(nao_aplica=True)
        categorias_disponiveis = sorted(set(cat for cat in todos_itens_obra.values_list('categoria', flat=True) if cat))
        fornecedores_disponiveis = sorted(set(forn for forn in todos_itens_obra.values_list('empresa_fornecedora', flat=True) if forn))
    
    # Serialização para lista virtual no mobile (evita 300+ cards no DOM)
    itens_pipeline_serializable = []
    if itens_pipeline:
        def _serialize(obj):
            if isinstance(obj, Decimal):
                return float(obj)
            if hasattr(obj, '__iter__') and not isinstance(obj, (str, dict)):
                return list(obj)
            return obj
        itens_pipeline_serializable = [{k: _serialize(v) for k, v in item.items()} for item in itens_pipeline]

    context = {
        'obras': obras,
        'obra_selecionada': obra_selecionada,
        'obra_id': obra_id,
        'locais_data': locais_data if obra_id else [],
        'local_selecionado_id': local_selecionado_id,
        'local_selecionado_obj': local_selecionado_obj,
        'itens_pipeline': itens_pipeline,
        'itens_pipeline_serializable': itens_pipeline_serializable,
        'item_selecionado_id': item_selecionado_id,
        'item_detalhes': item_detalhes,
        'busca_query': busca_query,
        'categoria_filtro': categoria_filtro if obra_id else '',
        'categorias_disponiveis': categorias_disponiveis if obra_id else [],
        'categorias_com_contagem': categorias_com_contagem if obra_id else [],
        'fornecedor_filtro': fornecedor_filtro if obra_id else '',
        'fornecedores_disponiveis': fornecedores_disponiveis if obra_id else [],
        'filtro_status': filtro_status if obra_id else '',
        # KPIs globais
        'kpi_total_itens': kpi_total_itens if obra_id else 0,
        'kpi_com_sc': kpi_com_sc if obra_id else 0,
        'kpi_sem_sc': kpi_sem_sc if obra_id else 0,
        'kpi_atrasados': kpi_atrasados if obra_id else 0,
        'kpi_total_locais': kpi_total_locais if obra_id else 0,
        'kpi_recebidos': kpi_recebidos if obra_id else 0,
        'kpi_alocados': kpi_alocados if obra_id else 0,
        'kpi_pct_sc': kpi_pct_sc if obra_id else 0,
        'kpi_pct_recebidos': kpi_pct_recebidos if obra_id else 0,
        'kpi_pct_alocados': kpi_pct_alocados if obra_id else 0,
        'drill_totals': drill_totals if obra_id else {'plano': 0.0, 'pedido': 0.0, 'patio': 0.0, 'destino': 0.0},
    }
    
    return render(request, 'suprimentos/dashboard_2.html', context)