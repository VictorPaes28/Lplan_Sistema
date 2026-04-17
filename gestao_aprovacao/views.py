from django.shortcuts import render, redirect, get_object_or_404
# auth import removido - autenticação centralizada no app 'accounts'
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import ensure_csrf_cookie
from django.contrib.auth.models import User, Group
from django.contrib import messages
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Q, Min, Value, Prefetch
from django.db.models.functions import Coalesce
from django.utils import timezone
from django.http import JsonResponse, HttpResponse
from django.conf import settings
from django.urls import reverse
from datetime import timedelta, datetime, time
from urllib.parse import urlencode
import os
import csv
import io
import logging
import uuid
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, LongTable
from xml.sax.saxutils import escape as xml_escape
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from .models import (
    Empresa, Obra, WorkOrder, Approval, Attachment, StatusHistory,
    WorkOrderPermission, UserEmpresa, UserProfile, Notificacao, Comment, Lembrete, TagErro, EmailLog,
    AprovacaoEmailDestinatario,
)
from .forms import EmpresaForm, ObraForm, WorkOrderForm, AttachmentForm, AprovacaoEmailDestinatarioForm
from .utils import (
    get_user_profile,
    is_engenheiro,
    is_gestor,
    is_admin,
    is_responsavel_empresa,
    is_aprovador,
    gestor_required,
    criar_notificacao,
    admin_required,
    usuario_pode_marcar_pedido_analisado,
)
from .email_utils import enviar_email_novo_pedido, enviar_email_aprovacao, enviar_email_reprovacao, enviar_email_credenciais_novo_usuario
from .services.user_governance import (
    TimelineOptions,
    build_audit_insights,
    build_critical_highlights,
    build_kpis,
    build_pending_queues,
    build_scope,
    build_strategic_insights,
    build_timeline_events,
    operational_alerts,
    usage_by_obra_gestao,
    viewer_can_see_target_user,
)
from accounts.groups import (
    GRUPOS,
    filtrar_grupos_post_atribuivel,
    grupos_modulos_para_atribuicao,
    grupos_ordenados_atribuivel,
    merge_grupos_legados_ocultos,
)
from accounts.models import UserSignupRequest
from accounts.painel_sistema_access import user_can_view_audit_events
from accounts.signup_services import (
    create_signup_request,
    is_allowed_signup_email,
    notify_signup_request_created,
)
from core.user_messages import flash_message


# --- Proteção contra duplo envio (double-submit) via token de idempotência ---
_FORM_TOKEN_SESSION_PREFIX = '_form_submit_token_'


def _generate_form_token(request, form_name: str) -> str:
    """Gera e guarda na sessão um token UUID único para o formulário dado.
    Use na view GET para injetar o token no contexto e depois no template como campo oculto."""
    token = str(uuid.uuid4())
    request.session[_FORM_TOKEN_SESSION_PREFIX + form_name] = token
    return token


def _consume_form_token(request, form_name: str, posted_token: str) -> bool:
    """Valida e consome (remove) o token de sessão.
    Retorna True se o token é válido (primeira submissão).
    Retorna False se o token já foi usado ou não existe (duplo envio)."""
    session_key = _FORM_TOKEN_SESSION_PREFIX + form_name
    expected = request.session.get(session_key)
    if not expected or not posted_token or expected != posted_token:
        return False
    # Consome o token — uma única vez
    del request.session[session_key]
    return True


# --- Evitar que o botão "Voltar" do navegador retorne a formulários já submetidos ---
def _redirect_if_back_after_post(request, session_key, default_redirect_to, redirect_args=None):
    """
    Se o usuário voltou (Back) para um formulário já submetido, redireciona para evitar reenvio/duplicação.
    default_redirect_to: nome da URL ou fallback se na sessão estiver True.
    Na sessão pode estar True ou um nome de URL (para redirect dinâmico, ex. central vs gestao).
    """
    val = request.session.pop(session_key, None)
    if val is None:
        return None
    to = val if isinstance(val, str) else default_redirect_to
    if redirect_args:
        return redirect(to, **redirect_args)
    return redirect(to)


def _mark_post_success_redirect(request, session_key, redirect_to=None):
    """
    Marca na sessão que o próximo GET neste formulário deve redirecionar (evitar Back).
    redirect_to: se informado, guarda esse nome de URL na sessão; senão guarda True.
    """
    request.session[session_key] = redirect_to if redirect_to is not None else True


def _no_cache_form_response(response):
    """Adiciona headers para não cachear a página; ao clicar Voltar o navegador refaz o GET e cai no redirect."""
    response['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    return response
from core.models import Project, ProjectMember

logger = logging.getLogger(__name__)

# --- Regras de acesso GestControll (simplificado) ---
# - Acesso é por OBRA: WorkOrderPermission (solicitante/aprovador) por obra.
# - Admin (staff) vê tudo. Aprovador vê pedidos das obras em que é aprovador.
# - Responsável por empresa vê pedidos das obras cujas empresas ele responde.
# - Solicitante vê apenas pedidos das obras em que está vinculado (sem fallback por UserEmpresa).
# - Criar/Editar usuário: única seção "Obras no GestControll"; empresas são derivadas das obras.

# Extensões permitidas para anexos (mesmas do AttachmentForm)
EXTENSOES_PERMITIDAS = [
    '.pdf', '.doc', '.docx', '.xls', '.xlsx',
    '.jpg', '.jpeg', '.png', '.gif',
    '.zip', '.rar', '.7z'
]


def validar_extensao_arquivo(nome_arquivo):
    """
    Valida se a extensão do arquivo está na lista de permitidas.
    Retorna True se válido, False caso contrário.
    """
    nome_arquivo = nome_arquivo.lower()
    return any(nome_arquivo.endswith(ext) for ext in EXTENSOES_PERMITIDAS)


def _get_approvers_for_exclusion(workorder):
    """
    Retorna usuários que efetivamente podem aprovar/rejeitar exclusão
    para o pedido informado, seguindo a mesma regra das views de aprovação.
    """
    if workorder.obra.empresa_id is None:
        approvers = User.objects.filter(
            permissoes_obra__obra=workorder.obra,
            permissoes_obra__tipo_permissao='aprovador',
            permissoes_obra__ativo=True,
        )
    else:
        approvers = User.objects.filter(
            permissoes_obra__obra__empresa_id=workorder.obra.empresa_id,
            permissoes_obra__tipo_permissao='aprovador',
            permissoes_obra__ativo=True,
        )

    admins = User.objects.filter(
        Q(is_superuser=True) | Q(groups__name='Administrador')
    )

    return (approvers | admins).distinct()


@login_required
def home(request):
    """View da home - mostra informações do sistema."""
    pode_criar_pedido = False
    ultimas_atualizacoes = []
    
    if request.user.is_authenticated:
        user = request.user
        user_profile = get_user_profile(user)
        
        # Usuário pode criar se:
        # 1. Está no grupo "Solicitante" OU
        # 2. Tem WorkOrderPermission de solicitante OU
        # 3. É admin
        pode_criar_pedido = (
            is_engenheiro(user) or  # Grupo "Solicitante"
            WorkOrderPermission.objects.filter(
                usuario=user,
                tipo_permissao='solicitante',
                ativo=True
            ).exists() or
            is_admin(user)
        )
        
        # Buscar últimas atualizações de pedidos relevantes para o usuário
        if is_admin(user):
            # Admins veem todos os pedidos
            workorders = WorkOrder.objects.select_related('obra', 'obra__empresa', 'criado_por').all()
        elif is_aprovador(user):
            # Aprovadores veem pedidos das empresas onde têm permissão + obras sem empresa onde têm permissão
            obras_ids = WorkOrderPermission.objects.filter(
                usuario=user,
                tipo_permissao='aprovador',
                ativo=True
            ).values_list('obra_id', flat=True).distinct()
            empresas_ids = [e for e in Obra.objects.filter(id__in=obras_ids).values_list('empresa_id', flat=True).distinct() if e is not None]
            obras_sem_empresa_ids = list(Obra.objects.filter(id__in=obras_ids, empresa_id__isnull=True).values_list('id', flat=True))
            workorders = WorkOrder.objects.filter(
                Q(obra__empresa_id__in=empresas_ids) | Q(obra_id__in=obras_sem_empresa_ids)
            ).select_related('obra', 'obra__empresa', 'criado_por')
        elif is_responsavel_empresa(user):
            # Responsável por empresa: vê pedidos das obras das empresas que gerencia
            empresas_resp = Empresa.objects.filter(responsavel=user, ativo=True)
            workorders = WorkOrder.objects.filter(
                obra__empresa__in=empresas_resp
            ).select_related('obra', 'obra__empresa', 'criado_por')
        elif is_engenheiro(user):
            # Solicitantes: só obras em que foram vinculados (WorkOrderPermission)
            obras_ids = WorkOrderPermission.objects.filter(
                usuario=user,
                tipo_permissao='solicitante',
                ativo=True
            ).values_list('obra_id', flat=True).distinct()
            if obras_ids:
                workorders = WorkOrder.objects.filter(
                    Q(criado_por=user) | Q(obra_id__in=obras_ids)
                ).select_related('obra', 'obra__empresa', 'criado_por').distinct()
            else:
                workorders = WorkOrder.objects.filter(criado_por=user).select_related('obra', 'obra__empresa', 'criado_por')
        else:
            workorders = WorkOrder.objects.filter(criado_por=user).select_related('obra', 'obra__empresa', 'criado_por')
        
        # Buscar os 5 pedidos mais recentemente atualizados (apenas dos últimos 7 dias)
        # Isso evita mostrar atualizações muito antigas que não são mais relevantes
        agora = timezone.now()
        data_limite = agora - timedelta(days=7)
        
        ultimas_atualizacoes = workorders.filter(
            updated_at__gte=data_limite
        ).order_by('-updated_at')[:5]
    
    context = {
        'title': 'GestControll',
        'user': request.user,
        'user_profile': get_user_profile(request.user) if request.user.is_authenticated else None,
        'pode_criar_pedido': pode_criar_pedido,
        'ultimas_atualizacoes': ultimas_atualizacoes,
        'is_admin': is_admin(request.user) if request.user.is_authenticated else False,
        'is_responsavel_empresa': is_responsavel_empresa(request.user) if request.user.is_authenticated else False,
        'pode_ver_desempenho': (is_admin(request.user) or is_responsavel_empresa(request.user)) if request.user.is_authenticated else False,
        'pode_ver_auditoria': user_can_view_audit_events(request.user) if request.user.is_authenticated else False,
    }
    return render(request, 'obras/home.html', context)


# ========== CRUD WorkOrder ==========
# Autenticação (login/logout) centralizada no app 'accounts'.

def _workorders_base_queryset_and_obras(user):
    """
    Retorna (queryset base, obras_disponiveis) conforme permissão multi-tenant.
    Mesma regra de list_workorders / exportação.
    """
    if is_admin(user):
        workorders = WorkOrder.objects.select_related('obra', 'obra__empresa', 'criado_por').all()
        obras_disponiveis = Obra.objects.filter(ativo=True)
    elif is_aprovador(user):
        obras_com_permissao = Obra.objects.filter(
            permissoes__usuario=user,
            permissoes__tipo_permissao='aprovador',
            permissoes__ativo=True,
            ativo=True
        ).distinct()
        empresas_ids = [e for e in obras_com_permissao.values_list('empresa_id', flat=True).distinct() if e is not None]
        obras_com_empresa = Obra.objects.filter(empresa_id__in=empresas_ids, ativo=True).distinct()
        obras_sem_empresa = obras_com_permissao.filter(empresa_id__isnull=True)
        obras_aprovador = (obras_com_empresa | obras_sem_empresa).distinct()
        workorders = WorkOrder.objects.filter(obra__in=obras_aprovador).select_related('obra', 'obra__empresa', 'criado_por')
        obras_disponiveis = obras_aprovador
    elif is_responsavel_empresa(user):
        empresas_resp = Empresa.objects.filter(responsavel=user, ativo=True)
        obras_disponiveis = Obra.objects.filter(empresa__in=empresas_resp, ativo=True).distinct()
        workorders = WorkOrder.objects.filter(obra__in=obras_disponiveis).select_related('obra', 'obra__empresa', 'criado_por')
    else:
        obras_solicitante = Obra.objects.filter(
            permissoes__usuario=user,
            permissoes__tipo_permissao='solicitante',
            permissoes__ativo=True,
            ativo=True
        ).distinct()
        if obras_solicitante.exists():
            outros_solicitantes_ids = WorkOrderPermission.objects.filter(
                obra__in=obras_solicitante,
                tipo_permissao='solicitante',
                ativo=True
            ).values_list('usuario_id', flat=True).distinct()
            workorders = WorkOrder.objects.filter(obra__in=obras_solicitante).filter(
                Q(criado_por=user) | Q(criado_por_id__in=outros_solicitantes_ids)
            ).select_related('obra', 'obra__empresa', 'criado_por')
            obras_disponiveis = obras_solicitante
        else:
            workorders = WorkOrder.objects.filter(criado_por=user).select_related('obra', 'obra__empresa', 'criado_por')
            obras_disponiveis = Obra.objects.filter(
                id__in=workorders.values_list('obra_id', flat=True).distinct(),
                ativo=True
            )
    return workorders, obras_disponiveis


def _parse_workorder_list_date(s):
    """Aceita YYYY-MM-DD (input date), DD/MM/YYYY ou DD-MM-YYYY."""
    s = (s or '').strip()
    if not s:
        return None
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _filter_workorders_by_envio_or_created_date(workorders, start_date, end_date):
    """
    Data de referência = data_envio quando preenchida; senão created_at.
    Equivalente ao Coalesce(data_envio, created_at) em filtros por dia.
    Evita lookups __date para não depender de tabelas de timezone no MySQL.
    """
    eff_end = end_date if end_date is not None else timezone.localdate()
    if start_date and start_date > eff_end:
        start_date, eff_end = eff_end, start_date

    current_tz = timezone.get_current_timezone()
    start_dt = None
    if start_date:
        start_dt = timezone.make_aware(datetime.combine(start_date, time.min), current_tz)
    end_exclusive = timezone.make_aware(
        datetime.combine(eff_end + timedelta(days=1), time.min),
        current_tz,
    )

    q_branch1 = Q(data_envio__isnull=False)
    if start_dt:
        q_branch1 &= Q(data_envio__gte=start_dt)
    q_branch1 &= Q(data_envio__lt=end_exclusive)

    q_branch2 = Q(data_envio__isnull=True)
    if start_dt:
        q_branch2 &= Q(created_at__gte=start_dt)
    q_branch2 &= Q(created_at__lt=end_exclusive)

    return workorders.filter(q_branch1 | q_branch2)


def _apply_workorder_list_filters(user, workorders, obras_disponiveis, get_params):
    """
    Aplica filtros GET (ORM apenas). Valida obra contra obras_disponiveis.
    Período (De/Até): usa data de envio; se não houver, a data de criação do pedido.
    Se data_fim vazia e houver filtro de datas, fim = hoje local.
    Se ambos os campos de data vazios, não filtra por data.
    """
    obra_filter = get_params.get('obra')
    obra_filter_out = ''
    if obra_filter:
        try:
            obra_id = int(obra_filter)
        except (ValueError, TypeError):
            obra_filter_out = ''
        else:
            allowed = set(obras_disponiveis.values_list('id', flat=True))
            if obra_id in allowed:
                workorders = workorders.filter(obra_id=obra_id)
                obra_filter_out = str(obra_id)
            else:
                workorders = workorders.none()
                obra_filter_out = ''

    status_filter = get_params.get('status')
    if status_filter:
        workorders = workorders.filter(status=status_filter)

    tipo_solicitacao_filter = get_params.get('tipo_solicitacao')
    if tipo_solicitacao_filter:
        workorders = workorders.filter(tipo_solicitacao=tipo_solicitacao_filter)

    credor_filter = (get_params.get('credor') or '').strip()
    if credor_filter:
        workorders = workorders.filter(nome_credor__icontains=credor_filter)

    engenheiro_filter = get_params.get('engenheiro')
    if engenheiro_filter and (is_aprovador(user) or is_admin(user)):
        workorders = workorders.filter(criado_por_id=engenheiro_filter)

    data_inicio_raw = (get_params.get('data_inicio') or '').strip()
    data_fim_raw = (get_params.get('data_fim') or '').strip()

    start_date = _parse_workorder_list_date(data_inicio_raw) if data_inicio_raw else None
    end_date = _parse_workorder_list_date(data_fim_raw) if data_fim_raw else None
    if data_inicio_raw and start_date is None:
        data_inicio_raw = ''
    if data_fim_raw and end_date is None:
        data_fim_raw = ''
    if start_date and end_date and start_date > end_date:
        start_date, end_date = end_date, start_date

    data_inicio = start_date.strftime('%Y-%m-%d') if start_date else ''
    data_fim = end_date.strftime('%Y-%m-%d') if end_date else ''

    if data_inicio_raw or data_fim_raw:
        workorders = _filter_workorders_by_envio_or_created_date(
            workorders, start_date, end_date
        )

    search_query = (get_params.get('search') or '').strip()
    if search_query:
        workorders = workorders.filter(
            Q(codigo__icontains=search_query) |
            Q(nome_credor__icontains=search_query) |
            Q(observacoes__icontains=search_query)
        )

    analisado_filter = get_params.get('analisado')
    if analisado_filter == 'sim':
        workorders = workorders.filter(marcado_para_deletar=True)
    elif analisado_filter == 'nao':
        workorders = workorders.filter(marcado_para_deletar=False)

    order_by = get_params.get('order_by', '-created_at')
    workorders = workorders.order_by(order_by)

    return workorders, {
        'obra_filter': obra_filter_out,
        'status_filter': status_filter or '',
        'tipo_solicitacao_filter': tipo_solicitacao_filter or '',
        'credor_filter': credor_filter,
        'engenheiro_filter': engenheiro_filter if engenheiro_filter and (is_aprovador(user) or is_admin(user)) else '',
        'analisado_filter': analisado_filter,
        'data_inicio': data_inicio,
        'data_fim': data_fim,
        'search_query': search_query,
        'order_by': order_by,
    }


@login_required
@ensure_csrf_cookie
def list_workorders(request):
    """
    Lista pedidos de obra. Regra única no GestControll:
    - Admin: vê todos.
    - Aprovador / Responsável: vê por permissão por obra ou por empresas que gerencia.
    - Solicitante: vê só as obras em que foi vinculado (WorkOrderPermission). Se não tiver nenhuma obra, vê só os pedidos que ele criou.
    """
    user = request.user
    user_profile = get_user_profile(user)

    workorders, obras_disponiveis = _workorders_base_queryset_and_obras(user)
    workorders, filter_ctx = _apply_workorder_list_filters(user, workorders, obras_disponiveis, request.GET)
    
    # Paginação
    paginator = Paginator(workorders, 15)  # 15 por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Lista de solicitantes para filtro (apenas para gestores/admins)
    engenheiros_list = None
    if is_aprovador(user) or is_admin(user):
        engenheiros_list = user.__class__.objects.filter(
            groups__name='Solicitante'
        ).distinct().order_by('username')
    
    # Verificar se o usuário pode criar pedidos (apenas solicitantes)
    pode_criar_pedido = (
        is_engenheiro(user) or  # Grupo "Solicitante"
        WorkOrderPermission.objects.filter(
            usuario=user,
            tipo_permissao='solicitante',
            ativo=True
        ).exists() or
        is_admin(user)
    )
    
    # Aviso para solicitante sem obras vinculadas (regra única: acesso é por obra)
    aviso_sem_obras = (
        user_profile == 'solicitante' and not is_admin(user) and not is_aprovador(user)
        and not obras_disponiveis.exists()
    )
    _gestao_marcar_pk_ph = 999001999
    context = {
        'page_obj': page_obj,
        'workorders': page_obj,
        'user_profile': user_profile,
        'obras_disponiveis': obras_disponiveis,
        'engenheiros_list': engenheiros_list,
        'pode_criar_pedido': pode_criar_pedido,
        'aviso_sem_obras': aviso_sem_obras,
        'status_choices': [
            ('pendente', 'Pendente Aprovação'),
            ('aprovado', 'Aprovado'),
            ('reprovado', 'Reprovado'),
            ('reaprovacao', 'Reaprovação'),
        ],
        'tipo_solicitacao_choices': WorkOrder.TIPO_SOLICITACAO_CHOICES,
        # Front (JS): URLs resolvidas pelo Django — evita path fixo errado em subpath/proxy
        'gestao_csrf_api_url': reverse('api_csrf_token'),
        'gestao_marcar_analisado_url_tpl': reverse(
            'gestao:marcar_pedido_analisado',
            args=[_gestao_marcar_pk_ph],
        ),
        'gestao_marcar_url_pk_placeholder': str(_gestao_marcar_pk_ph),
        **filter_ctx,
    }
    return render(request, 'obras/list_workorders.html', context)


def _pdf_esc(text):
    return xml_escape(str(text if text is not None else ''), {'"': '&quot;', "'": '&#39;'})


def _valor_medicao_relatorio(wo):
    """Texto para PDF/lista; vazio quando não é medição ou sem valor."""
    if getattr(wo, 'tipo_solicitacao', None) != 'medicao' or wo.valor_medicao is None:
        return '—'
    v = wo.valor_medicao
    # Formato legível em pt-BR (milhar . decimal ,)
    s = f'{v:.2f}'
    if '.' in s:
        inteiro, dec = s.split('.', 1)
    else:
        inteiro, dec = s, '00'
    neg = inteiro.startswith('-')
    if neg:
        inteiro = inteiro[1:]
    try:
        n = int(inteiro)
    except ValueError:
        return f'R$ {v}'
    inteiro_fmt = f'{n:,}'.replace(',', '.')
    if neg:
        inteiro_fmt = '-' + inteiro_fmt
    return f'R$ {inteiro_fmt},{dec}'


def _motivo_reprovacao_listagem(wo):
    """
    Texto do motivo da última reprovação (tags de erro + comentário), para PDF/Excel.
    Só preenche quando o pedido está reprovado e há registro de Approval correspondente.
    """
    if getattr(wo, 'status', None) != 'reprovado':
        return '—'
    rel = getattr(wo, 'prefetched_reprovacoes', None) or []
    if not rel:
        return '—'
    apr = rel[0]
    parts = []
    tags = list(apr.tags_erro.all())
    if tags:
        parts.append(', '.join(t.nome for t in tags))
    c = (apr.comentario or '').strip()
    if c:
        parts.append(c)
    return ' | '.join(parts) if parts else '—'


@login_required
def export_list_workorders_pdf(request):
    """
    Exporta PDF da listagem de pedidos com os mesmos filtros da tela (GET).
    Limite de linhas para evitar PDF excessivo.
    """
    user = request.user
    workorders, obras_disponiveis = _workorders_base_queryset_and_obras(user)
    workorders, filter_ctx = _apply_workorder_list_filters(user, workorders, obras_disponiveis, request.GET)
    workorders = workorders.select_related('obra', 'criado_por').prefetch_related(
        Prefetch(
            'approvals',
            queryset=Approval.objects.filter(decisao='reprovado')
            .prefetch_related('tags_erro')
            .order_by('-created_at'),
            to_attr='prefetched_reprovacoes',
        ),
    )

    MAX_ROWS = 3000
    total_count = workorders.count()
    items = list(workorders[:MAX_ROWS])

    color_primary = colors.HexColor('#1A3A5C')
    color_accent = colors.HexColor('#E8F0F8')
    color_border = colors.HexColor('#D0D9E3')
    color_text = colors.HexColor('#1C1C1C')

    def _logo_path():
        logo_dir = os.path.join(settings.BASE_DIR, 'core', 'static', 'core', 'images')
        for name in (
            'lpla-logo-pdf-transparent.png', 'lpla-logo-pdf.png', 'lplan-logo2.png',
            'lplan_logo.png', 'lplan_logo.jpg', 'lplan_logo.jpeg',
        ):
            p = os.path.join(logo_dir, name)
            if os.path.exists(p):
                return p
        return None

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=1.7 * cm,
        leftMargin=1.7 * cm,
        topMargin=1.2 * cm,
        bottomMargin=1.2 * cm,
    )
    styles = getSampleStyleSheet()
    normal = ParagraphStyle('ListPdfNormal', parent=styles['Normal'], fontName='Helvetica', fontSize=7.5, textColor=color_text, leading=9)
    th = ParagraphStyle('ListPdfTH', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=7.5, textColor=colors.white, alignment=TA_CENTER)
    title_style = ParagraphStyle('ListPdfTitle', parent=styles['Heading1'], fontName='Helvetica-Bold', fontSize=13, textColor=color_primary, alignment=TA_CENTER)
    sub_style = ParagraphStyle('ListPdfSub', parent=styles['Normal'], fontName='Helvetica', fontSize=8, textColor=color_primary, alignment=TA_CENTER)

    story = []
    title_main = Paragraph("<font color='#1A3A5C'><b>RELATÓRIO DE PEDIDOS DE OBRA</b></font>", title_style)
    filtros_bits = [
        f"Gerado em {timezone.now().strftime('%d/%m/%Y %H:%M')}",
        f"Registros: {total_count}" + (f" (exportando até {MAX_ROWS})" if total_count > MAX_ROWS else ''),
    ]
    if filter_ctx.get('search_query'):
        filtros_bits.append(f"Busca: {_pdf_esc(filter_ctx['search_query'])}")
    if filter_ctx.get('obra_filter'):
        filtros_bits.append(f"Obra ID: {_pdf_esc(filter_ctx['obra_filter'])}")
    if filter_ctx.get('status_filter'):
        filtros_bits.append(f"Status: {_pdf_esc(filter_ctx['status_filter'])}")
    if filter_ctx.get('tipo_solicitacao_filter'):
        filtros_bits.append(f"Tipo: {_pdf_esc(filter_ctx['tipo_solicitacao_filter'])}")
    if filter_ctx.get('engenheiro_filter'):
        filtros_bits.append(f"Solicitante ID: {_pdf_esc(filter_ctx['engenheiro_filter'])}")
    if filter_ctx.get('credor_filter'):
        filtros_bits.append(f"Credor: {_pdf_esc(filter_ctx['credor_filter'])}")
    if filter_ctx.get('data_inicio') or filter_ctx.get('data_fim'):
        filtros_bits.append(
            f"Período (envio ou criação): {_pdf_esc(filter_ctx.get('data_inicio') or '—')} a {_pdf_esc(filter_ctx.get('data_fim') or timezone.localdate().strftime('%Y-%m-%d'))}"
        )
    title_sub = Paragraph(
        "<font size='8' color='#1A3A5C'>GestControll · " + " · ".join(filtros_bits) + "</font>",
        sub_style,
    )

    lp = _logo_path()
    if lp:
        from reportlab.platypus import Image as RLImage
        logo = RLImage(lp, width=4.8 * cm, height=1.15 * cm)
        text_block = Table([[title_main], [title_sub]], colWidths=[11.2 * cm])
        text_block.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        hdr = Table([[logo, text_block]], colWidths=[4.8 * cm, 11.2 * cm])
    else:
        hdr = Table([[title_main], [title_sub]], colWidths=[16 * cm])
    hdr.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.white),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LINEBELOW', (0, -1), (-1, -1), 1.0, color_primary),
    ]))
    story.append(hdr)
    story.append(Spacer(1, 0.35 * cm))

    tipo_dict = dict(WorkOrder.TIPO_SOLICITACAO_CHOICES)
    status_dict = dict(WorkOrder.STATUS_CHOICES)

    data_rows = [[
        Paragraph('Código', th),
        Paragraph('Obra', th),
        Paragraph('Credor', th),
        Paragraph('Tipo', th),
        Paragraph('Valor medição', th),
        Paragraph('Status', th),
        Paragraph('Motivo reprovação', th),
        Paragraph('Solicitante', th),
        Paragraph('Data envio', th),
        Paragraph('Data aprovação', th),
    ]]
    for wo in items:
        sol = wo.criado_por.get_full_name() or wo.criado_por.username if wo.criado_por else '—'
        obra_txt = f"{wo.obra.codigo}" if wo.obra else '—'
        motivo_repr = _motivo_reprovacao_listagem(wo)
        vmed = _valor_medicao_relatorio(wo)
        data_rows.append([
            Paragraph(_pdf_esc(wo.codigo), normal),
            Paragraph(_pdf_esc(obra_txt), normal),
            Paragraph(_pdf_esc((wo.nome_credor or '')[:32]), normal),
            Paragraph(_pdf_esc(tipo_dict.get(wo.tipo_solicitacao, wo.tipo_solicitacao or '')), normal),
            Paragraph(_pdf_esc(vmed), normal),
            Paragraph(_pdf_esc(status_dict.get(wo.status, wo.status or '')), normal),
            Paragraph(_pdf_esc(motivo_repr[:500]), normal),
            Paragraph(_pdf_esc(sol[:22]), normal),
            Paragraph(_pdf_esc(wo.data_envio.strftime('%d/%m/%Y %H:%M') if wo.data_envio else '—'), normal),
            Paragraph(_pdf_esc(wo.data_aprovacao.strftime('%d/%m/%Y %H:%M') if wo.data_aprovacao else '—'), normal),
        ])

    tbl = LongTable(
        data_rows,
        colWidths=[
            1.65 * cm, 1.85 * cm, 2.1 * cm, 1.55 * cm, 1.55 * cm,
            1.45 * cm, 2.45 * cm, 1.65 * cm, 1.45 * cm, 1.45 * cm,
        ],
        repeatRows=1,
    )
    tbl.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), color_primary),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.25, color_border),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, color_accent]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('FONTSIZE', (0, 0), (-1, -1), 7.5),
    ]))
    story.append(tbl)
    if total_count > MAX_ROWS:
        story.append(Spacer(1, 0.2 * cm))
        story.append(Paragraph(
            f"<i>Lista truncada: {total_count} pedidos no filtro; exportados os primeiros {MAX_ROWS}.</i>",
            ParagraphStyle('Trunc', parent=normal, fontSize=8, textColor=color_text),
        ))

    doc.build(story)
    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    fname = f"relatorio_pedidos_{timezone.now().strftime('%Y%m%d_%H%M')}.pdf"
    disp = 'inline' if request.GET.get('inline') == '1' else 'attachment'
    response['Content-Disposition'] = f'{disp}; filename="{fname}"'
    return response


@login_required
def export_list_workorders_excel(request):
    """
    Exporta Excel (.xlsx) da listagem de pedidos com os mesmos filtros da tela (GET).
    Limite de linhas para evitar ficheiros excessivos.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        messages.error(
            request,
            'Exportação Excel não está disponível neste servidor (dependência openpyxl).',
        )
        return redirect('gestao:list_workorders')

    user = request.user
    workorders, obras_disponiveis = _workorders_base_queryset_and_obras(user)
    workorders, filter_ctx = _apply_workorder_list_filters(user, workorders, obras_disponiveis, request.GET)
    workorders = workorders.select_related('obra', 'criado_por').prefetch_related(
        Prefetch(
            'approvals',
            queryset=Approval.objects.filter(decisao='reprovado')
            .prefetch_related('tags_erro')
            .order_by('-created_at'),
            to_attr='prefetched_reprovacoes',
        ),
    )

    MAX_ROWS = 10000
    total_count = workorders.count()
    items = list(workorders[:MAX_ROWS])

    tipo_dict = dict(WorkOrder.TIPO_SOLICITACAO_CHOICES)
    status_dict = dict(WorkOrder.STATUS_CHOICES)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Pedidos'

    title_font = Font(bold=True, size=14, color='1A3A5C')
    meta_font = Font(size=9, color='1A3A5C')
    header_fill = PatternFill(start_color='1A3A5C', end_color='1A3A5C', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    cell_align = Alignment(vertical='center', wrap_text=True)

    ws['A1'] = 'RELATÓRIO DE PEDIDOS DE OBRA — GestControll'
    ws['A1'].font = title_font
    ws.merge_cells('A1:J1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    filtros_bits = [
        f"Gerado em {timezone.now().strftime('%d/%m/%Y %H:%M')}",
        f"Registros: {total_count}" + (f" (exportando até {MAX_ROWS})" if total_count > MAX_ROWS else ''),
    ]
    if filter_ctx.get('search_query'):
        filtros_bits.append(f"Busca: {filter_ctx['search_query']}")
    if filter_ctx.get('obra_filter'):
        filtros_bits.append(f"Obra ID: {filter_ctx['obra_filter']}")
    if filter_ctx.get('status_filter'):
        filtros_bits.append(f"Status: {filter_ctx['status_filter']}")
    if filter_ctx.get('tipo_solicitacao_filter'):
        filtros_bits.append(f"Tipo: {filter_ctx['tipo_solicitacao_filter']}")
    if filter_ctx.get('engenheiro_filter'):
        filtros_bits.append(f"Solicitante ID: {filter_ctx['engenheiro_filter']}")
    if filter_ctx.get('credor_filter'):
        filtros_bits.append(f"Credor: {filter_ctx['credor_filter']}")
    if filter_ctx.get('data_inicio') or filter_ctx.get('data_fim'):
        filtros_bits.append(
            f"Período: {filter_ctx.get('data_inicio') or '—'} a "
            f"{filter_ctx.get('data_fim') or timezone.localdate().strftime('%Y-%m-%d')}"
        )

    ws['A2'] = ' · '.join(filtros_bits)
    ws['A2'].font = meta_font
    ws.merge_cells('A2:J2')
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    headers = [
        'Código',
        'Obra',
        'Credor',
        'Tipo',
        'Valor medição (R$)',
        'Status',
        'Motivo reprovação',
        'Solicitante',
        'Data envio',
        'Data aprovação',
    ]
    start_row = 4
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    def _fmt_dt(dt):
        if not dt:
            return '—'
        return dt.strftime('%d/%m/%Y %H:%M')

    row = start_row + 1
    for wo in items:
        sol = (
            (wo.criado_por.get_full_name() or wo.criado_por.username)
            if wo.criado_por
            else '—'
        )
        obra_txt = f"{wo.obra.codigo}" if wo.obra else '—'
        motivo_repr = _motivo_reprovacao_listagem(wo)
        ws.cell(row=row, column=1, value=wo.codigo or '').alignment = cell_align
        ws.cell(row=row, column=2, value=obra_txt).alignment = cell_align
        ws.cell(row=row, column=3, value=(wo.nome_credor or '')[:200]).alignment = cell_align
        ws.cell(row=row, column=4, value=tipo_dict.get(wo.tipo_solicitacao, wo.tipo_solicitacao or '')).alignment = cell_align
        if wo.tipo_solicitacao == 'medicao' and wo.valor_medicao is not None:
            ws.cell(row=row, column=5, value=float(wo.valor_medicao)).alignment = cell_align
        else:
            ws.cell(row=row, column=5, value='').alignment = cell_align
        ws.cell(row=row, column=6, value=status_dict.get(wo.status, wo.status or '')).alignment = cell_align
        ws.cell(row=row, column=7, value=motivo_repr if motivo_repr != '—' else '').alignment = cell_align
        ws.cell(row=row, column=8, value=str(sol)[:120]).alignment = cell_align
        ws.cell(row=row, column=9, value=_fmt_dt(wo.data_envio)).alignment = cell_align
        ws.cell(row=row, column=10, value=_fmt_dt(wo.data_aprovacao)).alignment = cell_align
        row += 1

    if total_count > MAX_ROWS:
        note_row = row + 1
        ws.cell(
            row=note_row,
            column=1,
            value=f'Lista truncada: {total_count} pedidos no filtro; exportados os primeiros {MAX_ROWS}.',
        )
        ws.merge_cells(f'A{note_row}:J{note_row}')
        ws.cell(row=note_row, column=1).font = Font(italic=True, size=9)

    widths = [14, 12, 22, 18, 14, 14, 36, 20, 16, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"relatorio_pedidos_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{fname}"'
    return response


@login_required
def create_workorder(request):
    """
    Cria um novo pedido de obra.
    Apenas solicitantes podem criar pedidos.
    Aprovadores e admins não criam pedidos, apenas aprovam.
    """
    # Apenas solicitantes podem criar pedidos
    # Verificar se é solicitante através do grupo OU WorkOrderPermission
    
    # Verificar se o usuário está no grupo "Solicitante" OU tem permissão de solicitante
    is_solicitante_group = is_engenheiro(request.user)  # Verifica grupo "Solicitante"
    tem_permissao_solicitante = WorkOrderPermission.objects.filter(
        usuario=request.user,
        tipo_permissao='solicitante',
        ativo=True
    ).exists()
    
    # Usuário pode criar se:
    # 1. Está no grupo "Solicitante" OU
    # 2. Tem WorkOrderPermission de solicitante OU
    # 3. É admin
    pode_criar = is_solicitante_group or tem_permissao_solicitante or is_admin(request.user)
    
    if not pode_criar:
        flash_message(request, "error", "gestao.create.not_allowed")
        return redirect('gestao:list_workorders')
    
    # is_solicitante_only: é solicitante (grupo ou permissão) mas NÃO é admin ou aprovador
    is_solicitante_only = (is_solicitante_group or tem_permissao_solicitante) and not (is_aprovador(request.user) or is_admin(request.user))
    
    if request.method == 'POST':
        posted_token = request.POST.get('_form_token', '')
        if not _consume_form_token(request, 'create_workorder', posted_token):
            messages.warning(
                request,
                'O pedido já foi enviado. Verifique a lista de pedidos antes de tentar novamente.'
            )
            return redirect('gestao:list_workorders')

        form = WorkOrderForm(request.POST, user=request.user, is_creating=True)
        
        # Validar anexos obrigatórios para solicitantes
        anexos_obrigatorios = False
        if is_solicitante_only:
            anexos_files = request.FILES.getlist('anexos')
            if not anexos_files or len(anexos_files) == 0:
                extensoes_str = ', '.join(EXTENSOES_PERMITIDAS)
                flash_message(
                    request,
                    "error",
                    "gestao.create.attachment_required",
                    {"extensoes": extensoes_str},
                )
                anexos_obrigatorios = True
        
        if form.is_valid() and not anexos_obrigatorios:
            workorder = form.save(commit=False)
            workorder.criado_por = request.user
            
            # Verificar se o usuário tem acesso à obra selecionada
            obra = form.cleaned_data.get('obra')
            if obra:
                if is_solicitante_only:
                    # Se o usuário está no grupo "Solicitante" mas não tem WorkOrderPermission específica,
                    # permitir criar em qualquer obra
                    # Se tem WorkOrderPermission específica, validar apenas para essas obras
                    if tem_permissao_solicitante:
                        # Verificar se tem permissão de solicitante na obra específica
                        tem_permissao_obra = WorkOrderPermission.objects.filter(
                            obra=obra,
                            usuario=request.user,
                            tipo_permissao='solicitante',
                            ativo=True
                        ).exists()
                        if not tem_permissao_obra:
                            flash_message(
                                request,
                                "error",
                                "gestao.create.permission_obra",
                                {"obra": obra.nome},
                            )
                            context = {
                                'form': form,
                                'title': 'Criar Novo Pedido de Obra',
                                'user_profile': get_user_profile(request.user),
                                'is_solicitante': is_solicitante_only,
                            }
                            return render(request, 'obras/workorder_form.html', context)
                    # Se está no grupo "Solicitante" mas não tem WorkOrderPermission específica,
                    # permitir criar em qualquer obra (não precisa validar)
            
            # Gerar código automaticamente se não fornecido (para solicitantes)
            if not workorder.codigo and obra:
                year = timezone.now().year
                prefixo_padrao = f'{obra.codigo}-{year}'
                
                # 1. Busca todos os códigos que existem para essa obra neste ano
                codigos_existentes = WorkOrder.objects.filter(
                    obra=obra,
                    codigo__contains=f'{year}-'  # Filtro mais amplo para garantir que acha
                ).values_list('codigo', flat=True)

                maior_numero = 0

                # 2. Descobre qual é o maior número real já usado
                for codigo_db in codigos_existentes:
                    try:
                        # Tenta quebrar o código "OBRA-ANO-NUMERO" e pegar a última parte
                        partes = codigo_db.split('-')
                        numero_str = partes[-1]
                        
                        # Garante que é um número (ignora sufixos estranhos)
                        if numero_str.isdigit():
                            numero_int = int(numero_str)
                            if numero_int > maior_numero:
                                maior_numero = numero_int
                    except (ValueError, IndexError, AttributeError):
                        continue

                # 3. Define o próximo número
                proximo_numero = maior_numero + 1
                novo_codigo = f'{prefixo_padrao}-{proximo_numero:03d}'

                # 4. Trava de Segurança (While Loop)
                # Verifica no banco se esse código já existe. Se existir, pula pro próximo.
                # Isso resolve o problema de "Duplicate entry" para sempre.
                while WorkOrder.objects.filter(obra_id=obra.id, codigo=novo_codigo).exists():
                    proximo_numero += 1
                    novo_codigo = f'{prefixo_padrao}-{proximo_numero:03d}'
                
                workorder.codigo = novo_codigo
            
            # Para solicitantes, o pedido deve ir direto para "pendente" para aprovação
            # Aprovadores não criam pedidos, apenas aprovam
            if is_solicitante_only:
                workorder.status = 'pendente'
                workorder.data_envio = timezone.now()
            elif not workorder.status:
                # Se for admin criando, pode escolher o status
                workorder.status = 'pendente'
            
            # Se status for "pendente", preencher data_envio
            if workorder.status == 'pendente' and not workorder.data_envio:
                workorder.data_envio = timezone.now()
            
            workorder.save()
            
            # Processar anexos
            anexos_files = request.FILES.getlist('anexos')
            for arquivo in anexos_files:
                # Validar extensão permitida
                if validar_extensao_arquivo(arquivo.name):
                    Attachment.objects.create(
                        work_order=workorder,
                        arquivo=arquivo,
                        nome=arquivo.name,
                        descricao=f'Anexo enviado por {request.user.username}',
                        enviado_por=request.user
                    )
                else:
                    extensoes_str = ', '.join(EXTENSOES_PERMITIDAS)
                    messages.warning(request, f'O arquivo "{arquivo.name}" não tem uma extensão permitida e foi ignorado. Extensões permitidas: {extensoes_str}')
            
            # Registrar no histórico de status
            StatusHistory.objects.create(
                work_order=workorder,
                status_anterior=None,
                status_novo=workorder.status,
                alterado_por=request.user,
                observacao='Pedido criado'
            )
            
            # Enviar e-mail se status for pendente
            if workorder.status == 'pendente':
                enviar_email_novo_pedido(workorder)
                
                # Criar notificações para aprovadores da empresa
                aprovadores = User.objects.filter(
                    permissoes_obra__obra__empresa=workorder.obra.empresa,
                    permissoes_obra__tipo_permissao='aprovador',
                    permissoes_obra__ativo=True
                ).distinct()
                
                # Adicionar admins também (podem aprovar qualquer pedido)
                admins = User.objects.filter(
                    Q(is_superuser=True) | Q(groups__name='Administrador')
                ).distinct()
                
                # Combinar aprovadores e admins (sem duplicatas)
                usuarios_notificar = set(list(aprovadores) + list(admins))
                
                for usuario in usuarios_notificar:
                    # Não notificar o próprio criador do pedido
                    if usuario != request.user:
                        criar_notificacao(
                            usuario=usuario,
                            tipo='pedido_criado',
                            titulo=f'Novo Pedido: {workorder.codigo}',
                            mensagem=f'Um novo pedido foi criado por {request.user.get_full_name() or request.user.username}: {workorder.codigo}',
                            work_order=workorder
                        )
            
            messages.success(request, f'Pedido de obra "{workorder.codigo}" criado com sucesso!')
            _mark_post_success_redirect(request, '_prevent_back_create_workorder')
            return redirect('gestao:detail_workorder', pk=workorder.pk)
    else:
        # Evitar que Voltar do navegador retorne ao formulário já submetido
        r = _redirect_if_back_after_post(request, '_prevent_back_create_workorder', 'gestao:list_workorders')
        if r:
            return r
        form = WorkOrderForm(user=request.user, is_creating=True)
    
    context = {
        'form': form,
        'title': 'Criar Novo Pedido de Obra',
        'user_profile': get_user_profile(request.user),
        'is_solicitante': is_solicitante_only,
        'form_token': _generate_form_token(request, 'create_workorder'),
    }
    response = render(request, 'obras/workorder_form.html', context)
    return _no_cache_form_response(response)


@login_required
def detail_workorder(request, pk):
    """
    Visualiza os detalhes de um pedido de obra.
    Solicitantes só veem os próprios pedidos.
    Gestores e admins veem todos.
    """
    workorder = get_object_or_404(WorkOrder, pk=pk)
    user = request.user
    user_profile = get_user_profile(user)
    
    # Verificar permissão de visualização baseada em obra
    tem_permissao = False
    if is_admin(user):
        tem_permissao = True
    elif is_aprovador(user):
        # Aprovador vê se tem permissão na obra (ou em obra da mesma empresa)
        if workorder.obra.empresa_id is None:
            tem_permissao = WorkOrderPermission.objects.filter(
                obra=workorder.obra, usuario=user, tipo_permissao='aprovador', ativo=True
            ).exists()
        else:
            empresas_ids = Empresa.objects.filter(
                obras__permissoes__usuario=user,
                obras__permissoes__tipo_permissao='aprovador',
                obras__permissoes__ativo=True
            ).values_list('id', flat=True).distinct()
            tem_permissao = workorder.obra.empresa_id in empresas_ids
    elif is_engenheiro(user):
        # Solicitante pode ver:
        # 1. Pedidos que criou
        # 2. Pedidos de outros solicitantes da mesma obra (se tiver permissão na obra)
        tem_permissao_obra = WorkOrderPermission.objects.filter(
            obra=workorder.obra,
            usuario=user,
            tipo_permissao='solicitante',
            ativo=True
        ).exists()
        
        # Verificar se está no grupo "Solicitante" (mesmo sem permissão específica)
        is_solicitante_group = user.groups.filter(name='Solicitante').exists()
        
        if workorder.criado_por == user:
            # Se criou o pedido, pode ver
            tem_permissao = True
        elif tem_permissao_obra or is_solicitante_group:
            # Se tem permissão na obra OU está no grupo "Solicitante", pode ver pedidos de outros solicitantes da mesma obra
            # Verificar se o criador do pedido também é solicitante (grupo ou permissão)
            outros_solicitantes_obra = WorkOrderPermission.objects.filter(
                obra=workorder.obra,
                tipo_permissao='solicitante',
                ativo=True
            ).values_list('usuario_id', flat=True)
            
            # Se o criador está no grupo "Solicitante" ou tem permissão na obra, pode ver
            if workorder.criado_por_id:
                criador_no_grupo = workorder.criado_por.groups.filter(name='Solicitante').exists()
                criador_tem_permissao = workorder.criado_por.id in outros_solicitantes_obra
            else:
                criador_no_grupo = False
                criador_tem_permissao = False
            
            if criador_no_grupo or criador_tem_permissao:
                tem_permissao = True
            else:
                tem_permissao = False
        else:
            # Não tem permissão na obra e não está no grupo - não pode ver
            tem_permissao = False
    
    if not tem_permissao:
        messages.error(request, 'Você não tem permissão para visualizar este pedido.')
        return redirect('gestao:list_workorders')
    
    # Verificar se pode editar (apenas criador, e apenas se pendente para aprovação)
    can_edit = workorder.pode_editar(user)
    
    # Verificar se pode aprovar (aprovador/admin da obra, e pedido pendente)
    can_approve = False
    if workorder.pode_aprovar(user):
        if is_admin(user):
            can_approve = True
        elif is_aprovador(user):
            if workorder.obra.empresa_id is None:
                can_approve = WorkOrderPermission.objects.filter(
                    obra=workorder.obra, usuario=user, tipo_permissao='aprovador', ativo=True
                ).exists()
            else:
                empresas_ids = Empresa.objects.filter(
                    obras__permissoes__usuario=user,
                    obras__permissoes__tipo_permissao='aprovador',
                    obras__permissoes__ativo=True
                ).values_list('id', flat=True).distinct()
                can_approve = workorder.obra.empresa_id in empresas_ids
    
    # Buscar histórico de aprovações
    approvals = Approval.objects.filter(work_order=workorder).order_by('-created_at')
    
    # Buscar histórico completo de status
    status_history = StatusHistory.objects.filter(work_order=workorder).order_by('-created_at')
    
    # Buscar anexos - separar por versão de reaprovação
    attachments = Attachment.objects.filter(work_order=workorder).order_by('versao_reaprovacao', '-created_at')
    attachments_originais = attachments.filter(versao_reaprovacao=0)
    attachments_reaprovacao = attachments.filter(versao_reaprovacao__gt=0)
    
    # Agrupar anexos de reaprovação por versão
    anexos_por_versao = {}
    for att in attachments_reaprovacao:
        versao = att.versao_reaprovacao
        if versao not in anexos_por_versao:
            anexos_por_versao[versao] = []
        anexos_por_versao[versao].append(att)
    
    # Criar lista de tuplas (versao, anexos) ordenada por versão para facilitar iteração no template
    anexos_por_versao_ordenado = sorted(anexos_por_versao.items()) if anexos_por_versao else []
    
    # Verificar se pode adicionar anexos
    # Aprovadores NÃO podem adicionar/deletar anexos
    # Solicitantes só podem adicionar/deletar anexos na tela de EDIÇÃO, não na tela de detalhes
    # Na tela de detalhes, solicitantes só podem adicionar/deletar se o pedido está em rascunho
    can_add_attachment = False
    can_delete_attachment = False
    
    if is_admin(user):
        # Admins podem sempre adicionar/deletar
        can_add_attachment = True
        can_delete_attachment = True
    elif is_engenheiro(user) and not (is_aprovador(user) or is_admin(user)):
        # É solicitante (não é aprovador nem admin)
        # Na tela de detalhes, só pode adicionar/deletar anexos se o pedido está em rascunho
        # Para reaprovação, deve usar a tela de EDIÇÃO
        if workorder.criado_por == user and workorder.status == 'rascunho':
            can_add_attachment = True
            can_delete_attachment = True
    # Aprovadores não podem adicionar/deletar anexos (já está False por padrão)
    
    # Verificar se pode solicitar exclusão (apenas criador, pedido pendente/reprovado, não já solicitado)
    can_solicitar_exclusao = (
        workorder.criado_por == user and
        workorder.status in ['pendente', 'reprovado'] and
        not workorder.solicitado_exclusao
    )
    
    # Verificar se pode aprovar/rejeitar exclusão (aprovador/admin, pedido solicitado e em status elegível)
    can_aprovar_exclusao = False
    if workorder.solicitado_exclusao and workorder.status in ['pendente', 'reprovado']:
        if is_admin(user):
            can_aprovar_exclusao = True
        elif is_aprovador(user):
            if workorder.obra.empresa_id is None:
                can_aprovar_exclusao = WorkOrderPermission.objects.filter(
                    obra=workorder.obra, usuario=user, tipo_permissao='aprovador', ativo=True
                ).exists()
            else:
                empresas_ids = Empresa.objects.filter(
                    obras__permissoes__usuario=user,
                    obras__permissoes__tipo_permissao='aprovador',
                    obras__permissoes__ativo=True
                ).values_list('id', flat=True).distinct()
                can_aprovar_exclusao = workorder.obra.empresa_id in empresas_ids
    
    # Buscar comentários do pedido
    comments = Comment.objects.filter(work_order=workorder).select_related('autor').order_by('created_at')
    
    # Verificar se pode comentar (solicitante, aprovador ou admin que tem acesso ao pedido)
    can_comment = tem_permissao
    
    context = {
        'workorder': workorder,
        'user': user,
        'user_profile': user_profile,
        'can_edit': can_edit,
        'can_approve': can_approve,
        'can_add_attachment': can_add_attachment,
        'can_delete_attachment': can_delete_attachment,
        'can_solicitar_exclusao': can_solicitar_exclusao,
        'can_aprovar_exclusao': can_aprovar_exclusao,
        'can_comment': can_comment,
        'approvals': approvals,
        'status_history': status_history,
        'attachments': attachments,
        'attachments_originais': attachments_originais,
        'anexos_por_versao': anexos_por_versao,
        'anexos_por_versao_ordenado': anexos_por_versao_ordenado,
        'comments': comments,
    }
    return render(request, 'obras/detail_workorder.html', context)


@login_required
def exportar_snapshot_workorder_pdf(request, pk):
    """Exporta um PDF rápido com os detalhes do pedido e linha do tempo."""
    workorder = get_object_or_404(WorkOrder, pk=pk)
    user = request.user

    tem_permissao = False
    if is_admin(user):
        tem_permissao = True
    elif is_aprovador(user):
        if workorder.obra.empresa_id is None:
            tem_permissao = WorkOrderPermission.objects.filter(
                obra=workorder.obra, usuario=user, tipo_permissao='aprovador', ativo=True
            ).exists()
        else:
            empresas_ids = Empresa.objects.filter(
                obras__permissoes__usuario=user,
                obras__permissoes__tipo_permissao='aprovador',
                obras__permissoes__ativo=True
            ).values_list('id', flat=True).distinct()
            tem_permissao = workorder.obra.empresa_id in empresas_ids
    elif is_engenheiro(user):
        tem_permissao_obra = WorkOrderPermission.objects.filter(
            obra=workorder.obra,
            usuario=user,
            tipo_permissao='solicitante',
            ativo=True
        ).exists()
        is_solicitante_group = user.groups.filter(name='Solicitante').exists()
        if workorder.criado_por == user:
            tem_permissao = True
        elif tem_permissao_obra or is_solicitante_group:
            outros_solicitantes_obra = WorkOrderPermission.objects.filter(
                obra=workorder.obra,
                tipo_permissao='solicitante',
                ativo=True
            ).values_list('usuario_id', flat=True)
            if workorder.criado_por_id:
                criador_no_grupo = workorder.criado_por.groups.filter(name='Solicitante').exists()
                criador_tem_permissao = workorder.criado_por.id in outros_solicitantes_obra
            else:
                criador_no_grupo = False
                criador_tem_permissao = False
            tem_permissao = bool(criador_no_grupo or criador_tem_permissao)

    if not tem_permissao:
        messages.error(request, 'Você não tem permissão para exportar este pedido.')
        return redirect('gestao:list_workorders')

    status_history = list(StatusHistory.objects.filter(work_order=workorder).order_by('created_at'))

    def _safe(value, default='-'):
        return value if value not in (None, '') else default

    def _dt(value, default='-'):
        if not value:
            return default
        return value.strftime('%d/%m/%Y %H:%M')

    def _money(value):
        if value in (None, ''):
            return '-'
        return f"R$ {value:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')

    def _get_logo_path():
        logo_dir = os.path.join(settings.BASE_DIR, 'core', 'static', 'core', 'images')
        for name in (
            'lpla-logo-pdf-transparent.png',
            'lpla-logo-pdf.png',
            'lplan-logo2.png',
            'lplan_logo.png',
            'lplan_logo.jpg',
            'lplan_logo.jpeg',
        ):
            path = os.path.join(logo_dir, name)
            if os.path.exists(path):
                return path
        return None

    # Paleta igual aos badges da UI (list_workorders.css)
    status_palette = {
        'criacao': {'bg': colors.HexColor('#e0e7ff'), 'fg': colors.HexColor('#3730a3')},      # rascunho
        'envio': {'bg': colors.HexColor('#fef3c7'), 'fg': colors.HexColor('#92400e')},        # pendente
        'reenviado': {'bg': colors.HexColor('#fed7aa'), 'fg': colors.HexColor('#9a3412')},    # reaprovacao
        'reprovado': {'bg': colors.HexColor('#fee2e2'), 'fg': colors.HexColor('#991b1b')},
        'aprovado': {'bg': colors.HexColor('#d1fae5'), 'fg': colors.HexColor('#065f46')},
    }

    color_primary = colors.HexColor('#1A3A5C')
    color_accent = colors.HexColor('#E8F0F8')
    color_border = colors.HexColor('#D0D9E3')
    color_text = colors.HexColor('#1C1C1C')
    color_sub = colors.HexColor('#5A5A5A')
    color_info_bg = colors.HexColor('#F4F8FC')
    status_labels = dict(WorkOrder.STATUS_CHOICES)
    tipo_labels = dict(WorkOrder.TIPO_SOLICITACAO_CHOICES)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=1.7 * cm,
        leftMargin=1.7 * cm,
        topMargin=1.7 * cm,
        bottomMargin=1.5 * cm,
    )
    styles = getSampleStyleSheet()
    elements = []

    normal = ParagraphStyle('SnapNormal', parent=styles['Normal'], fontName='Helvetica', fontSize=9.2, textColor=color_text, leading=12)
    label = ParagraphStyle('SnapLabel', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=9, textColor=color_sub)
    h2 = ParagraphStyle('SnapH2', parent=styles['Heading2'], fontName='Helvetica-Bold', fontSize=11, textColor=color_primary, spaceAfter=6)
    table_head = ParagraphStyle('SnapTH', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=8.5, textColor=colors.white, alignment=TA_CENTER)
    badge_style = ParagraphStyle('SnapBadge', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=8.1, alignment=TA_CENTER, leading=10)
    def _build_status_badge(label_text, palette_key, align='CENTER', with_bullet=False, width_cm=None):
        pal = status_palette.get(palette_key, status_palette['envio'])
        badge_text = _safe(label_text, '-').upper()
        if with_bullet:
            badge_text = f"● {badge_text}"
        align_value = 'LEFT' if align == 'LEFT' else 'CENTER'
        p_style = ParagraphStyle(
            f"SnapBadge{palette_key}{align}",
            parent=badge_style,
            textColor=pal['fg'],
            alignment=TA_LEFT if align == 'LEFT' else TA_CENTER,
        )
        badge_para = Paragraph(badge_text, p_style)
        if width_cm is not None:
            col_widths = [width_cm * cm]
        else:
            col_widths = [2.45 * cm] if align == 'CENTER' else [3.1 * cm]
        badge_tbl = Table([[badge_para]], colWidths=col_widths, hAlign=align_value)
        badge_tbl.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), pal['bg']),
            ('BOX', (0, 0), (-1, -1), 0.5, pal['bg']),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ('ALIGN', (0, 0), (-1, -1), align_value),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        return badge_tbl


    # Cabeçalho no padrão RDO (logo à esquerda, texto central, fundo branco e linha azul)
    title_main = Paragraph(
        "<font color='#1A3A5C' size='14'><b>DETALHES PEDIDO</b></font>",
        ParagraphStyle('SnapTitle', parent=styles['Heading1'], fontName='Helvetica-Bold', fontSize=14, textColor=color_primary, alignment=TA_CENTER, leading=16, spaceAfter=2),
    )
    title_sub = Paragraph(
        f"<font color='#1A3A5C' size='9'>Pedido: {_safe(workorder.codigo)} · Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}</font>",
        ParagraphStyle('SnapSub', parent=styles['Normal'], fontName='Helvetica', fontSize=9, textColor=color_primary, alignment=TA_CENTER, leading=11),
    )
    logo_path = _get_logo_path()
    if logo_path:
        from reportlab.platypus import Image as RLImage
        logo = RLImage(logo_path, width=4.8 * cm, height=1.15 * cm)
        text_block = Table([[title_main], [title_sub]], colWidths=[11.2 * cm])
        text_block.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
            ('TOPPADDING', (0, 0), (-1, -1), 0),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        header = Table([[logo, text_block]], colWidths=[4.8 * cm, 11.2 * cm])
    else:
        header = Table([[title_main], [title_sub]], colWidths=[16 * cm])
    header.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.white),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LINEBELOW', (0, -1), (-1, -1), 1.2, color_primary),
    ]))
    elements.append(header)
    elements.append(Spacer(1, 0.28 * cm))

    # Informações do pedido
    status_badge_key_map = {
        'rascunho': 'criacao',
        'pendente': 'envio',
        'reaprovacao': 'reenviado',
        'reprovado': 'reprovado',
        'aprovado': 'aprovado',
    }
    status_badge_key = status_badge_key_map.get(workorder.status, 'envio')

    info_rows = [
        [Paragraph('<b>Código</b>', label), Paragraph(_safe(workorder.codigo), normal)],
        [Paragraph('<b>Obra</b>', label), Paragraph(f"{_safe(getattr(workorder.obra, 'codigo', '-'))} - {_safe(getattr(workorder.obra, 'nome', '-'))}", normal)],
        [Paragraph('<b>Credor</b>', label), Paragraph(_safe(workorder.nome_credor), normal)],
        [Paragraph('<b>Tipo de Solicitação</b>', label), Paragraph(_safe(tipo_labels.get(workorder.tipo_solicitacao, workorder.tipo_solicitacao)), normal)],
        [Paragraph('<b>Valor de medição (R$)</b>', label), Paragraph(_safe(_valor_medicao_relatorio(workorder)), normal)],
        [Paragraph('<b>Status Atual</b>', label), _build_status_badge(status_labels.get(workorder.status, workorder.status), status_badge_key, align='LEFT', with_bullet=True, width_cm=11.25)],
        [Paragraph('<b>Valor Estimado</b>', label), Paragraph(_money(workorder.valor_estimado), normal)],
        [Paragraph('<b>Prazo Estimado</b>', label), Paragraph(f"{workorder.prazo_estimado} dia(s)" if workorder.prazo_estimado else '-', normal)],
        [Paragraph('<b>Local</b>', label), Paragraph(_safe(workorder.local), normal)],
        [Paragraph('<b>Solicitante</b>', label), Paragraph(_safe(workorder.criado_por.get_full_name() or workorder.criado_por.username if workorder.criado_por else '-'), normal)],
        [Paragraph('<b>Data de Criação</b>', label), Paragraph(_dt(workorder.created_at), normal)],
        [Paragraph('<b>Data de Envio</b>', label), Paragraph(_dt(workorder.data_envio, 'Ainda não enviado'), normal)],
        [Paragraph('<b>Data da Decisão</b>', label), Paragraph(_dt(workorder.data_aprovacao, 'Ainda sem decisão'), normal)],
    ]
    if workorder.observacoes:
        info_rows.append([Paragraph('<b>Observações</b>', label), Paragraph(str(workorder.observacoes).replace('\n', '<br/>'), normal)])
    info_tbl = Table(info_rows, colWidths=[4.2 * cm, 11.8 * cm])
    info_tbl.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), color_accent),
        ('BOX', (0, 0), (-1, -1), 0.6, color_border),
        ('INNERGRID', (0, 0), (-1, -1), 0.25, color_border),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 7),
        ('RIGHTPADDING', (0, 0), (-1, -1), 7),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    elements.append(Paragraph('INFORMAÇÕES DO PEDIDO', h2))
    elements.append(info_tbl)
    elements.append(Spacer(1, 0.28 * cm))

    # Linha do tempo imutável (ordem natural dos fatos)
    elements.append(Paragraph('LINHA DO TEMPO DO FLUXO', h2))
    timeline_rows = [[
        Paragraph('Descrição', table_head),
        Paragraph('Etapa', table_head),
        Paragraph('Responsável', table_head),
        Paragraph('Data e Hora', table_head),
    ]]

    timeline_events = []
    has_previous_reprovacao = False
    for item in status_history:
        etapa_key = 'envio'
        descricao = _safe(item.observacao, '-')
        if item.status_anterior is None:
            etapa_key = 'criacao'
            descricao = descricao if descricao != '-' else 'Pedido criado no sistema.'
        elif item.status_novo == 'reprovado':
            etapa_key = 'reprovado'
            has_previous_reprovacao = True
            if descricao == '-':
                descricao = 'Pedido reprovado pelo gestor.'
        elif item.status_novo == 'aprovado':
            etapa_key = 'aprovado'
            if descricao == '-':
                descricao = 'Pedido aprovado pelo gestor.'
        elif item.status_novo == 'reaprovacao':
            etapa_key = 'reenviado' if has_previous_reprovacao else 'envio'
            if descricao == '-':
                descricao = 'Pedido reenviado para reaprovação.'
        elif item.status_novo == 'pendente':
            etapa_key = 'envio'
            if descricao == '-':
                descricao = 'Pedido enviado para análise.'
        elif item.status_anterior == item.status_novo:
            etapa_key = 'envio'
            if descricao == '-':
                descricao = 'Edição/atualização intermediária.'

        responsavel = _safe(item.alterado_por.get_full_name() or item.alterado_por.username if item.alterado_por else '-')
        timeline_events.append({
            'descricao': descricao,
            'etapa': etapa_key,
            'responsavel': responsavel,
            'data': item.created_at,
        })

    if not timeline_events:
        timeline_events.append({
            'descricao': 'Pedido criado no sistema.',
            'etapa': 'criacao',
            'responsavel': _safe(workorder.criado_por.get_full_name() or workorder.criado_por.username if workorder.criado_por else '-'),
            'data': workorder.created_at,
        })

    timeline_events.sort(key=lambda x: x['data'] or datetime.min)
    etapa_labels = {
        'criacao': 'Criação',
        'envio': 'Envio',
        'reenviado': 'Reenviado',
        'reprovado': 'Reprovado',
        'aprovado': 'Aprovado',
    }

    for index, ev in enumerate(timeline_events, start=1):
        etapa = ev['etapa']
        badge_label = etapa_labels.get(etapa, etapa.title())
        timeline_rows.append([
            Paragraph(_safe(ev['descricao']).replace('\n', '<br/>'), normal),
            _build_status_badge(badge_label, etapa, align='CENTER', with_bullet=False),
            Paragraph(_safe(ev['responsavel']), normal),
            Paragraph(_dt(ev['data']), normal),
        ])

    timeline_tbl = Table(timeline_rows, colWidths=[7.6 * cm, 2.6 * cm, 3.8 * cm, 2.0 * cm], repeatRows=1)
    timeline_style = [
        ('BACKGROUND', (0, 0), (-1, 0), color_primary),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('ALIGN', (1, 1), (1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOX', (0, 0), (-1, -1), 0.6, color_border),
        ('INNERGRID', (0, 0), (-1, -1), 0.25, color_border),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, color_info_bg]),
        ('LEFTPADDING', (0, 0), (-1, -1), 5),
        ('RIGHTPADDING', (0, 0), (-1, -1), 5),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]
    timeline_tbl.setStyle(TableStyle(timeline_style))
    elements.append(timeline_tbl)
    elements.append(Spacer(1, 0.28 * cm))

    doc.build(elements)
    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    nome_arquivo = f'pedido_{workorder.codigo}_{datetime.now().strftime("%Y%m%d_%H%M")}.pdf'
    disp = 'inline' if request.GET.get('inline') == '1' else 'attachment'
    response['Content-Disposition'] = f'{disp}; filename="{nome_arquivo}"'
    return response


@login_required
def leitura_lista_pedidos_pdf(request):
    """Modo leitura web: PDF da listagem de pedidos (mesmos filtros GET) em iframe."""
    return render(request, 'obras/pdf_reader_pedidos_lista.html', {
        'query_string': request.GET.urlencode(),
    })


@login_required
def leitura_pedido_pdf(request, pk):
    """Modo leitura web: PDF de snapshot de um pedido em iframe."""
    workorder = get_object_or_404(WorkOrder, pk=pk)
    return render(request, 'obras/pdf_reader_pedido.html', {
        'workorder': workorder,
    })


@login_required
def edit_workorder(request, pk):
    """
    Edita um pedido de obra.
    Apenas o criador pode editar (se ainda estiver em rascunho ou pendente).
    """
    workorder = get_object_or_404(WorkOrder, pk=pk)
    user = request.user
    
    # Verificar permissão de edição
    if not workorder.pode_editar(user):
        messages.error(request, 'Você não tem permissão para editar este pedido ou ele não pode mais ser editado.')
        return redirect('gestao:detail_workorder', pk=workorder.pk)
    
    # Verificar se é solicitante
    is_solicitante_group = is_engenheiro(user)  # Grupo "Solicitante"
    tem_permissao_solicitante = WorkOrderPermission.objects.filter(
        usuario=user,
        tipo_permissao='solicitante',
        ativo=True
    ).exists()
    is_solicitante_only = (is_solicitante_group or tem_permissao_solicitante) and not (is_aprovador(user) or is_admin(user))
    
    if request.method == 'POST':
        # Preservar obra quando disabled
        obra_id = request.POST.get('obra') or workorder.obra_id
        form = WorkOrderForm(request.POST, instance=workorder, user=user, is_creating=False)
        if form.is_valid():
            # Salvar valores anteriores para comparação
            status_anterior = workorder.status
            obra_anterior = workorder.obra
            nome_credor_anterior = workorder.nome_credor
            tipo_solicitacao_anterior = workorder.tipo_solicitacao
            observacoes_anterior = workorder.observacoes or ''
            valor_anterior = workorder.valor_estimado
            prazo_anterior = workorder.prazo_estimado
            local_anterior = workorder.local or ''
            
            workorder = form.save(commit=False)
            
            # Preservar obra se estava disabled
            if obra_id:
                workorder.obra_id = obra_id
            
            # Lógica especial para pedidos reprovados sendo reenviados por solicitantes
            # Contar quantas vezes este pedido já foi reenviado (contar quantas vezes mudou de reprovado para reaprovacao)
            versao_reaprovacao_atual = 0
            
            # Se o pedido está ou vai estar em reaprovação, buscar a versão atual
            if status_anterior == 'reaprovacao' or status_anterior == 'reprovado':
                versao_existente = Attachment.objects.filter(
                    work_order=workorder,
                    versao_reaprovacao__gt=0
                ).values_list('versao_reaprovacao', flat=True)
                if versao_existente:
                    versao_reaprovacao_atual = max(versao_existente)
                else:
                    versao_reaprovacao_atual = 0  # Ainda não há versões de reaprovação
            
            if is_solicitante_only:
                # Se o pedido estava reprovado e o solicitante está editando, mudar para "reaprovação"
                if status_anterior == 'reprovado':
                    # Se já existe versão, incrementar. Senão, criar versão 1
                    if versao_reaprovacao_atual > 0:
                        versao_reaprovacao_atual = versao_reaprovacao_atual + 1
                    else:
                        versao_reaprovacao_atual = 1
                    
                    workorder.status = 'reaprovacao'
                    # Atualizar data_envio para a nova submissão
                    workorder.data_envio = timezone.now()
                else:
                    # Para outros status, preservar o status atual
                    workorder.status = status_anterior
                    # Se já está em reaprovação, usar a versão atual (não criar nova)
                    if workorder.status == 'reaprovacao':
                        # versao_reaprovacao_atual já foi definida acima
                        pass
            
            # Se status mudou para "pendente" ou "reaprovação", preencher/atualizar data_envio
            if workorder.status in ['pendente', 'reaprovacao']:
                if not workorder.data_envio or status_anterior == 'reprovado':
                    workorder.data_envio = timezone.now()
            
            # Salvar o pedido
            workorder.save()
            
            # Processar exclusão de anexos existentes
            # Solicitantes podem excluir anexos se o pedido estiver em rascunho ou pendente
            anexos_excluidos_ids = request.POST.getlist('excluir_anexos')
            anexos_excluidos = []
            if anexos_excluidos_ids:
                # Verificar permissão: solicitantes podem excluir se status = 'rascunho' ou 'pendente'
                pode_excluir = True
                if is_solicitante_only and workorder.status not in ['rascunho', 'pendente']:
                    pode_excluir = False
                    messages.warning(request, 'Você não tem permissão para excluir anexos de um pedido que já foi aprovado ou reprovado.')
                
                if pode_excluir:
                    for anexo_id in anexos_excluidos_ids:
                        try:
                            anexo = Attachment.objects.get(id=int(anexo_id), work_order=workorder)
                            # Verificar permissão adicional: apenas o criador do anexo, criador do pedido, aprovadores ou admins podem excluir
                            can_delete_anexo = (
                                anexo.enviado_por == user or
                                workorder.criado_por == user or
                                is_aprovador(user) or
                                is_admin(user)
                            )
                            if can_delete_anexo:
                                nome_anexo = anexo.nome or anexo.arquivo.name.split('/')[-1]
                                anexos_excluidos.append(nome_anexo)
                                # Deletar o arquivo físico também
                                if anexo.arquivo:
                                    anexo.arquivo.delete(save=False)
                                anexo.delete()
                                
                                # Criar notificação para usuários relevantes sobre a exclusão do anexo
                                from django.db.models import Q
                                
                                # Notificar o criador do pedido (se não for quem excluiu)
                                if workorder.criado_por and workorder.criado_por != user:
                                    criar_notificacao(
                                        usuario=workorder.criado_por,
                                        tipo='anexo_removido',
                                        titulo=f'Anexo Removido: {workorder.codigo}',
                                        mensagem=f'O anexo "{nome_anexo}" foi removido do pedido {workorder.codigo} por {user.get_full_name() or user.username}.',
                                        work_order=workorder
                                    )
                                
                                # Notificar aprovadores da empresa
                                aprovadores = User.objects.filter(
                                    permissoes_obra__obra__empresa=workorder.obra.empresa,
                                    permissoes_obra__tipo_permissao='aprovador',
                                    permissoes_obra__ativo=True
                                ).distinct()
                                
                                # Adicionar admins também
                                admins = User.objects.filter(
                                    Q(is_superuser=True) | Q(groups__name='Administrador')
                                ).distinct()
                                
                                usuarios_notificar = set(list(aprovadores) + list(admins))
                                
                                for usuario in usuarios_notificar:
                                    # Não notificar o próprio usuário que excluiu
                                    if usuario != user:
                                        criar_notificacao(
                                            usuario=usuario,
                                            tipo='anexo_removido',
                                            titulo=f'Anexo Removido: {workorder.codigo}',
                                            mensagem=f'O anexo "{nome_anexo}" foi removido do pedido {workorder.codigo} por {user.get_full_name() or user.username}.',
                                            work_order=workorder
                                        )
                        except (Attachment.DoesNotExist, ValueError):
                            pass
            
            # Processar novos anexos (adicionar aos existentes, não substituir)
            anexos_files = request.FILES.getlist('anexos')
            novos_anexos = []
            
            # Determinar se estamos em contexto de reaprovação
            # IMPORTANTE: usar o status ANTES de salvar, pois pode ter mudado
            is_reaprovacao = (workorder.status == 'reaprovacao' or status_anterior == 'reaprovacao')
            
            # Se está em reaprovação mas versao_reaprovacao_atual não foi definida ou é 0, buscar a versão atual
            if is_reaprovacao:
                if versao_reaprovacao_atual == 0:
                    versao_existente = Attachment.objects.filter(
                        work_order=workorder,
                        versao_reaprovacao__gt=0
                    ).values_list('versao_reaprovacao', flat=True)
                    if versao_existente:
                        versao_reaprovacao_atual = max(versao_existente)
                    else:
                        # Se não há versões ainda, mas está em reaprovação, criar versão 1
                        versao_reaprovacao_atual = 1
            
            for arquivo in anexos_files:
                # Validar extensão permitida
                if validar_extensao_arquivo(arquivo.name):
                    # Determinar descrição e versão de reaprovação
                    if is_reaprovacao and versao_reaprovacao_atual > 0:
                        descricao = f'Anexo adicionado na reaprovação v{versao_reaprovacao_atual} por {user.username}'
                        versao = versao_reaprovacao_atual
                    else:
                        descricao = f'Anexo adicionado durante edição por {user.username}'
                        versao = 0
                    
                    try:
                        attachment = Attachment.objects.create(
                            work_order=workorder,
                            arquivo=arquivo,
                            nome=arquivo.name,
                            descricao=descricao,
                            enviado_por=user,
                            versao_reaprovacao=versao
                        )
                        novos_anexos.append(arquivo.name)
                    except Exception as e:
                        messages.error(request, f'Erro ao salvar o arquivo "{arquivo.name}": {str(e)}')
                else:
                    extensoes_str = ', '.join(EXTENSOES_PERMITIDAS)
                    messages.warning(request, f'O arquivo "{arquivo.name}" não tem uma extensão permitida e foi ignorado. Extensões permitidas: {extensoes_str}')
            
            # Detectar alterações nos campos
            alteracoes = []
            if obra_anterior != workorder.obra:
                alteracoes.append(f'Obra: {obra_anterior.codigo} → {workorder.obra.codigo}')
            if nome_credor_anterior != workorder.nome_credor:
                alteracoes.append(f'Nome do Credor: "{nome_credor_anterior}" → "{workorder.nome_credor}"')
            if tipo_solicitacao_anterior != workorder.tipo_solicitacao:
                alteracoes.append(f'Tipo: {workorder.get_tipo_solicitacao_display()}')
            if observacoes_anterior != (workorder.observacoes or ''):
                alteracoes.append('Observações atualizadas')
            if valor_anterior != workorder.valor_estimado:
                if workorder.valor_estimado:
                    # Formatar valor em reais (R$ 1.234,56)
                    valor_formatado = f'{workorder.valor_estimado:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
                    alteracoes.append(f'Valor Estimado: R$ {valor_formatado}')
                else:
                    alteracoes.append('Valor Estimado removido')
            if prazo_anterior != workorder.prazo_estimado:
                if workorder.prazo_estimado:
                    alteracoes.append(f'Prazo Estimado: {workorder.prazo_estimado} dias')
                else:
                    alteracoes.append('Prazo Estimado removido')
            if local_anterior != (workorder.local or ''):
                if workorder.local:
                    alteracoes.append(f'Local: {workorder.local}')
                else:
                    alteracoes.append('Local removido')
            if anexos_excluidos:
                if len(anexos_excluidos) == 1:
                    alteracoes.append(f'Anexo excluído: {anexos_excluidos[0]}')
                else:
                    alteracoes.append(f'{len(anexos_excluidos)} anexos excluídos: {", ".join(anexos_excluidos)}')
            
            if novos_anexos:
                if len(novos_anexos) == 1:
                    alteracoes.append(f'Novo anexo adicionado: {novos_anexos[0]}')
                    messages.success(request, f'Anexo "{novos_anexos[0]}" adicionado com sucesso!')
                else:
                    alteracoes.append(f'{len(novos_anexos)} novos anexos adicionados: {", ".join(novos_anexos)}')
                    messages.success(request, f'{len(novos_anexos)} anexos adicionados com sucesso!')
            
            # Registrar mudança de status no histórico (se o status mudou)
            if status_anterior != workorder.status:
                # Obter display do status anterior
                status_anterior_display = dict(WorkOrder.STATUS_CHOICES).get(status_anterior, status_anterior or 'N/A') if status_anterior else 'N/A'
                
                # Mensagem especial para reaprovação
                if workorder.status == 'reaprovacao' and status_anterior == 'reprovado':
                    observacao = f'Pedido reenviado para reaprovação (versão {versao_reaprovacao_atual}) por {user.get_full_name() or user.username}'
                    if alteracoes:
                        observacao += f'. Alterações realizadas: {"; ".join(alteracoes)}'
                else:
                    observacao = f'Status alterado de {status_anterior_display} para {workorder.get_status_display()}'
                    if alteracoes:
                        observacao += f'. Alterações: {"; ".join(alteracoes)}'
                
                StatusHistory.objects.create(
                    work_order=workorder,
                    status_anterior=status_anterior,
                    status_novo=workorder.status,
                    alterado_por=user,
                    observacao=observacao
                )
                
                # Enviar e-mail se mudou para pendente ou reaprovação
                if workorder.status == 'pendente':
                    enviar_email_novo_pedido(workorder)
                elif workorder.status == 'reaprovacao':
                    # Criar notificações para aprovadores quando pedido é reenviado para reaprovação
                    from django.db.models import Q
                    aprovadores = User.objects.filter(
                        permissoes_obra__obra__empresa=workorder.obra.empresa,
                        permissoes_obra__tipo_permissao='aprovador',
                        permissoes_obra__ativo=True
                    ).distinct()
                    
                    # Adicionar admins também
                    admins = User.objects.filter(
                        Q(is_superuser=True) | Q(groups__name='Administrador')
                    ).distinct()
                    
                    usuarios_notificar = set(list(aprovadores) + list(admins))
                    
                    for usuario in usuarios_notificar:
                        if usuario != user:  # Não notificar o próprio criador
                            criar_notificacao(
                                usuario=usuario,
                                tipo='pedido_atualizado',
                                titulo=f'Pedido {workorder.codigo} Enviado para Reaprovação',
                                mensagem=f'O pedido {workorder.codigo} foi reenviado para reaprovação (versão {versao_reaprovacao_atual}) por {user.get_full_name() or user.username}.',
                                work_order=workorder
                            )
            else:
                # Se o status não mudou mas o pedido foi editado, registrar edição no histórico
                if alteracoes:
                    observacao = f'Pedido editado por {user.get_full_name() or user.username}. Alterações: {"; ".join(alteracoes)}'
                else:
                    observacao = f'Pedido editado por {user.get_full_name() or user.username}.'
                
                StatusHistory.objects.create(
                    work_order=workorder,
                    status_anterior=workorder.status,
                    status_novo=workorder.status,
                    alterado_por=user,
                    observacao=observacao
                )
                
                # Criar notificações para aprovadores quando o pedido é editado
                if workorder.status == 'pendente':
                    from django.db.models import Q
                    aprovadores = User.objects.filter(
                        permissoes_obra__obra__empresa=workorder.obra.empresa,
                        permissoes_obra__tipo_permissao='aprovador',
                        permissoes_obra__ativo=True
                    ).distinct()
                    
                    # Adicionar admins também
                    admins = User.objects.filter(
                        Q(is_superuser=True) | Q(groups__name='Administrador')
                    ).distinct()
                    
                    usuarios_notificar = set(list(aprovadores) + list(admins))
                    
                    for usuario in usuarios_notificar:
                        # Não notificar o próprio editor
                        if usuario != user:
                            criar_notificacao(
                                usuario=usuario,
                                tipo='pedido_atualizado',
                                titulo=f'Pedido {workorder.codigo} Editado',
                                mensagem=f'O pedido {workorder.codigo} foi editado por {user.get_full_name() or user.username}.',
                                work_order=workorder
                            )
            
            messages.success(request, f'Pedido de obra "{workorder.codigo}" atualizado com sucesso!')
            request.session['_prevent_back_edit_workorder_pk'] = str(workorder.pk)
            return redirect('gestao:detail_workorder', pk=workorder.pk)
    else:
        # GET: evitar que Voltar do navegador retorne ao formulário já submetido
        if request.session.pop('_prevent_back_edit_workorder_pk', None) == str(workorder.pk):
            return redirect('gestao:detail_workorder', pk=workorder.pk)
        form = WorkOrderForm(instance=workorder, user=user, is_creating=False)
    
    context = {
        'form': form,
        'workorder': workorder,
        'title': f'Editar Pedido: {workorder.codigo}',
        'user_profile': get_user_profile(user),
        'is_solicitante': is_solicitante_only,
    }
    response = render(request, 'obras/workorder_form.html', context)
    return _no_cache_form_response(response)


# ========== Aprovação ==========

@login_required
@gestor_required
def approve_workorder(request, pk):
    """
    Aprova um pedido de obra.
    Apenas gestores e admins podem aprovar.
    """
    workorder = get_object_or_404(WorkOrder, pk=pk)
    
    # Verificar se pode aprovar (status pendente E gestor da obra)
    if not workorder.pode_aprovar(request.user):
        flash_message(request, "error", "gestao.approval.not_allowed_now")
        return redirect('gestao:detail_workorder', pk=workorder.pk)
    
    # Verificar se aprovador tem acesso (por empresa ou por permissão na obra quando obra sem empresa)
    if not is_admin(request.user):
        if workorder.obra.empresa_id is None:
            tem_permissao = WorkOrderPermission.objects.filter(
                obra=workorder.obra, usuario=request.user, tipo_permissao='aprovador', ativo=True
            ).exists()
        else:
            empresas_ids = Empresa.objects.filter(
                obras__permissoes__usuario=request.user,
                obras__permissoes__tipo_permissao='aprovador',
                obras__permissoes__ativo=True
            ).values_list('id', flat=True).distinct()
            tem_permissao = workorder.obra.empresa_id in empresas_ids
        if not tem_permissao:
            flash_message(request, "error", "gestao.approval.no_scope")
            return redirect('gestao:detail_workorder', pk=workorder.pk)
    
    if request.method == 'POST':
        comentario = request.POST.get('comentario', '').strip()
        
        # Usar transaction.atomic + select_for_update para evitar race condition
        # (dois aprovadores aprovando ao mesmo tempo)
        with transaction.atomic():
            # Re-buscar com lock para garantir consistência
            workorder = WorkOrder.objects.select_for_update().get(pk=pk)
            
            # Re-verificar status dentro da transação
            if not workorder.pode_aprovar(request.user):
                flash_message(request, "error", "gestao.approval.race_conflict")
                return redirect('gestao:detail_workorder', pk=workorder.pk)
            
            status_anterior = workorder.status
            
            # Criar registro de aprovação
            Approval.objects.create(
                work_order=workorder,
                aprovado_por=request.user,
                decisao='aprovado',
                comentario=comentario if comentario else None
            )
            
            # Atualizar status do pedido
            workorder.status = 'aprovado'
            workorder.data_aprovacao = timezone.now()
            workorder.save()
            
            # Registrar no histórico de status
            StatusHistory.objects.create(
                work_order=workorder,
                status_anterior=status_anterior,
                status_novo='aprovado',
                alterado_por=request.user,
                observacao=f'Aprovado. {comentario}' if comentario else 'Aprovado'
            )
        
        # Enviar e-mail FORA da transação (on_commit pattern)
        enviar_email_aprovacao(workorder, request.user, comentario)
        
        # Criar notificação para o criador do pedido
        if workorder.criado_por and workorder.criado_por != request.user:
            criar_notificacao(
                usuario=workorder.criado_por,
                tipo='pedido_aprovado',
                titulo=f'Pedido {workorder.codigo} Aprovado',
                mensagem=f'O pedido {workorder.codigo} foi aprovado por {request.user.get_full_name() or request.user.username}.',
                work_order=workorder
            )
        
        messages.success(request, f'Pedido "{workorder.codigo}" aprovado com sucesso!')
        return redirect('gestao:detail_workorder', pk=workorder.pk)
    
    # GET - mostrar formulário de confirmação
    context = {
        'workorder': workorder,
        'action': 'aprovar',
        'user_profile': get_user_profile(request.user),
    }
    return render(request, 'obras/approval_form.html', context)


@login_required
@gestor_required
def reject_workorder(request, pk):
    """
    Reprova um pedido de obra.
    Apenas gestores e admins podem reprovar.
    """
    workorder = get_object_or_404(WorkOrder, pk=pk)
    
    # Verificar se pode reprovar (status pendente E gestor da obra)
    if not workorder.pode_aprovar(request.user):
        flash_message(request, "error", "gestao.reject.not_allowed_now")
        return redirect('gestao:detail_workorder', pk=workorder.pk)
    
    # Verificar se aprovador tem acesso (por empresa ou por permissão na obra quando obra sem empresa)
    if not is_admin(request.user):
        if workorder.obra.empresa_id is None:
            tem_permissao = WorkOrderPermission.objects.filter(
                obra=workorder.obra, usuario=request.user, tipo_permissao='aprovador', ativo=True
            ).exists()
        else:
            empresas_ids = Empresa.objects.filter(
                obras__permissoes__usuario=request.user,
                obras__permissoes__tipo_permissao='aprovador',
                obras__permissoes__ativo=True
            ).values_list('id', flat=True).distinct()
            tem_permissao = workorder.obra.empresa_id in empresas_ids
        if not tem_permissao:
            flash_message(request, "error", "gestao.approval.no_scope")
            return redirect('gestao:detail_workorder', pk=workorder.pk)
    
    if request.method == 'POST':
        comentario = request.POST.get('comentario', '').strip()
        tags_selecionadas = request.POST.getlist('tags_erro')  # Lista de IDs das tags selecionadas
        novas_tags_raw = (request.POST.get('novas_tags') or '').strip()

        def _parse_novas_tags(raw_value):
            # Aceita tags separadas por vírgula, ponto e vírgula ou quebra de linha.
            if not raw_value:
                return []
            normalized = raw_value.replace(';', ',').replace('\n', ',')
            tokens = [t.strip() for t in normalized.split(',')]
            out = []
            seen = set()
            for token in tokens:
                if not token:
                    continue
                key = token.lower()
                if key in seen:
                    continue
                seen.add(key)
                out.append(token[:200])  # limite do campo nome
            return out

        novas_tags_nomes = _parse_novas_tags(novas_tags_raw)
        
        # Validar: pelo menos uma tag OU comentário deve ser fornecido
        if not tags_selecionadas and not novas_tags_nomes and not comentario:
            flash_message(request, "error", "gestao.reject.tags_or_comment_required")
            # Buscar tags disponíveis para este tipo de solicitação
            from .models import TagErro
            tags_disponiveis = TagErro.objects.filter(
                tipo_solicitacao=workorder.tipo_solicitacao,
                ativo=True
            ).order_by('ordem', 'nome')
            context = {
                'workorder': workorder,
                'action': 'reprovar',
                'user_profile': get_user_profile(request.user),
                'comentario': comentario,
                'tags_disponiveis': tags_disponiveis,
                'tags_selecionadas': [int(t) for t in tags_selecionadas] if tags_selecionadas else [],
                'novas_tags': novas_tags_raw,
            }
            return render(request, 'obras/approval_form.html', context)
        
        # Usar transaction.atomic + select_for_update para evitar race condition
        with transaction.atomic():
            # Re-buscar com lock para garantir consistência
            workorder = WorkOrder.objects.select_for_update().get(pk=pk)
            
            # Re-verificar status dentro da transação
            if not workorder.pode_aprovar(request.user):
                flash_message(request, "error", "gestao.approval.race_conflict")
                return redirect('gestao:detail_workorder', pk=workorder.pk)
            
            status_anterior = workorder.status
            
            # Criar registro de reprovação
            approval = Approval.objects.create(
                work_order=workorder,
                aprovado_por=request.user,
                decisao='reprovado',
                comentario=comentario if comentario else None
            )
            
            # Adicionar tags selecionadas
            all_tags = []
            if tags_selecionadas:
                from .models import TagErro
                tags_ids = [int(tag_id) for tag_id in tags_selecionadas]
                tags = list(
                    TagErro.objects.filter(
                        id__in=tags_ids,
                        tipo_solicitacao=workorder.tipo_solicitacao,
                        ativo=True,
                    )
                )
                all_tags.extend(tags)

            # Criar novas tags (quando aprovador informar no momento da reprovação)
            if novas_tags_nomes:
                from .models import TagErro
                try:
                    max_ordem = TagErro.objects.filter(
                        tipo_solicitacao=workorder.tipo_solicitacao
                    ).order_by('-ordem').values_list('ordem', flat=True).first()
                    next_ordem = int(max_ordem or 0)
                except Exception:
                    next_ordem = 0

                for nome_tag in novas_tags_nomes:
                    existing = TagErro.objects.filter(
                        tipo_solicitacao=workorder.tipo_solicitacao,
                        nome__iexact=nome_tag,
                    ).first()
                    if existing:
                        if not existing.ativo:
                            existing.ativo = True
                            existing.save(update_fields=['ativo', 'updated_at'])
                        all_tags.append(existing)
                        continue

                    next_ordem += 1
                    nova_tag = TagErro.objects.create(
                        nome=nome_tag,
                        tipo_solicitacao=workorder.tipo_solicitacao,
                        ativo=True,
                        ordem=next_ordem,
                        descricao='Tag criada durante reprovação',
                    )
                    all_tags.append(nova_tag)

            if all_tags:
                # remove duplicadas mantendo ordem estável
                unique_by_id = {}
                for t in all_tags:
                    unique_by_id[t.id] = t
                approval.tags_erro.set(list(unique_by_id.values()))
            
            # Atualizar status do pedido
            workorder.status = 'reprovado'
            workorder.save()
            
            # Registrar no histórico de status
            tags_nomes = ', '.join([tag.nome for tag in approval.tags_erro.all()]) if approval.tags_erro.exists() else ''
            observacao_parts = []
            if tags_nomes:
                observacao_parts.append(f'Tags: {tags_nomes}')
            if comentario:
                observacao_parts.append(f'Comentário: {comentario}')
            observacao = 'Reprovado. ' + ' | '.join(observacao_parts) if observacao_parts else 'Reprovado.'
            
            StatusHistory.objects.create(
                work_order=workorder,
                status_anterior=status_anterior,
                status_novo='reprovado',
                alterado_por=request.user,
                observacao=observacao
            )
        
        # Enviar e-mail FORA da transação
        enviar_email_reprovacao(workorder, request.user, comentario)
        
        # Criar notificação para o criador do pedido
        if workorder.criado_por and workorder.criado_por != request.user:
            criar_notificacao(
                usuario=workorder.criado_por,
                tipo='pedido_reprovado',
                titulo=f'Pedido {workorder.codigo} Reprovado',
                mensagem=f'O pedido {workorder.codigo} foi reprovado por {request.user.get_full_name() or request.user.username}. Motivo: {comentario}',
                work_order=workorder
            )
        
        messages.success(request, f'Pedido "{workorder.codigo}" reprovado.')
        return redirect('gestao:detail_workorder', pk=workorder.pk)
    
    # GET - mostrar formulário de confirmação
    # Buscar tags disponíveis para este tipo de solicitação
    from .models import TagErro
    tags_disponiveis = TagErro.objects.filter(
        tipo_solicitacao=workorder.tipo_solicitacao,
        ativo=True
    ).order_by('ordem', 'nome')
    
    context = {
        'workorder': workorder,
        'action': 'reprovar',
        'user_profile': get_user_profile(request.user),
        'tags_disponiveis': tags_disponiveis,
        'tags_selecionadas': [],
        'comentario': '',
        'novas_tags': '',
    }
    return render(request, 'obras/approval_form.html', context)


# ========== Anexos ==========

@login_required
def upload_attachment(request, pk):
    """
    Faz upload de um anexo para um pedido de obra.
    Apenas usuários autenticados podem fazer upload.
    Solicitantes só podem adicionar anexos se o pedido ainda estiver em rascunho.
    """
    workorder = get_object_or_404(WorkOrder, pk=pk)
    user = request.user
    
    # Verificar se é solicitante (não é aprovador nem admin)
    is_solicitante_only = is_engenheiro(user) and not (is_aprovador(user) or is_admin(user))
    
    # Solicitantes só podem adicionar anexos via upload_attachment se o pedido está em rascunho
    # Para reaprovação, devem usar a tela de EDIÇÃO do pedido
    if is_solicitante_only and workorder.status != 'rascunho':
        messages.error(request, 'Você não pode adicionar anexos a um pedido que já foi enviado. Use a opção de editar o pedido para adicionar novos anexos.')
        return redirect('gestao:detail_workorder', pk=workorder.pk)
    
    # Verificar permissão de visualização do pedido
    # Aprovadores NÃO podem adicionar anexos
    if not (is_admin(user) or workorder.criado_por == user):
        messages.error(request, 'Você não tem permissão para anexar arquivos a este pedido.')
        return redirect('gestao:detail_workorder', pk=workorder.pk)
    
    if request.method == 'POST':
        form = AttachmentForm(request.POST, request.FILES)
        if form.is_valid():
            attachment = form.save(commit=False)
            attachment.work_order = workorder
            attachment.enviado_por = user
            
            # Se não informou nome, usar o nome do arquivo
            if not attachment.nome:
                attachment.nome = os.path.basename(attachment.arquivo.name)
            
            # Se o pedido está em reaprovação, marcar o anexo com a versão atual
            if workorder.status == 'reaprovacao':
                # Buscar a versão atual de reaprovação
                versao_atual = Attachment.objects.filter(
                    work_order=workorder,
                    versao_reaprovacao__gt=0
                ).values_list('versao_reaprovacao', flat=True)
                if versao_atual:
                    versao = max(versao_atual)
                else:
                    versao = 1
                attachment.versao_reaprovacao = versao
                attachment.descricao = f'Anexo adicionado na reaprovação v{versao} por {user.username}'
            else:
                attachment.versao_reaprovacao = 0
            
            attachment.save()
            messages.success(request, f'Anexo "{attachment.get_nome_display()}" enviado com sucesso!')
            return redirect('gestao:detail_workorder', pk=workorder.pk)
    else:
        form = AttachmentForm()
    
    context = {
        'form': form,
        'workorder': workorder,
        'user_profile': get_user_profile(user),
    }
    return render(request, 'obras/upload_attachment.html', context)


@login_required
def delete_attachment(request, pk):
    """
    Deleta um anexo.
    Apenas o criador do anexo, o criador do pedido, gestores ou admins podem deletar.
    Solicitantes só podem deletar anexos se o pedido ainda estiver em rascunho.
    """
    attachment = get_object_or_404(Attachment, pk=pk)
    workorder = attachment.work_order
    user = request.user
    
    # Verificar se é solicitante (não é aprovador nem admin)
    is_solicitante_only = is_engenheiro(user) and not (is_aprovador(user) or is_admin(user))
    
    # Solicitantes NÃO podem deletar anexos após enviar o pedido
    if is_solicitante_only and workorder.status != 'rascunho':
        messages.error(request, 'Você não pode deletar anexos de um pedido que já foi enviado.')
        return redirect('gestao:detail_workorder', pk=workorder.pk)
    
    # Verificar permissão de deleção
    # Aprovadores NÃO podem deletar anexos
    can_delete = (
        attachment.enviado_por == user or
        workorder.criado_por == user or
        is_admin(user)
    )
    
    if not can_delete:
        messages.error(request, 'Você não tem permissão para deletar este anexo.')
        return redirect('gestao:detail_workorder', pk=workorder.pk)
    
    if request.method == 'POST':
        nome_arquivo = attachment.get_nome_display()
        
        # Deletar arquivo físico se existir
        if attachment.arquivo:
            try:
                if os.path.isfile(attachment.arquivo.path):
                    os.remove(attachment.arquivo.path)
            except Exception as e:
                # Log do erro, mas continua com a deleção do registro
                pass
        
        # Deletar registro
        attachment.delete()
        messages.success(request, f'Anexo "{nome_arquivo}" deletado com sucesso!')
        return redirect('gestao:detail_workorder', pk=workorder.pk)
    
    # GET - mostrar confirmação
    context = {
        'attachment': attachment,
        'workorder': workorder,
        'user_profile': get_user_profile(user),
    }
    return render(request, 'obras/delete_attachment.html', context)


# ========== CRUD Obra ==========

@login_required
def list_obras(request):
    """
    Lista todas as obras.
    Apenas administradores e responsáveis por empresa podem acessar.
    """
    if not (is_admin(request.user) or is_responsavel_empresa(request.user)):
        messages.error(request, 'Você não tem permissão para gerenciar obras.')
        return redirect('gestao:home')
    
    obras = Obra.objects.all().select_related('empresa')
    
    # Filtrar por empresa se for responsável
    if is_responsavel_empresa(request.user) and not is_admin(request.user):
        obras = obras.filter(empresa__responsavel=request.user)
    
    obras = obras.order_by('empresa', 'codigo')
    
    # Filtros
    ativo_filter = request.GET.get('ativo')
    if ativo_filter is not None and ativo_filter != '':
        obras = obras.filter(ativo=ativo_filter == '1')
    
    search_query = request.GET.get('search')
    if search_query:
        obras = obras.filter(
            Q(codigo__icontains=search_query) |
            Q(nome__icontains=search_query) |
            Q(descricao__icontains=search_query)
        )
    
    # Paginação
    paginator = Paginator(obras, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'obras': page_obj,
        'user_profile': get_user_profile(request.user),
        'ativo_filter': ativo_filter,
        'search_query': search_query,
    }
    return render(request, 'obras/list_obras.html', context)


@login_required
def create_obra(request):
    """
    Cria uma nova obra.
    Apenas administradores podem criar.
    """
    if not is_admin(request.user):
        messages.error(request, 'Apenas administradores podem criar obras.')
        return redirect('gestao:list_obras')
    
    # Evitar que Voltar do navegador retorne ao formulário já submetido
    if request.method == 'GET':
        r = _redirect_if_back_after_post(request, '_prevent_back_create_obra', 'gestao:list_obras')
        if r:
            return r
    
    if request.method == 'POST':
        form = ObraForm(request.POST, user=request.user)
        if form.is_valid():
            obra = form.save()
            from gestao_aprovacao.services.entity_audit import record_obra_created

            record_obra_created(request, request.user, obra)
            messages.success(request, f'Obra "{obra.codigo}" criada com sucesso!')
            _mark_post_success_redirect(request, '_prevent_back_create_obra')
            return redirect('gestao:detail_obra', pk=obra.pk)
    else:
        form = ObraForm(user=request.user)
    
    context = {
        'form': form,
        'title': 'Criar Nova Obra',
        'user_profile': get_user_profile(request.user),
    }
    response = render(request, 'obras/obra_form.html', context)
    return _no_cache_form_response(response)


@login_required
def detail_obra(request, pk):
    """
    Visualiza os detalhes de uma obra.
    Apenas administradores e responsáveis por empresa podem acessar.
    """
    obra = get_object_or_404(Obra, pk=pk)
    
    # Verificar permissão
    if not (is_admin(request.user) or 
            (is_responsavel_empresa(request.user) and obra.empresa and obra.empresa.responsavel == request.user)):
        messages.error(request, 'Você não tem permissão para visualizar esta obra.')
        return redirect('gestao:list_obras')
    
    # Contar pedidos relacionados
    workorders_count = obra.work_orders.count()
    workorders_pendentes = obra.work_orders.filter(status='pendente').count()
    
    context = {
        'obra': obra,
        'user_profile': get_user_profile(request.user),
        'workorders_count': workorders_count,
        'workorders_pendentes': workorders_pendentes,
    }
    return render(request, 'obras/detail_obra.html', context)


@login_required
def edit_obra(request, pk):
    """
    Edita uma obra.
    Apenas administradores e responsáveis por empresa podem editar.
    """
    obra = get_object_or_404(Obra, pk=pk)
    
    # Verificar permissão
    if not (is_admin(request.user) or 
            (is_responsavel_empresa(request.user) and obra.empresa and obra.empresa.responsavel == request.user)):
        messages.error(request, 'Você não tem permissão para editar esta obra.')
        return redirect('gestao:list_obras')
    
    # GET: evitar que Voltar do navegador retorne ao formulário já submetido
    if request.method == 'GET':
        if request.session.pop('_prevent_back_edit_obra_pk', None) == str(obra.pk):
            return redirect('gestao:detail_obra', pk=obra.pk)
    
    if request.method == 'POST':
        from gestao_aprovacao.services.entity_audit import record_obra_updated, snapshot_obra

        before = snapshot_obra(Obra.objects.get(pk=obra.pk))
        form = ObraForm(request.POST, instance=obra, user=request.user)
        if form.is_valid():
            saved = form.save()
            record_obra_updated(request, request.user, before, snapshot_obra(saved))
            messages.success(request, f'Obra "{saved.codigo}" atualizada com sucesso!')
            request.session['_prevent_back_edit_obra_pk'] = str(saved.pk)
            return redirect('gestao:detail_obra', pk=saved.pk)
    else:
        form = ObraForm(instance=obra, user=request.user)
        # Garantir que empresa seja obrigatória mesmo na edição
        form.fields['empresa'].required = True
    
    context = {
        'form': form,
        'obra': obra,
        'title': f'Editar Obra: {obra.codigo}',
        'user_profile': get_user_profile(request.user),
    }
    response = render(request, 'obras/obra_form.html', context)
    return _no_cache_form_response(response)


# ========== GERENCIAMENTO DE USUÁRIOS (central em /central/usuarios/) ==========

@login_required
def list_users_or_redirect_central(request):
    if request.user.is_staff or request.user.is_superuser:
        return redirect('central_list_users')
    return list_users(request)


@login_required
def create_user_or_redirect_central(request):
    if request.user.is_staff or request.user.is_superuser:
        return redirect('central_create_user')
    return create_user(request)


@login_required
def edit_user_or_redirect_central(request, pk):
    if request.user.is_staff or request.user.is_superuser:
        return redirect('central_edit_user', pk=pk)
    return edit_user(request, pk=pk)


@login_required
def delete_user_or_redirect_central(request, pk):
    if request.user.is_staff or request.user.is_superuser:
        return redirect('central_delete_user', pk=pk)
    return delete_user(request, pk=pk)


def _user_governance_resolve_href(ev):
    r = ev.get('reverse')
    if not r:
        return None
    name, kw = r
    try:
        return reverse(name, kwargs=kw)
    except Exception:
        return None


@login_required
def user_governance_or_redirect_central(request, pk):
    if request.user.is_staff or request.user.is_superuser:
        return redirect('central_user_governance', pk=pk)
    return user_governance_panel(request, pk=pk)


@login_required
def user_governance_panel(request, pk):
    """
    Painel de governança operacional do usuário (administração).
    Leitura de dados já persistidos; filtros por período, módulo e obra.
    """
    if not (is_admin(request.user) or is_responsavel_empresa(request.user)):
        messages.error(request, 'Você não tem permissão para acessar esta página.')
        return redirect('gestao:home')

    target = get_object_or_404(
        User.objects.prefetch_related('groups', 'perfil'),
        pk=pk,
    )
    if not viewer_can_see_target_user(request.user, target):
        messages.error(request, 'Você não tem permissão para visualizar este usuário.')
        return redirect('central_list_users' if getattr(request, '_central_redirect', False) else 'gestao:list_users')

    period_raw = (request.GET.get('period') or '90').strip()
    try:
        period_days = int(period_raw)
    except ValueError:
        period_days = 90
    if period_days not in (0, 30, 90, 365):
        period_days = 90

    module = (request.GET.get('module') or '').strip()
    if module not in ('', 'gestao', 'diario', 'contas', 'admin'):
        module = ''

    obra_filter = (request.GET.get('obra') or '').strip()
    obra_filter_id = None
    if obra_filter.isdigit():
        oid = int(obra_filter)
        if is_responsavel_empresa(request.user) and not is_admin(request.user):
            _emp_filt = Empresa.objects.filter(responsavel=request.user, ativo=True)
            obras_qs = Obra.objects.filter(empresa__in=_emp_filt, ativo=True)
        else:
            obras_qs = Obra.objects.filter(ativo=True)
        if obras_qs.filter(pk=oid).exists():
            obra_filter_id = oid

    kpi_window = 30
    kpis = build_kpis(target, days_window=kpi_window)
    scope = build_scope(target)
    pending = build_pending_queues(target)
    critical = build_critical_highlights(target, limit=20)
    for ev in critical:
        ev['href'] = _user_governance_resolve_href(ev)

    timeline_opts = TimelineOptions(
        period_days=period_days,
        module=module,
        obra_id=obra_filter_id,
    )
    timeline_events = build_timeline_events(target, timeline_opts)
    for ev in timeline_events:
        ev['href'] = _user_governance_resolve_href(ev)

    paginator = Paginator(timeline_events, 25)
    page_number = request.GET.get('page')
    timeline_page = paginator.get_page(page_number)

    if is_responsavel_empresa(request.user) and not is_admin(request.user):
        _emp_filt = Empresa.objects.filter(responsavel=request.user, ativo=True)
        obras_filtro = Obra.objects.filter(empresa__in=_emp_filt, ativo=True).order_by('nome', 'codigo')
    else:
        obras_filtro = Obra.objects.filter(ativo=True).order_by('nome', 'codigo')

    alerts = operational_alerts(target, kpis)
    obra_usage = usage_by_obra_gestao(target, limit=10)
    audit_insights = build_audit_insights(target)
    strategic_insights = build_strategic_insights(target, kpis, audit_insights)

    _central = getattr(request, '_central_redirect', False)
    list_url = reverse('central_list_users' if _central else 'gestao:list_users')

    context = {
        'target_user': target,
        'user_profile': get_user_profile(request.user),
        'kpis': kpis,
        'audit_insights': audit_insights,
        'strategic_insights': strategic_insights,
        'scope': scope,
        'pending': pending,
        'critical': critical,
        'timeline_page': timeline_page,
        'timeline_total': paginator.count,
        'period_days': period_days,
        'module_filter': module,
        'obra_filter': str(obra_filter_id) if obra_filter_id else '',
        'obras_filtro': obras_filtro,
        'alerts': alerts,
        'obra_usage': obra_usage,
        'use_central_urls': _central,
        'list_users_url': list_url,
    }
    return render(request, 'obras/user_governance.html', context)


@login_required
def list_users(request):
    """
    Lista todos os usuários.
    Administradores veem todos os usuários.
    Responsáveis por empresa veem apenas usuários vinculados às suas empresas.
    """
    if not (is_admin(request.user) or is_responsavel_empresa(request.user)):
        messages.error(request, 'Você não tem permissão para acessar esta página.')
        return redirect('gestao:home')
    
    # Se for responsável por empresa (não admin), mostrar apenas usuários de suas empresas
    if is_responsavel_empresa(request.user) and not is_admin(request.user):
        # Buscar empresas onde o usuário é responsável
        empresas_responsavel = Empresa.objects.filter(responsavel=request.user, ativo=True)
        # Buscar usuários vinculados a essas empresas
        users = User.objects.filter(
            empresas_vinculadas__empresa__in=empresas_responsavel,
            empresas_vinculadas__ativo=True
        ).distinct().order_by('username')
    else:
        # Admin vê todos
        users = User.objects.all().order_by('username')
    
    # Filtros
    grupo_filter = request.GET.get('grupo')
    if grupo_filter:
        try:
            grupo = Group.objects.get(name=grupo_filter)
            users = users.filter(groups=grupo)
        except Group.DoesNotExist:
            pass
    
    search_query = request.GET.get('search')
    if search_query:
        users = users.filter(
            Q(username__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query)
        )

    # Obras disponíveis no filtro (mesmo recorte de permissão da tela)
    if is_responsavel_empresa(request.user) and not is_admin(request.user):
        _emp_filt = Empresa.objects.filter(responsavel=request.user, ativo=True)
        obras_queryset = Obra.objects.filter(empresa__in=_emp_filt, ativo=True)
    else:
        obras_queryset = Obra.objects.filter(ativo=True)

    obra_filter = (request.GET.get('obra') or '').strip()
    obra_filter_id = None
    if obra_filter.isdigit():
        oid = int(obra_filter)
        if obras_queryset.filter(pk=oid).exists():
            obra_filter_id = oid
            users = users.filter(
                permissoes_obra__obra_id=oid,
                permissoes_obra__ativo=True,
            ).distinct()

    users = users.annotate(
        _obra_sort=Min(
            'permissoes_obra__obra__nome',
            filter=Q(permissoes_obra__ativo=True),
        )
    ).order_by(Coalesce('_obra_sort', Value('ZZZZZZZZZZ')), 'username')

    users = users.prefetch_related(
        Prefetch(
            'permissoes_obra',
            queryset=WorkOrderPermission.objects.filter(ativo=True).select_related('obra'),
        )
    )

    # Paginação
    paginator = Paginator(users, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Obter grupos e obras (GestControll) para cada usuário
    users_with_groups = []
    for user in page_obj:
        grupos = user.groups.all()
        obras_vistas = []
        _seen = set()
        for p in user.permissoes_obra.all():
            if not p.obra_id or p.obra_id in _seen:
                continue
            _seen.add(p.obra_id)
            obras_vistas.append(p.obra)
        obras_vistas.sort(key=lambda o: (o.nome.lower(), o.codigo.lower()))

        users_with_groups.append({
            'user': user,
            'groups': grupos,
            'is_engenheiro': grupos.filter(name='Solicitante').exists(),
            'is_gestor': grupos.filter(name='Gestor').exists(),
            'is_admin': grupos.filter(name='Administrador').exists() or user.is_superuser,
            'obras_list': obras_vistas,
        })
    
    # Verificar se é responsável por empresa para mostrar apenas empresas dele
    empresas_disponiveis = None
    if is_responsavel_empresa(request.user) and not is_admin(request.user):
        empresas_disponiveis = Empresa.objects.filter(responsavel=request.user, ativo=True)
    elif is_admin(request.user):
        empresas_disponiveis = Empresa.objects.filter(ativo=True)
    
    # Grupos oficiais para o filtro (ordem por módulo; sem perfis ocultos na UI)
    grupos_filtro = grupos_ordenados_atribuivel()
    _central = getattr(request, '_central_redirect', False)
    context = {
        'page_obj': page_obj,
        'users_with_groups': users_with_groups,
        'user_profile': get_user_profile(request.user),
        'grupo_filter': grupo_filter or '',
        'search_query': search_query or '',
        'grupos': grupos_filtro,
        'empresas_disponiveis': empresas_disponiveis,
        'is_responsavel_empresa': is_responsavel_empresa(request.user) and not is_admin(request.user),
        'is_admin': is_admin(request.user),
        'use_central_urls': _central,
        'obras_filtro': obras_queryset.order_by('nome', 'codigo'),
        'obra_filter': str(obra_filter_id) if obra_filter_id is not None else '',
    }
    return render(request, 'obras/list_users.html', context)


@login_required
def create_user(request):
    """
    Cria um novo usuário.
    Administradores podem criar qualquer usuário.
    Responsáveis por empresa podem criar usuários e vinculá-los às suas empresas.
    """
    if not (is_admin(request.user) or is_responsavel_empresa(request.user)):
        messages.error(request, 'Você não tem permissão para criar usuários.')
        return redirect('gestao:list_users')
    
    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')
        first_name = request.POST.get('first_name', '')
        last_name = request.POST.get('last_name', '')
        grupo = request.POST.get('grupo')
        create_as_pending = request.POST.get('create_as_pending') == 'on'
        grupos_selecionados = filtrar_grupos_post_atribuivel(request.POST.getlist('grupos'))
        project_ids = request.POST.getlist('projects')
        
        if create_as_pending and not email:
            messages.error(request, 'Para cadastro pendente, informe o e-mail corporativo.')
        elif create_as_pending and not is_allowed_signup_email(email):
            messages.error(request, 'E-mail fora dos domínios permitidos para auto cadastro/aprovação.')
        elif not create_as_pending and (not username or not password):
            messages.error(request, 'Username e senha são obrigatórios.')
        elif not grupos_selecionados:
            messages.error(request, 'Selecione pelo menos um grupo para o usuário acessar algum sistema.')
        elif not create_as_pending and User.objects.filter(username=username).exists():
            messages.error(request, 'Já existe um usuário com este username.')
        else:
            if create_as_pending:
                if UserSignupRequest.objects.filter(email__iexact=(email or '').strip(), status=UserSignupRequest.STATUS_PENDENTE).exists():
                    messages.info(request, 'Já existe uma solicitação pendente para este e-mail.')
                    return redirect('central_list_users' if getattr(request, '_central_redirect', False) else 'gestao:list_users')

                full_name = f"{first_name} {last_name}".strip() or username or (email or '').split('@')[0]
                signup_req = create_signup_request(
                    full_name=full_name,
                    email=(email or '').strip().lower(),
                    username_suggestion=(username or '').strip(),
                    notes='Solicitação criada pelo cadastro interno (pendente de aprovação).',
                    requested_groups=grupos_selecionados,
                    requested_project_ids=project_ids,
                    origem=UserSignupRequest.ORIGEM_INTERNO,
                    requested_by=request.user,
                )
                notify_signup_request_created(signup_req)
                messages.success(request, 'Solicitação criada como pendente. O aprovador foi alertado no sistema.')
                from accounts.audit_signup import record_signup_request_internal

                record_signup_request_internal(request, request.user, signup_req)
                _mark_post_success_redirect(
                    request, '_prevent_back_create_user',
                    'central_list_users' if getattr(request, '_central_redirect', False) else 'gestao:list_users'
                )
                return redirect('central_list_users' if getattr(request, '_central_redirect', False) else 'gestao:list_users')

            user = User.objects.create_user(
                username=username,
                email=email,
                password=password,
                first_name=first_name,
                last_name=last_name
            )
            
            # Adicionar aos grupos selecionados (usuário pode ter vários: Gestão + Mapa + Diário)
            for grupo_name in grupos_selecionados:
                if grupo_name not in GRUPOS.TODOS:
                    continue
                try:
                    grupo_obj = Group.objects.get(name=grupo_name)
                    user.groups.add(grupo_obj)
                except Group.DoesNotExist:
                    pass
            
            # Lista única de obras = core.Project. Vincular ao Diário (ProjectMember) e ao GestControll (WorkOrderPermission quando Obra.project existe).
            for pid in project_ids:
                try:
                    project_id = int(pid)
                    project = Project.objects.filter(pk=project_id, is_active=True).first()
                    if not project:
                        continue
                    ProjectMember.objects.get_or_create(user=user, project_id=project_id)
                    obra = Obra.objects.filter(project_id=project_id, ativo=True).first()
                    if obra:
                        if 'Solicitante' in grupos_selecionados:
                            WorkOrderPermission.objects.get_or_create(
                                usuario=user, obra=obra, tipo_permissao='solicitante',
                                defaults={'ativo': True}
                            )
                        if 'Aprovador' in grupos_selecionados:
                            WorkOrderPermission.objects.get_or_create(
                                usuario=user, obra=obra, tipo_permissao='aprovador',
                                defaults={'ativo': True}
                            )
                        if obra.empresa_id and (is_admin(request.user) or (is_responsavel_empresa(request.user) and obra.empresa.responsavel == request.user)):
                            UserEmpresa.objects.get_or_create(
                                usuario=user,
                                empresa=obra.empresa,
                                defaults={'ativo': True}
                            )
                except (ValueError, TypeError):
                    pass
            
            # Criar perfil de usuário e fazer upload da foto se fornecida
            perfil, created = UserProfile.objects.get_or_create(usuario=user)
            if 'foto_perfil' in request.FILES:
                perfil.foto_perfil = request.FILES['foto_perfil']
                perfil.save()
            
            # Enviar e-mail com login e senha para o novo usuário (se tiver e-mail)
            if email and email.strip():
                nome_completo = f"{first_name} {last_name}".strip() or username
                site_url = request.build_absolute_uri('/').rstrip('/')
                enviado = enviar_email_credenciais_novo_usuario(
                    email_destino=email.strip(),
                    username=username,
                    senha_plana=password,
                    nome_completo=nome_completo,
                    site_url=site_url,
                )
                if not enviado:
                    messages.warning(
                        request,
                        'Usuário criado, mas o e-mail com login e senha não pôde ser enviado. '
                        'Informe as credenciais manualmente ao usuário.'
                    )
            else:
                messages.warning(request, 'E-mail não informado: o usuário não receberá as credenciais por e-mail.')

            from gestao_aprovacao.services.user_admin_audit import record_user_created

            _pids = []
            for _pid in project_ids:
                try:
                    _pids.append(int(_pid))
                except (TypeError, ValueError):
                    pass
            record_user_created(
                request,
                request.user,
                user,
                sorted(grupos_selecionados),
                sorted(_pids),
            )
            messages.success(request, f'Usuário "{username}" criado com sucesso!')
            _mark_post_success_redirect(
                request, '_prevent_back_create_user',
                'central_list_users' if getattr(request, '_central_redirect', False) else 'gestao:list_users'
            )
            return redirect('central_list_users' if getattr(request, '_central_redirect', False) else 'gestao:list_users')
    
    # GET: evitar que Voltar do navegador retorne ao formulário já submetido
    if request.method == 'GET':
        r = _redirect_if_back_after_post(request, '_prevent_back_create_user', 'gestao:list_users')
        if r:
            return r
    
    # Filtrar empresas baseado no usuário
    if is_responsavel_empresa(request.user) and not is_admin(request.user):
        empresas_disponiveis = Empresa.objects.filter(responsavel=request.user, ativo=True).order_by('codigo')
    else:
        empresas_disponiveis = Empresa.objects.filter(ativo=True).order_by('codigo')
    
    grupos_modulos = grupos_modulos_para_atribuicao()
    
    # Lista única de obras do sistema = core.Project (mesma lista do Diário de Obra)
    projects = Project.objects.filter(is_active=True).order_by('name')
    context = {
        'grupos_modulos': grupos_modulos,
        'empresas': empresas_disponiveis,
        'projects': projects,
        'user_profile': get_user_profile(request.user),
        'is_responsavel_empresa': is_responsavel_empresa(request.user) and not is_admin(request.user),
        'use_central_urls': getattr(request, '_central_redirect', False),
    }
    response = render(request, 'obras/create_user.html', context)
    return _no_cache_form_response(response)


@login_required
def edit_user(request, pk):
    """
    Edita um usuário (grupos e informações básicas).
    Administradores podem editar qualquer usuário.
    Responsáveis por empresa podem editar apenas usuários vinculados às suas empresas.
    """
    user = get_object_or_404(User, pk=pk)
    
    # Verificar permissão
    if is_admin(request.user):
        # Admin pode editar qualquer um
        pass
    elif is_responsavel_empresa(request.user):
        # Responsável só pode editar usuários de suas empresas
        empresas_responsavel = Empresa.objects.filter(responsavel=request.user, ativo=True)
        usuarios_permitidos = User.objects.filter(
            empresas_vinculadas__empresa__in=empresas_responsavel,
            empresas_vinculadas__ativo=True
        )
        if user not in usuarios_permitidos:
            messages.error(request, 'Você não tem permissão para editar este usuário.')
            return redirect('gestao:list_users')
    else:
        messages.error(request, 'Você não tem permissão para editar usuários.')
        return redirect('gestao:list_users')
    
    if request.method == 'POST':
        from gestao_aprovacao.services.user_admin_audit import (
            record_user_updated,
            snapshot_user_admin_state,
        )

        before = snapshot_user_admin_state(user)
        new_password = request.POST.get('password')
        password_changed = bool(new_password)

        user.email = request.POST.get('email', user.email)
        user.first_name = request.POST.get('first_name', user.first_name)
        user.last_name = request.POST.get('last_name', user.last_name)
        
        # Atualizar senha se fornecida
        if new_password:
            user.set_password(new_password)
        
        user.save()
        
        # Atualizar grupos (mantém grupos ocultos legados que não vêm no POST)
        post_grupos = filtrar_grupos_post_atribuivel(request.POST.getlist('grupos'))
        grupos_selecionados = merge_grupos_legados_ocultos(user, post_grupos)
        user.groups.clear()
        for grupo_name in grupos_selecionados:
            if grupo_name not in GRUPOS.TODOS:
                continue
            try:
                grupo = Group.objects.get(name=grupo_name)
                user.groups.add(grupo)
            except Group.DoesNotExist:
                pass
        
        # Lista única de obras = core.Project. Definir Diário (ProjectMember) = exatamente os projetos selecionados.
        project_ids = request.POST.getlist('projects')
        ProjectMember.objects.filter(user=user).delete()
        for pid in project_ids:
            try:
                project_id = int(pid)
                if Project.objects.filter(pk=project_id, is_active=True).exists():
                    ProjectMember.objects.get_or_create(user=user, project_id=project_id)
            except (ValueError, TypeError):
                pass
        WorkOrderPermission.objects.filter(usuario=user).delete()
        if is_admin(request.user):
            UserEmpresa.objects.filter(usuario=user).update(ativo=False)
        else:
            empresas_responsavel = Empresa.objects.filter(responsavel=request.user, ativo=True)
            UserEmpresa.objects.filter(usuario=user, empresa__in=empresas_responsavel).update(ativo=False)
        for pid in project_ids:
            try:
                project_id = int(pid)
                obra = Obra.objects.filter(project_id=project_id, ativo=True).first()
                if not obra:
                    continue
                if 'Solicitante' in grupos_selecionados:
                    WorkOrderPermission.objects.create(
                        usuario=user, obra=obra, tipo_permissao='solicitante', ativo=True
                    )
                if 'Aprovador' in grupos_selecionados:
                    WorkOrderPermission.objects.create(
                        usuario=user, obra=obra, tipo_permissao='aprovador', ativo=True
                    )
                if obra.empresa_id and (is_admin(request.user) or (is_responsavel_empresa(request.user) and obra.empresa.responsavel == request.user)):
                    UserEmpresa.objects.update_or_create(
                        usuario=user,
                        empresa=obra.empresa,
                        defaults={'ativo': True}
                    )
            except (ValueError, TypeError):
                pass
        
        # Atualizar foto de perfil se fornecida
        perfil, created = UserProfile.objects.get_or_create(usuario=user)
        if 'foto_perfil' in request.FILES:
            perfil.foto_perfil = request.FILES['foto_perfil']
            perfil.save()
        
        messages.success(request, f'Usuário "{user.username}" atualizado com sucesso!')
        user.refresh_from_db()
        record_user_updated(
            request,
            request.user,
            user,
            before,
            snapshot_user_admin_state(user),
            password_changed,
        )
        request.session['_prevent_back_edit_user_pk'] = str(user.pk)
        return redirect('central_list_users' if getattr(request, '_central_redirect', False) else 'gestao:list_users')
    
    # GET: evitar que Voltar do navegador retorne ao formulário já submetido
    if request.session.pop('_prevent_back_edit_user_pk', None) == str(user.pk):
        return redirect('central_list_users' if getattr(request, '_central_redirect', False) else 'gestao:list_users')
    
    # Obter empresas disponíveis baseado no usuário
    if is_responsavel_empresa(request.user) and not is_admin(request.user):
        empresas_disponiveis = Empresa.objects.filter(responsavel=request.user, ativo=True).order_by('codigo')
        user_empresas = Empresa.objects.filter(
            usuarios_vinculados__usuario=user,
            usuarios_vinculados__ativo=True,
            responsavel=request.user
        )
    else:
        empresas_disponiveis = Empresa.objects.filter(ativo=True).order_by('codigo')
        user_empresas = Empresa.objects.filter(usuarios_vinculados__usuario=user, usuarios_vinculados__ativo=True)
    
    try:
        user_perfil = UserProfile.objects.get(usuario=user)
    except UserProfile.DoesNotExist:
        user_perfil = None
    
    grupos_modulos = grupos_modulos_para_atribuicao()
    
    projects = Project.objects.filter(is_active=True).order_by('name')
    user_project_ids = list(ProjectMember.objects.filter(user=user).values_list('project_id', flat=True))
    # Sincronizar Diário: se o usuário tem obras no GestControll mas não tem ProjectMember, criar a partir das Obras vinculadas
    if not user_project_ids:
        project_ids_from_obras = list(
            Obra.objects.filter(
                permissoes__usuario=user,
                permissoes__ativo=True,
            ).exclude(project__isnull=True).values_list('project_id', flat=True).distinct()
        )
        for project_id in project_ids_from_obras:
            ProjectMember.objects.get_or_create(user=user, project_id=project_id)
        if project_ids_from_obras:
            user_project_ids = list(ProjectMember.objects.filter(user=user).values_list('project_id', flat=True))
    
    context = {
        'user_obj': user,
        'user_perfil': user_perfil,
        'grupos_modulos': grupos_modulos,
        'user_grupos': user.groups.all(),
        'empresas': empresas_disponiveis,
        'user_empresas': user_empresas,
        'projects': projects,
        'user_project_ids': user_project_ids,
        'user_profile': get_user_profile(request.user),
        'is_responsavel_empresa': is_responsavel_empresa(request.user) and not is_admin(request.user),
        'use_central_urls': getattr(request, '_central_redirect', False),
    }
    response = render(request, 'obras/edit_user.html', context)
    return _no_cache_form_response(response)


@login_required
def delete_user(request, pk):
    """
    Exclui um usuário e todos os seus dados relacionados.
    Apenas admins podem excluir usuários.
    """
    if not is_admin(request.user):
        messages.error(request, 'Você não tem permissão para excluir usuários.')
        _r = 'central_list_users' if getattr(request, '_central_redirect', False) else 'gestao:list_users'
        return redirect(_r)
    
    user = get_object_or_404(User, pk=pk)
    
    # Não permitir excluir a si mesmo
    if user == request.user:
        messages.error(request, 'Você não pode excluir seu próprio usuário.')
        _r = 'central_list_users' if getattr(request, '_central_redirect', False) else 'gestao:list_users'
        return redirect(_r)
    
    # Verificar se há relacionamentos que impedem a exclusão (PROTECT)
    bloqueios = []
    
    # Verificar se é responsável por alguma empresa
    if Empresa.objects.filter(responsavel=user, ativo=True).exists():
        bloqueios.append('É responsável por uma ou mais empresas')
    
    # Verificar se criou pedidos
    if WorkOrder.objects.filter(criado_por=user).exists():
        bloqueios.append('Criou pedidos de obra')
    
    # Verificar se aprovou pedidos
    if Approval.objects.filter(aprovado_por=user).exists():
        bloqueios.append('Aprovou/reprovou pedidos')
    
    # Verificar se enviou anexos
    if Attachment.objects.filter(enviado_por=user).exists():
        bloqueios.append('Enviou anexos')
    
    # Verificar se alterou status
    if StatusHistory.objects.filter(alterado_por=user).exists():
        bloqueios.append('Alterou status de pedidos')
    
    # Verificar se fez comentários
    if Comment.objects.filter(autor=user).exists():
        bloqueios.append('Fez comentários em pedidos')
    
    if request.method == 'POST':
        # Confirmar exclusão
        if bloqueios:
            messages.error(request, f'Não é possível excluir este usuário porque: {", ".join(bloqueios)}. Remova esses dados primeiro.')
            return redirect('central_list_users' if getattr(request, '_central_redirect', False) else 'gestao:list_users')
        
        username = user.username
        from gestao_aprovacao.services.user_admin_audit import record_user_deleted

        record_user_deleted(
            request,
            request.user,
            {
                'user_id': user.pk,
                'username': user.username,
                'email': user.email or '',
                'group_names': list(user.groups.values_list('name', flat=True)),
                'empresa_ids_vinculadas': list(
                    UserEmpresa.objects.filter(usuario=user, ativo=True).values_list(
                        'empresa_id', flat=True
                    )
                ),
            },
        )

        # Excluir dados relacionados (CASCADE será feito automaticamente, mas vamos garantir)
        # Notificações
        Notificacao.objects.filter(usuario=user).delete()
        
        # Lembretes
        Lembrete.objects.filter(enviado_para=user).delete()
        
        # Permissões de obra
        WorkOrderPermission.objects.filter(usuario=user).delete()
        
        # Vínculos com empresas
        UserEmpresa.objects.filter(usuario=user).delete()
        
        # Perfil do usuário
        UserProfile.objects.filter(usuario=user).delete()
        
        # Limpar grupos
        user.groups.clear()
        
        # Excluir o usuário
        user.delete()
        
        messages.success(request, f'Usuário "{username}" e todos os seus dados foram excluídos com sucesso!')
        return redirect('central_list_users' if getattr(request, '_central_redirect', False) else 'gestao:list_users')
    
    # GET - mostrar página de confirmação
    context = {
        'user': user,
        'bloqueios': bloqueios,
        'user_profile': get_user_profile(request.user),
        'use_central_urls': getattr(request, '_central_redirect', False),
    }
    return render(request, 'obras/delete_user_confirm.html', context)


@login_required
def edit_my_profile(request):
    """
    Permite que o usuário edite seu próprio perfil (nome, email, senha e foto).
    """
    user = request.user
    
    if request.method == 'POST':
        # Atualizar informações básicas
        user.email = request.POST.get('email', user.email)
        user.first_name = request.POST.get('first_name', user.first_name)
        user.last_name = request.POST.get('last_name', user.last_name)
        
        # Atualizar senha se fornecida
        new_password = request.POST.get('password')
        if new_password:
            user.set_password(new_password)
        
        user.save()
        
        # Atualizar foto de perfil se fornecida
        perfil, created = UserProfile.objects.get_or_create(usuario=user)
        if 'foto_perfil' in request.FILES:
            perfil.foto_perfil = request.FILES['foto_perfil']
            perfil.save()
        
        messages.success(request, 'Seu perfil foi atualizado com sucesso!')
        return redirect('gestao:edit_my_profile')
    
    # Obter perfil do usuário
    try:
        user_perfil = UserProfile.objects.get(usuario=user)
    except UserProfile.DoesNotExist:
        user_perfil = None
    
    context = {
        'user_obj': user,
        'user_perfil': user_perfil,
        'user_profile': get_user_profile(request.user),
    }
    return render(request, 'obras/edit_my_profile.html', context)


@login_required
def list_notificacoes(request):
    """
    Lista todas as notificações do usuário logado.
    """
    notificacoes = Notificacao.objects.filter(usuario=request.user)
    
    # Filtros
    tipo_filter = request.GET.get('tipo')
    if tipo_filter:
        notificacoes = notificacoes.filter(tipo=tipo_filter)
    
    lida_filter = request.GET.get('lida')
    if lida_filter == 'true':
        notificacoes = notificacoes.filter(lida=True)
    elif lida_filter == 'false':
        notificacoes = notificacoes.filter(lida=False)
    
    # Ordenação
    order_by = request.GET.get('order_by', '-created_at')
    notificacoes = notificacoes.order_by(order_by)
    
    # Paginação
    paginator = Paginator(notificacoes, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Contar não lidas
    nao_lidas_count = Notificacao.objects.filter(usuario=request.user, lida=False).count()
    
    # Marcar como lidas apenas se o usuário clicar no botão
    marcar_todas_lidas = request.GET.get('marcar_todas_lidas')
    if marcar_todas_lidas == 'true':
        Notificacao.objects.filter(usuario=request.user, lida=False).update(lida=True)
        messages.success(request, 'Todas as notificações foram marcadas como lidas.')
        # Preservar filtros ao redirecionar sem quebrar reverse com querystring.
        params = {}
        if tipo_filter:
            params['tipo'] = tipo_filter
        if lida_filter:
            params['lida'] = lida_filter
        if order_by:
            params['order_by'] = order_by
        query = urlencode(params)
        if query:
            return redirect(f'{request.path}?{query}')
        return redirect(request.path)
    
    context = {
        'notificacoes': page_obj,
        'page_obj': page_obj,
        'user_profile': get_user_profile(request.user),
        'nao_lidas_count': nao_lidas_count,
        'tipo_filter': tipo_filter,
        'lida_filter': lida_filter,
        'order_by': order_by,
        'tipos_notificacao': Notificacao.TIPO_CHOICES,
    }
    return render(request, 'obras/list_notificacoes.html', context)


@login_required
def marcar_notificacao_lida(request, pk):
    """
    Marca uma notificação específica como lida.
    """
    notificacao = get_object_or_404(Notificacao, pk=pk, usuario=request.user)
    notificacao.marcar_como_lida()
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        # Se for requisição AJAX, retorna JSON
        from django.http import JsonResponse
        return JsonResponse({'success': True})
    
    return redirect('gestao:list_notificacoes')


@login_required
def get_notificacoes_count(request):
    """
    Retorna o número de notificações não lidas do usuário (para AJAX).
    """
    count = Notificacao.objects.filter(usuario=request.user, lida=False).count()
    from django.http import JsonResponse
    return JsonResponse({'count': count})


# ========== CRUD EMPRESA ==========

@login_required
def list_empresas(request):
    """
    Lista todas as empresas.
    Apenas administradores e responsáveis por empresa podem acessar.
    """
    if not (is_admin(request.user) or is_responsavel_empresa(request.user)):
        messages.error(request, 'Você não tem permissão para acessar esta página.')
        return redirect('gestao:home')
    
    empresas = Empresa.objects.all()
    
    # Filtrar por responsável se não for admin
    if is_responsavel_empresa(request.user) and not is_admin(request.user):
        empresas = empresas.filter(responsavel=request.user)
    
    empresas = empresas.order_by('codigo')
    
    # Filtros
    ativo_filter = request.GET.get('ativo')
    if ativo_filter is not None and ativo_filter != '':
        empresas = empresas.filter(ativo=ativo_filter == '1')
    
    search_query = request.GET.get('search')
    if search_query:
        empresas = empresas.filter(
            Q(codigo__icontains=search_query) |
            Q(nome__icontains=search_query) |
            Q(razao_social__icontains=search_query) |
            Q(cnpj__icontains=search_query)
        )
    
    # Paginação
    paginator = Paginator(empresas, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'empresas': page_obj,
        'user_profile': get_user_profile(request.user),
        'ativo_filter': ativo_filter,
        'search_query': search_query,
    }
    return render(request, 'obras/list_empresas.html', context)


@login_required
def create_empresa(request):
    """
    Cria uma nova empresa.
    Apenas administradores podem criar.
    """
    if not is_admin(request.user):
        messages.error(request, 'Apenas administradores podem criar empresas.')
        return redirect('gestao:list_empresas')
    
    # Evitar que Voltar do navegador retorne ao formulário já submetido
    if request.method == 'GET':
        r = _redirect_if_back_after_post(request, '_prevent_back_create_empresa', 'gestao:list_empresas')
        if r:
            return r
    
    if request.method == 'POST':
        form = EmpresaForm(request.POST, user=request.user)
        if form.is_valid():
            empresa = form.save()
            from gestao_aprovacao.services.entity_audit import record_empresa_created

            record_empresa_created(request, request.user, empresa)
            messages.success(request, f'Empresa "{empresa.codigo}" criada com sucesso!')
            _mark_post_success_redirect(request, '_prevent_back_create_empresa')
            return redirect('gestao:detail_empresa', pk=empresa.pk)
    else:
        form = EmpresaForm(user=request.user)
    
    context = {
        'form': form,
        'title': 'Criar Nova Empresa',
        'user_profile': get_user_profile(request.user),
    }
    response = render(request, 'obras/empresa_form.html', context)
    return _no_cache_form_response(response)


@login_required
def detail_empresa(request, pk):
    """
    Exibe detalhes de uma empresa.
    """
    empresa = get_object_or_404(Empresa, pk=pk)
    
    # Verificar permissão
    if not (is_admin(request.user) or (is_responsavel_empresa(request.user) and empresa.responsavel == request.user)):
        messages.error(request, 'Você não tem permissão para acessar esta empresa.')
        return redirect('gestao:list_empresas')
    
    # Estatísticas
    obras_count = empresa.obras.count()
    obras_ativas = empresa.obras.filter(ativo=True).count()
    usuarios_count = empresa.usuarios_vinculados.filter(ativo=True).count()
    usuarios_vinculados = empresa.usuarios_vinculados.filter(ativo=True).select_related('usuario')
    
    context = {
        'empresa': empresa,
        'obras_count': obras_count,
        'obras_ativas': obras_ativas,
        'usuarios_count': usuarios_count,
        'usuarios_vinculados': usuarios_vinculados,
        'user_profile': get_user_profile(request.user),
    }
    return render(request, 'obras/detail_empresa.html', context)


@login_required
def edit_empresa(request, pk):
    """
    Edita uma empresa.
    Apenas administradores podem editar.
    """
    if not is_admin(request.user):
        messages.error(request, 'Apenas administradores podem editar empresas.')
        return redirect('gestao:list_empresas')
    
    empresa = get_object_or_404(Empresa, pk=pk)
    
    # GET: evitar que Voltar do navegador retorne ao formulário já submetido
    if request.method == 'GET':
        if request.session.pop('_prevent_back_edit_empresa_pk', None) == str(empresa.pk):
            return redirect('gestao:detail_empresa', pk=empresa.pk)
    
    if request.method == 'POST':
        from gestao_aprovacao.services.entity_audit import record_empresa_updated, snapshot_empresa

        before = snapshot_empresa(Empresa.objects.get(pk=empresa.pk))
        form = EmpresaForm(request.POST, instance=empresa, user=request.user)
        if form.is_valid():
            saved = form.save()
            record_empresa_updated(request, request.user, before, snapshot_empresa(saved))
            messages.success(request, f'Empresa "{saved.codigo}" atualizada com sucesso!')
            request.session['_prevent_back_edit_empresa_pk'] = str(saved.pk)
            return redirect('gestao:detail_empresa', pk=saved.pk)
    else:
        form = EmpresaForm(instance=empresa, user=request.user)
    
    context = {
        'form': form,
        'empresa': empresa,
        'title': f'Editar Empresa: {empresa.codigo}',
        'user_profile': get_user_profile(request.user),
    }
    response = render(request, 'obras/empresa_form.html', context)
    return _no_cache_form_response(response)


# ========== GERENCIAMENTO DE PERMISSÕES POR OBRA ==========

@login_required
def manage_obra_permissions(request, pk):
    """
    Gerencia permissões de usuários por obra.
    Permite adicionar/remover solicitantes e aprovadores de uma obra.
    """
    obra = get_object_or_404(Obra, pk=pk)
    
    # Verificar permissão
    if not (is_admin(request.user) or 
            (is_responsavel_empresa(request.user) and obra.empresa and obra.empresa.responsavel == request.user)):
        messages.error(request, 'Você não tem permissão para gerenciar permissões desta obra.')
        return redirect('gestao:list_obras')
    
    if request.method == 'POST':
        from audit.action_codes import AuditAction
        from audit.recording import record_audit_event

        action = request.POST.get('action')
        
        if action == 'add':
            usuario_id = request.POST.get('usuario')
            tipo_permissao = request.POST.get('tipo_permissao')
            
            if usuario_id and tipo_permissao:
                try:
                    usuario = User.objects.get(pk=usuario_id)
                    # Verificar se já existe
                    perm, created = WorkOrderPermission.objects.get_or_create(
                        obra=obra,
                        usuario=usuario,
                        tipo_permissao=tipo_permissao,
                        defaults={'ativo': True}
                    )
                    if not created:
                        perm.ativo = True
                        perm.save()
                    record_audit_event(
                        actor=request.user,
                        subject_user=usuario,
                        action_code=AuditAction.OBRA_WORKORDER_PERM_ADD,
                        summary=f'Permissão {perm.get_tipo_permissao_display()} na obra {obra.codigo} — {usuario.username}',
                        payload={
                            'obra_id': obra.pk,
                            'obra_codigo': obra.codigo,
                            'tipo_permissao': perm.tipo_permissao,
                            'permission_id': perm.pk,
                        },
                        module='gestao',
                        request=request,
                    )
                    messages.success(request, f'Permissão adicionada para {usuario.username}.')
                except User.DoesNotExist:
                    messages.error(request, 'Usuário não encontrado.')

        elif action == 'remove':
            perm_id = request.POST.get('permission_id')
            if perm_id:
                try:
                    perm = WorkOrderPermission.objects.select_related('usuario').get(pk=perm_id, obra=obra)
                    target_u = perm.usuario
                    tipo_ant = perm.tipo_permissao
                    perm.delete()
                    record_audit_event(
                        actor=request.user,
                        subject_user=target_u,
                        action_code=AuditAction.OBRA_WORKORDER_PERM_REMOVE,
                        summary=f'Permissão removida ({tipo_ant}) na obra {obra.codigo} — {target_u.username}',
                        payload={
                            'obra_id': obra.pk,
                            'obra_codigo': obra.codigo,
                            'tipo_permissao': tipo_ant,
                        },
                        module='gestao',
                        request=request,
                    )
                    messages.success(request, 'Permissão removida com sucesso.')
                except WorkOrderPermission.DoesNotExist:
                    messages.error(request, 'Permissão não encontrada.')

        elif action == 'toggle':
            perm_id = request.POST.get('permission_id')
            if perm_id:
                try:
                    perm = WorkOrderPermission.objects.select_related('usuario').get(pk=perm_id, obra=obra)
                    perm.ativo = not perm.ativo
                    perm.save()
                    record_audit_event(
                        actor=request.user,
                        subject_user=perm.usuario,
                        action_code=AuditAction.OBRA_WORKORDER_PERM_TOGGLE,
                        summary=f'Permissão {"ativada" if perm.ativo else "desativada"} ({perm.tipo_permissao}) — {obra.codigo} / {perm.usuario.username}',
                        payload={
                            'obra_id': obra.pk,
                            'obra_codigo': obra.codigo,
                            'tipo_permissao': perm.tipo_permissao,
                            'ativo': perm.ativo,
                            'permission_id': perm.pk,
                        },
                        module='gestao',
                        request=request,
                    )
                    messages.success(request, f'Permissão {"ativada" if perm.ativo else "desativada"}.')
                except WorkOrderPermission.DoesNotExist:
                    messages.error(request, 'Permissão não encontrada.')
        
        return redirect('gestao:manage_obra_permissions', pk=obra.pk)
    
    # Obter permissões existentes
    permissoes = WorkOrderPermission.objects.filter(obra=obra).select_related('usuario').order_by('tipo_permissao', 'usuario__username')
    
    # Obter usuários disponíveis (solicitantes e aprovadores)
    grupo_solicitante = Group.objects.filter(name='Solicitante').first()
    grupo_aprovador = Group.objects.filter(name='Aprovador').first()
    
    usuarios_solicitantes = User.objects.filter(groups=grupo_solicitante).order_by('username') if grupo_solicitante else User.objects.none()
    usuarios_aprovadores = User.objects.filter(groups=grupo_aprovador).order_by('username') if grupo_aprovador else User.objects.none()
    
    # Filtrar usuários vinculados à empresa da obra (quando obra tem empresa)
    if obra.empresa_id:
        usuarios_empresa = User.objects.filter(
            empresas_vinculadas__empresa=obra.empresa,
            empresas_vinculadas__ativo=True
        ).distinct()
        usuarios_solicitantes = usuarios_solicitantes.filter(id__in=usuarios_empresa)
        usuarios_aprovadores = usuarios_aprovadores.filter(id__in=usuarios_empresa)
    # Obra sem empresa: mantém todos do grupo Solicitante/Aprovador para atribuir permissão por obra
    
    context = {
        'obra': obra,
        'permissoes': permissoes,
        'usuarios_solicitantes': usuarios_solicitantes,
        'usuarios_aprovadores': usuarios_aprovadores,
        'user_profile': get_user_profile(request.user),
    }
    return render(request, 'obras/manage_obra_permissions.html', context)


@login_required
def solicitar_exclusao(request, pk):
    """
    Permite que o solicitante solicite a exclusão de um pedido pendente ou reprovado.
    A exclusão só será efetivada após aprovação do aprovador.
    """
    workorder = get_object_or_404(WorkOrder, pk=pk)
    user = request.user
    
    # Verificar se é o criador do pedido
    if workorder.criado_por != user:
        flash_message(request, "error", "gestao.delete_request.only_creator")
        return redirect('gestao:detail_workorder', pk=workorder.pk)
    
    # Verificar se o pedido está em status elegível
    if workorder.status not in ['pendente', 'reprovado']:
        flash_message(request, "error", "gestao.delete_request.invalid_status")
        return redirect('gestao:detail_workorder', pk=workorder.pk)
    
    # Verificar se já foi solicitado
    if workorder.solicitado_exclusao:
        messages.warning(request, 'A exclusão deste pedido já foi solicitada e está aguardando aprovação.')
        return redirect('gestao:detail_workorder', pk=workorder.pk)
    
    if request.method == 'POST':
        motivo = request.POST.get('motivo', '').strip()
        
        # Validar motivo
        if not motivo:
            flash_message(request, "error", "gestao.delete_request.reason_required")
            context = {
                'workorder': workorder,
                'user_profile': get_user_profile(user),
            }
            return render(request, 'obras/solicitar_exclusao.html', context)
        
        # Marcar como solicitado para exclusão
        workorder.solicitado_exclusao = True
        workorder.solicitado_exclusao_por = user
        workorder.solicitado_exclusao_em = timezone.now()
        workorder.motivo_exclusao = motivo
        workorder.save()
        
        # Notificar exatamente quem pode aprovar/rejeitar essa exclusão.
        usuarios_notificar = _get_approvers_for_exclusion(workorder)

        for usuario in usuarios_notificar:
            criar_notificacao(
                usuario=usuario,
                tipo='exclusao_solicitada',
                titulo=f'Exclusão Solicitada: {workorder.codigo}',
                mensagem=f'O solicitante {user.get_full_name() or user.username} solicitou a exclusão do pedido {workorder.codigo}. Motivo: {motivo}',
                work_order=workorder
            )
        
        messages.success(request, 'Solicitação de exclusão enviada. Aguardando aprovação do aprovador.')
        return redirect('gestao:detail_workorder', pk=workorder.pk)
    
    # GET - mostrar confirmação
    context = {
        'workorder': workorder,
        'user_profile': get_user_profile(user),
    }
    return render(request, 'obras/solicitar_exclusao.html', context)


@login_required
def aprovar_exclusao(request, pk):
    """
    Permite que o aprovador aprove a exclusão de um pedido.
    """
    workorder = get_object_or_404(WorkOrder, pk=pk)
    user = request.user
    
    # Verificar se foi solicitado
    if not workorder.solicitado_exclusao:
        flash_message(request, "error", "gestao.delete_approve.not_requested")
        return redirect('gestao:detail_workorder', pk=workorder.pk)
    
    # Verificar se é aprovador ou admin
    if not (is_aprovador(user) or is_admin(user)):
        flash_message(request, "error", "gestao.delete_approve.no_role")
        return redirect('gestao:detail_workorder', pk=workorder.pk)
    
    # Verificar se o aprovador tem permissão (por empresa ou por permissão na obra quando obra sem empresa)
    if is_aprovador(user) and not is_admin(user):
        if workorder.obra.empresa_id is None:
            tem_permissao = WorkOrderPermission.objects.filter(
                obra=workorder.obra, usuario=user, tipo_permissao='aprovador', ativo=True
            ).exists()
        else:
            empresas_ids = Empresa.objects.filter(
                obras__permissoes__usuario=user,
                obras__permissoes__tipo_permissao='aprovador',
                obras__permissoes__ativo=True
            ).values_list('id', flat=True).distinct()
            tem_permissao = workorder.obra.empresa_id in empresas_ids
        if not tem_permissao:
            flash_message(request, "error", "gestao.delete_approve.no_scope")
            return redirect('gestao:detail_workorder', pk=workorder.pk)
    
    if request.method == 'POST':
        # Salvar dados da solicitação antes de limpar
        solicitante_nome = workorder.solicitado_exclusao_por.get_full_name() if workorder.solicitado_exclusao_por else "N/A"
        motivo_texto = f' Motivo: {workorder.motivo_exclusao}' if workorder.motivo_exclusao else ''
        
        # Registrar no histórico
        StatusHistory.objects.create(
            work_order=workorder,
            status_anterior=workorder.status,
            status_novo='cancelado',
            alterado_por=user,
            observacao=f'Pedido excluído após solicitação do solicitante {solicitante_nome}.{motivo_texto} Aprovação realizada por {user.get_full_name() or user.username}.'
        )
        
        # Mudar status para cancelado e limpar campos de solicitação de exclusão
        workorder.status = 'cancelado'
        workorder.solicitado_exclusao = False
        workorder.solicitado_exclusao_por = None
        workorder.solicitado_exclusao_em = None
        workorder.motivo_exclusao = None
        workorder.save()
        
        # Notificar o solicitante
        if workorder.criado_por:
            criar_notificacao(
                usuario=workorder.criado_por,
                tipo='exclusao_aprovada',
                titulo=f'Exclusão Aprovada: {workorder.codigo}',
                mensagem=f'A exclusão do pedido {workorder.codigo} foi aprovada por {user.get_full_name() or user.username}.',
                work_order=workorder
            )
        
        messages.success(request, f'Exclusão do pedido {workorder.codigo} aprovada e efetivada.')
        return redirect('gestao:list_workorders')
    
    # GET - mostrar confirmação
    context = {
        'workorder': workorder,
        'user_profile': get_user_profile(user),
    }
    return render(request, 'obras/aprovar_exclusao.html', context)


@login_required
def rejeitar_exclusao(request, pk):
    """
    Permite que o aprovador rejeite a solicitação de exclusão de um pedido.
    """
    workorder = get_object_or_404(WorkOrder, pk=pk)
    user = request.user
    
    # Verificar se foi solicitado
    if not workorder.solicitado_exclusao:
        flash_message(request, "error", "gestao.delete_approve.not_requested")
        return redirect('gestao:detail_workorder', pk=workorder.pk)
    
    # Verificar se é aprovador ou admin
    if not (is_aprovador(user) or is_admin(user)):
        flash_message(request, "error", "gestao.delete_reject.no_role")
        return redirect('gestao:detail_workorder', pk=workorder.pk)
    
    # Verificar se o aprovador tem permissão (por empresa ou por permissão na obra quando obra sem empresa)
    if is_aprovador(user) and not is_admin(user):
        if workorder.obra.empresa_id is None:
            tem_permissao = WorkOrderPermission.objects.filter(
                obra=workorder.obra, usuario=user, tipo_permissao='aprovador', ativo=True
            ).exists()
        else:
            empresas_ids = Empresa.objects.filter(
                obras__permissoes__usuario=user,
                obras__permissoes__tipo_permissao='aprovador',
                obras__permissoes__ativo=True
            ).values_list('id', flat=True).distinct()
            tem_permissao = workorder.obra.empresa_id in empresas_ids
        if not tem_permissao:
            flash_message(request, "error", "gestao.delete_reject.no_scope")
            return redirect('gestao:detail_workorder', pk=workorder.pk)
    
    if request.method == 'POST':
        motivo = request.POST.get('motivo', '')
        
        # Remover solicitação de exclusão
        workorder.solicitado_exclusao = False
        workorder.solicitado_exclusao_por = None
        workorder.solicitado_exclusao_em = None
        workorder.motivo_exclusao = None
        workorder.save()
        
        # Notificar o solicitante
        if workorder.criado_por:
            mensagem = f'A solicitação de exclusão do pedido {workorder.codigo} foi rejeitada por {user.get_full_name() or user.username}.'
            if motivo:
                mensagem += f' Motivo: {motivo}'
            
            criar_notificacao(
                usuario=workorder.criado_por,
                tipo='exclusao_rejeitada',
                titulo=f'Exclusão Rejeitada: {workorder.codigo}',
                mensagem=mensagem,
                work_order=workorder
            )
        
        messages.success(request, f'Solicitação de exclusão do pedido {workorder.codigo} rejeitada.')
        return redirect('gestao:detail_workorder', pk=workorder.pk)
    
    # GET - mostrar formulário
    context = {
        'workorder': workorder,
        'user_profile': get_user_profile(user),
    }
    return render(request, 'obras/rejeitar_exclusao.html', context)


@login_required
def desempenho_equipe(request):
    """
    Página de desempenho da equipe (tempo médio de resposta).
    Exclusivo para administradores e responsáveis pela empresa.
    """
    # Verificar permissão: apenas admin ou responsável pela empresa
    if not (is_admin(request.user) or is_responsavel_empresa(request.user)):
        messages.error(request, 'Você não tem permissão para acessar esta página.')
        return redirect('gestao:home')
    
    context = {
        'title': 'Desempenho da Equipe',
        'user': request.user,
        'user_profile': get_user_profile(request.user),
    }
    return render(request, 'obras/desempenho_equipe.html', context)


@login_required
def desempenho_equipe_api(request):
    """
    API endpoint para retornar dados de desempenho da equipe (tempo médio de resposta).
    Exclusivo para administradores e responsáveis pela empresa.
    Retorna JSON com tempo médio de resposta por aprovador.
    """
    # Verificar permissão: apenas admin ou responsável pela empresa
    if not (is_admin(request.user) or is_responsavel_empresa(request.user)):
        return JsonResponse({
            'erro': 'Você não tem permissão para acessar esta página.',
            'dados': [],
            'tempo_medio_geral': 0,
            'total_pedidos': 0,
            'periodo': 'Últimos 30 dias'
        }, status=403)
    
    try:
        # Obter período da requisição (padrão: 30 dias)
        dias_periodo = int(request.GET.get('dias', 30))
        if dias_periodo not in [7, 15, 30, 60, 90]:
            dias_periodo = 30  # Garantir valor válido
        
        # Período atual
        agora = timezone.now()
        data_inicio_atual = agora - timedelta(days=dias_periodo)
        
        # Período anterior equivalente (mesmo número de dias antes do período atual)
        data_fim_anterior = data_inicio_atual
        data_inicio_anterior = data_fim_anterior - timedelta(days=dias_periodo)
        
        # Limitar número de aprovações para evitar timeout
        MAX_APPROVALS = 500  # Reduzido de 1000 para melhor performance
        
        # Função auxiliar para buscar aprovações de um período
        def buscar_aprovacoes(data_inicio_periodo, data_fim_periodo=None):
            if data_fim_periodo is None:
                data_fim_periodo = agora
            
            if is_responsavel_empresa(request.user) and not is_admin(request.user):
                # Buscar empresas onde o usuário é responsável
                empresas_ids = list(Empresa.objects.filter(
                    responsavel=request.user
                ).values_list('id', flat=True))
                
                if not empresas_ids:
                    return Approval.objects.none()
                
                # Buscar aprovações apenas de pedidos dessas empresas
                return Approval.objects.filter(
                    created_at__gte=data_inicio_periodo,
                    created_at__lte=data_fim_periodo,
                    work_order__obra__empresa_id__in=empresas_ids
                ).select_related('aprovado_por', 'work_order').order_by('-created_at')[:MAX_APPROVALS]
            else:
                # Admin vê todas as aprovações
                return Approval.objects.filter(
                    created_at__gte=data_inicio_periodo,
                    created_at__lte=data_fim_periodo
                ).select_related('aprovado_por', 'work_order').order_by('-created_at')[:MAX_APPROVALS]
        
        # Buscar aprovações do período atual
        approvals = buscar_aprovacoes(data_inicio_atual)
        
        # Buscar aprovações do período anterior para comparação
        approvals_anterior = buscar_aprovacoes(data_inicio_anterior, data_fim_anterior)
        
        # Função auxiliar para processar aprovações e calcular estatísticas
        def processar_aprovacoes(approvals_list):
            desempenho_por_usuario = {}
            
            for approval in approvals_list:
                try:
                    if not approval.aprovado_por_id:
                        continue
                    usuario_id = approval.aprovado_por.id
                    usuario_nome = approval.aprovado_por.get_full_name() or approval.aprovado_por.username
                    
                    # Inicializar estrutura se não existir
                    if usuario_id not in desempenho_por_usuario:
                        desempenho_por_usuario[usuario_id] = {
                            'usuario': usuario_nome,
                            'tempos': [],
                            'total_decisoes': 0,
                            'aprovados': 0,
                            'reprovados': 0
                        }
                    
                    # Contar decisões
                    desempenho_por_usuario[usuario_id]['total_decisoes'] += 1
                    if approval.decisao == 'aprovado':
                        desempenho_por_usuario[usuario_id]['aprovados'] += 1
                    elif approval.decisao == 'reprovado':
                        desempenho_por_usuario[usuario_id]['reprovados'] += 1
                    
                    # Calcular tempo de resposta
                    # Usar data_envio se disponível, senão usar created_at do pedido
                    data_inicio_pedido = approval.work_order.data_envio or approval.work_order.created_at
                    data_fim = approval.created_at  # Data da decisão
                    
                    if data_inicio_pedido and data_fim:
                        tempo_resposta = (data_fim - data_inicio_pedido).total_seconds() / 3600  # Em horas
                        
                        # Ignorar tempos negativos (erro de dados)
                        if tempo_resposta >= 0:
                            desempenho_por_usuario[usuario_id]['tempos'].append(tempo_resposta)
                except Exception as e:
                    # Continuar processamento mesmo se houver erro em um registro
                    continue
            
            return desempenho_por_usuario
        
        # SLA padrão: 24 horas
        SLA_HORAS = 24
        LIMITE_CRITICO_HORAS = 24
        
        # Buscar pedidos recebidos (não apenas aprovados) para calcular capacidade
        def buscar_pedidos_recebidos(data_inicio_periodo, data_fim_periodo=None):
            if data_fim_periodo is None:
                data_fim_periodo = agora
            
            if is_responsavel_empresa(request.user) and not is_admin(request.user):
                empresas_ids = list(Empresa.objects.filter(
                    responsavel=request.user
                ).values_list('id', flat=True))
                if not empresas_ids:
                    return WorkOrder.objects.none()
                return WorkOrder.objects.filter(
                    data_envio__gte=data_inicio_periodo,
                    data_envio__lte=data_fim_periodo,
                    obra__empresa_id__in=empresas_ids
                ).select_related('criado_por', 'obra')
            else:
                return WorkOrder.objects.filter(
                    data_envio__gte=data_inicio_periodo,
                    data_envio__lte=data_fim_periodo
                ).select_related('criado_por', 'obra')
        
        # Buscar StatusHistory para rastrear retrabalho
        def buscar_retrabalho(data_inicio_periodo, data_fim_periodo=None):
            if data_fim_periodo is None:
                data_fim_periodo = agora
            
            if is_responsavel_empresa(request.user) and not is_admin(request.user):
                empresas_ids = list(Empresa.objects.filter(
                    responsavel=request.user
                ).values_list('id', flat=True))
                if not empresas_ids:
                    return StatusHistory.objects.none()
                return StatusHistory.objects.filter(
                    created_at__gte=data_inicio_periodo,
                    created_at__lte=data_fim_periodo,
                    work_order__obra__empresa_id__in=empresas_ids,
                    status_anterior='reprovado',
                    status_novo='reaprovacao'
                ).select_related('work_order')
            else:
                return StatusHistory.objects.filter(
                    created_at__gte=data_inicio_periodo,
                    created_at__lte=data_fim_periodo,
                    status_anterior='reprovado',
                    status_novo='reaprovacao'
                ).select_related('work_order')
        
        # Processar período atual
        desempenho_por_usuario = processar_aprovacoes(approvals)
        pedidos_recebidos = buscar_pedidos_recebidos(data_inicio_atual)
        retrabalhos = buscar_retrabalho(data_inicio_atual)
        
        # Processar período de 7 dias para tendência
        data_inicio_7d = agora - timedelta(days=7)
        desempenho_7d = processar_aprovacoes(buscar_aprovacoes(data_inicio_7d))
        
        # Processar período anterior para comparação
        desempenho_por_usuario_anterior = processar_aprovacoes(approvals_anterior)
        
        # Calcular estatísticas acionáveis
        resultado = []
        for usuario_id, dados in desempenho_por_usuario.items():
            if not dados['tempos']:
                continue
                
            tempos = dados['tempos']
            n = len(tempos)
            tempo_medio = sum(tempos) / n
            
            # 1. MÉTRICAS DE SLA
            fora_sla = sum(1 for t in tempos if t > SLA_HORAS)
            acima_critico = sum(1 for t in tempos if t > LIMITE_CRITICO_HORAS)
            pct_fora_sla = round((fora_sla / n * 100) if n > 0 else 0, 1)
            pct_acima_critico = round((acima_critico / n * 100) if n > 0 else 0, 1)
            
            # 2. CAPACIDADE DO APROVADOR
            # BMs aprovadas/dia no período
            dias_periodo_float = max(dias_periodo, 1)
            bms_aprovadas_dia = round(dados['aprovados'] / dias_periodo_float, 2)
            
            # BMs recebidas/dia (pedidos que chegaram para este aprovador)
            # Contar pedidos que foram aprovados/reprovados por este usuário
            bms_recebidas = dados['total_decisoes']
            bms_recebidas_dia = round(bms_recebidas / dias_periodo_float, 2)
            
            # Backlog implícito (diferença entre recebidas e processadas)
            backlog = max(0, bms_recebidas - dados['aprovados'])
            saldo_capacidade = round(bms_aprovadas_dia - bms_recebidas_dia, 2)
            
            # 3. TAXA DE REPROVAÇÃO NORMALIZADA
            # Reprovações por BM aprovada
            taxa_reprovacao_normalizada = round((dados['reprovados'] / dados['aprovados']) if dados['aprovados'] > 0 else dados['reprovados'], 2)
            
            # 4. RETRABALHO
            # Contar BMs que foram reprovadas e depois reapresentadas
            work_orders_aprovados = set()
            for approval in approvals:
                if approval.aprovado_por_id and approval.aprovado_por.id == usuario_id and approval.decisao == 'aprovado':
                    work_orders_aprovados.add(approval.work_order.id)
            
            retrabalhos_aprovador = 0
            for retrabalho in retrabalhos:
                # Verificar se este pedido foi aprovado por este aprovador após retrabalho
                if retrabalho.work_order.id in work_orders_aprovados:
                    # Verificar se houve reprovação anterior por este aprovador
                    reprovacoes_anteriores = Approval.objects.filter(
                        work_order=retrabalho.work_order,
                        aprovado_por_id=usuario_id,
                        decisao='reprovado',
                        created_at__lt=retrabalho.created_at
                    ).exists()
                    if reprovacoes_anteriores:
                        retrabalhos_aprovador += 1
            
            # 5. TENDÊNCIA TEMPORAL (7d vs 30d)
            dados_7d = desempenho_7d.get(usuario_id, {})
            tempos_7d = dados_7d.get('tempos', [])
            tempo_medio_7d = sum(tempos_7d) / len(tempos_7d) if tempos_7d and len(tempos_7d) > 0 else None
            if tempo_medio_7d and tempo_medio_7d > 0:
                variacao_tendencia = round(((tempo_medio - tempo_medio_7d) / tempo_medio_7d * 100), 1)
            else:
                variacao_tendencia = None
            
            # 6. ÍNDICE SINTÉTICO DE RISCO OPERACIONAL
            # Combina: atraso (tempo médio vs SLA), reprovação e violação de SLA
            # Normalizar cada componente de 0 a 1
            risco_atraso = min(1.0, tempo_medio / SLA_HORAS)  # 1.0 se tempo médio >= SLA
            risco_reprovacao = min(1.0, taxa_reprovacao_normalizada / 2.0)  # 1.0 se taxa >= 2.0
            risco_sla = pct_fora_sla / 100.0  # Já é 0-1
            
            # Peso dos componentes (ajustável)
            peso_atraso = 0.3
            peso_reprovacao = 0.3
            peso_sla = 0.4
            
            indice_risco = round(
                (risco_atraso * peso_atraso + 
                 risco_reprovacao * peso_reprovacao + 
                 risco_sla * peso_sla) * 100, 1
            )
            
            # Classificação de risco
            if indice_risco < 30:
                classificacao_risco = 'baixo'
                cor_risco = '#28a745'
            elif indice_risco < 60:
                classificacao_risco = 'moderado'
                cor_risco = '#ffc107'
            else:
                classificacao_risco = 'alto'
                cor_risco = '#dc3545'
            
            # 7. DIAGNÓSTICO (linguagem simples para quem olha rápido)
            diagnosticos = []
            if tempo_medio > SLA_HORAS:
                diagnosticos.append(f"Está demorando em média mais de {SLA_HORAS}h para decidir (meta é {SLA_HORAS}h)")
            if pct_fora_sla > 30:
                diagnosticos.append(f"{pct_fora_sla}% das decisões atrasaram (passaram de 24h)")
            if saldo_capacidade < 0:
                diagnosticos.append("Está chegando mais pedido do que dá para analisar — a fila está crescendo")
            if variacao_tendencia and variacao_tendencia > 20:
                diagnosticos.append("Nos últimos 7 dias está demorando mais que antes")
            if taxa_reprovacao_normalizada > 1.0:
                diagnosticos.append(f"Mais reprovações que aprovações (proporção {taxa_reprovacao_normalizada:.1f})")
            if retrabalhos_aprovador > dados['aprovados'] * 0.2:
                diagnosticos.append("Vários pedidos foram reapresentados para nova análise")
            
            if not diagnosticos:
                diagnostico_texto = "Tudo dentro do esperado no período"
            else:
                diagnostico_texto = ". ".join(diagnosticos)
            
            resultado.append({
                'usuario': dados['usuario'],
                'usuario_id': usuario_id,
                # Métricas principais
                'tempo_medio_horas': round(tempo_medio, 2),
                'sla_horas': SLA_HORAS,
                'pct_fora_sla': pct_fora_sla,
                'pct_acima_critico': pct_acima_critico,
                # Capacidade
                'bms_aprovadas_dia': bms_aprovadas_dia,
                'bms_recebidas_dia': bms_recebidas_dia,
                'saldo_capacidade': saldo_capacidade,
                'backlog': backlog,
                # Qualidade
                'taxa_reprovacao_normalizada': taxa_reprovacao_normalizada,
                'retrabalhos': retrabalhos_aprovador,
                # Tendência
                'tempo_medio_7d': round(tempo_medio_7d, 2) if tempo_medio_7d else None,
                'variacao_tendencia': variacao_tendencia,
                # Risco
                'indice_risco': indice_risco,
                'classificacao_risco': classificacao_risco,
                'cor_risco': cor_risco,
                # Diagnóstico
                'diagnostico': diagnostico_texto,
                # Dados auxiliares
                'total_decisoes': dados['total_decisoes'],
                'aprovados': dados['aprovados'],
                'reprovados': dados['reprovados'],
            })
        
        # Ordenar por índice de risco (maior primeiro = pior desempenho)
        resultado.sort(key=lambda x: x['indice_risco'], reverse=True)
        
        # Calcular estatísticas gerais
        tempo_medio_geral = 0
        total_decisoes = 0
        total_aprovados = 0
        total_reprovados = 0
        total_fora_sla = 0
        total_acima_critico = 0
        
        if resultado:
            for item in resultado:
                tempo_medio_geral += item['tempo_medio_horas'] * item['total_decisoes']
                total_decisoes += item['total_decisoes']
                total_aprovados += item['aprovados']
                total_reprovados += item['reprovados']
                total_fora_sla += round(item['pct_fora_sla'] * item['total_decisoes'] / 100)
                total_acima_critico += round(item['pct_acima_critico'] * item['total_decisoes'] / 100)
            
            if total_decisoes > 0:
                tempo_medio_geral = round(tempo_medio_geral / total_decisoes, 2)
        
        pct_fora_sla_geral = round((total_fora_sla / total_decisoes * 100) if total_decisoes > 0 else 0, 1)
        pct_acima_critico_geral = round((total_acima_critico / total_decisoes * 100) if total_decisoes > 0 else 0, 1)
        taxa_aprovacao_geral = round((total_aprovados / total_decisoes * 100) if total_decisoes > 0 else 0, 1)
        
        # Determinar período texto
        periodo_texto = f'Últimos {dias_periodo} dias'
        
        return JsonResponse({
            'dados': resultado,
            'tempo_medio_geral': tempo_medio_geral,
            'sla_horas': SLA_HORAS,
            'pct_fora_sla_geral': pct_fora_sla_geral,
            'pct_acima_critico_geral': pct_acima_critico_geral,
            'total_decisoes': total_decisoes,
            'total_aprovados': total_aprovados,
            'total_reprovados': total_reprovados,
            'taxa_aprovacao_geral': taxa_aprovacao_geral,
            'periodo': periodo_texto,
            'dias_periodo': dias_periodo,
        })
    except Exception as e:
        # Log do erro para debug
        import traceback
        error_trace = traceback.format_exc()
        logger.error(f"Erro em desempenho_equipe_api: {str(e)}", exc_info=True)
        
        # Retornar erro em formato JSON
        return JsonResponse({
            'erro': f'Erro ao processar dados: {str(e)}',
            'dados': [],
            'tempo_medio_geral': 0,
            'total_pedidos': 0,
            'periodo': 'Últimos 30 dias'
        }, status=500)


@login_required
def desempenho_solicitantes_api(request):
    """
    API endpoint para retornar dados de desempenho dos solicitantes (reprovações com tags).
    Exclusivo para administradores e responsáveis pela empresa.
    Retorna JSON com análise de reprovações por solicitante, incluindo tags mais frequentes.
    """
    # Verificar permissão: apenas admin ou responsável pela empresa
    if not (is_admin(request.user) or is_responsavel_empresa(request.user)):
        return JsonResponse({
            'erro': 'Você não tem permissão para acessar esta página.',
            'dados': [],
            'total_reprovacoes': 0,
            'periodo': 'Últimos 30 dias'
        }, status=403)
    
    try:
        # Obter parâmetros da requisição
        dias_periodo = int(request.GET.get('dias', 30))
        if dias_periodo not in [7, 15, 30, 60, 90]:
            dias_periodo = 30
        
        tipo_solicitacao = request.GET.get('tipo_solicitacao', None)  # Filtro opcional por tipo
        
        # Período atual
        agora = timezone.now()
        data_inicio = agora - timedelta(days=dias_periodo)
        
        # Buscar reprovações do período
        reprovacoes = Approval.objects.filter(
            decisao='reprovado',
            created_at__gte=data_inicio,
            created_at__lte=agora
        ).select_related('work_order', 'work_order__criado_por', 'work_order__obra')
        
        # Filtrar por tipo de solicitação se fornecido
        if tipo_solicitacao:
            reprovacoes = reprovacoes.filter(work_order__tipo_solicitacao=tipo_solicitacao)
        
        # Filtrar por empresa se for responsável (não admin)
        if is_responsavel_empresa(request.user) and not is_admin(request.user):
            empresas_ids = list(Empresa.objects.filter(
                responsavel=request.user
            ).values_list('id', flat=True))
            
            if empresas_ids:
                reprovacoes = reprovacoes.filter(work_order__obra__empresa_id__in=empresas_ids)
            else:
                reprovacoes = Approval.objects.none()
        
        # Carregar tags para evitar N+1 queries
        reprovacoes = reprovacoes.prefetch_related('tags_erro')
        
        # PRIMEIRO: Buscar TODOS os solicitantes que criaram pedidos no período
        # Isso garante que apareçam todos, mesmo os que não tiveram reprovações
        pedidos_periodo = WorkOrder.objects.filter(
            created_at__gte=data_inicio,
            created_at__lte=agora
        ).select_related('criado_por')
        
        if tipo_solicitacao:
            pedidos_periodo = pedidos_periodo.filter(tipo_solicitacao=tipo_solicitacao)
        
        if is_responsavel_empresa(request.user) and not is_admin(request.user):
            empresas_ids = list(Empresa.objects.filter(
                responsavel=request.user
            ).values_list('id', flat=True))
            if empresas_ids:
                pedidos_periodo = pedidos_periodo.filter(obra__empresa_id__in=empresas_ids)
            else:
                pedidos_periodo = WorkOrder.objects.none()
        
        # Inicializar estrutura para TODOS os solicitantes que criaram pedidos
        dados_por_solicitante = {}
        total_pedidos_por_solicitante = {}
        
        for pedido in pedidos_periodo:
            if not pedido.criado_por:
                continue
                
            solicitante_id = pedido.criado_por.id
            solicitante_nome = pedido.criado_por.get_full_name() or pedido.criado_por.username
            
            # Inicializar estrutura do solicitante (mesmo sem reprovações)
            if solicitante_id not in dados_por_solicitante:
                dados_por_solicitante[solicitante_id] = {
                    'solicitante': solicitante_nome,
                    'solicitante_id': solicitante_id,
                    'total_reprovacoes': 0,
                    'reprovacoes_por_tipo': {},
                    'tags_contagem': {},  # {tag_id: {'nome': '...', 'count': X}}
                    'tags_por_tipo': {}   # {tipo_solicitacao: {tag_id: count}}
                }
            
            # Contar total de pedidos
            if solicitante_id not in total_pedidos_por_solicitante:
                total_pedidos_por_solicitante[solicitante_id] = 0
            total_pedidos_por_solicitante[solicitante_id] += 1
        
        # SEGUNDO: Processar reprovações e preencher dados
        tags_geral = {}  # Contador global de tags
        
        for reprovacao in reprovacoes:
            solicitante = reprovacao.work_order.criado_por
            if not solicitante:
                continue
            
            solicitante_id = solicitante.id
            tipo_sol = reprovacao.work_order.tipo_solicitacao
            
            # Se o solicitante não está na lista (não deveria acontecer, mas por segurança)
            if solicitante_id not in dados_por_solicitante:
                solicitante_nome = solicitante.get_full_name() or solicitante.username
                dados_por_solicitante[solicitante_id] = {
                    'solicitante': solicitante_nome,
                    'solicitante_id': solicitante_id,
                    'total_reprovacoes': 0,
                    'reprovacoes_por_tipo': {},
                    'tags_contagem': {},
                    'tags_por_tipo': {}
                }
            
            dados = dados_por_solicitante[solicitante_id]
            dados['total_reprovacoes'] += 1
            
            # Contar por tipo de solicitação
            if tipo_sol not in dados['reprovacoes_por_tipo']:
                dados['reprovacoes_por_tipo'][tipo_sol] = 0
            dados['reprovacoes_por_tipo'][tipo_sol] += 1
            
            # Processar tags desta reprovação
            tags_desta_reprovacao = reprovacao.tags_erro.all()
            for tag in tags_desta_reprovacao:
                tag_id = tag.id
                tag_nome = tag.nome
                
                # Contagem geral de tags do solicitante
                if tag_id not in dados['tags_contagem']:
                    dados['tags_contagem'][tag_id] = {
                        'nome': tag_nome,
                        'count': 0
                    }
                dados['tags_contagem'][tag_id]['count'] += 1
                
                # Contagem por tipo de solicitação
                if tipo_sol not in dados['tags_por_tipo']:
                    dados['tags_por_tipo'][tipo_sol] = {}
                if tag_id not in dados['tags_por_tipo'][tipo_sol]:
                    dados['tags_por_tipo'][tipo_sol][tag_id] = {
                        'nome': tag_nome,
                        'count': 0
                    }
                dados['tags_por_tipo'][tipo_sol][tag_id]['count'] += 1
                
                # Contagem global de tags (para top geral)
                if tag_id not in tags_geral:
                    tags_geral[tag_id] = {
                        'nome': tag_nome,
                        'tipo_solicitacao': tipo_sol,
                        'count': 0
                    }
                tags_geral[tag_id]['count'] += 1
        
        # Calcular métricas de tempo por solicitante
        def calcular_metricas_tempo(solicitante_id):
            """
            Calcula métricas de tempo úteis para avaliar o desempenho do solicitante:
            - Tempo para corrigir: quanto tempo leva para reenviar após reprovação
            - Tempo total até aprovação: tempo desde criação até aprovação final (incluindo correções)
            """
            metricas = {
                'tempo_medio_para_corrigir': None,  # Tempo médio entre reprovação e reenvio (mostra proatividade)
                'tempo_medio_total_aprovacao': None,  # Tempo médio total desde criação até aprovação final
                'tempos_para_corrigir': [],
                'tempos_total_aprovacao': [],
            }
            
            # Buscar todos os pedidos do solicitante no período
            pedidos_solicitante = WorkOrder.objects.filter(
                criado_por_id=solicitante_id,
                created_at__gte=data_inicio,
                created_at__lte=agora
            ).select_related('obra').prefetch_related('approvals')
            
            if tipo_solicitacao:
                pedidos_solicitante = pedidos_solicitante.filter(tipo_solicitacao=tipo_solicitacao)
            
            if is_responsavel_empresa(request.user) and not is_admin(request.user):
                empresas_ids = list(Empresa.objects.filter(
                    responsavel=request.user
                ).values_list('id', flat=True))
                if empresas_ids:
                    pedidos_solicitante = pedidos_solicitante.filter(obra__empresa_id__in=empresas_ids)
            
            tempos_para_corrigir = []  # Tempo entre reprovação e reenvio
            tempos_total_aprovacao = []  # Tempo total desde criação até aprovação final
            
            for pedido in pedidos_solicitante:
                # Buscar aprovações do pedido (pode ter múltiplas se foi reprovado e reaprovado)
                approvals = pedido.approvals.all().order_by('created_at')
                
                if not approvals.exists():
                    continue
                
                # 1. TEMPO PARA CORRIGIR: Tempo entre reprovação e reenvio
                # Isso mostra se o solicitante está sendo proativo em corrigir erros
                if approvals.count() > 1:
                    for i in range(1, len(approvals)):
                        aprova_anterior = approvals[i-1]
                        aprova_atual = approvals[i]
                        
                        if aprova_anterior.decisao == 'reprovado':
                            # Buscar quando o pedido foi reenviado (StatusHistory)
                            # Pode haver múltiplos StatusHistory se foi reenviado várias vezes
                            # Precisamos pegar o que corresponde a esta reprovação específica
                            status_hist = StatusHistory.objects.filter(
                                work_order=pedido,
                                status_anterior='reprovado',
                                status_novo='reaprovacao',
                                created_at__gte=aprova_anterior.created_at,  # Depois da reprovação
                                created_at__lte=aprova_atual.created_at if aprova_atual else timezone.now()  # Antes da próxima aprovação ou agora
                            ).order_by('created_at').first()
                            
                            if status_hist:
                                # Tempo entre reprovação e reenvio (quando mudou para reaprovacao)
                                tempo_corrigir = (status_hist.created_at - aprova_anterior.created_at).total_seconds() / 3600
                                # Validar: tempo deve ser razoável (entre 0.1h e 720h = 30 dias)
                                # Ignorar valores muito pequenos (< 0.1h = 6 minutos) que podem ser erros de dados
                                if 0.1 <= tempo_corrigir <= 720:
                                    tempos_para_corrigir.append(tempo_corrigir)
                            elif aprova_atual and aprova_atual.decisao == 'aprovado' and aprova_atual.created_at > aprova_anterior.created_at:
                                # Fallback para dados antigos: usar diferença entre reprovação e próxima aprovação
                                # Mas apenas se a próxima aprovação for depois da reprovação
                                # Isso pode não ser 100% preciso, mas é melhor que nada
                                tempo_corrigir = (aprova_atual.created_at - aprova_anterior.created_at).total_seconds() / 3600
                                # Validar: tempo deve ser razoável (entre 0.1h e 720h = 30 dias)
                                if 0.1 <= tempo_corrigir <= 720:
                                    tempos_para_corrigir.append(tempo_corrigir)
                            elif pedido.data_envio and pedido.data_envio > aprova_anterior.created_at:
                                # Último fallback: usar data_envio se existir e for depois da reprovação
                                # Isso é menos preciso, mas melhor que nada para dados muito antigos
                                tempo_corrigir = (pedido.data_envio - aprova_anterior.created_at).total_seconds() / 3600
                                # Validar: tempo deve ser razoável (entre 0.1h e 720h = 30 dias)
                                if 0.1 <= tempo_corrigir <= 720:
                                    tempos_para_corrigir.append(tempo_corrigir)
                
                # 2. TEMPO TOTAL ATÉ APROVAÇÃO: Tempo desde criação até aprovação final
                # Isso mostra quanto tempo total leva para o pedido ser aprovado (incluindo correções)
                if pedido.status == 'aprovado':
                    # Buscar a última aprovação (não reprovação) para ter a data correta
                    # Isso é mais confiável que data_aprovacao que pode estar incorreta
                    ultima_aprovacao_aprovada = approvals.filter(decisao='aprovado').order_by('created_at').last()
                    
                    if ultima_aprovacao_aprovada:
                        # Usar data da última aprovação (mais confiável)
                        data_aprovacao_final = ultima_aprovacao_aprovada.created_at
                    elif pedido.data_aprovacao:
                        # Fallback: usar data_aprovacao se não houver Approval
                        data_aprovacao_final = pedido.data_aprovacao
                    else:
                        # Se não tem nenhuma data, pular este pedido
                        data_aprovacao_final = None
                    
                    if data_aprovacao_final and pedido.created_at:
                        tempo_total = (data_aprovacao_final - pedido.created_at).total_seconds() / 3600
                        # Validar: tempo deve ser razoável (mínimo 0h, máximo 2160h = 90 dias)
                        # Permitir 0h se realmente foi aprovado na mesma hora (pode acontecer)
                        if 0 <= tempo_total <= 2160:
                            tempos_total_aprovacao.append(tempo_total)
            
            # Calcular médias (com tratamento de outliers para dados mais confiáveis)
            if tempos_para_corrigir:
                # Remover outliers se houver muitos dados
                tempos_validos = tempos_para_corrigir
                if len(tempos_para_corrigir) > 3:
                    import statistics
                    try:
                        media = statistics.mean(tempos_para_corrigir)
                        if len(tempos_para_corrigir) > 1:
                            desvio = statistics.stdev(tempos_para_corrigir)
                            # Remover valores que estão muito fora (mais de 2 desvios padrão)
                            tempos_validos = [t for t in tempos_para_corrigir if abs(t - media) <= 2 * desvio]
                            if not tempos_validos:  # Se todos foram removidos, usar todos
                                tempos_validos = tempos_para_corrigir
                    except Exception:
                        tempos_validos = tempos_para_corrigir
                
                metricas['tempo_medio_para_corrigir'] = round(sum(tempos_validos) / len(tempos_validos), 2)
                metricas['tempos_para_corrigir'] = sorted(tempos_para_corrigir)
            
            if tempos_total_aprovacao:
                # Remover outliers se houver muitos dados
                tempos_validos = tempos_total_aprovacao
                if len(tempos_total_aprovacao) > 3:
                    import statistics
                    try:
                        media = statistics.mean(tempos_total_aprovacao)
                        if len(tempos_total_aprovacao) > 1:
                            desvio = statistics.stdev(tempos_total_aprovacao)
                            tempos_validos = [t for t in tempos_total_aprovacao if abs(t - media) <= 2 * desvio]
                            if not tempos_validos:
                                tempos_validos = tempos_total_aprovacao
                    except Exception:
                        tempos_validos = tempos_total_aprovacao
                
                metricas['tempo_medio_total_aprovacao'] = round(sum(tempos_validos) / len(tempos_validos), 2)
                metricas['tempos_total_aprovacao'] = sorted(tempos_total_aprovacao)
            
            return metricas
        
        # Converter para lista e ordenar por total de reprovações (maior primeiro)
        resultado = []
        for solicitante_id, dados in dados_por_solicitante.items():
            # Top 5 tags do solicitante
            tags_list = list(dados['tags_contagem'].values())
            tags_list.sort(key=lambda x: x['count'], reverse=True)
            top_tags = tags_list[:5]
            
            # Top tags por tipo
            top_tags_por_tipo = {}
            for tipo_sol, tags_dict in dados['tags_por_tipo'].items():
                tags_list_tipo = list(tags_dict.values())
                tags_list_tipo.sort(key=lambda x: x['count'], reverse=True)
                top_tags_por_tipo[tipo_sol] = tags_list_tipo[:3]  # Top 3 por tipo
            
            # Calcular métricas adicionais
            total_pedidos = total_pedidos_por_solicitante.get(solicitante_id, 0)
            # Taxa de erro: reprovações / total de pedidos (máximo 100%)
            if total_pedidos > 0:
                taxa_erro = round((dados['total_reprovacoes'] / total_pedidos * 100), 1)
                # Garantir que não ultrapasse 100% (caso haja dados inconsistentes)
                taxa_erro = min(taxa_erro, 100.0)
            else:
                taxa_erro = 0.0
            total_tags = sum(tag['count'] for tag in dados['tags_contagem'].values())
            media_tags_por_reprovacao = round(total_tags / dados['total_reprovacoes'], 1) if dados['total_reprovacoes'] > 0 else 0
            
            # Calcular métricas de tempo
            metricas_tempo = calcular_metricas_tempo(solicitante_id)
            
            # Calcular taxa de aprovação (pedidos aprovados / total de pedidos)
            pedidos_aprovados = WorkOrder.objects.filter(
                criado_por_id=solicitante_id,
                created_at__gte=data_inicio,
                created_at__lte=agora,
                status='aprovado'
            )
            if tipo_solicitacao:
                pedidos_aprovados = pedidos_aprovados.filter(tipo_solicitacao=tipo_solicitacao)
            if is_responsavel_empresa(request.user) and not is_admin(request.user):
                empresas_ids = list(Empresa.objects.filter(
                    responsavel=request.user
                ).values_list('id', flat=True))
                if empresas_ids:
                    pedidos_aprovados = pedidos_aprovados.filter(obra__empresa_id__in=empresas_ids)
            
            total_aprovados = pedidos_aprovados.count()
            taxa_aprovacao = round((total_aprovados / total_pedidos * 100), 1) if total_pedidos > 0 else 0.0
            
            resultado.append({
                'solicitante': dados['solicitante'],
                'solicitante_id': dados['solicitante_id'],
                'total_reprovacoes': dados['total_reprovacoes'],
                'total_pedidos': total_pedidos,
                'total_aprovados': total_aprovados,
                'taxa_erro': taxa_erro,
                'taxa_aprovacao': taxa_aprovacao,
                'media_tags_por_reprovacao': media_tags_por_reprovacao,
                'reprovacoes_por_tipo': dados['reprovacoes_por_tipo'],
                'top_tags': top_tags,
                'top_tags_por_tipo': top_tags_por_tipo,
                'tempo_medio_para_corrigir': metricas_tempo.get('tempo_medio_para_corrigir'),
                'tempo_medio_total_aprovacao': metricas_tempo.get('tempo_medio_total_aprovacao')
            })
        
        # Ordenar por total de reprovações (maior primeiro)
        resultado.sort(key=lambda x: x['total_reprovacoes'], reverse=True)
        
        # Adicionar ranking
        for idx, item in enumerate(resultado, 1):
            item['ranking'] = idx
        
        # Top tags geral (todas as tags mais frequentes)
        top_tags_geral = list(tags_geral.values())
        top_tags_geral.sort(key=lambda x: x['count'], reverse=True)
        top_tags_geral = top_tags_geral[:10]  # Top 10 tags gerais
        
        # Estatísticas gerais
        total_reprovacoes_geral = sum(d['total_reprovacoes'] for d in resultado)
        
        # Calcular total de pedidos geral (para taxa de reprovação)
        total_pedidos_geral = pedidos_periodo.count()
        taxa_reprovacao_geral = round((total_reprovacoes_geral / total_pedidos_geral * 100) if total_pedidos_geral > 0 else 0, 1)
        
        # Contagem por tipo de solicitação (geral)
        reprovacoes_por_tipo_geral = {}
        for item in resultado:
            for tipo, count in item['reprovacoes_por_tipo'].items():
                if tipo not in reprovacoes_por_tipo_geral:
                    reprovacoes_por_tipo_geral[tipo] = 0
                reprovacoes_por_tipo_geral[tipo] += count
        
        # Determinar período texto
        periodo_texto = f'Últimos {dias_periodo} dias'
        
        return JsonResponse({
            'dados': resultado,
            'total_reprovacoes': total_reprovacoes_geral,
            'total_pedidos_geral': total_pedidos_geral,
            'taxa_reprovacao_geral': taxa_reprovacao_geral,
            'reprovacoes_por_tipo_geral': reprovacoes_por_tipo_geral,
            'top_tags_geral': top_tags_geral,
            'periodo': periodo_texto,
            'dias_periodo': dias_periodo,
            'tipo_solicitacao_filtro': tipo_solicitacao
        })
    except Exception as e:
        logger.error(f"Erro em desempenho_solicitantes_api: {str(e)}", exc_info=True)
        
        return JsonResponse({
            'erro': f'Erro ao processar dados: {str(e)}',
            'dados': [],
            'total_reprovacoes': 0,
            'periodo': 'Últimos 30 dias'
        }, status=500)


@login_required
def add_comment(request, pk):
    """
    Adiciona um comentário a um pedido de obra.
    Permite comunicação entre solicitantes e aprovadores durante a análise.
    """
    workorder = get_object_or_404(WorkOrder, pk=pk)
    user = request.user
    
    # Verificar permissão de visualização (mesma lógica do detail_workorder)
    tem_permissao = False
    if is_admin(user):
        tem_permissao = True
    elif is_aprovador(user):
        if workorder.obra.empresa_id is None:
            tem_permissao = WorkOrderPermission.objects.filter(
                obra=workorder.obra, usuario=user, tipo_permissao='aprovador', ativo=True
            ).exists()
        else:
            empresas_ids = Empresa.objects.filter(
                obras__permissoes__usuario=user,
                obras__permissoes__tipo_permissao='aprovador',
                obras__permissoes__ativo=True
            ).values_list('id', flat=True).distinct()
            tem_permissao = workorder.obra.empresa_id in empresas_ids
    elif is_engenheiro(user):
        if workorder.criado_por == user:
            tem_permissao = True
        else:
            tem_permissao_obra = WorkOrderPermission.objects.filter(
                obra=workorder.obra,
                usuario=user,
                tipo_permissao='solicitante',
                ativo=True
            ).exists()
            is_solicitante_group = user.groups.filter(name='Solicitante').exists()
            if tem_permissao_obra or is_solicitante_group:
                tem_permissao = True
    
    if not tem_permissao:
        messages.error(request, 'Você não tem permissão para comentar neste pedido.')
        return redirect('gestao:detail_workorder', pk=workorder.pk)
    
    if request.method == 'POST':
        texto = request.POST.get('texto', '').strip()
        
        if not texto:
            messages.error(request, 'O comentário não pode estar vazio.')
            return redirect('gestao:detail_workorder', pk=workorder.pk)
        
        # Criar comentário
        comment = Comment.objects.create(
            work_order=workorder,
            autor=user,
            texto=texto
        )
        
        # Criar notificações para os outros usuários envolvidos
        usuarios_notificar = set()
        
        # Notificar o criador do pedido (se não for quem comentou)
        if workorder.criado_por and workorder.criado_por != user:
            usuarios_notificar.add(workorder.criado_por)
        
        # Notificar aprovadores da obra (se não for quem comentou)
        if is_aprovador(user) or is_admin(user):
            # Se quem comentou é aprovador/admin, notificar o solicitante
            if workorder.criado_por and workorder.criado_por != user:
                usuarios_notificar.add(workorder.criado_por)
        else:
            # Se quem comentou é solicitante, notificar aprovadores
            aprovadores = WorkOrderPermission.objects.filter(
                obra=workorder.obra,
                tipo_permissao='aprovador',
                ativo=True
            ).select_related('usuario')
            
            for perm in aprovadores:
                if perm.usuario != user and perm.usuario.is_active:
                    usuarios_notificar.add(perm.usuario)
            
            # Também notificar admins
            admins = User.objects.filter(
                groups__name='Administrador',
                is_active=True
            ).exclude(id=user.id)
            
            for admin_user in admins:
                usuarios_notificar.add(admin_user)
        
        # Criar notificações
        autor_nome = user.get_full_name() or user.username
        for usuario_notificar in usuarios_notificar:
            criar_notificacao(
                usuario=usuario_notificar,
                tipo='comentario_adicionado',
                titulo=f'Novo Comentário: {workorder.codigo}',
                mensagem=f'{autor_nome} comentou no pedido {workorder.codigo}: {texto[:100]}{"..." if len(texto) > 100 else ""}',
                work_order=workorder
            )
        
        messages.success(request, 'Comentário adicionado com sucesso!')
        return redirect('gestao:detail_workorder', pk=workorder.pk)
    
    # GET - redirecionar para detalhes
    return redirect('gestao:detail_workorder', pk=workorder.pk)


def _gerar_csv_historico(solicitante, reprovacoes, dias_periodo, tipo_solicitacao, total_reprovacoes, tags_count):
    """Gera CSV melhorado e formatado com visual mais profissional."""
    output = io.StringIO()
    writer = csv.writer(output, delimiter=';', quotechar='"', quoting=csv.QUOTE_MINIMAL)
    
    # Cabeçalho do relatório (mais espaçado e organizado)
    writer.writerow(['=' * 100])
    writer.writerow(['RELATÓRIO DE HISTÓRICO DE REPROVAÇÕES - QUALIDADE DAS SOLICITAÇÕES'])
    writer.writerow(['=' * 100])
    writer.writerow([])
    writer.writerow([f'Solicitante: {solicitante.get_full_name() or solicitante.username}'])
    writer.writerow([f'E-mail: {solicitante.email or "Não informado"}'])
    writer.writerow([f'Período Analisado: Últimos {dias_periodo} dias'])
    if tipo_solicitacao:
        tipo_labels_header = {
            'contrato': 'Contrato',
            'medicao': 'Medição',
            'ordem_servico': 'Ordem de Serviço (OS)',
            'mapa_cotacao': 'Mapa de Cotação',
        }
        writer.writerow([f'Filtro de Tipo: {tipo_labels_header.get(tipo_solicitacao, tipo_solicitacao)}'])
    else:
        writer.writerow(['Filtro de Tipo: Todos os tipos'])
    writer.writerow([f'Data de Geração: {datetime.now().strftime("%d/%m/%Y às %H:%M")}'])
    writer.writerow([])
    writer.writerow(['RESUMO ESTATÍSTICO'])
    writer.writerow(['-' * 100])
    writer.writerow([f'Total de Reprovações: {total_reprovacoes}'])
    if tags_count:
        tag_mais_frequente = max(tags_count.items(), key=lambda x: x[1])
        writer.writerow([f'Tag Mais Frequente: {tag_mais_frequente[0]} ({tag_mais_frequente[1]} ocorrências)'])
    writer.writerow([])
    writer.writerow(['=' * 100])
    writer.writerow([])
    
    # Cabeçalho da tabela
    writer.writerow([
        'Data/Hora da Reprovação',
        'Código do Pedido',
        'Obra',
        'Tipo de Solicitação',
        'Nome do Credor',
        'Aprovador',
        'Tags de Erro',
        'Comentário do Aprovador',
        'Valor Estimado (R$)',
        'Prazo Estimado (dias)',
        'Data de Criação',
        'Data de Envio',
    ])
    
    # Dados formatados
    tipo_labels = {
        'contrato': 'Contrato',
        'medicao': 'Medição',
        'ordem_servico': 'Ordem de Serviço (OS)',
        'mapa_cotacao': 'Mapa de Cotação',
    }
    
    for reprovacao in reprovacoes:
        tags_nomes = ', '.join([tag.nome for tag in reprovacao.tags_erro.all()])
        if not tags_nomes:
            tags_nomes = 'Nenhuma tag registrada'
        
        valor_str = '-'
        if reprovacao.work_order.valor_estimado:
            valor_str = f"R$ {reprovacao.work_order.valor_estimado:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
        
        prazo_str = str(reprovacao.work_order.prazo_estimado) if reprovacao.work_order.prazo_estimado else '-'
        
        data_criacao = reprovacao.work_order.created_at.strftime('%d/%m/%Y %H:%M') if reprovacao.work_order.created_at else '-'
        data_envio = reprovacao.work_order.data_envio.strftime('%d/%m/%Y %H:%M') if reprovacao.work_order.data_envio else 'Não enviado'
        
        writer.writerow([
            reprovacao.created_at.strftime('%d/%m/%Y %H:%M'),
            reprovacao.work_order.codigo or '-',
            f"{reprovacao.work_order.obra.codigo} - {reprovacao.work_order.obra.nome}" if reprovacao.work_order.obra else '-',
            tipo_labels.get(reprovacao.work_order.tipo_solicitacao, reprovacao.work_order.tipo_solicitacao),
            reprovacao.work_order.nome_credor or '-',
            reprovacao.aprovado_por.get_full_name() or reprovacao.aprovado_por.username if reprovacao.aprovado_por else '-',
            tags_nomes,
            reprovacao.comentario or 'Sem comentário',
            valor_str,
            prazo_str,
            data_criacao,
            data_envio,
        ])
    
    output.seek(0)
    csv_content = output.getvalue()
    
    response = HttpResponse(
        csv_content.encode('utf-8-sig'),
        content_type='text/csv; charset=utf-8-sig'
    )
    nome_arquivo = f'historico_{solicitante.username}_{dias_periodo}dias_{datetime.now().strftime("%Y%m%d")}.csv'
    response['Content-Disposition'] = f'attachment; filename="{nome_arquivo}"'
    
    return response

def _gerar_pdf_historico(solicitante, reprovacoes, dias_periodo, tipo_solicitacao, total_reprovacoes, tags_count):
    """Gera PDF corporativo de histórico de reprovações (padrão visual LPLAN)."""
    from reportlab.platypus import Image as RLImage

    def _safe(v, default='-'):
        return v if v not in (None, '') else default

    def _money(v):
        if not v:
            return '-'
        return f"R$ {v:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')

    def _dt(v, default='-'):
        if not v:
            return default
        return v.strftime('%d/%m/%Y %H:%M')

    def _get_logo_path():
        base = settings.BASE_DIR
        logo_dir = os.path.join(base, 'core', 'static', 'core', 'images')
        for name in ('lplan-logo2.png', 'lplan_logo.png', 'lplan_logo.jpg', 'lplan_logo.jpeg'):
            p = os.path.join(logo_dir, name)
            if os.path.exists(p):
                return p
        return None

    color_primary = colors.HexColor('#1A3A5C')
    color_accent = colors.HexColor('#E8F0F8')
    color_border = colors.HexColor('#D0D9E3')
    color_text = colors.HexColor('#1C1C1C')
    color_sub = colors.HexColor('#5A5A5A')
    color_warn_bg = colors.HexColor('#FFF3E0')
    color_info_bg = colors.HexColor('#F4F8FC')

    tipo_labels = {
        'contrato': 'Contrato',
        'medicao': 'Medição',
        'ordem_servico': 'Ordem de Serviço (OS)',
        'mapa_cotacao': 'Mapa de Cotação',
    }
    recommendation_keywords = [
        (('document', 'anexo', 'arquivo', 'comprovante'), 'Reforçar checklist de anexos obrigatórios antes do envio.'),
        (('valor', 'preço', 'orc', 'cotacao', 'cotação'), 'Padronizar revisão de valores e cotações com dupla conferência.'),
        (('prazo', 'data', 'cronograma'), 'Conferir prazos e datas com o planejamento antes da submissão.'),
        (('fornecedor', 'credor', 'empresa'), 'Validar dados cadastrais de fornecedor/credor na abertura da solicitação.'),
        (('assinatura', 'assin'), 'Orientar assinatura e conferência final antes de concluir o fluxo.'),
    ]

    reprovacoes_list = list(reprovacoes)
    tipos_count = {}
    comentarios_registrados = 0
    for item in reprovacoes_list:
        tipo_label = tipo_labels.get(item.work_order.tipo_solicitacao, item.work_order.tipo_solicitacao or 'Não informado')
        tipos_count[tipo_label] = tipos_count.get(tipo_label, 0) + 1
        if item.comentario:
            comentarios_registrados += 1

    top_tags = sorted(tags_count.items(), key=lambda x: x[1], reverse=True)[:5]
    taxa_comentarios = round((comentarios_registrados / total_reprovacoes * 100), 1) if total_reprovacoes else 0
    media_diaria = round((total_reprovacoes / dias_periodo), 2) if dias_periodo else 0
    tag_frequente = f'{top_tags[0][0]} ({top_tags[0][1]}x)' if top_tags else '-'

    recomendacoes = []
    for tag, _count in top_tags:
        tag_lower = (tag or '').lower()
        for keywords, message in recommendation_keywords:
            if any(k in tag_lower for k in keywords) and message not in recomendacoes:
                recomendacoes.append(message)
                break
        if len(recomendacoes) >= 4:
            break
    if tipos_count:
        tipo_mais_critico, qtd_critica = max(tipos_count.items(), key=lambda x: x[1])
        recomendacoes.append(f'Priorizar treinamento do time em solicitações de "{tipo_mais_critico}" ({qtd_critica} reprovações no período).')
    if not recomendacoes:
        recomendacoes.append('Sem padrão crítico no período; manter monitoramento e feedback contínuo.')

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=1.7 * cm,
        leftMargin=1.7 * cm,
        topMargin=1.7 * cm,
        bottomMargin=1.5 * cm,
    )
    styles = getSampleStyleSheet()
    elements = []

    normal = ParagraphStyle('HistNormal', parent=styles['Normal'], fontName='Helvetica', fontSize=9.2, textColor=color_text, leading=12)
    label = ParagraphStyle('HistLabel', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=9, textColor=color_sub)
    h2 = ParagraphStyle('HistH2', parent=styles['Heading2'], fontName='Helvetica-Bold', fontSize=11, textColor=color_primary, spaceAfter=6)
    bullet = ParagraphStyle('HistBullet', parent=styles['Normal'], fontName='Helvetica', fontSize=9.2, textColor=color_text, leftIndent=10, bulletIndent=0, leading=12)
    table_head = ParagraphStyle('HistTH', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=8.5, textColor=colors.white, alignment=TA_CENTER)

    # Header com logo + identificação
    title_main = Paragraph("RELATÓRIO DE HISTÓRICO DE REPROVAÇÕES", ParagraphStyle('HistTitle', parent=styles['Heading1'], fontName='Helvetica-Bold', fontSize=13, textColor=colors.white, alignment=TA_LEFT))
    title_sub = Paragraph(
        f"<font color='white' size='9'>GesttControll · Solicitante: {_safe(solicitante.get_full_name() or solicitante.username)} · Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}</font>",
        ParagraphStyle('HistSub', parent=styles['Normal'], fontName='Helvetica', fontSize=9, textColor=colors.white),
    )

    logo_path = _get_logo_path()
    if logo_path:
        logo = RLImage(logo_path, width=2.2 * cm, height=1.0 * cm)
        text_block = Table([[title_main], [title_sub]], colWidths=[13.4 * cm])
        text_block.setStyle(TableStyle([
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
            ('TOPPADDING', (0, 0), (-1, -1), 0),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
        ]))
        header = Table([[logo, text_block]], colWidths=[2.6 * cm, 13.4 * cm])
    else:
        header = Table([[title_main], [title_sub]], colWidths=[16 * cm])
    header.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), color_primary),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(header)
    elements.append(Spacer(1, 0.28 * cm))

    tipo_desc = tipo_labels.get(tipo_solicitacao, tipo_solicitacao) if tipo_solicitacao else 'Todos os tipos'
    info_rows = [
        [Paragraph('<b>Solicitante</b>', label), Paragraph(_safe(solicitante.get_full_name() or solicitante.username), normal)],
        [Paragraph('<b>E-mail</b>', label), Paragraph(_safe(solicitante.email, 'Não informado'), normal)],
        [Paragraph('<b>Período analisado</b>', label), Paragraph(f'Últimos {dias_periodo} dias', normal)],
        [Paragraph('<b>Filtro de tipo</b>', label), Paragraph(tipo_desc, normal)],
    ]
    info_tbl = Table(info_rows, colWidths=[4.2 * cm, 11.8 * cm])
    info_tbl.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), color_accent),
        ('BOX', (0, 0), (-1, -1), 0.6, color_border),
        ('INNERGRID', (0, 0), (-1, -1), 0.25, color_border),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 7),
        ('RIGHTPADDING', (0, 0), (-1, -1), 7),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    elements.append(info_tbl)
    elements.append(Spacer(1, 0.28 * cm))

    elements.append(Paragraph('RESUMO EXECUTIVO', h2))
    resumo_tbl = Table([
        [
            Paragraph('<b>Total de reprovações</b><br/><font size="11"><b>%s</b></font>' % total_reprovacoes, normal),
            Paragraph('<b>Tag mais frequente</b><br/>%s' % _safe(tag_frequente), normal),
            Paragraph('<b>Média diária</b><br/><font size="11"><b>%s</b></font> reprovações/dia' % media_diaria, normal),
            Paragraph('<b>Com comentário</b><br/><font size="11"><b>%s%%</b></font>' % str(taxa_comentarios).replace('.', ','), normal),
        ]
    ], colWidths=[4 * cm, 4 * cm, 4 * cm, 4 * cm])
    resumo_tbl.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), color_warn_bg),
        ('BOX', (0, 0), (-1, -1), 0.6, color_border),
        ('INNERGRID', (0, 0), (-1, -1), 0.4, color_border),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('LEFTPADDING', (0, 0), (-1, -1), 7),
        ('RIGHTPADDING', (0, 0), (-1, -1), 7),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(resumo_tbl)
    elements.append(Spacer(1, 0.28 * cm))

    elements.append(Paragraph('TOP TAGS DE REPROVAÇÃO', h2))
    if top_tags:
        tags_data = [[Paragraph('Tag', table_head), Paragraph('Ocorrências', table_head)]]
        for tag_name, qtd in top_tags:
            tags_data.append([Paragraph(_safe(tag_name), normal), Paragraph(str(qtd), normal)])
        tags_tbl = Table(tags_data, colWidths=[12.2 * cm, 3.8 * cm], repeatRows=1)
        tags_tbl.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), color_primary),
            ('ALIGN', (1, 1), (1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOX', (0, 0), (-1, -1), 0.6, color_border),
            ('INNERGRID', (0, 0), (-1, -1), 0.25, color_border),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, color_info_bg]),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ]))
        elements.append(tags_tbl)
    else:
        elements.append(Paragraph('Sem tags registradas no período selecionado.', normal))
    elements.append(Spacer(1, 0.28 * cm))

    elements.append(Paragraph('REPROVAÇÕES POR TIPO DE SOLICITAÇÃO', h2))
    if tipos_count:
        tipos_data = [[Paragraph('Tipo', table_head), Paragraph('Quantidade', table_head)]]
        for tipo_nome, quantidade in sorted(tipos_count.items(), key=lambda x: x[1], reverse=True):
            tipos_data.append([Paragraph(_safe(tipo_nome), normal), Paragraph(str(quantidade), normal)])
        tipos_tbl = Table(tipos_data, colWidths=[12.2 * cm, 3.8 * cm], repeatRows=1)
        tipos_tbl.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), color_primary),
            ('ALIGN', (1, 1), (1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOX', (0, 0), (-1, -1), 0.6, color_border),
            ('INNERGRID', (0, 0), (-1, -1), 0.25, color_border),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, color_accent]),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ]))
        elements.append(tipos_tbl)
    else:
        elements.append(Paragraph('Sem dados por tipo para o período selecionado.', normal))
    elements.append(Spacer(1, 0.28 * cm))

    elements.append(Paragraph('AÇÕES RECOMENDADAS', h2))
    acoes_data = [[Paragraph('Plano de melhoria sugerido', table_head)]]
    for acao in recomendacoes[:5]:
        acoes_data.append([Paragraph(acao, bullet, bulletText='•')])
    acoes_tbl = Table(acoes_data, colWidths=[16 * cm], repeatRows=1)
    acoes_tbl.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, 0), color_primary),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('BOX', (0, 0), (-1, -1), 0.6, color_border),
        ('INNERGRID', (0, 1), (-1, -1), 0.25, color_border),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, color_info_bg]),
        ('LEFTPADDING', (0, 0), (-1, -1), 7),
        ('RIGHTPADDING', (0, 0), (-1, -1), 7),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(acoes_tbl)
    elements.append(Spacer(1, 0.32 * cm))

    elements.append(Paragraph('DETALHAMENTO DAS REPROVAÇÕES', h2))
    if total_reprovacoes <= 0:
        elements.append(Paragraph('Nenhuma reprovação encontrada no período selecionado.', normal))
    else:
        table_data = [[
            Paragraph('Data', table_head),
            Paragraph('Pedido/Tipo', table_head),
            Paragraph('Obra/Credor', table_head),
            Paragraph('Tags e Comentário', table_head),
        ]]
        for r in reprovacoes_list:
            wo = r.work_order
            obra = f"{wo.obra.codigo} - {wo.obra.nome}" if wo.obra else '-'
            aprovador = r.aprovado_por.get_full_name() or r.aprovado_por.username if r.aprovado_por else '-'
            tags = ', '.join([t.nome for t in r.tags_erro.all()]) or 'Sem tags'
            comentario = r.comentario or 'Sem comentário'

            col_a = Paragraph(_dt(r.created_at), normal)
            col_b = Paragraph(f"<b>{_safe(wo.codigo)}</b><br/>{tipo_labels.get(wo.tipo_solicitacao, wo.tipo_solicitacao)}<br/>Aprovador: {_safe(aprovador)}", normal)
            col_c = Paragraph(f"{_safe(obra)}<br/>Credor: {_safe(wo.nome_credor)}<br/>Valor: {_money(wo.valor_estimado)} | Prazo: {_safe(wo.prazo_estimado)} dia(s)", normal)
            col_d = Paragraph(f"<b>Tags:</b> {tags}<br/><b>Comentário:</b> {comentario}", normal)
            table_data.append([col_a, col_b, col_c, col_d])

        detail_tbl = Table(table_data, colWidths=[2.2 * cm, 4.1 * cm, 4.5 * cm, 5.2 * cm], repeatRows=1)
        detail_tbl.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), color_primary),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('BOX', (0, 0), (-1, -1), 0.6, color_border),
            ('INNERGRID', (0, 0), (-1, -1), 0.25, color_border),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, color_accent]),
            ('LEFTPADDING', (0, 0), (-1, -1), 5),
            ('RIGHTPADDING', (0, 0), (-1, -1), 5),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ]))
        elements.append(detail_tbl)

    doc.build(elements)
    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    nome_arquivo = f'historico_{solicitante.username}_{dias_periodo}dias_{datetime.now().strftime("%Y%m%d")}.pdf'
    response['Content-Disposition'] = f'attachment; filename="{nome_arquivo}"'
    return response

@login_required
def exportar_historico_solicitante(request, solicitante_id):
    """
    Exporta o histórico detalhado de reprovações de um solicitante em CSV ou PDF.
    Exclusivo para administradores e responsáveis pela empresa.
    """
    # Verificar permissão
    if not (is_admin(request.user) or is_responsavel_empresa(request.user)):
        messages.error(request, 'Você não tem permissão para exportar este relatório.')
        return redirect('gestao:desempenho_equipe')
    
    try:
        solicitante = get_object_or_404(User, id=solicitante_id)
        
        # Obter parâmetros
        dias_periodo = int(request.GET.get('dias', 30))
        if dias_periodo not in [7, 15, 30, 60, 90]:
            dias_periodo = 30
        
        tipo_solicitacao = request.GET.get('tipo_solicitacao', None)
        formato = request.GET.get('formato', 'csv')  # csv ou pdf
        
        # Período
        agora = timezone.now()
        data_inicio = agora - timedelta(days=dias_periodo)
        
        # Buscar reprovações do solicitante
        reprovacoes = Approval.objects.filter(
            decisao='reprovado',
            work_order__criado_por=solicitante,
            created_at__gte=data_inicio,
            created_at__lte=agora
        ).select_related('work_order', 'work_order__obra', 'aprovado_por').prefetch_related('tags_erro')
        
        # Filtrar por tipo se fornecido
        if tipo_solicitacao:
            reprovacoes = reprovacoes.filter(work_order__tipo_solicitacao=tipo_solicitacao)
        
        # Filtrar por empresa se for responsável (não admin)
        if is_responsavel_empresa(request.user) and not is_admin(request.user):
            empresas_ids = list(Empresa.objects.filter(
                responsavel=request.user
            ).values_list('id', flat=True))
            
            if empresas_ids:
                reprovacoes = reprovacoes.filter(work_order__obra__empresa_id__in=empresas_ids)
            else:
                reprovacoes = Approval.objects.none()
        
        total_reprovacoes = reprovacoes.count()
        
        # Calcular estatísticas
        tags_count = {}
        if total_reprovacoes > 0:
            for reprovacao in reprovacoes:
                for tag in reprovacao.tags_erro.all():
                    if tag.nome not in tags_count:
                        tags_count[tag.nome] = 0
                    tags_count[tag.nome] += 1
        
        # Escolher formato de exportação
        if formato.lower() == 'pdf':
            return _gerar_pdf_historico(
                solicitante, reprovacoes, dias_periodo, tipo_solicitacao, 
                total_reprovacoes, tags_count
            )
        else:
            return _gerar_csv_historico(
                solicitante, reprovacoes, dias_periodo, tipo_solicitacao,
                total_reprovacoes, tags_count
            )
    
    except Exception as e:
        messages.error(request, f'Erro ao exportar relatório: {str(e)}')
        return redirect('gestao:desempenho_equipe')


def serve_media_file(request, path):
    """
    View para servir arquivos media em produção (quando DEBUG=False).
    Trata encoding corretamente para arquivos com caracteres especiais.
    """
    from django.conf import settings
    from django.http import FileResponse, Http404
    from django.views.decorators.cache import cache_control
    import os
    
    # Construir caminho completo do arquivo
    file_path = os.path.join(settings.MEDIA_ROOT, path)
    
    # Verificar se o arquivo existe
    if not os.path.exists(file_path) or not os.path.isfile(file_path):
        raise Http404("Arquivo não encontrado")
    
    # Verificar se o caminho está dentro de MEDIA_ROOT (segurança)
    file_path = os.path.abspath(file_path)
    media_root = os.path.abspath(settings.MEDIA_ROOT)
    if not file_path.startswith(media_root):
        raise Http404("Acesso negado")
    
    # Determinar content-type baseado na extensão
    ext = os.path.splitext(file_path)[1].lower()
    content_types = {
        '.pdf': 'application/pdf',
        '.jpg': 'image/jpeg',
        '.jpeg': 'image/jpeg',
        '.png': 'image/png',
        '.gif': 'image/gif',
        '.doc': 'application/msword',
        '.docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
        '.xls': 'application/vnd.ms-excel',
        '.xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        '.zip': 'application/zip',
        '.rar': 'application/x-rar-compressed',
    }
    content_type = content_types.get(ext, 'application/octet-stream')
    
    # Servir o arquivo
    response = FileResponse(open(file_path, 'rb'), content_type=content_type)
    response['Content-Disposition'] = f'inline; filename="{os.path.basename(file_path)}"'
    
    # Cache control (opcional - pode ser ajustado)
    response = cache_control(private=True, max_age=3600)(lambda r: response)(request)
    
    return response


@admin_required
def manage_aprovacao_destinatarios(request):
    """
    E-mails que sempre recebem o PDF/notificação de pedido aprovado (GestControll).
    Substitui a lista fixa no código; apenas administradores.
    """
    if request.method == 'POST':
        if request.POST.get('action') == 'delete':
            pk = request.POST.get('pk')
            try:
                row = AprovacaoEmailDestinatario.objects.get(pk=pk)
                em = row.email
                row.delete()
                messages.success(request, f'Destinatário removido: {em}.')
            except AprovacaoEmailDestinatario.DoesNotExist:
                messages.error(request, 'Registro não encontrado.')
            except Exception as e:
                logger.exception('Erro ao excluir destinatário de aprovação')
                messages.error(request, f'Não foi possível excluir: {e}')
            return redirect('gestao:manage_aprovacao_destinatarios')

        form = AprovacaoEmailDestinatarioForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Destinatário adicionado com sucesso.')
            return redirect('gestao:manage_aprovacao_destinatarios')
        messages.error(request, 'Corrija os erros abaixo.')
    else:
        form = AprovacaoEmailDestinatarioForm()

    destinatarios = AprovacaoEmailDestinatario.objects.order_by('ordem', 'email')
    return render(
        request,
        'obras/aprovacao_destinatarios.html',
        {'form': form, 'destinatarios': destinatarios},
    )


@admin_required
def list_email_logs(request):
    """
    Lista todos os logs de email do sistema.
    Apenas administradores podem acessar.
    """
    from django.core.paginator import Paginator
    
    # Filtros
    status_filter = request.GET.get('status', '')
    tipo_filter = request.GET.get('tipo', '')
    search = request.GET.get('search', '')
    
    # Query base
    logs = EmailLog.objects.select_related('work_order').all()
    
    # Aplicar filtros
    if status_filter:
        logs = logs.filter(status=status_filter)
    
    if tipo_filter:
        logs = logs.filter(tipo_email=tipo_filter)
    
    if search:
        logs = logs.filter(
            Q(assunto__icontains=search) |
            Q(destinatarios__icontains=search) |
            Q(work_order__codigo__icontains=search)
        )
    
    # Ordenar por data mais recente
    logs = logs.order_by('-criado_em')
    
    # Paginação
    paginator = Paginator(logs, 50)  # 50 por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Estatísticas (count() já é otimizado - não carrega objetos, apenas conta)
    total_logs = EmailLog.objects.count()
    enviados = EmailLog.objects.filter(status='enviado').count()
    falhados = EmailLog.objects.filter(status='falhou').count()
    pendentes = EmailLog.objects.filter(status='pendente').count()
    
    # Taxa de sucesso
    taxa_sucesso = round((enviados / total_logs * 100) if total_logs > 0 else 0, 1)
    
    context = {
        'page_obj': page_obj,
        'status_filter': status_filter,
        'tipo_filter': tipo_filter,
        'search': search,
        'total_logs': total_logs,
        'enviados': enviados,
        'falhados': falhados,
        'pendentes': pendentes,
        'taxa_sucesso': taxa_sucesso,
        'email_type_choices': EmailLog.TIPO_EMAIL_CHOICES,
        'reenviar_tipos_suportados': ['novo_pedido', 'aprovacao', 'reprovacao'],
    }
    
    return render(request, 'obras/email_logs.html', context)


@admin_required
def reenviar_email(request, log_id):
    """
    Reenvia um email que falhou.
    Apenas administradores podem reenviar.
    """
    from django.core.mail import EmailMultiAlternatives
    from django.conf import settings
    import os
    
    try:
        email_log = EmailLog.objects.get(pk=log_id)
    except EmailLog.DoesNotExist:
        messages.error(request, 'Log de email não encontrado.')
        return redirect('gestao:list_email_logs')
    
    if email_log.status == 'enviado':
        messages.warning(request, 'Este email já foi enviado com sucesso. Não é necessário reenviar.')
        return redirect('gestao:list_email_logs')
    
    # Verificar se email está configurado
    if not settings.EMAIL_HOST_USER or not settings.EMAIL_HOST_PASSWORD:
        messages.error(request, 'Email não está configurado no sistema.')
        return redirect('gestao:list_email_logs')
    
    try:
        # Recriar o email baseado no tipo
        destinatarios = [email.strip() for email in email_log.destinatarios.split(',')]
        
        if email_log.tipo_email == 'novo_pedido':
            # Reenviar email de novo pedido
            from .email_utils import enviar_email_novo_pedido
            if email_log.work_order:
                sucesso = enviar_email_novo_pedido(email_log.work_order)
            else:
                messages.error(request, 'Pedido não encontrado para este log.')
                return redirect('gestao:list_email_logs')
                
        elif email_log.tipo_email == 'aprovacao':
            # Reenviar email de aprovação
            from .email_utils import enviar_email_aprovacao
            if email_log.work_order:
                # Buscar último aprovador
                approval = Approval.objects.filter(work_order=email_log.work_order).order_by('-created_at').first()
                if approval and approval.aprovado_por:
                    sucesso = enviar_email_aprovacao(email_log.work_order, approval.aprovado_por, None)
                else:
                    messages.error(request, 'Não foi possível identificar o aprovador.')
                    return redirect('gestao:list_email_logs')
            else:
                messages.error(request, 'Pedido não encontrado para este log.')
                return redirect('gestao:list_email_logs')
                
        elif email_log.tipo_email == 'reprovacao':
            # Reenviar email de reprovação
            from .email_utils import enviar_email_reprovacao
            if email_log.work_order:
                approval = Approval.objects.filter(
                    work_order=email_log.work_order,
                    decisao='reprovado',
                ).order_by('-created_at').first()
                if approval and approval.aprovado_por:
                    comentario = approval.comentario or 'Sem comentário'
                    sucesso = enviar_email_reprovacao(email_log.work_order, approval.aprovado_por, comentario)
                else:
                    messages.error(request, 'Não foi possível identificar o aprovador ou comentário.')
                    return redirect('gestao:list_email_logs')
            else:
                messages.error(request, 'Pedido não encontrado para este log.')
                return redirect('gestao:list_email_logs')
        else:
            messages.error(request, 'Tipo de email não suportado para reenvio.')
            return redirect('gestao:list_email_logs')
        
        if sucesso:
            messages.success(request, f'Email reenviado com sucesso!')
        else:
            messages.warning(request, 'Tentativa de reenvio realizada, mas pode ter falhado. Verifique os logs.')
        
    except Exception as e:
        logger.error(f"Erro ao reenviar email log {log_id}: {e}", exc_info=True)
        messages.error(request, f'Erro ao reenviar email: {str(e)}')
    
    return redirect('gestao:list_email_logs')


@login_required
def marcar_pedido_analisado(request, pk):
    """
    View para marcar/desmarcar pedido como analisado via AJAX.
    Apenas e-mails em usuario_pode_marcar_pedido_analisado (utils) ou superuser.
    """
    if not usuario_pode_marcar_pedido_analisado(request.user):
        return JsonResponse({
            'success': False, 
            'error': 'Você não tem permissão para usar esta funcionalidade.'
        }, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método não permitido.'}, status=405)
    
    try:
        pedido = get_object_or_404(WorkOrder, pk=pk)
        
        # Verificar se o pedido está aprovado
        if pedido.status != 'aprovado':
            return JsonResponse({
                'success': False, 
                'error': 'Apenas pedidos com status "Aprovado" podem ser marcados como analisados.'
            }, status=400)
        
        # Obter valor do checkbox
        marcado = request.POST.get('marcado', 'false').lower() == 'true'
        
        # Atualizar pedido (usando o mesmo campo marcado_para_deletar, mas com significado diferente)
        pedido.marcado_para_deletar = marcado
        if marcado:
            pedido.marcado_para_deletar_por = request.user
            pedido.marcado_para_deletar_em = timezone.now()
        else:
            pedido.marcado_para_deletar_por = None
            pedido.marcado_para_deletar_em = None
        pedido.save()
        
        return JsonResponse({
            'success': True,
            'marcado': marcado,
            'message': 'Pedido marcado como analisado.' if marcado else 'Pedido desmarcado.'
        })
        
    except Exception as e:
        logger.error(f"Erro ao marcar pedido {pk} como analisado: {e}", exc_info=True)
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

