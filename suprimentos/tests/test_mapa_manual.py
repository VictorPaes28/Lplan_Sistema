"""Mapa de suprimentos em modo manual (sem Sienge/SC)."""
import json
from datetime import date
from decimal import Decimal

from django.contrib.auth.models import Group, User
from django.test import TestCase, override_settings
from django.urls import reverse

from accounts.groups import GRUPOS
from core.models import Project
from mapa_obras.models import LocalObra, Obra
from suprimentos.models import AlocacaoRecebimento, Insumo, ItemMapa


@override_settings(MAPA_SUPRIMENTOS_MANUAL=True)
class TestMapaManual(TestCase):
    def setUp(self):
        Project.objects.create(
            name='Obra Manual',
            code='OBR-MAN',
            is_active=True,
            start_date=date(2025, 1, 1),
            end_date=date(2026, 12, 31),
        )
        self.obra = Obra.objects.create(codigo_sienge='OBR-MAN', nome='Obra Manual', ativa=True)
        self.local = LocalObra.objects.create(obra=self.obra, nome='Bloco A', tipo='BLOCO')
        self.insumo = Insumo.objects.create(codigo_sienge='2001', descricao='Cimento', unidade='KG')
        self.user = User.objects.create_superuser('eng', 'eng@test', 'x')
        Group.objects.get_or_create(name=GRUPOS.ENGENHARIA)
        self.user.groups.add(Group.objects.get(name=GRUPOS.ENGENHARIA))
        self.client.force_login(self.user)
        session = self.client.session
        session['obra_id'] = self.obra.id
        session.save()
        self.item = ItemMapa.objects.create(
            obra=self.obra,
            insumo=self.insumo,
            local_aplicacao=self.local,
            quantidade_planejada=Decimal('10'),
        )

    def test_status_etapa_manual(self):
        self.assertEqual(self.item.status_etapa, 'LEVANTAMENTO')
        AlocacaoRecebimento.objects.create(
            obra=self.obra,
            insumo=self.insumo,
            local_aplicacao=self.local,
            item_mapa=self.item,
            quantidade_alocada=Decimal('4'),
            criado_por=self.user,
        )
        item = ItemMapa.objects.get(pk=self.item.pk)
        self.assertEqual(item.status_etapa, 'PARCIAL')
        item.quantidade_planejada = Decimal('4')
        item.save()
        item = ItemMapa.objects.get(pk=self.item.pk)
        self.assertEqual(item.status_etapa, 'ENTREGUE')

    def test_import_sienge_redirect(self):
        r = self.client.get(reverse('engenharia:importar_sienge'))
        self.assertEqual(r.status_code, 302)
        self.assertIn('/engenharia/mapa/', r.url)

    def test_numero_sc_bloqueado_na_api(self):
        url = reverse('suprimentos:item_atualizar_campo')
        r = self.client.post(
            url,
            data=json.dumps({'item_id': self.item.id, 'field': 'numero_sc', 'value': '99'}),
            content_type='application/json',
        )
        self.assertEqual(r.status_code, 403)

    def test_row_patch_ao_alterar_planejado(self):
        AlocacaoRecebimento.objects.create(
            obra=self.obra,
            insumo=self.insumo,
            local_aplicacao=self.local,
            item_mapa=self.item,
            quantidade_alocada=Decimal('4'),
            criado_por=self.user,
        )
        url = reverse('suprimentos:item_atualizar_campo')
        r = self.client.post(
            url,
            data=json.dumps({
                'item_id': self.item.id,
                'field': 'quantidade_planejada',
                'value': '20',
            }),
            content_type='application/json',
        )
        self.assertEqual(r.status_code, 200)
        data = r.json()
        self.assertTrue(data.get('success'))
        patch = data.get('row_patch') or {}
        self.assertEqual(patch.get('quantidade_planejada'), '20,00')
        self.assertEqual(patch.get('quantidade_alocada'), '4,00')
        self.assertIn('PARCIAL', patch.get('status_etapa', ''))
        self.assertEqual(patch.get('percentual_pct'), 20)

    def test_alocacao_respeita_planejado(self):
        url = reverse('suprimentos:item_alocar', kwargs={'item_id': self.item.id})
        r = self.client.post(
            url,
            data=json.dumps({'quantidade_alocada': '15'}),
            content_type='application/json',
        )
        self.assertEqual(r.status_code, 400)
        self.assertFalse(r.json().get('success'))

    def test_planejado_zero_com_alocacao_bloqueado(self):
        AlocacaoRecebimento.objects.create(
            obra=self.obra,
            insumo=self.insumo,
            local_aplicacao=self.local,
            item_mapa=self.item,
            quantidade_alocada=Decimal('2'),
            criado_por=self.user,
        )
        url = reverse('suprimentos:item_atualizar_campo')
        r = self.client.post(
            url,
            data=json.dumps({
                'item_id': self.item.id,
                'field': 'quantidade_planejada',
                'value': '0',
            }),
            content_type='application/json',
        )
        self.assertEqual(r.status_code, 400)
        self.assertFalse(r.json().get('success'))

    def test_planejado_menor_que_alocado_bloqueado(self):
        AlocacaoRecebimento.objects.create(
            obra=self.obra,
            insumo=self.insumo,
            local_aplicacao=self.local,
            item_mapa=self.item,
            quantidade_alocada=Decimal('6'),
            criado_por=self.user,
        )
        url = reverse('suprimentos:item_atualizar_campo')
        r = self.client.post(
            url,
            data=json.dumps({
                'item_id': self.item.id,
                'field': 'quantidade_planejada',
                'value': '4',
            }),
            content_type='application/json',
        )
        self.assertEqual(r.status_code, 400)
        self.assertFalse(r.json().get('success'))

    def test_saldo_negativo_sem_planejado(self):
        AlocacaoRecebimento.objects.create(
            obra=self.obra,
            insumo=self.insumo,
            local_aplicacao=self.local,
            item_mapa=self.item,
            quantidade_alocada=Decimal('3'),
            criado_por=self.user,
        )
        self.item.quantidade_planejada = Decimal('0')
        self.item.save()
        item = ItemMapa.objects.get(pk=self.item.pk)
        self.assertTrue(item.saldo_negativo)

    def test_filtro_status_levantamento(self):
        url = reverse('engenharia:mapa') + f'?obra={self.obra.id}&status=LEVANTAMENTO'
        r = self.client.get(url)
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, 'Cimento')

    def test_duplicar_item(self):
        url = reverse('suprimentos:item_duplicar', kwargs={'item_id': self.item.id})
        antes = ItemMapa.objects.filter(obra=self.obra).count()
        r = self.client.post(url, data='{}', content_type='application/json')
        self.assertEqual(r.status_code, 200)
        self.assertTrue(r.json().get('success'))
        self.assertEqual(ItemMapa.objects.filter(obra=self.obra).count(), antes + 1)
        copia = ItemMapa.objects.exclude(pk=self.item.pk).get(obra=self.obra)
        self.assertEqual(copia.quantidade_planejada, self.item.quantidade_planejada)
        self.assertEqual(copia.quantidade_alocada_local, Decimal('0'))

    def test_paginacao_mapa(self):
        for i in range(85):
            ins = Insumo.objects.create(
                codigo_sienge=f'SM-T-{i}',
                descricao=f'Item teste {i}',
                unidade='UND',
            )
            ItemMapa.objects.create(obra=self.obra, insumo=ins, quantidade_planejada=Decimal('1'))
        r = self.client.get(reverse('engenharia:mapa') + f'?obra={self.obra.id}')
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, 'Página 1 de')

    def test_filtro_pendencia_sem_local(self):
        ItemMapa.objects.create(
            obra=self.obra,
            insumo=Insumo.objects.create(codigo_sienge='SM-LEV-X', descricao='Sem local', unidade='UND'),
            quantidade_planejada=Decimal('5'),
            local_aplicacao=None,
        )
        url = reverse('engenharia:mapa') + f'?obra={self.obra.id}&pendencia=SEM_LOCAL'
        r = self.client.get(url)
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, 'Sem local')
        self.assertContains(r, 'data-remove-param="pendencia"')

    def test_filtro_pendencia_incompleto(self):
        url = reverse('engenharia:mapa') + f'?obra={self.obra.id}&pendencia=INCOMPLETO'
        r = self.client.get(url)
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, 'mapa-chip-filtro')

    def test_ordenar_por_prazo(self):
        item2 = ItemMapa.objects.create(
            obra=self.obra,
            insumo=Insumo.objects.create(codigo_sienge='SM-T-PRZ', descricao='Com prazo', unidade='UND'),
            quantidade_planejada=Decimal('1'),
            prazo_necessidade=date(2026, 1, 15),
        )
        self.item.prazo_necessidade = date(2026, 6, 1)
        self.item.save()
        url = reverse('engenharia:mapa') + f'?obra={self.obra.id}&ordenar=prazo'
        r = self.client.get(url)
        self.assertEqual(r.status_code, 200)
        content = r.content.decode('utf-8')
        start = content.find('<table class="tabela-mapa">')
        self.assertGreater(start, -1)
        end = content.find('</table>', start)
        table_html = content[start:end]
        pos_item2 = table_html.find(f'data-item-id="{item2.id}"')
        pos_item1 = table_html.find(f'data-item-id="{self.item.id}"')
        self.assertGreater(pos_item2, -1)
        self.assertGreater(pos_item1, -1)
        self.assertLess(pos_item2, pos_item1)

    def test_novo_levantamento_ajax(self):
        url = reverse('engenharia:novo_levantamento')
        antes = ItemMapa.objects.filter(obra=self.obra).count()
        r = self.client.post(
            url,
            data={
                'obra': self.obra.id,
                'descricao_insumo': 'Tijolo cerâmico',
                'quantidade_planejada': '100',
                'unidade': 'UND',
                'local_aplicacao': self.local.id,
            },
            HTTP_X_REQUESTED_WITH='XMLHttpRequest',
        )
        self.assertEqual(r.status_code, 200)
        data = r.json()
        self.assertTrue(data.get('success'))
        self.assertIn('scroll_item', data.get('redirect_url', ''))
        self.assertIn('row_html', data)
        self.assertIn('mobile_card_html', data)
        self.assertEqual(ItemMapa.objects.filter(obra=self.obra).count(), antes + 1)

    def test_item_restaurar_after_delete(self):
        self.item.responsavel = 'eng'
        self.item.save()
        AlocacaoRecebimento.objects.create(
            obra=self.obra,
            insumo=self.insumo,
            local_aplicacao=self.local,
            item_mapa=self.item,
            quantidade_alocada=Decimal('2'),
            criado_por=self.user,
        )
        url_del = reverse('suprimentos:item_excluir', kwargs={'item_id': self.item.id})
        r = self.client.post(url_del)
        self.assertEqual(r.status_code, 200)
        snapshot = r.json().get('undo_snapshot')
        self.assertIsNotNone(snapshot)
        self.assertFalse(ItemMapa.objects.filter(pk=self.item.pk).exists())

        url_restore = reverse('suprimentos:item_restaurar')
        r2 = self.client.post(
            url_restore,
            data=json.dumps({'undo_snapshot': snapshot}),
            content_type='application/json',
        )
        self.assertEqual(r2.status_code, 200)
        self.assertTrue(r2.json().get('success'))
        restaurado = ItemMapa.objects.filter(obra=self.obra, responsavel='eng').first()
        self.assertIsNotNone(restaurado)
        self.assertEqual(restaurado.alocacoes.count(), 1)

    def test_quick_filter_meus_itens(self):
        self.user.first_name = 'João'
        self.user.last_name = 'Silva'
        self.user.save()
        self.item.responsavel = 'João Silva'
        self.item.save()
        outro = ItemMapa.objects.create(
            obra=self.obra,
            insumo=Insumo.objects.create(codigo_sienge='SM-OUT', descricao='Outro item', unidade='UND'),
            quantidade_planejada=Decimal('1'),
            responsavel='Outra pessoa',
        )
        url = reverse('engenharia:mapa') + f'?obra={self.obra.id}&quick=MEUS_ITENS'
        r = self.client.get(url)
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, 'Cimento')
        self.assertNotContains(r, outro.insumo.descricao)

    def test_mapa_engenharia_fragment_returns_html(self):
        url = reverse('suprimentos:mapa_engenharia_fragment') + f'?obra={self.obra.id}'
        r = self.client.get(url)
        self.assertEqual(r.status_code, 200)
        data = r.json()
        self.assertTrue(data.get('success'))
        self.assertIn(f'data-item-id="{self.item.id}"', data.get('tbody_html', ''))
        self.assertIn('supply-card', data.get('mobile_cards_html', ''))
        self.assertIn('total', data.get('kpis', {}))
        self.assertIn('page', data.get('pagination', {}))

    def test_export_respects_hidden_cols(self):
        from io import BytesIO

        from openpyxl import load_workbook

        url = (
            reverse('engenharia:exportar_excel')
            + f'?obra={self.obra.id}&hidden_cols=col-resp,col-prazo'
        )
        r = self.client.get(url)
        self.assertEqual(r.status_code, 200)
        wb = load_workbook(BytesIO(r.content))
        ws = wb.active
        headers = [cell.value for cell in ws[3] if cell.value]
        self.assertNotIn('5. RESPONSÁVEL', headers)
        self.assertNotIn('6. PRAZO', headers)
        self.assertIn('3. DESCRIÇÃO DO ITEM', headers)

    def test_item_detalhe_inclui_historico(self):
        from suprimentos.models import HistoricoAlteracao

        HistoricoAlteracao.registrar(
            obra=self.obra,
            usuario=self.user,
            tipo='EDICAO',
            descricao='Prazo alterado para 2026-01-15',
            item_mapa=self.item,
            campo_alterado='prazo_necessidade',
        )
        url = reverse('suprimentos:item_detalhe', kwargs={'item_id': self.item.id})
        r = self.client.get(url)
        self.assertEqual(r.status_code, 200)
        html = r.json().get('html', '')
        self.assertIn('Histórico Recente', html)
        self.assertIn('Prazo alterado', html)
